import os
import sys
from datetime import datetime
from pathlib import Path

import django
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'cofluhab.settings')
django.setup()

from principal.models import Contrato  # noqa: E402
from principal.ficha_generators import gerar_lote_fh1_separado  # noqa: E402


LAYOUT_FH1 = [
    (1, 'UFS', 'NUM', 1, 2, '2 NUM', ''),
    (2, 'MAT. AG. FINANC. /DV', 'NUM', 3, 8, '6 NUM', ''),
    (3, 'N.º CONTRATO DO MUT. NO AGENTE', 'ALFA', 9, 21, '13 ALFA', ''),
    (4, 'HIPOTECA', 'NUM', 22, 22, '1 NUM', ''),
    (5, 'TIPO DE REGISTRO', 'NUM', 23, 23, '1 NUM', '1'),
    (6, 'SEQUENCIAL', 'NUM', 24, 25, '2 NUM', '0'),
    (7, 'CONSTANTE', 'NUM', 26, 26, '1 NUM', '0'),
    (8, 'NOME DO MUT. PRINCIPAL', 'ALFA', 27, 66, '40 ALFA', ''),
    (9, 'TIPO', 'NUM', 67, 67, '1 NUM', ''),
    (10, 'CPF/CI', 'ALFA', 68, 84, '17 ALFA', ''),
    (11, 'DATA DE NASCIMENTO', 'NUM', 85, 90, '6 NUM', '(DDMMAA)'),
    (12, 'CODIGO DO MUNICÍPIO', 'NUM', 91, 95, '5 NUM', ''),
    (13, 'UF', 'ALFA', 96, 97, '2 ALFA', ''),
    (14, 'ENDEREÇO DO IMÓVEL', 'ALFA', 98, 135, '38 ALFA', ''),
    (15, 'DATA DO CONTRATO', 'NUM', 136, 141, '6 NUM', '(DDMMAA)'),
    (16, 'VALOR DA GARANTIA', 'NUM', 142, 153, '12 NUM', '10 INT. e 2 DEC.'),
    (17, 'IM', 'NUM', 154, 155, '2 NUM', ''),
    (18, 'DATA DA LEGISLAÇÃO', 'NUM', 156, 161, '6 NUM', '(DDMMAA)'),
    (19, 'VALOR FINANCIAMENTO CONTRATADO', 'NUM', 162, 173, '12 NUM', '10 INT. e 2 DEC.'),
    (20, 'VALOR FINANC. PADRÃO FCVS', 'NUM', 174, 185, '12 NUM', '10 INT. e 2 DEC.'),
    (21, 'CÓDIGO DA CATEG. PROFISSIONAL', 'ALFA', 186, 190, '5 ALFA', ''),
    (22, 'SEGURO DE CRÉDITO', 'ALFA', 191, 191, '1 ALFA', ''),
    (23, 'CARÊNCIA NO 1O VENCIMENTO', 'ALFA', 192, 192, '1 ALFA', ''),
    (24, 'SEGURO DFI POR LOTES URBANIZADOS', 'ALFA', 193, 193, '1 ALFA', ''),
    (25, 'CRÉDITOS ADQUIRIDOS PELA CAIXA COM RECURSOS DO PROER', 'ALFA', 194, 194, '1 ALFA', ''),
    (26, 'VAGO', 'ALFA', 195, 195, '1 ALFA', 'Não preencher'),
    (27, 'PRAZO CONTRATADO', 'NUM', 196, 198, '3 NUM', ''),
    (28, 'TAXA JUROS CONTRATADO', 'NUM', 199, 204, '6 NUM', '2 INT. e 4 DEC.'),
    (29, 'CES CONTRATUAL', 'NUM', 205, 208, '4 NUM', 'Manual cita 5 NUM, mas o intervalo tem 4 posições'),
    (30, 'PLANO', 'ALFA', 209, 211, '3 ALFA', ''),
    (31, 'ST', 'NUM', 212, 212, '1 NUM', ''),
    (32, 'RJ', 'ALFA', 213, 213, '1 ALFA', ''),
    (33, 'RR', 'NUM', 214, 215, '2 NUM', ''),
    (34, 'INDEX', 'ALFA', 216, 218, '3 ALFA', ''),
    (35, 'PRAZO FCVS', 'NUM', 219, 221, '3 NUM', ''),
    (36, 'TAXA JUROS PARA FCVS', 'NUM', 222, 227, '6 NUM', '2 INT. e 4 DEC.'),
    (37, 'CES PARA FCVS', 'NUM', 228, 231, '4 NUM', '1 INT. e 3 DEC.'),
    (38, 'PLANO', 'ALFA', 232, 234, '3 ALFA', ''),
    (39, 'ST', 'NUM', 235, 235, '1 NUM', ''),
    (40, 'RJ', 'ALFA', 236, 236, '1 ALFA', ''),
    (41, 'RR', 'NUM', 237, 238, '2 NUM', ''),
    (42, 'INDEX', 'ALFA', 239, 241, '3 ALFA', ''),
    (43, 'DATA SALDO CONSTRUÇÃO', 'NUM', 242, 247, '6 NUM', '(DDMMAA)'),
    (44, 'SALDO DEVEDOR', 'NUM', 248, 259, '12 NUM', '10 INT. e 2 DEC.'),
    (45, '1o VENCIMENTO', 'NUM', 260, 265, '6 NUM', '(DDMMAA)'),
    (46, 'SEGURO CREDITO / MIP / DFI', 'NUM', 266, 273, '8 NUM', '6 INT. e 2 DEC.'),
    (47, 'VALOR DA PRESTAÇÃO', 'NUM', 274, 283, '10 NUM', '8 INT. e 2 DEC.'),
    (49, 'TCA/TAC', 'NUM', 284, 291, '8 NUM', '6 INT. e 2 DEC.'),
    (50, 'FCVS MENSAL', 'NUM', 292, 299, '8 NUM', '6 INT. e 2 DEC.'),
    (51, 'RAZÃO ACRES/ DECRES.', 'NUM', 300, 307, '8 NUM', '6 INT. e 2 DEC.'),
    (52, 'TIPO EVENTO', 'ALFA', 308, 310, '3 ALFA', ''),
    (53, 'DATA DO EVENTO', 'NUM', 311, 316, '6 NUM', '(DDMMAA)'),
    (54, 'OR/CO', 'NUM', 317, 318, '2 NUM', ''),
    (55, '% CAIXA', 'NUM', 319, 322, '4 NUM', '100'),
    (56, 'N.º CONTR. EMPR. CAIXA', 'NUM', 323, 340, '18 NUM', ''),
    (57, 'TAXA JUROS EVENTO', 'NUM', 341, 346, '6 NUM', '2 INT. e 4 DEC.'),
    (58, 'VAF1 – VALOR BÁSICO', 'NUM', 347, 360, '14 NUM', '12 INT. e 2 DEC.'),
    (59, 'VAF2 – VALOR COMPLEMENTAR', 'NUM', 361, 374, '14 NUM', '12 INT. e 2 DEC.'),
    (60, 'VAF3 – VALOR RESIDUAL', 'NUM', 375, 388, '14 NUM', '12 INT. e 2 DEC.'),
    (61, 'JUROS CALCULADOS PELO AGENTE FINANCEIRO', 'NUM', 389, 402, '14 NUM', ''),
    (62, 'DEBITO/CRÉDITO', 'ALFA', 403, 403, '1 ALFA', 'D ou C'),
    (63, 'QUANTIDADE DE ALTERAÇÕES', 'NUM', 404, 405, '2 NUM', ''),
    (64, 'UFS', 'NUM', 406, 407, '2 NUM', 'Código da UFS'),
    (65, 'MAT. AG. FINANC.', 'NUM', 408, 413, '6 NUM', 'Matrícula do Agente Financeiro'),
    (66, 'DATA GERAÇÃO', 'NUM', 414, 419, '6 NUM', 'Data geração do lote (DDMMAA)'),
    (67, 'NÚMERO', 'NUM', 420, 422, '3 NUM', 'Número do lote'),
    (68, 'FORMA DE ENVIO', 'ALFA', 423, 423, '1 ALFA', '= S (FCVS 2000)'),
    (69, 'TIPO MOVIMENTO', 'ALFA', 424, 424, '1 ALFA', 'I'),
    (70, 'FILLER', '-', 425, 430, '6', 'BRANCOS'),
]


def interpretar_valor(tipo, formato, valor_bruto):
    valor_limpo = valor_bruto.rstrip()
    if not valor_limpo:
        return ''

    if tipo != 'NUM':
        return valor_limpo

    if 'DDMMAA' in formato and len(valor_limpo) == 6 and valor_limpo.isdigit():
        return f'{valor_limpo[0:2]}/{valor_limpo[2:4]}/{valor_limpo[4:6]}'

    if 'DEC.' in formato and valor_limpo.isdigit():
        casas = 2
        inteiro = valor_limpo[:-casas] or '0'
        decimal = valor_limpo[-casas:]
        return f'{int(inteiro):,}.{decimal}'.replace(',', 'X').replace('.', ',').replace('X', '.')

    return valor_limpo


def gerar_planilha(codigo_contrato='1075', matricula='00044', numero_lote='001'):
    contrato = Contrato.objects.filter(codigo=str(codigo_contrato)).first()
    if not contrato:
        raise SystemExit(f'Contrato {codigo_contrato} não encontrado.')

    resultado = gerar_lote_fh1_separado([contrato], matricula=matricula, numero_lote=numero_lote)
    linhas = [ln for ln in (resultado.get('dados_conteudo') or '').splitlines() if ln]
    if not linhas:
        raise SystemExit(f'Nenhuma linha FH1 gerada para o contrato {codigo_contrato}.')

    linha = linhas[0]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Decomposicao FH1'
    horizontal = wb.create_sheet('Horizontal CEF')
    raw = wb.create_sheet('Linha Bruta')
    meta = wb.create_sheet('Metadados')

    headers = ['SEQ', 'CAMPO', 'TIPO', 'INICIO', 'FIM', 'TAMANHO', 'FORMATO', 'OBSERVACOES', 'VALOR BRUTO', 'VALOR INTERPRETADO']
    ws.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    warn_fill = PatternFill(fill_type='solid', fgColor='FFF2CC')

    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    horizontal['A1'] = 'CAMPO'
    horizontal['A2'] = 'VALOR ENCONTRADO'
    horizontal['A3'] = 'POSIÇÃO'
    horizontal['A4'] = 'FORMATO'
    horizontal['A5'] = 'VALOR BRUTO'
    for cell_name in ('A1', 'A2', 'A3', 'A4', 'A5'):
        horizontal[cell_name].fill = header_fill
        horizontal[cell_name].font = header_font
        horizontal[cell_name].alignment = Alignment(horizontal='center', vertical='center')

    for seq, nome, tipo, inicio, fim, formato, obs in LAYOUT_FH1:
        valor_bruto = linha[inicio - 1:fim]
        valor_interpretado = interpretar_valor(tipo, formato, valor_bruto)
        row = [seq, nome, tipo, inicio, fim, fim - inicio + 1, formato, obs, valor_bruto, valor_interpretado]
        ws.append(row)
        if obs:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = warn_fill

        col_idx = seq + 1
        horizontal.cell(row=1, column=col_idx, value=nome)
        horizontal.cell(row=2, column=col_idx, value=valor_interpretado)
        horizontal.cell(row=3, column=col_idx, value=f'{inicio}-{fim}')
        horizontal.cell(row=4, column=col_idx, value=formato)
        horizontal.cell(row=5, column=col_idx, value=valor_bruto)
        for row_idx in range(1, 6):
            horizontal.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if obs:
            horizontal.cell(row=1, column=col_idx).fill = warn_fill
            horizontal.cell(row=2, column=col_idx).fill = warn_fill
            horizontal.cell(row=3, column=col_idx).fill = warn_fill
            horizontal.cell(row=4, column=col_idx).fill = warn_fill
            horizontal.cell(row=5, column=col_idx).fill = warn_fill

    widths = {
        'A': 8, 'B': 42, 'C': 10, 'D': 8, 'E': 8, 'F': 10, 'G': 18, 'H': 40, 'I': 22, 'J': 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'

    horizontal.freeze_panes = 'B1'
    horizontal.row_dimensions[1].height = 54
    for row_idx in range(2, 6):
        horizontal.row_dimensions[row_idx].height = 28
    horizontal.column_dimensions['A'].width = 18
    for seq, _, _, _, _, _, _ in LAYOUT_FH1:
        horizontal.column_dimensions[get_column_letter(seq + 1)].width = 20

    raw['A1'] = 'Linha FH1 bruta'
    raw['A1'].font = Font(bold=True)
    raw['A2'] = linha
    raw.column_dimensions['A'].width = 140
    raw['A2'].alignment = Alignment(wrap_text=True)

    meta_rows = [
        ('Contrato', contrato.codigo),
        ('Contrato ID', contrato.id),
        ('Matricula', matricula),
        ('Numero lote', numero_lote),
        ('Tamanho linha', len(linha)),
        ('Total fichas sucesso', resultado.get('total_fichas_sucesso', 0)),
        ('Total fichas erro', resultado.get('total_fichas_erro', 0)),
        ('Gerado em', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
    ]
    for item in resultado.get('detalhes', []):
        if str(item.get('contrato')) == str(contrato.codigo):
            avisos = '; '.join(av.get('mensagem', '') for av in item.get('avisos', []))
            if avisos:
                meta_rows.append(('Avisos', avisos))

    for idx, (chave, valor) in enumerate(meta_rows, start=1):
        meta.cell(row=idx, column=1, value=chave).font = Font(bold=True)
        meta.cell(row=idx, column=2, value=valor)
    meta.column_dimensions['A'].width = 22
    meta.column_dimensions['B'].width = 120

    out_dir = PROJECT_ROOT.parent / 'exports'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f'decomposicao_fh1_horizontal_{contrato.codigo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(out_path)
    print(out_path)


if __name__ == '__main__':
    codigo = sys.argv[1] if len(sys.argv) > 1 else '1075'
    matricula = sys.argv[2] if len(sys.argv) > 2 else '00044'
    numero_lote = sys.argv[3] if len(sys.argv) > 3 else '001'
    gerar_planilha(codigo, matricula, numero_lote)