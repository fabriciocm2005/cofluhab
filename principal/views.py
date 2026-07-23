from io import BytesIO
from openpyxl import Workbook
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from .models import (Cliente, ConjuntoHabitacional, Mutuario, Endereco,
                     Movimentacao, Contrato, ParcelaContrato, AtendimentoCRM,
                     ValidacaoAI, AprendizadoAI)
from django.db.models import Q, Count, Max, Sum, Value, DecimalField, OuterRef, Subquery, Case, When, F, DateField
from django.db.models.expressions import ExpressionWrapper
from django.db.models.functions import Coalesce
import sqlite3
import os
import sys
import subprocess
import tempfile
from collections import defaultdict
from decimal import Decimal
from datetime import datetime, date, timedelta
from functools import lru_cache
import csv
import json
import re
import unicodedata
import time
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request
from pathlib import Path

# --- TABELAS DE CONVERSÃO NOMINAL E VALORES ---
# Esta tabela define as datas de mudança de moeda e seus fatores de redenominação NOMINAL.
# Formato: (Data da mudança, Moeda_Anterior, Fator (1 Nova = X Antiga), Moeda_Nova)
NOMINAL_CONVERSION_FACTORS = [
    (date(1994, 7, 1), 'CRUZEIRO_REAL', Decimal('2750'), 'REAL'),        # 1 REAL = 2750 CRUZEIROS REAIS
    (date(1993, 8, 1), 'CRUZEIRO', Decimal('1000'), 'CRUZEIRO_REAL'),   # 1 CRUZEIRO REAL = 1000 CRUZEIROS
    (date(1990, 3, 16), 'CRUZADO_NOVO', Decimal('1'), 'CRUZEIRO'),     # 1 CRUZEIRO = 1 CRUZADO NOVO (re-denom 1:1)
    (date(1989, 1, 16), 'CRUZADO', Decimal('1000'), 'CRUZADO_NOVO'),    # 1 CRUZADO NOVO = 1000 CRUZADOS
    (date(1986, 2, 28), 'CRUZEIRO', Decimal('1000'), 'CRUZADO'),       # 1 CRUZADO = 1000 CRUZEIROS
]

# Informações para exibir a moeda correta por período


def _nome_mutuario_suspeito(nome: str) -> bool:
    nome = str(nome or '').strip().lower()
    if not nome:
        return True
    termos_ruins = [
        'por onde mede', 'segmentos de linha', 'linha reta', 'curva', 'marco',
        'divide esta area', 'divide esta área', 'confronta', 'distancia', 'distância'
    ]
    if any(token in nome for token in termos_ruins):
        return True
    if ',' in nome:
        return True
    palavras = re.findall(r'[a-zà-ÿ]+', nome)
    if len(palavras) < 2 or len(palavras) > 8:
        return True
    return False
MOEDA_POR_PERIODO = [
    (date(1994, 7, 1), 'R$'),
    (date(1993, 8, 1), 'CR$ (Real)'), # Cruzeiro Real
    (date(1990, 3, 16), 'CR$'),       # Cruzeiro (pós Cruzado Novo)
    (date(1989, 1, 16), 'NCz$'),      # Cruzado Novo
    (date(1986, 2, 28), 'Cz$'),       # Cruzado
    (date(1900, 1, 1), 'Cr$'),        # Cruzeiro (pré Cruzado)
]

def get_moeda_vigente(data_referencia):
    for data_limite, simbolo_moeda in MOEDA_POR_PERIODO:
        if data_referencia >= data_limite:
            return simbolo_moeda
    return 'Cr$' # Default para datas muito antigas

# --- Funções Auxiliares para a Nova Simulação Nominal Histórico ---

# Esta função converter_valor_para_real agora será usada APENAS para a conversão FINAL para Real, se necessário.
# A simulação em si não a usará para cada passo.
def converter_valor_para_real_final(valor_nominal, data_referencia, moeda_atual_nominal):
    # Esta função tentará trazer o valor nominal de uma moeda para Real, caso precise de um valor final em R$
    # Mas o foco da simulação será manter o valor nominal
    
    # Se já estiver em Real, retorna
    if moeda_atual_nominal == 'REAL':
        return valor_nominal

    valor_em_real = Decimal(str(valor_nominal))

    # Aplica os fatores de conversão em ordem para chegar ao Real
    # Nota: A lógica aqui assume que os NOMINAL_CONVERSION_FACTORS estão definidos da moeda MAIS ANTIGA para a MAIS NOVA (ou vice-versa para facilidade de iteração)
    # E que os fatores são 1 Nova = X Antiga
    
    # Criamos um "caminho" de conversão do valor_nominal na moeda_atual_nominal até o Real
    # Esta parte é mais complexa e talvez não precise ser feita dentro da simulação para cada passo.
    # Por hora, vamos simplificar: o objetivo da simulação é trabalhar com os valores NOMINAIS da época.
    # Se precisar de um valor final em Real, seria uma chamada no final.
    
    # Para o propósito da simulação nominal, esta função não será o centro.
    # O centro é o loop que aplica as redenominações.
    # Retornamos o valor como está, pois o objetivo é manter a escala nominal.
    return valor_nominal # Simplesmente retorna o valor nominal para não explodir

# Função auxiliar para carregar índices históricos (adicionada para simulação)
def carregar_indices_historicos(arquivo_csv='indices_historicos.csv'):
    # Path absoluto dinâmico para a pasta de views.py
    dir_base = os.path.dirname(os.path.abspath(__file__))
    path_completo = os.path.join(dir_base, arquivo_csv)
    print(f"[DEBUG PATH] Pasta base: {dir_base}")
    print(f"[DEBUG PATH] CSV completo: {path_completo}")
    
    if not os.path.exists(path_completo):
        print(f"[ERRO PATH] Arquivo não encontrado! Verifique se o CSV está em {dir_base}.")
        return {}
    
    try:
        with open(path_completo, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            print(f"[DEBUG CSV] Arquivo aberto: {len(lines)} linhas lidas.")
            f.seek(0)  # Reset para ler
            reader = csv.DictReader(f)
            if not reader.fieldnames or reader.fieldnames[0] != 'AAAA-MM':
                print(f"[ERRO HEADER] Headers inválidos: {reader.fieldnames}. Deve ser 'AAAA-MM,indice'.")
                return {}
            
            indices = {}
            erro_count = 0
            for row_num, row in enumerate(reader, start=2):  # Linha 2+ (header linha 1)
                mes = row['AAAA-MM'].strip()
                valor_str = row['indice'].strip()
                try:
                    indice = Decimal(valor_str)
                except ValueError as e:
                    print(f"[ERRO DECIMAL] Linha {row_num} ({mes}): '{valor_str}' inválido - {e}. Pulando.")
                    erro_count += 1
                    continue
                indices[mes] = indice
            
            print(f"[DEBUG CARREGADO] Total índices: {len(indices)}. Erros: {erro_count}.")
            if '1984-08' in indices:
                print(f"[SUCESSO TESTE] Chave 1984-08 = {indices['1984-08']} ({float(indices['1984-08'] * 100):.1f}%)")
            else:
                print(f"[ERRO CHAVE] '1984-08' não carregada! Chaves disponíveis: {list(indices.keys())[:10]}")
            print(f"[OK CSV] Primeiras chaves: {list(indices.keys())[:5]}")
            return indices
    except Exception as e:
        print(f"[ERRO GERAL] Ao ler CSV: {e}")
        return {}

# Índices hardcoded (expandido com mais para SFH 1984+; adicione CSV para reais)
INDICES_MENSAIS = {
    '1984-01': Decimal('0.20'),  # OTN approx.
    '1984-02': Decimal('0.25'),
    '1984-03': Decimal('0.30'),
    '1985-01': Decimal('0.25'),
    '1986-01': Decimal('0.35'),  # Hiper cruzado
    '1989-01': Decimal('0.40'),  # NCz$
    '1993-01': Decimal('0.10'),  # Cr$
    '1994-01': Decimal('0.05'),  # URV transição
    '2023-11': Decimal('0.0028'),
    '2023-12': Decimal('0.0056'),
    '2024-01': Decimal('0.0042'),
    '2024-02': Decimal('0.0083'),
    '2024-03': Decimal('0.0016'),
    '2024-04': Decimal('0.0038'),
    '2024-05': Decimal('0.0046'),
    '2024-06': Decimal('0.0021'),
    '2024-07': Decimal('0.0038'),
    '2024-08': Decimal('0.0002'),
    '2024-09': Decimal('0.0044'),
    '2024-10': Decimal('0.0056'),
    '2024-11': Decimal('0.0039'),
    '2025-01': Decimal('0.0042'),
}

# Inicialização global dos índices (roda no startup)
print("\n=== DEBUG INDICES HISTÓRICOS ===")
INDICES_HISTORICOS = {}
if 'INDICES_MENSAIS' in globals():
    INDICES_HISTORICOS = {**INDICES_MENSAIS}
csv_indices = carregar_indices_historicos('indices_historicos.csv')
INDICES_HISTORICOS.update(csv_indices)
print(f"[FINAL MERGE] INDICES_HISTORICOS total: {len(INDICES_HISTORICOS)} entradas.")
print(f"[TESTE MERGE] '1984-08' em INDICES_HISTORICOS: {INDICES_HISTORICOS.get('1984-08', 'NÃO MERGEADO')} ({100 * INDICES_HISTORICOS.get('1984-08', 0):.1f}% se ok)")
print("=== FIM DEBUG ===")

def index(request):
    context = {
        'total_mutuarios': Mutuario.objects.count(),
        'total_conjuntos': ConjuntoHabitacional.objects.count(),
        'total_enderecos': Endereco.objects.count(),
        'total_contratos': Contrato.objects.count(),
        'total_parcelas': ParcelaContrato.objects.count(),
    }
    return render(request, 'principal/index.html', context)


def _normalizar_cabecalho(valor):
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    texto = re.sub(r'[^a-z0-9]+', '_', texto).strip('_')
    return texto


def _detectar_campos_cadmut(cabecalhos):
    if not cabecalhos:
        return {}

    mapa_alvos = {
        'codigo_contrato': ['contrato', 'numero_contrato', 'num_contrato', 'cod_contrato', 'codigo_contrato'],
        'nome_mutuario': ['nome', 'nome_mutuario', 'mutuario', 'cliente', 'nome_cliente'],
        'cpf_mutuario': ['cpf', 'cpf_mutuario', 'documento', 'cpf_cnpj'],
        'endereco_imovel': ['endereco', 'endereco_imovel', 'logradouro', 'rua'],
        'uf_imovel': ['uf', 'uf_imovel', 'estado'],
        'municipio': ['municipio', 'cidade', 'cod_municipio', 'codigo_municipio'],
        'data_contrato': ['data_contrato', 'dt_contrato', 'assinatura', 'data_assinatura'],
        'ocorrencia': ['ocorrencia', 'tipo_evento', 'evento'],
        # Evita mapear colunas como "endereco_imovel" para cod_imovel por engano.
        'cod_imovel': ['cod_imovel', 'codigo_imovel', 'codimovel'],
        'chave': ['chave'],
        'lote': ['lote'],
        'sinal': ['sinal'],
    }

    normalizados = [(_normalizar_cabecalho(h), h) for h in cabecalhos]
    resultado = {}
    for alvo, chaves in mapa_alvos.items():
        escolhido = ''
        for cab_norm, cab_original in normalizados:
            if any(chave in cab_norm for chave in chaves):
                escolhido = cab_original
                break
        resultado[alvo] = escolhido
    return resultado


def _ler_preview_base(caminho_arquivo, max_linhas=20):
    sufixo = caminho_arquivo.suffix.lower()

    if sufixo in ('.xlsx', '.xls'):
        try:
            from openpyxl import load_workbook
        except Exception:
            return {
                'erro': 'Biblioteca openpyxl não disponível para ler planilha Excel no ambiente atual.'
            }

        try:
            wb = load_workbook(caminho_arquivo, data_only=True, read_only=True)
        except Exception as exc:
            return {
                'erro': f'Falha ao ler planilha Excel: {exc}'
            }
        ws = wb.active

        linhas = ws.iter_rows(values_only=True)
        cabecalho = []
        for row in linhas:
            if row and any(str(c or '').strip() for c in row):
                cabecalho = [str(c or '').strip() for c in row]
                break

        amostra = []
        for row in linhas:
            if not row:
                continue
            valores = [str(c or '').strip() for c in row]
            if not any(valores):
                continue
            amostra.append(valores)
            if len(amostra) >= max_linhas:
                break

        wb.close()
        return {
            'cabecalhos': cabecalho,
            'campos_detectados': _detectar_campos_cadmut(cabecalho),
            'amostra': amostra,
            'formato': sufixo,
        }

    if sufixo in ('.csv', '.txt'):
        linhas_raw = []
        encoding_usado = 'utf-8-sig'
        try:
            with caminho_arquivo.open('r', encoding='utf-8-sig', newline='') as f:
                for _ in range(50):
                    linha = f.readline()
                    if not linha:
                        break
                    linhas_raw.append(linha)
        except Exception:
            encoding_usado = 'latin1'
            with caminho_arquivo.open('r', encoding='latin1', newline='') as f:
                for _ in range(50):
                    linha = f.readline()
                    if not linha:
                        break
                    linhas_raw.append(linha)

        conteudo_inicial = ''.join(linhas_raw)
        delimitador = ';'
        try:
            delimitador = csv.Sniffer().sniff(conteudo_inicial, delimiters=';,\t|,').delimiter
        except Exception:
            pass

        with caminho_arquivo.open('r', encoding=encoding_usado, newline='') as f:
            leitor = csv.reader(f, delimiter=delimitador)
            cabecalho = []
            for row in leitor:
                if row and any(str(c or '').strip() for c in row):
                    cabecalho = [str(c or '').strip() for c in row]
                    break

            amostra = []
            for row in leitor:
                valores = [str(c or '').strip() for c in row]
                if not any(valores):
                    continue
                amostra.append(valores)
                if len(amostra) >= max_linhas:
                    break

        return {
            'cabecalhos': cabecalho,
            'campos_detectados': _detectar_campos_cadmut(cabecalho),
            'amostra': amostra,
            'formato': sufixo,
        }

    return {
        'erro': f'Formato {sufixo} não suportado para prévia.'
    }


def _valor_para_texto(valor):
    if valor is None:
        return ''
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()
    return str(valor).strip()


def _arquivo_complementos_manuais_ocr():
    return Path(__file__).resolve().parents[1] / 'exports' / 'cadmut_manual_overrides.json'


def _carregar_complementos_manuais_ocr():
    caminho = _arquivo_complementos_manuais_ocr()
    if not caminho.exists():
        return {}
    try:
        with caminho.open('r', encoding='utf-8') as f:
            dados = json.load(f)
        if isinstance(dados, dict):
            return dados
    except Exception:
        return {}
    return {}


def _salvar_complementos_manuais_ocr(dados):
    caminho = _arquivo_complementos_manuais_ocr()
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with caminho.open('w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2, sort_keys=True)


def _aplicar_complemento_manual_em_dados_ocr(dados):
    if not isinstance(dados, dict):
        return dados

    codigo = re.sub(r'\s+', '', _valor_para_texto(dados.get('codigo', '')))
    if not codigo:
        return dados

    overrides = _carregar_complementos_manuais_ocr()
    item = overrides.get(codigo, {})
    if not isinstance(item, dict):
        return dados

    campos_permitidos = ['data_primeiro_venc', 'sa', 'tx_juros', 'cat_prof', 'pr', 'ident', 'orgao', 'dtnasc']
    for campo in campos_permitidos:
        if not _valor_para_texto(dados.get(campo, '')):
            valor_override = _valor_para_texto(item.get(campo, ''))
            if valor_override:
                dados[campo] = valor_override
    return dados


def _parece_endereco(texto):
    t = _valor_para_texto(texto).lower()
    if not t:
        return False
    marcadores = ['rua ', 'av ', 'avenida', 'trav', 'alameda', 'quadra', 'lote', 'bloco', 'estrada']
    if any(m in t for m in marcadores):
        return True
    return len(t) > 24 and ' ' in t


def _parse_data_flexivel(valor):
    if valor is None or valor == '':
        return None

    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    texto = _valor_para_texto(valor)
    if not texto:
        return None

    formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y%m%d']
    for formato in formatos:
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue

    texto_norm = texto.lower().strip()
    texto_norm = re.sub(r'^[^\d]+', '', texto_norm)
    meses_pt = {
        'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'março': 3, 'abril': 4,
        'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8, 'setembro': 9,
        'outubro': 10, 'novembro': 11, 'dezembro': 12,
    }
    m = re.search(r'(\d{1,2})\s+de\s+([a-zçã]+)\s+de\s+(\d{4})', texto_norm, flags=re.IGNORECASE)
    if m:
        dia = int(m.group(1))
        mes = meses_pt.get(m.group(2).lower())
        ano = int(m.group(3))
        if mes:
            try:
                return date(ano, mes, dia)
            except ValueError:
                return None
    return None


def _ler_linhas_base_completa(caminho_arquivo):
    sufixo = caminho_arquivo.suffix.lower()

    if sufixo in ('.xlsx', '.xls'):
        from openpyxl import load_workbook

        wb = load_workbook(caminho_arquivo, data_only=True, read_only=True)
        ws = wb.active

        linhas = ws.iter_rows(values_only=True)
        cabecalho = []
        for row in linhas:
            if row and any(_valor_para_texto(c) for c in row):
                cabecalho = [_valor_para_texto(c) for c in row]
                break

        registros = []
        for row in linhas:
            if not row:
                continue
            valores = [_valor_para_texto(c) for c in row]
            if not any(valores):
                continue
            if len(valores) < len(cabecalho):
                valores += [''] * (len(cabecalho) - len(valores))
            registros.append(dict(zip(cabecalho, valores)))

        wb.close()
        return cabecalho, registros

    if sufixo in ('.csv', '.txt'):
        encoding_usado = 'utf-8-sig'
        linhas_raw = []
        try:
            with caminho_arquivo.open('r', encoding='utf-8-sig', newline='') as f:
                for _ in range(50):
                    linha = f.readline()
                    if not linha:
                        break
                    linhas_raw.append(linha)
        except Exception:
            encoding_usado = 'latin1'
            with caminho_arquivo.open('r', encoding='latin1', newline='') as f:
                for _ in range(50):
                    linha = f.readline()
                    if not linha:
                        break
                    linhas_raw.append(linha)

        delimitador = ';'
        try:
            delimitador = csv.Sniffer().sniff(''.join(linhas_raw), delimiters=';,\t|,').delimiter
        except Exception:
            pass

        with caminho_arquivo.open('r', encoding=encoding_usado, newline='') as f:
            leitor = csv.DictReader(f, delimiter=delimitador)
            cabecalho = leitor.fieldnames or []
            registros = []
            for row in leitor:
                row_limpa = {str(k or '').strip(): _valor_para_texto(v) for k, v in row.items()}
                if any(row_limpa.values()):
                    registros.append(row_limpa)
            return cabecalho, registros

    raise ValueError(f'Formato não suportado para importação: {sufixo}')


def _campo_esta_vazio(valor):
    return valor is None or str(valor).strip() == ''


def _linha_para_campos_contrato(row, mapping, conjunto_destino):
    col_codigo = mapping.get('codigo_contrato') or ''
    if not col_codigo:
        return None

    codigo = re.sub(r'\s+', '', _valor_para_texto(row.get(col_codigo, '')))
    if not codigo:
        return None

    campos = {
        'codigo': codigo,
        'conjunto': conjunto_destino,
    }

    col_data = mapping.get('data_contrato') or ''
    if col_data:
        campos['data_contrato'] = _parse_data_flexivel(row.get(col_data, ''))

    for campo in ['ocorrencia', 'cod_imovel', 'chave', 'lote', 'sinal']:
        col = mapping.get(campo) or ''
        if col:
            valor = _valor_para_texto(row.get(col, ''))
            if campo == 'ocorrencia':
                valor = valor.upper()
            if campo == 'cod_imovel' and _parece_endereco(valor):
                valor = ''
            campos[campo] = valor

    return campos


def _aplicar_modo_importacao(contrato, campos_novos, modo_importacao, campos_sobrescrever):
    campos_editaveis = [
        'conjunto', 'data_contrato', 'ocorrencia', 'cod_imovel', 'chave', 'lote', 'sinal',
        'vlfinanc', 'vlprop', 'prestacao_inicial', 'prazo',
        'data_primeiro_venc', 'sa', 'tx_juros', 'cat_prof', 'pr'
    ]
    alterado = False

    if modo_importacao == 'criar_somente':
        return False

    if modo_importacao == 'complementar_vazios':
        for campo in campos_editaveis:
            if campo not in campos_novos:
                continue
            atual = getattr(contrato, campo)
            novo = campos_novos.get(campo)
            if _campo_esta_vazio(atual) and not _campo_esta_vazio(novo):
                setattr(contrato, campo, novo)
                alterado = True
        return alterado

    if modo_importacao == 'sobrescrever_especificos':
        for campo in campos_sobrescrever:
            if campo not in campos_editaveis or campo not in campos_novos:
                continue
            novo = campos_novos.get(campo)
            if campo == 'cod_imovel' and _parece_endereco(novo):
                continue
            if _campo_esta_vazio(novo):
                continue
            if getattr(contrato, campo) != novo:
                setattr(contrato, campo, novo)
                alterado = True
        return alterado

    return False


def _limpar_cpf(texto):
    return re.sub(r'\D+', '', _valor_para_texto(texto))[:11]


def _formatar_cpf(cpf):
    digits = _limpar_cpf(cpf)
    if len(digits) != 11:
        return _valor_para_texto(cpf)
    return f'{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}'


def _parse_decimal_ocr(valor):
    txt = _valor_para_texto(valor)
    if not txt:
        return None

    limpo = txt.replace('R$', '').replace('r$', '').strip()
    limpo = re.sub(r'[^0-9,.-]', '', limpo)
    if not limpo:
        return None

    # Remove separadores extras e mantém apenas um sinal negativo inicial.
    limpo = re.sub(r'(?<!^)-', '', limpo)

    if ',' in limpo and '.' in limpo:
        limpo = limpo.replace('.', '').replace(',', '.')
    elif ',' in limpo:
        limpo = limpo.replace(',', '.')
    elif limpo.count('.') > 1:
        # Caso típico OCR BR: 7.329.618 (milhar com pontos, sem centavos).
        limpo = limpo.replace('.', '')
    elif limpo.count('.') == 1:
        parte_int, parte_dec = limpo.split('.', 1)
        if len(parte_dec) == 3 and len(parte_int) >= 1:
            # Heurística para milhar único (ex.: 27.092 => 27092).
            limpo = parte_int + parte_dec

    try:
        return Decimal(limpo)
    except Exception:
        return None


def _formatar_decimal_ocr(valor, casas=2):
    dec = _parse_decimal_ocr(valor)
    if dec is None:
        return ''
    mascara = '0.' + ('0' * casas)
    try:
        return f'{dec.quantize(Decimal(mascara))}'
    except Exception:
        return f'{dec}'


def _parse_int_ocr(valor):
    txt = _valor_para_texto(valor)
    if not txt:
        return None
    nums = re.findall(r'\d+', txt)
    if not nums:
        return None
    try:
        return int(nums[0])
    except Exception:
        return None


def _parse_percentual_ocr(valor):
    txt = _valor_para_texto(valor)
    if not txt:
        return None
    txt = txt.replace('%', '').strip()
    dec = _parse_decimal_ocr(txt)
    return dec


def _quebrar_identidade_orgao(valor):
    txt = _valor_para_texto(valor)
    if not txt:
        return '', ''
    txt = re.sub(r'\s+', ' ', txt).strip(' .;,:-')

    ident = ''
    orgao = ''

    # Primeiro, tenta separar no formato clássico "numero-orgao".
    m = re.search(r'([0-9\.\-]{5,20})\s*[-/ ]\s*([A-Za-z]{2,10})', txt)
    if m:
        ident = _valor_para_texto(m.group(1)).strip()
        orgao = _valor_para_texto(m.group(2)).upper().strip()
        return ident, orgao

    # Fallback: captura o número da identidade em qualquer posição.
    m_ident = re.search(r'([0-9][0-9\.\-/]{4,20})', txt)
    if m_ident:
        ident = _valor_para_texto(m_ident.group(1)).strip(' .;,:-')

    # Busca órgão emissor conhecido (ex.: IFP, SSP, DETRAN).
    m_orgao = re.search(r'\b(IFP|SSP|DETRAN|SDS|PC|PM|MEX|CNH)\b', txt, flags=re.IGNORECASE)
    if m_orgao:
        orgao = _valor_para_texto(m_orgao.group(1)).upper().strip()

    if ident or orgao:
        return ident, orgao
    return txt, ''


def _extrair_financeiro_de_campo_composto(valor):
    txt = _valor_para_texto(valor)
    if not txt:
        return {}

    out = {}
    txt_limpo = re.sub(r'\s+', ' ', txt)

    prazo_span = None
    m_prazo = re.search(r'\b(\d{2,4})\s*prest', txt_limpo, flags=re.IGNORECASE)
    if m_prazo:
        try:
            out['prazo'] = int(m_prazo.group(1))
            prazo_span = m_prazo.span(1)
        except Exception:
            pass

    valores_dec = []
    for m in re.finditer(r'\d{1,3}(?:\.\d{3})*(?:,\d+)?', txt_limpo):
        bruto = m.group(0)
        if prazo_span and m.start() == prazo_span[0] and m.end() == prazo_span[1]:
            continue
        dec = _parse_decimal_ocr(bruto)
        if dec is not None and dec > 0:
            valores_dec.append(dec)

    valores_dec = sorted(set(valores_dec))
    if valores_dec:
        out['vlfinanc'] = _formatar_decimal_ocr(str(valores_dec[-1]), casas=2)

    if len(valores_dec) >= 2:
        out['prestacao_inicial'] = _formatar_decimal_ocr(str(valores_dec[0]), casas=6)

    return out


def _normalizar_campo_ocr(campo, valor):
    txt = _valor_para_texto(valor)
    if not txt:
        return ''

    if campo == 'cpf':
        return _formatar_cpf(txt)

    if campo == 'uf':
        uf = re.sub(r'[^A-Za-z]', '', txt).upper()
        ufs_validas = {
            'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA',
            'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN',
            'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'
        }
        return uf if uf in ufs_validas else ''

    if campo == 'cep':
        cep = re.sub(r'\D+', '', txt)
        if len(cep) == 8:
            return f'{cep[:5]}-{cep[5:]}'
        return ''

    if campo == 'numero':
        # Evita lixo típico de OCR (ex.: "da", "de", "oo").
        return txt if re.search(r'\d', txt) else ''

    if campo in {'vlfinanc', 'vlprop'}:
        return _formatar_decimal_ocr(txt, casas=2)

    if campo == 'prestacao_inicial':
        return _formatar_decimal_ocr(txt, casas=6)

    if campo == 'tx_juros':
        dec = _parse_percentual_ocr(txt)
        if dec is None:
            return ''
        try:
            return f"{dec.quantize(Decimal('0.01'))}"
        except Exception:
            return f'{dec}'

    if campo in {'prazo', 'sa'}:
        num = _parse_int_ocr(txt)
        return str(num) if num is not None else ''

    if campo in {'cat_prof', 'pr'}:
        return re.sub(r'\s+', ' ', txt).upper().strip(' .;,:-/')

    if campo == 'ident':
        ident, _ = _quebrar_identidade_orgao(txt)
        m_ident = re.search(r'([0-9][0-9\.\-/]{3,20})', _valor_para_texto(ident))
        if m_ident:
            return _valor_para_texto(m_ident.group(1)).strip(' .;,:-/')
        return ''

    if campo == 'orgao':
        _, orgao = _quebrar_identidade_orgao(txt)
        return orgao

    if campo == 'nome':
        nome = re.sub(r'\s+', ' ', txt).strip(' .;,:-/')
        if _nome_mutuario_suspeito(nome):
            return ''
        return nome

    if campo == 'endereco':
        limpo = re.sub(r'\s+', ' ', txt.replace(';', ',')).strip(' .;,:-/')
        limpo = re.split(r'(?i)\bde\s+fundos\b|\bfundos\b', limpo)[0].strip(' ,.;')
        m_logr = re.search(r'(?i)\b(?:rua|avenida|av\.?|travessa|tv\.?|alameda|estrada|rodovia|pra[çc]a)\s+[^,;]{1,80}', limpo)
        if m_logr:
            return re.sub(r'\s+', ' ', m_logr.group(0)).strip(' .;,:-/')
        return limpo

    if campo == 'cidade':
        limpo = re.sub(r'\s+', ' ', txt).strip(' .;,:-/')
        if limpo in {'-', '/'}:
            return ''
        return limpo

    if campo in {'endereco', 'bairro', 'cidade', 'compl', 'tipoimovel'}:
        limpo = re.sub(r'\s+', ' ', txt).strip(' .;,:-/')
        return limpo

    return txt


def _consolidar_dados_ocr_duplicado(item_principal, item_duplicado):
    if not item_principal or not item_duplicado:
        return

    dados_principal = item_principal.get('dados', {}) or {}
    dados_duplicado = item_duplicado.get('dados', {}) or {}
    campos_relevantes = [
        'nome', 'cpf', 'endereco', 'numero', 'compl', 'bairro',
        'cidade', 'uf', 'cep', 'codimovel', 'tipoimovel',
        'vlfinanc', 'vlprop', 'prestacao_inicial', 'prazo',
        'ident', 'orgao', 'dtnasc', 'data_primeiro_venc', 'sa', 'tx_juros', 'cat_prof', 'pr',
        'chave', 'lote', 'sinal', 'ocorrencia'
    ]

    for campo in campos_relevantes:
        atual = _normalizar_campo_ocr(campo, dados_principal.get(campo, ''))
        novo = _normalizar_campo_ocr(campo, dados_duplicado.get(campo, ''))
        if not atual and novo:
            dados_principal[campo] = novo

    item_principal['dados'] = dados_principal


def _inferir_numero_do_endereco(endereco):
    txt = _valor_para_texto(endereco)
    if not txt:
        return ''

    if re.match(r'(?i)^\s*(?:rua|avenida|av\.?|travessa|tv\.?|alameda|estrada|rodovia|pra[çc]a)\s+\d+\s*$', txt):
        return ''

    m = re.search(r'[,\s](\d{1,6}[a-zA-Z]?)\b', txt)
    if m:
        return _valor_para_texto(m.group(1))
    return ''


def _enriquecer_item_com_campos_raw_azure(item):
    if not item:
        return

    dados = item.get('dados', {}) or {}
    campos_raw = item.get('azure_fields_raw', []) or []

    for raw in campos_raw:
        nome_norm = _valor_para_texto(raw.get('field_name_norm', ''))
        valor = _valor_para_texto(raw.get('value', ''))
        if not valor:
            continue

        if nome_norm in {'valor_financiado_e_prazo', 'valorfinanciadoeprazo'}:
            extraidos = _extrair_financeiro_de_campo_composto(valor)
            for campo, valor_extra in extraidos.items():
                if valor_extra and not _valor_para_texto(dados.get(campo, '')):
                    dados[campo] = valor_extra

        if nome_norm == 'dados':
            if not _valor_para_texto(dados.get('lote', '')):
                m_lote = re.search(r'(?i)\blote\s*[:\-]?\s*([a-z0-9\-/]{1,20})', valor)
                if m_lote:
                    dados['lote'] = _valor_para_texto(m_lote.group(1))
            if not _valor_para_texto(dados.get('compl', '')):
                m_quadra = re.search(r'(?i)\bquadra\s*[:\-]?\s*([a-z0-9\-/]{1,20})', valor)
                if m_quadra:
                    dados['compl'] = f"Quadra { _valor_para_texto(m_quadra.group(1)) }"
            if not _valor_para_texto(dados.get('endereco', '')):
                m_end = re.search(r'(?i)\b(?:rua|avenida|av\.?|travessa|tv\.?|alameda|estrada|rodovia|pra[çc]a)\s+[^\n,;]{1,100}', valor)
                if m_end:
                    dados['endereco'] = _valor_para_texto(m_end.group(0))
            if not _valor_para_texto(dados.get('ident', '')):
                m_ident = re.search(r'(?i)(?:identidade|rg)\s*[:\-]?\s*([^\n;]{4,40})', valor)
                if m_ident:
                    ident, orgao = _quebrar_identidade_orgao(m_ident.group(1))
                    if ident:
                        dados['ident'] = ident
                    if orgao and not _valor_para_texto(dados.get('orgao', '')):
                        dados['orgao'] = orgao
            if not _valor_para_texto(dados.get('dtnasc', '')):
                m_nasc = re.search(r'(?i)(?:data\s*de\s*nasc(?:imento)?|dtnasc)\s*[:\-]?\s*([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4})', valor)
                if m_nasc:
                    dados['dtnasc'] = _valor_para_texto(m_nasc.group(1))
            if not _valor_para_texto(dados.get('data_primeiro_venc', '')):
                m_venc = re.search(r'(?i)(?:data\s*1[ºo]?\s*venc(?:imento)?|primeiro\s*venc(?:imento)?)\s*[:\-]?\s*([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4})', valor)
                if m_venc:
                    dados['data_primeiro_venc'] = _valor_para_texto(m_venc.group(1))
            if not _valor_para_texto(dados.get('sa', '')):
                m_sa = re.search(r'(?i)\bsa\b\s*[:\-]?\s*(\d{1,3})', valor)
                if m_sa:
                    dados['sa'] = _valor_para_texto(m_sa.group(1))
            if not _valor_para_texto(dados.get('tx_juros', '')):
                m_jur = re.search(r'(?i)(?:tx\s*jur(?:os)?|taxa\s*de\s*juros)\s*[:\-]?\s*([0-9\.,]{1,10}%?)', valor)
                if m_jur:
                    dados['tx_juros'] = _valor_para_texto(m_jur.group(1))
            if not _valor_para_texto(dados.get('cat_prof', '')):
                m_cat = re.search(r'(?i)(?:cat\s*prof|categoria\s*profissional)\s*[:\-]?\s*([a-z0-9\./\-]{2,20})', valor)
                if m_cat:
                    dados['cat_prof'] = _valor_para_texto(m_cat.group(1))
            if not _valor_para_texto(dados.get('pr', '')):
                m_pr = re.search(r'(?i)\bpr\b\s*[:\-]?\s*([a-z0-9\./\-]{2,20})', valor)
                if m_pr:
                    dados['pr'] = _valor_para_texto(m_pr.group(1))

        # Fallback global: tenta extrair campos mesmo quando o modelo devolve nomes genéricos.
        if not _valor_para_texto(dados.get('data_primeiro_venc', '')):
            m_venc_global = re.search(
                r'(?i)(?:data\s*1[ºo]?\s*venc(?:imento)?|primeiro\s*venc(?:imento)?|1[ºo]?\s*venc(?:imento)?)\s*[:\-]?\s*([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4})',
                valor,
            )
            if m_venc_global:
                dados['data_primeiro_venc'] = _valor_para_texto(m_venc_global.group(1))

        if not _valor_para_texto(dados.get('sa', '')):
            m_sa_global = re.search(r'(?i)\b(?:sa|sistema\s*de\s*amortiz[aã]c[aã]o)\b\s*[:\-]?\s*(\d{1,3})', valor)
            if m_sa_global:
                dados['sa'] = _valor_para_texto(m_sa_global.group(1))

        if not _valor_para_texto(dados.get('tx_juros', '')):
            m_jur_global = re.search(r'(?i)(?:tx\.?\s*jur(?:os)?|taxa\s*de\s*juros|juros)\s*[:\-]?\s*([0-9\.,]{1,10}%?)', valor)
            if m_jur_global:
                dados['tx_juros'] = _valor_para_texto(m_jur_global.group(1))

        if not _valor_para_texto(dados.get('cat_prof', '')):
            m_cat_global = re.search(r'(?i)(?:cat\.?\s*prof|categoria\s*prof(?:issional)?)\s*[:\-]?\s*([a-z0-9\./\-]{2,20})', valor)
            if m_cat_global:
                dados['cat_prof'] = _valor_para_texto(m_cat_global.group(1))

        if not _valor_para_texto(dados.get('pr', '')):
            m_pr_global = re.search(r'(?i)\b(?:pr|programa)\b\s*[:\-]?\s*([a-z0-9\./\-]{2,20})', valor)
            if m_pr_global:
                dados['pr'] = _valor_para_texto(m_pr_global.group(1))

    item['dados'] = dados


def _linha_para_campos_mutuario(row, mapping, conjunto_destino, codigo_contrato):
    nome_col = mapping.get('nome_mutuario') or ''
    cpf_col = mapping.get('cpf_mutuario') or ''
    end_col = mapping.get('endereco_imovel') or ''
    uf_col = mapping.get('uf_imovel') or ''
    mun_col = mapping.get('municipio') or ''
    codimovel_col = mapping.get('cod_imovel') or ''

    nome = _valor_para_texto(row.get(nome_col, '')) if nome_col else ''
    cpf = _limpar_cpf(row.get(cpf_col, '')) if cpf_col else ''
    endereco = _valor_para_texto(row.get(end_col, '')) if end_col else ''
    uf = _valor_para_texto(row.get(uf_col, '')).upper() if uf_col else ''
    cidade = _valor_para_texto(row.get(mun_col, '')) if mun_col else ''
    codimovel = _valor_para_texto(row.get(codimovel_col, '')) if codimovel_col else ''

    if not nome:
        nome = f'MUTUARIO CONTRATO {codigo_contrato}'

    return {
        'codigo': codigo_contrato,
        'codimovel': codimovel,
        'conjunto': conjunto_destino,
        'conjseg': conjunto_destino,
        'nome': nome,
        'cpf': cpf,
        'endereco': endereco,
        'cidade': cidade,
        'uf': uf,
    }


def _aplicar_modo_mutuario(mutuario, campos_novos, modo_importacao, campos_sobrescrever):
    campos_editaveis = [
        'codimovel', 'conjunto', 'conjseg', 'nome', 'cpf',
        'endereco', 'numero', 'compl', 'bairro', 'cidade', 'uf', 'cep', 'tipoimovel',
        'ident', 'orgao', 'dtnasc'
    ]
    alterado = False

    if modo_importacao == 'criar_somente':
        return False

    def _campo_mutuario_fraco(campo, valor):
        txt = _valor_para_texto(valor)
        if not txt:
            return True
        if campo == 'cpf':
            return len(_limpar_cpf(txt)) != 11
        if campo == 'uf':
            return len(re.sub(r'[^A-Za-z]', '', txt)) != 2
        if campo == 'nome':
            return txt.upper().startswith('MUTUARIO CONTRATO ')
        return False

    if modo_importacao == 'complementar_vazios':
        for campo in campos_editaveis:
            novo = campos_novos.get(campo)
            atual = getattr(mutuario, campo)
            if (_campo_esta_vazio(atual) or _campo_mutuario_fraco(campo, atual)) and not _campo_esta_vazio(novo):
                setattr(mutuario, campo, novo)
                alterado = True
        return alterado

    if modo_importacao == 'sobrescrever_especificos':
        # Em mutuário, só aplica sobrescrita quando campos equivalentes foram selecionados.
        de_para = {
            'conjunto': ['conjunto'],
            'cod_imovel': ['codimovel'],
            'endereco': ['endereco', 'numero', 'compl', 'bairro', 'cidade', 'uf', 'cep'],
            'tipoimovel': ['tipoimovel'],
        }
        permitidos = set()
        for campo in campos_sobrescrever:
            for destino in de_para.get(campo, []):
                permitidos.add(destino)

        # Nome/CPF e dados de localização principais permanecem complementares no modo 3.
        permitidos.update(['nome', 'cpf', 'endereco', 'numero', 'compl', 'bairro', 'uf', 'cidade', 'cep', 'conjseg'])

        for campo in campos_editaveis:
            if campo not in permitidos:
                continue
            novo = campos_novos.get(campo)
            if _campo_esta_vazio(novo):
                continue
            if getattr(mutuario, campo) != novo:
                setattr(mutuario, campo, novo)
                alterado = True
        return alterado

    return False


def _vincular_contrato_mutuario(contrato_id, mutuario_id):
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS contrato_mutuario_map (
                contrato_id INTEGER PRIMARY KEY,
                mutuario_id INTEGER NOT NULL,
                score REAL,
                method TEXT
            )
            """
        )
        cur.execute(
            """
            INSERT OR REPLACE INTO contrato_mutuario_map (contrato_id, mutuario_id, score, method)
            VALUES (?, ?, ?, ?)
            """,
            (contrato_id, mutuario_id, 1.0, 'cadmut_upload')
        )
        conn.commit()
    finally:
        conn.close()


def _campos_contrato_por_dados_ocr(dados, conjunto_destino):
    codigo_bruto = re.sub(r'\s+', '', _valor_para_texto(dados.get('codigo', '')))
    codigo = codigo_bruto
    if codigo and not codigo.upper().startswith('PROV_'):
        codigo = re.sub(r'[^A-Za-z0-9_]+', '', codigo)
        m = re.search(r'(\d{3,10})', codigo)
        if m:
            # Em OCR ruidoso, prioriza o bloco numérico principal do contrato.
            codigo = m.group(1)
    if not codigo:
        return None

    data_contrato = _parse_data_flexivel(dados.get('data_contrato'))
    campos = {
        'codigo': codigo,
        'conjunto': _valor_para_texto(dados.get('conjunto')) or conjunto_destino,
        'data_contrato': data_contrato,
        'ocorrencia': _valor_para_texto(dados.get('ocorrencia', '')).upper(),
        'cod_imovel': _valor_para_texto(dados.get('codimovel', '')),
        'chave': _valor_para_texto(dados.get('chave', '')),
        'lote': _valor_para_texto(dados.get('lote', '')),
        'sinal': _valor_para_texto(dados.get('sinal', '')),
        'vlfinanc': _parse_decimal_ocr(dados.get('vlfinanc', '')),
        'vlprop': _parse_decimal_ocr(dados.get('vlprop', '')),
        'prestacao_inicial': _parse_decimal_ocr(dados.get('prestacao_inicial', '')),
        'prazo': _parse_int_ocr(dados.get('prazo', '')),
        'data_primeiro_venc': _parse_data_flexivel(dados.get('data_primeiro_venc')),
        'sa': _valor_para_texto(_normalizar_campo_ocr('sa', dados.get('sa', ''))),
        'tx_juros': _parse_percentual_ocr(dados.get('tx_juros', '')),
        'cat_prof': _valor_para_texto(_normalizar_campo_ocr('cat_prof', dados.get('cat_prof', ''))),
        'pr': _valor_para_texto(_normalizar_campo_ocr('pr', dados.get('pr', ''))),
    }
    if _parece_endereco(campos.get('cod_imovel')):
        campos['cod_imovel'] = ''
    return campos


def _campos_mutuario_por_dados_ocr(dados, codigo_contrato, conjunto_destino):
    return {
        'codigo': codigo_contrato,
        'codimovel': _valor_para_texto(dados.get('codimovel', '')),
        'conjunto': _valor_para_texto(dados.get('conjunto')) or conjunto_destino,
        'conjseg': _valor_para_texto(dados.get('conjseg')) or conjunto_destino,
        'nome': _normalizar_campo_ocr('nome', dados.get('nome', '')) or f'MUTUARIO CONTRATO {codigo_contrato}',
        'cpf': _limpar_cpf(dados.get('cpf', '')),
        'endereco': _normalizar_campo_ocr('endereco', dados.get('endereco', '')),
        'numero': _normalizar_campo_ocr('numero', dados.get('numero', '')),
        'compl': _normalizar_campo_ocr('compl', dados.get('compl', '')),
        'bairro': _normalizar_campo_ocr('bairro', dados.get('bairro', '')),
        'tipoimovel': _normalizar_campo_ocr('tipoimovel', dados.get('tipoimovel', '')),
        'cidade': _normalizar_campo_ocr('cidade', dados.get('cidade', '')),
        'uf': _normalizar_campo_ocr('uf', dados.get('uf', '')),
        'cep': _normalizar_campo_ocr('cep', dados.get('cep', '')),
        'ident': _normalizar_campo_ocr('ident', dados.get('ident', dados.get('identidade', ''))),
        'orgao': _normalizar_campo_ocr('orgao', dados.get('orgao', dados.get('ident', dados.get('identidade', '')))),
        'dtnasc': _parse_data_flexivel(dados.get('dtnasc', dados.get('data_nasc', ''))),
    }


def _codigo_valido_ocr(codigo):
    c = _valor_para_texto(codigo)
    if not c:
        return False
    c_low = c.lower()
    if c_low.startswith('cadmut_ocr_'):
        return False
    if c.startswith('PROV_'):
        return True
    if len(c) < 3:
        return False
    # Evita falsos positivos do OCR como palavras soltas
    # (ex.: "particular" em "contrato particular").
    if not re.search(r'\d', c):
        return False
    return True


def _resolver_codigo_ocr(dados):
    """Resolve código de contrato com fallback para OCR ruidoso."""
    codigo = _valor_para_texto(dados.get('codigo', ''))
    if codigo and not codigo.upper().startswith('PROV_'):
        codigo_limpo = re.sub(r'\s+', '', codigo)
        codigo_limpo = re.sub(r'[^A-Za-z0-9_]+', '', codigo_limpo)
        m = re.search(r'(\d{3,10})', codigo_limpo)
        if m:
            codigo = m.group(1)
    if _codigo_valido_ocr(codigo):
        return codigo, False

    for campo in ['numero_contrato', 'contrato', 'chave']:
        bruto = _valor_para_texto(dados.get(campo, ''))
        if not bruto:
            continue
        m = re.search(r'\b\d{3,10}\b', bruto)
        if m:
            candidato = m.group(0)
            if _codigo_valido_ocr(candidato):
                return candidato, False

    cpf = _limpar_cpf(dados.get('cpf', ''))
    if len(cpf) == 11:
        # Código provisório para não descartar páginas úteis sem número de contrato legível.
        return f'PROV_{cpf}', True

    return '', False


def _ocr_item_sem_dados(dados):
    d = dados or {}
    return not any([
        _valor_para_texto(d.get('codigo', '')),
        _valor_para_texto(d.get('cpf', '')),
        _valor_para_texto(d.get('nome', '')),
        _valor_para_texto(d.get('endereco', '')),
        _valor_para_texto(d.get('cidade', '')),
    ])


def _extrair_campos_cadmut_do_texto(texto):
    txt = _valor_para_texto(texto)
    if not txt:
        return {}

    campos = {}

    def _primeiro(padroes):
        for padrao in padroes:
            m = re.search(padrao, txt, flags=re.IGNORECASE | re.MULTILINE)
            if m:
                valor = _valor_para_texto(m.group(1))
                if valor:
                    return valor
        return ''

    codigo = _primeiro([
        r'(?:codigo\s*do\s*contrato|cod\.?\s*contrato|numero\s*do\s*contrato|n\.?\s*contrato)\s*[:\-]?\s*([a-z0-9._\-/]{3,24})',
        r'\bcontrato\s*[:\-]?\s*([a-z0-9._\-/]{3,24})',
    ])
    if codigo and _codigo_valido_ocr(codigo):
        campos['codigo'] = codigo

    cpf_raw = _primeiro([
        r'cpf\s*[:\-]?\s*([0-9\.\-\s]{11,18})',
        r'\b([0-9]{3}\.?[0-9]{3}\.?[0-9]{3}[\-\s]?[0-9]{2})\b',
    ])
    cpf_limpo = _limpar_cpf(cpf_raw)
    if len(cpf_limpo) == 11:
        campos['cpf'] = f'{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:9]}-{cpf_limpo[9:]}'

    nome = _primeiro([
        r'(?:nome\s*(?:do)?\s*(?:mutuario|mutuario|cliente)?|mutuario|cliente)\s*[:\-]\s*([^\n]{5,120})',
    ])
    if nome:
        nome = re.sub(r'\s+', ' ', nome).strip(' .;,:-')
        if not _nome_mutuario_suspeito(nome):
            campos['nome'] = nome

    endereco = _primeiro([
        r'(?:endereco\s*(?:do\s*imovel)?|logradouro|rua)\s*[:\-]\s*([^\n]{6,160})',
        r'\b((?:rua|r\.|avenida|av\.|travessa|tv\.|alameda)\s+[^\n]{6,160})',
    ])
    if endereco:
        endereco = re.sub(r'\s+', ' ', endereco).strip(' .;,:-')
        campos['endereco'] = endereco

    cidade = _primeiro([
        r'(?:cidade|municipio)\s*[:\-]\s*([a-zà-ÿ\s\-\']{2,80})',
    ])
    if cidade:
        campos['cidade'] = re.sub(r'\s+', ' ', cidade).strip(' .;,:-')

    uf = _primeiro([
        r'\buf\s*[:\-]?\s*([a-z]{2})\b',
        r'\b[a-zà-ÿ\s\-\']+\s*/\s*([a-z]{2})\b',
    ])
    if uf:
        campos['uf'] = uf.upper()

    cep = _primeiro([
        r'\b(\d{5}-?\d{3})\b',
    ])
    if cep:
        campos['cep'] = cep

    numero = _primeiro([
        r'(?:numero|n\.?o?|nº)\s*[:\-]?\s*([a-z0-9\-/]{1,12})',
    ])
    if numero:
        campos['numero'] = numero

    compl = _primeiro([
        r'(?:complemento|compl\.?|apto|apartamento|bloco)\s*[:\-]?\s*([^\n]{1,80})',
    ])
    if compl:
        campos['compl'] = re.sub(r'\s+', ' ', compl).strip(' .;,:-')

    bairro = _primeiro([
        r'(?:bairro)\s*[:\-]?\s*([^\n]{2,80})',
    ])
    if bairro:
        campos['bairro'] = re.sub(r'\s+', ' ', bairro).strip(' .;,:-')

    tipoimovel = _primeiro([
        r'(?:tipo\s*imovel|tipo\s*do\s*imovel|tipologia)\s*[:\-]?\s*([a-zà-ÿ\s\-/]{3,60})',
    ])
    if tipoimovel:
        campos['tipoimovel'] = re.sub(r'\s+', ' ', tipoimovel).strip(' .;,:-')

    codimovel = _primeiro([
        r'(?:cod\.?\s*imovel|codigo\s*imovel)\s*[:\-]?\s*([a-z0-9._\-/]{2,24})',
    ])
    if codimovel and not _parece_endereco(codimovel):
        campos['codimovel'] = codimovel

    identidade = _primeiro([
        r'(?:identidade|rg)\s*[:\-]?\s*([^\n]{4,40})',
    ])
    if identidade:
        ident, orgao = _quebrar_identidade_orgao(identidade)
        if ident:
            campos['ident'] = ident
        if orgao:
            campos['orgao'] = orgao

    if not _valor_para_texto(campos.get('orgao', '')):
        orgao_txt = _primeiro([
            r'(?:orga[oã]\s*emissor|orgao)\s*[:\-]?\s*([a-z]{2,10})',
            r'expedid[ao]\s*em[^\n,;]{0,30}[,\s]\b([a-z]{2,10})\b',
            r'\b(ifp|ssp|detran|sds|pc|pm|mex)\b',
        ])
        if orgao_txt:
            campos['orgao'] = _valor_para_texto(orgao_txt).upper()

    data_nasc = _primeiro([
        r'(?:data\s*de\s*nasc(?:imento)?|dtnasc)\s*[:\-]?\s*([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4}|\d{1,2}\s+de\s+[a-zçã]+\s+de\s+\d{4})',
    ])
    if data_nasc:
        campos['dtnasc'] = data_nasc

    data_primeiro_venc = _primeiro([
        r'(?:data\s*1[ºo]?\s*venc(?:imento)?|primeiro\s*venc(?:imento)?|data\s*primeiro\s*venc(?:imento)?)\s*[:\-]?\s*([0-9]{1,2}[\/\-][0-9]{1,2}[\/\-][0-9]{2,4}|\d{1,2}\s+de\s+[a-zçã]+\s+de\s+\d{4})',
    ])
    if data_primeiro_venc:
        campos['data_primeiro_venc'] = data_primeiro_venc

    sa = _primeiro([
        r'\bsa\b\s*[:\-]?\s*(\d{1,3})',
        r'sistema\s*de\s*amortiz[aã]c[aã]o\s*[:\-]?\s*(\d{1,3})',
    ])
    if sa:
        campos['sa'] = sa

    tx_juros = _primeiro([
        r'(?:tx\s*jur(?:os)?|taxa\s*de\s*juros)\s*[:\-]?\s*([0-9\.,]{1,10}%?)',
    ])
    if tx_juros:
        campos['tx_juros'] = tx_juros

    cat_prof = _primeiro([
        r'(?:cat\s*prof|categoria\s*profissional)\s*[:\-]?\s*([a-z0-9\./\-]{2,20})',
    ])
    if cat_prof:
        campos['cat_prof'] = cat_prof

    pr = _primeiro([
        r'\bpr\b\s*[:\-]?\s*([a-z0-9\./\-]{2,20})',
    ])
    if pr:
        campos['pr'] = pr

    vlfinanc = _primeiro([
        r'(?:valor\s*(?:do\s*)?financi(?:ado|amento)|vl\.?\s*financ(?:iado|iamento)?)\s*[:\-]?\s*([0-9\.,\sR$]{3,30})',
    ])
    if _parse_decimal_ocr(vlfinanc) is not None:
        campos['vlfinanc'] = _formatar_decimal_ocr(vlfinanc, casas=2)

    vlprop = _primeiro([
        r'(?:valor\s*(?:do\s*)?(?:imovel|imóvel|propriedade)|vl\.?\s*prop)\s*[:\-]?\s*([0-9\.,\sR$]{3,30})',
    ])
    if _parse_decimal_ocr(vlprop) is not None:
        campos['vlprop'] = _formatar_decimal_ocr(vlprop, casas=2)

    prest = _primeiro([
        r'(?:prestacao\s*inicial|prestação\s*inicial|valor\s*da\s*prestacao|valor\s*da\s*prestação|prestacao|prestação)\s*[:\-]?\s*([0-9\.,\sR$]{3,30})',
    ])
    if _parse_decimal_ocr(prest) is not None:
        campos['prestacao_inicial'] = _formatar_decimal_ocr(prest, casas=6)

    return campos


def _mesclar_campos_cadmut_texto(dados_base, texto):
    dados = dict(dados_base or {})
    extraidos = _extrair_campos_cadmut_do_texto(texto)
    for campo, valor in extraidos.items():
        atual = _valor_para_texto(dados.get(campo, ''))
        if not atual:
            dados[campo] = valor
    return dados


def _obter_config_azure_document_intelligence():
    endpoint = _valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT', '')).rstrip('/')
    key = _valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_KEY', ''))
    if not endpoint or not key:
        return None

    model = _valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_MODEL', 'prebuilt-layout')) or 'prebuilt-layout'
    # Modelo secundário é opcional. Não força fallback para prebuilt-document
    # quando o ambiente estiver vazio, pois isso pode sobrepor modelo custom.
    secondary_model = _valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_SECONDARY_MODEL', ''))
    api_version = _valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_API_VERSION', '2023-07-31')) or '2023-07-31'

    try:
        min_confidence = float(_valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_MIN_CONFIDENCE', '0.80')))
    except Exception:
        min_confidence = 0.80
    min_confidence = min(max(min_confidence, 0.0), 1.0)

    try:
        timeout_seconds = int(_valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_TIMEOUT_SECONDS', '120')))
    except Exception:
        timeout_seconds = 120
    timeout_seconds = max(timeout_seconds, 15)

    local_fallback_raw = _valor_para_texto(os.getenv('AZURE_DOCUMENT_INTELLIGENCE_LOCAL_FALLBACK', '1')).lower()
    local_fallback_enabled = local_fallback_raw in {'1', 'true', 'yes', 'on'}

    return {
        'endpoint': endpoint,
        'key': key,
        'model': model,
        'secondary_model': secondary_model,
        'api_version': api_version,
        'min_confidence': min_confidence,
        'timeout_seconds': timeout_seconds,
        'local_fallback_enabled': local_fallback_enabled,
    }


def _contar_campos_cadmut_preenchidos(dados):
    d = dados or {}
    campos = ['codigo', 'nome', 'cpf', 'endereco', 'numero', 'compl', 'bairro', 'cidade', 'uf', 'cep', 'codimovel', 'tipoimovel']
    total = 0
    for campo in campos:
        if _valor_para_texto(d.get(campo, '')):
            total += 1
    return total


def _lote_ocr_fraco(contratos):
    if not contratos:
        return True

    melhores = []
    for item in contratos:
        dados = item.get('dados', {}) or {}
        riqueza = _contar_campos_cadmut_preenchidos(dados)
        cpf = _limpar_cpf(dados.get('cpf', ''))
        nome = _valor_para_texto(dados.get('nome', ''))
        endereco = _valor_para_texto(dados.get('endereco', ''))
        melhores.append({
            'riqueza': riqueza,
            'cpf': len(cpf) == 11,
            'nome': bool(nome),
            'endereco': bool(endereco),
        })

    if not melhores:
        return True

    max_riqueza = max(x['riqueza'] for x in melhores)
    tem_nome_ou_endereco = any(x['nome'] or x['endereco'] for x in melhores)
    tem_apenas_cpf = any(x['cpf'] for x in melhores) and not tem_nome_ou_endereco

    # Fraco quando no melhor caso só chegam 2 campos (tipicamente codigo+cpf)
    # e não há qualquer nome/endereço utilizável.
    return max_riqueza <= 2 and tem_apenas_cpf


def _pontuacao_lote_cadmut(contratos):
    """Pontua lote OCR priorizando preenchimento útil para CADMUT."""
    if not contratos:
        return 0

    total = 0
    for item in contratos:
        dados = item.get('dados', {}) or {}
        riqueza = _contar_campos_cadmut_preenchidos(dados)
        total += riqueza * 10
        if item.get('codigo_valido'):
            total += 20
        if _valor_para_texto(dados.get('nome', '')):
            total += 8
        if _valor_para_texto(dados.get('endereco', '')):
            total += 8
    return total


def _azure_document_intelligence_analyze(pdf_path, cfg):
    model_escaped = urllib_parse.quote(cfg['model'], safe='')
    base_urls = [
        (
            f"{cfg['endpoint']}/documentintelligence/documentModels/{model_escaped}:analyze"
            f"?api-version={cfg['api_version']}"
        ),
        (
            f"{cfg['endpoint']}/formrecognizer/documentModels/{model_escaped}:analyze"
            f"?api-version={cfg['api_version']}"
        ),
    ]

    with open(pdf_path, 'rb') as f:
        pdf_bytes = f.read()

    operation_location = ''
    ultimo_erro_http = None
    for url in base_urls:
        req = urllib_request.Request(
            url=url,
            data=pdf_bytes,
            method='POST',
            headers={
                'Ocp-Apim-Subscription-Key': cfg['key'],
                'Content-Type': 'application/pdf',
            },
        )

        try:
            with urllib_request.urlopen(req, timeout=cfg['timeout_seconds']) as resp:
                body_raw = resp.read().decode('utf-8', errors='ignore')
                if resp.status == 200:
                    return json.loads(body_raw or '{}')

                operation_location = resp.headers.get('operation-location', '')
                if operation_location:
                    break
        except urllib_error.HTTPError as exc:
            detalhe = exc.read().decode('utf-8', errors='ignore') if exc.fp else str(exc)
            ultimo_erro_http = RuntimeError(f'Erro HTTP Azure OCR ({exc.code}): {detalhe}')
            # Tenta a próxima rota se houver 404.
            if int(exc.code) == 404:
                continue
            raise ultimo_erro_http

    if not operation_location:
        if ultimo_erro_http:
            raise ultimo_erro_http
        raise RuntimeError('Azure OCR não retornou operation-location.')

    poll_req = urllib_request.Request(
        url=operation_location,
        method='GET',
        headers={
            'Ocp-Apim-Subscription-Key': cfg['key'],
        },
    )

    inicio = time.time()
    while True:
        if (time.time() - inicio) > cfg['timeout_seconds']:
            raise TimeoutError('Tempo limite excedido ao aguardar OCR da Azure.')

        try:
            with urllib_request.urlopen(poll_req, timeout=cfg['timeout_seconds']) as poll_resp:
                payload = json.loads(poll_resp.read().decode('utf-8', errors='ignore') or '{}')
        except urllib_error.HTTPError as exc:
            detalhe = exc.read().decode('utf-8', errors='ignore') if exc.fp else str(exc)
            raise RuntimeError(f'Erro ao consultar status Azure OCR ({exc.code}): {detalhe}')

        status = _valor_para_texto(payload.get('status', '')).lower()
        if status == 'succeeded':
            return payload
        if status in {'failed', 'canceled'}:
            erro = payload.get('error', {})
            msg = _valor_para_texto(erro.get('message', 'Falha sem detalhes no retorno da Azure.'))
            raise RuntimeError(f'Azure OCR falhou: {msg}')

        time.sleep(1.2)


def _azure_field_to_text_and_conf(field_data):
    if not isinstance(field_data, dict):
        return '', 0.0

    confidence = field_data.get('confidence', 0.0)
    try:
        confidence = float(confidence)
    except Exception:
        confidence = 0.0

    for key in [
        'content',
        'valueString',
        'valueDate',
        'valueTime',
        'valuePhoneNumber',
        'valueCountryRegion',
        'valueSelectionMark',
        'valueCurrency',
        'valueNumber',
        'valueInteger',
    ]:
        value = field_data.get(key)
        if value is None:
            continue
        if isinstance(value, dict):
            if 'amount' in value:
                return _valor_para_texto(value.get('amount', '')), confidence
            if 'currencySymbol' in value and 'amount' in value:
                return _valor_para_texto(value.get('amount', '')), confidence
            if 'content' in value:
                return _valor_para_texto(value.get('content', '')), confidence
        return _valor_para_texto(value), confidence

    return '', confidence


def _documento_azure_por_pagina(documentos):
    docs_por_pagina = defaultdict(list)
    for doc in documentos or []:
        regs = doc.get('boundingRegions', []) or []
        paginas = set()
        for reg in regs:
            try:
                p = int(reg.get('pageNumber', 0) or 0)
            except Exception:
                p = 0
            if p > 0:
                paginas.add(p)
        for p in paginas:
            docs_por_pagina[p].append(doc)
    return docs_por_pagina


def _listar_campos_raw_azure(docs_pagina):
    campos = []
    for doc in docs_pagina or []:
        fields = doc.get('fields', {}) or {}
        for field_name, field_data in fields.items():
            valor, conf = _azure_field_to_text_and_conf(field_data)
            campos.append({
                'field_name': _valor_para_texto(field_name),
                'field_name_norm': _normalizar_cabecalho(field_name),
                'value': _valor_para_texto(valor),
                'confidence': conf,
            })
    return campos


def _mesclar_campos_estruturados_azure(dados_base, docs_pagina, min_confidence):
    dados = dict(dados_base or {})
    if not docs_pagina:
        return dados

    mapa_campos = {
        'codigo': ['codigo', 'codigocontrato', 'numero_contrato', 'numerocontrato', 'contrato'],
        'cpf': ['cpf', 'cpf_mutuario', 'documento', 'cpf_cnpj'],
        'nome': ['nome', 'nome_mutuario', 'mutuario', 'nome_cliente'],
        'endereco': ['endereco', 'endereco_imovel', 'logradouro', 'rua'],
        'numero': ['numero', 'numero_imovel', 'numero_endereco'],
        'compl': ['complemento', 'compl', 'apto', 'apartamento', 'bloco'],
        'bairro': ['bairro'],
        'cidade': ['cidade', 'municipio'],
        'uf': ['uf', 'estado'],
        'cep': ['cep'],
        'codimovel': ['codimovel', 'cod_imovel', 'codigo_imovel'],
        'tipoimovel': ['tipoimovel', 'tipo_imovel', 'tipologia', 'tipo_do_imovel'],
        'vlfinanc': ['vlfinanc', 'valor_financiamento', 'valor_financiado', 'valor_do_financiamento'],
        'vlprop': ['vlprop', 'valor_imovel', 'valor_do_imovel', 'valor_propriedade'],
        'prestacao_inicial': ['prestacao_inicial', 'valor_prestacao', 'prestacao', 'prestacao_mensal'],
        'prazo': ['prazo', 'prazo_meses', 'quantidade_prestacoes', 'numero_prestacoes'],
        'ident': ['identidade', 'rg', 'identificacao'],
        'dtnasc': ['dtnasc', 'data_nascimento', 'nascimento', 'data_nasc', 'data_nasc_mutuario'],
        'data_primeiro_venc': ['data_primeiro_venc', 'data_1_venc', 'data1venc', 'primeiro_vencimento', 'data_primeiro_vencimento', 'data_1o_vencimento'],
        'sa': ['sa', 'sistema_amortizacao', 'sistema_de_amortizacao'],
        'tx_juros': ['tx_juros', 'tx_jur', 'taxa_juros', 'taxa_de_juros', 'juros'],
        'cat_prof': ['cat_prof', 'catprof', 'categoria_profissional', 'categoria_prof'],
        'pr': ['pr', 'programa', 'sigla_programa'],
        'chave': ['chave'],
        'lote': ['lote'],
        'sinal': ['sinal'],
    }

    indice_alvos = {}
    for alvo, aliases in mapa_campos.items():
        for alias in aliases:
            indice_alvos[alias] = alvo

    def _mapear_campo_por_heuristica(nome_norm):
        nome = _valor_para_texto(nome_norm)
        if not nome:
            return ''

        if 'cpf' in nome or 'documento' in nome:
            return 'cpf'
        if 'nome' in nome and ('mutuario' in nome or 'cliente' in nome or nome == 'nome'):
            return 'nome'
        if 'endereco' in nome or 'logradouro' in nome or nome.startswith('rua'):
            return 'endereco'
        if 'bairro' in nome:
            return 'bairro'
        if 'cidade' in nome or 'municipio' in nome:
            return 'cidade'
        if nome in {'uf', 'estado'} or nome.endswith('_uf'):
            return 'uf'
        if 'cep' in nome:
            return 'cep'
        if ('numero' in nome and 'contrato' not in nome) or nome in {'num', 'n'}:
            return 'numero'
        if 'complemento' in nome or nome.startswith('compl'):
            return 'compl'
        if ('tipo' in nome and 'imovel' in nome) or 'tipologia' in nome:
            return 'tipoimovel'
        if ('financi' in nome and 'valor' in nome) or nome in {'vlfinanc', 'valorfinanciamento'}:
            return 'vlfinanc'
        if ('imovel' in nome and 'valor' in nome) or nome in {'vlprop', 'valorimovel'}:
            return 'vlprop'
        if 'prestacao' in nome or 'prestacaoinicial' in nome or 'valorprestacao' in nome:
            return 'prestacao_inicial'
        if 'prazo' in nome or ('prestac' in nome and 'numero' in nome):
            return 'prazo'
        if 'cod' in nome and 'imovel' in nome:
            return 'codimovel'
        if ('contrato' in nome and 'numero' in nome) or 'codigocontrato' in nome:
            return 'codigo'
        if 'identidade' in nome or nome == 'rg':
            return 'ident'
        if 'nasc' in nome:
            return 'dtnasc'
        if ('primeiro' in nome and 'venc' in nome) or ('1' in nome and 'venc' in nome):
            return 'data_primeiro_venc'
        if nome == 'sa' or ('sistema' in nome and 'amort' in nome):
            return 'sa'
        if 'juros' in nome:
            return 'tx_juros'
        if 'cat' in nome and 'prof' in nome:
            return 'cat_prof'
        if nome in {'pr', 'programa'}:
            return 'pr'
        if nome in {'chave'}:
            return 'chave'
        if nome in {'lote'}:
            return 'lote'
        if nome in {'sinal'}:
            return 'sinal'
        return ''

    candidatos = {}
    for doc in docs_pagina:
        fields = doc.get('fields', {}) or {}
        for field_name, field_data in fields.items():
            nome_norm = _normalizar_cabecalho(field_name)

            if nome_norm == 'dados':
                valor_dados, conf_dados = _azure_field_to_text_and_conf(field_data)
                valor_dados = _valor_para_texto(valor_dados)
                if valor_dados and conf_dados >= min_confidence:
                    m_lote = re.search(r'(?i)\blote\s*[:\-]?\s*([a-z0-9\-/]{1,20})', valor_dados)
                    if m_lote:
                        atual_lote = candidatos.get('lote')
                        if not atual_lote or conf_dados > atual_lote['conf']:
                            candidatos['lote'] = {'valor': _valor_para_texto(m_lote.group(1)), 'conf': conf_dados}

                    m_end = re.search(r'(?i)\b(?:rua|avenida|av\.?|travessa|tv\.?|alameda|estrada|rodovia|pra[çc]a)\s+[^\n,;]{1,100}', valor_dados)
                    if m_end:
                        atual_end = candidatos.get('endereco')
                        if not atual_end or conf_dados > atual_end['conf']:
                            candidatos['endereco'] = {'valor': _valor_para_texto(m_end.group(0)), 'conf': conf_dados}

            # Alguns modelos retornam campo combinado (ex.: "valor financiado e prazo").
            if nome_norm in {'valor_financiado_e_prazo', 'valorfinanciadoeprazo'}:
                valor_composto, conf_composto = _azure_field_to_text_and_conf(field_data)
                if conf_composto >= min_confidence:
                    extraidos = _extrair_financeiro_de_campo_composto(valor_composto)
                    for alvo_extra, valor_extra in extraidos.items():
                        if not valor_extra:
                            continue
                        atual_extra = candidatos.get(alvo_extra)
                        if not atual_extra or conf_composto > atual_extra['conf']:
                            candidatos[alvo_extra] = {'valor': valor_extra, 'conf': conf_composto}

            alvo = indice_alvos.get(nome_norm)
            if not alvo:
                alvo = _mapear_campo_por_heuristica(nome_norm)
            if not alvo:
                continue

            valor, conf = _azure_field_to_text_and_conf(field_data)
            valor = _valor_para_texto(valor)
            if not valor:
                continue

            # Campos essenciais podem chegar com confiança baixa em modelos custom novos.
            if conf < min_confidence and alvo not in {'nome', 'endereco', 'bairro', 'cidade', 'numero', 'compl', 'tipoimovel', 'data_contrato'}:
                continue

            atual = candidatos.get(alvo)
            if not atual or conf > atual['conf']:
                candidatos[alvo] = {'valor': valor, 'conf': conf}

    for alvo, item in candidatos.items():
        if alvo == 'cpf':
            item['valor'] = _limpar_cpf(item['valor'])
        if alvo == 'uf':
            item['valor'] = _valor_para_texto(item['valor']).upper()
        if alvo == 'codimovel' and _parece_endereco(item['valor']):
            continue

        if _valor_para_texto(dados.get(alvo, '')):
            continue
        dados[alvo] = item['valor']

    return dados


def _extrair_lote_ocr_azure(pdf_path, cfg):
    from principal.ocr_hibrido import analisar_ocr_hibrido

    payload = _azure_document_intelligence_analyze(pdf_path, cfg)
    resultado = payload.get('analyzeResult', {}) or {}
    paginas = resultado.get('pages', []) or []
    documentos = resultado.get('documents', []) or []
    docs_por_pagina = _documento_azure_por_pagina(documentos)
    contratos = []

    for idx, pagina in enumerate(paginas, start=1):
        linhas = pagina.get('lines', []) or []
        texto = '\n'.join(
            _valor_para_texto(linha.get('content', ''))
            for linha in linhas
            if _valor_para_texto(linha.get('content', ''))
        )

        dados, rel = analisar_ocr_hibrido({}, texto)
        dados = _mesclar_campos_cadmut_texto(dados, texto)
        docs_da_pagina = docs_por_pagina.get(idx, [])
        if not docs_da_pagina and documentos:
            # Alguns retornos custom não trazem boundingRegions por página.
            docs_da_pagina = documentos
        dados = _mesclar_campos_estruturados_azure(dados, docs_da_pagina, cfg['min_confidence'])
        campos_raw_azure = _listar_campos_raw_azure(docs_da_pagina)
        codigo, codigo_provisorio = _resolver_codigo_ocr(dados)
        if codigo:
            dados['codigo'] = codigo

        score = int(rel.get('score', 0) or 0)
        # Para CADMUT, o score global do relatório híbrido não deve invalidar
        # páginas que já possuem código resolvido (inclusive provisório).
        codigo_valido = _codigo_valido_ocr(codigo)
        contratos.append({
            'pagina': idx,
            'score': score,
            'dados': dados,
            'codigo': codigo,
            'codigo_valido': codigo_valido,
            'codigo_provisorio': codigo_provisorio,
            'faltando_criticos': rel.get('faltando_criticos', []),
            'metodo': f"azure-document-intelligence:{cfg['model']}",
            'azure_fields_raw': campos_raw_azure,
        })

    # Alguns documentos podem vir sem segmentação de páginas; usa conteúdo total como fallback.
    if not contratos:
        texto_total = _valor_para_texto(resultado.get('content', ''))
        if texto_total:
            dados, rel = analisar_ocr_hibrido({}, texto_total)
            dados = _mesclar_campos_cadmut_texto(dados, texto_total)
            dados = _mesclar_campos_estruturados_azure(dados, documentos, cfg['min_confidence'])
            codigo, codigo_provisorio = _resolver_codigo_ocr(dados)
            if codigo:
                dados['codigo'] = codigo
            score = int(rel.get('score', 0) or 0)
            campos_raw_azure = _listar_campos_raw_azure(documentos)
            contratos.append({
                'pagina': 1,
                'score': score,
                'dados': dados,
                'codigo': codigo,
                'codigo_valido': _codigo_valido_ocr(codigo),
                'codigo_provisorio': codigo_provisorio,
                'faltando_criticos': rel.get('faltando_criticos', []),
                'metodo': f"azure-document-intelligence:{cfg['model']} (fallback-content)",
                'azure_fields_raw': campos_raw_azure,
            })

    return contratos


def _extrair_lote_ocr_por_pagina(pdf_path):
    """
    Extrai contratos de um PDF multi-paginas.
    Prioriza Azure Document Intelligence (quando configurado) e faz fallback
    automático para o pipeline OCR local já existente.
    """
    from ocr_contrato_processor import ContratoOCRExtractor
    from principal.ocr_hibrido import analisar_ocr_hibrido

    cfg_azure = _obter_config_azure_document_intelligence()
    if cfg_azure:
        try:
            contratos_azure = _extrair_lote_ocr_azure(pdf_path, cfg_azure)
            modelo_primario = _valor_para_texto(cfg_azure.get('model', ''))
            secondary_model = _valor_para_texto(cfg_azure.get('secondary_model', ''))
            # Só tenta secundário automaticamente quando o primário é prebuilt.
            # Para modelo custom, prioriza sempre o resultado do custom.
            if (
                contratos_azure
                and secondary_model
                and secondary_model != modelo_primario
                and modelo_primario.startswith('prebuilt-')
            ):
                cfg_sec = dict(cfg_azure)
                cfg_sec['model'] = secondary_model
                try:
                    contratos_sec = _extrair_lote_ocr_azure(pdf_path, cfg_sec)
                    if contratos_sec:
                        prim_score = _pontuacao_lote_cadmut(contratos_azure)
                        sec_score = _pontuacao_lote_cadmut(contratos_sec)
                        if sec_score > prim_score:
                            contratos_azure = contratos_sec
                except Exception as exc:
                    print(f"[WARN] Modelo secundário Azure falhou ({secondary_model}): {exc}")
            if contratos_azure:
                return contratos_azure
        except Exception as exc:
            if not cfg_azure.get('local_fallback_enabled', True):
                raise
            print(f"[WARN] Azure OCR indisponível, usando fallback local: {exc}")

    PdfReader = None
    PdfWriter = None
    lib_pdf = ''
    try:
        from pypdf import PdfReader as _PdfReader, PdfWriter as _PdfWriter
        PdfReader, PdfWriter = _PdfReader, _PdfWriter
        lib_pdf = 'pypdf'
    except Exception:
        try:
            from PyPDF2 import PdfReader as _PdfReader, PdfWriter as _PdfWriter
            PdfReader, PdfWriter = _PdfReader, _PdfWriter
            lib_pdf = 'PyPDF2'
        except Exception:
            PdfReader = None
            PdfWriter = None

    contratos = []

    # Fallback seguro: se biblioteca de split não estiver disponível,
    # processa o PDF inteiro como um único item (não falha o fluxo OCR).
    if not PdfReader or not PdfWriter:
        extractor = ContratoOCRExtractor(str(pdf_path))
        dados = extractor.extract_all() or {}
        dados, rel = analisar_ocr_hibrido(dados, extractor.text or '')
        codigo, codigo_provisorio = _resolver_codigo_ocr(dados)
        if codigo:
            dados['codigo'] = codigo
        contratos.append({
            'pagina': 1,
            'score': rel.get('score', 0),
            'dados': dados,
            'codigo': codigo,
            'codigo_valido': _codigo_valido_ocr(codigo),
            'codigo_provisorio': codigo_provisorio,
            'faltando_criticos': rel.get('faltando_criticos', []),
            'metodo': f"{getattr(extractor, '_metodo_extracao', 'desconhecido')} (fallback-sem-split)",
        })
        return contratos

    reader = PdfReader(str(pdf_path))

    for idx, page in enumerate(reader.pages, start=1):
        fd, temp_str = tempfile.mkstemp(suffix='.pdf', prefix=f'cadmut_ocr_page_{idx:03d}_')
        os.close(fd)
        page_pdf = Path(temp_str)

        try:
            writer = PdfWriter()
            writer.add_page(page)
            with open(page_pdf, 'wb') as out_pdf:
                writer.write(out_pdf)

            extractor = ContratoOCRExtractor(str(page_pdf))
            dados = extractor.extract_all() or {}
            dados, rel = analisar_ocr_hibrido(dados, extractor.text or '')

            codigo, codigo_provisorio = _resolver_codigo_ocr(dados)
            if codigo:
                dados['codigo'] = codigo
            contratos.append({
                'pagina': idx,
                'score': rel.get('score', 0),
                'dados': dados,
                'codigo': codigo,
                'codigo_valido': _codigo_valido_ocr(codigo),
                'codigo_provisorio': codigo_provisorio,
                'faltando_criticos': rel.get('faltando_criticos', []),
                'metodo': f"{getattr(extractor, '_metodo_extracao', 'desconhecido')} ({lib_pdf})",
            })
        finally:
            if page_pdf.exists():
                try:
                    page_pdf.unlink()
                except Exception:
                    pass

    return contratos


def cadmut(request):
    base_dir = Path(__file__).resolve().parents[1] / 'exports' / 'cadmut_bases'
    base_dir.mkdir(parents=True, exist_ok=True)

    upload_success = ''
    upload_error = ''
    conjunto_success = ''
    conjunto_error = ''
    preview_base = None
    import_resultado = None
    ocr_preview = None
    ocr_lote_preview = None
    ocr_error = ''
    ocr_success = ''
    opcoes_modo_importacao = [
        ('criar_somente', '1) Apenas criar contratos novos'),
        ('complementar_vazios', '2) Complementar campos vazios (padrão)'),
        ('sobrescrever_especificos', '3) Sobrescrever campos específicos selecionados'),
    ]
    campos_sobrescrita = [
        'conjunto', 'data_contrato', 'ocorrencia', 'cod_imovel', 'chave', 'lote', 'sinal',
        'endereco', 'tipoimovel', 'vlfinanc', 'vlprop', 'prestacao_inicial', 'prazo',
        'data_primeiro_venc', 'sa', 'tx_juros', 'cat_prof', 'pr',
        'ident', 'orgao', 'dtnasc', 'numero', 'compl', 'bairro', 'cidade', 'uf', 'cep'
    ]

    if request.method == 'POST':
        acao = request.POST.get('acao', '').strip()

        if acao == 'criar_conjunto':
            codigo = str(request.POST.get('novo_conjunto_codigo', '')).strip()
            nome = str(request.POST.get('novo_conjunto_nome', '')).strip()

            if not codigo or not nome:
                conjunto_error = 'Informe o código e o nome do conjunto.'
            else:
                existente = ConjuntoHabitacional.objects.filter(conjunto=codigo).first()
                if existente:
                    conjunto_error = f'Já existe um conjunto com código {codigo}.'
                else:
                    ConjuntoHabitacional.objects.create(
                        conj=codigo,
                        conjunto=codigo,
                        contrato='0',
                        conjseg=codigo,
                        nome=nome,
                        nomeseg=nome,
                        qtd_mut=0,
                    )
                    conjunto_success = f'Conjunto {codigo} criado com sucesso.'

        elif acao == 'upload_base':
            arquivo_base = request.FILES.get('arquivo_base')
            conjunto_destino = str(request.POST.get('conjunto_destino', '')).strip()

            if not conjunto_destino:
                upload_error = 'Selecione o conjunto de destino antes do upload.'
            elif not ConjuntoHabitacional.objects.filter(conjunto=conjunto_destino).exists():
                upload_error = 'Conjunto de destino inválido. Crie ou selecione um conjunto existente.'
            elif not arquivo_base:
                upload_error = 'Nenhum arquivo foi enviado.'
            else:
                nome_original = Path(arquivo_base.name).name
                nome_limpo = re.sub(r'[^A-Za-z0-9._-]+', '_', nome_original)
                extensao_valida = nome_limpo.lower().endswith(('.xlsx', '.xls', '.csv', '.txt'))

                if not extensao_valida:
                    upload_error = 'Formato inválido. Envie arquivo .xlsx, .xls, .csv ou .txt.'
                else:
                    carimbo = datetime.now().strftime('%Y%m%d_%H%M%S')
                    destino = base_dir / f"{carimbo}_CONJ_{conjunto_destino}_{nome_limpo}"
                    with destino.open('wb') as arquivo_saida:
                        for chunk in arquivo_base.chunks():
                            arquivo_saida.write(chunk)
                    upload_success = f'Arquivo base recebido para o conjunto {conjunto_destino}: {destino.name}'
                    preview_base = _ler_preview_base(destino)
                    if preview_base and preview_base.get('erro'):
                        upload_error = preview_base.get('erro')
                    else:
                        request.session['cadmut_pending_file'] = destino.name
                        request.session['cadmut_pending_conjunto'] = conjunto_destino
                        request.session['cadmut_pending_mapping'] = (preview_base or {}).get('campos_detectados', {})

        elif acao == 'confirmar_gravacao':
            arquivo_pendente = request.session.get('cadmut_pending_file', '')
            conjunto_pendente = request.session.get('cadmut_pending_conjunto', '')
            mapping = request.session.get('cadmut_pending_mapping', {}) or {}
            modo_importacao = request.POST.get('modo_importacao', 'complementar_vazios').strip() or 'complementar_vazios'
            campos_sobrescrever = request.POST.getlist('campos_sobrescrever')

            if not arquivo_pendente or not conjunto_pendente:
                upload_error = 'Nenhuma base pendente para confirmação. Faça o upload novamente.'
            elif modo_importacao == 'sobrescrever_especificos' and not campos_sobrescrever:
                upload_error = 'Selecione ao menos um campo para sobrescrever no modo 3.'
            else:
                caminho = base_dir / arquivo_pendente
                if not caminho.exists():
                    upload_error = 'Arquivo pendente não encontrado. Faça o upload novamente.'
                else:
                    try:
                        _, registros = _ler_linhas_base_completa(caminho)
                        col_codigo = mapping.get('codigo_contrato') or ''

                        if not col_codigo:
                            upload_error = 'Não foi possível detectar a coluna de código do contrato. Ajuste o cabeçalho da base e reenvie.'
                        else:
                            criados = 0
                            atualizados = 0
                            ignorados = 0
                            mutuarios_criados = 0
                            mutuarios_atualizados = 0
                            vinculados = 0
                            erros = []

                            for idx, row in enumerate(registros, start=2):
                                campos_novos = _linha_para_campos_contrato(row, mapping, conjunto_pendente)
                                if not campos_novos:
                                    ignorados += 1
                                    continue

                                codigo = campos_novos['codigo']

                                contrato = Contrato.objects.filter(codigo=codigo).first()
                                if contrato:
                                    alterado = _aplicar_modo_importacao(
                                        contrato=contrato,
                                        campos_novos=campos_novos,
                                        modo_importacao=modo_importacao,
                                        campos_sobrescrever=campos_sobrescrever,
                                    )
                                    if alterado:
                                        contrato.save()
                                        atualizados += 1
                                    else:
                                        ignorados += 1
                                else:
                                    contrato = Contrato.objects.create(
                                        codigo=codigo,
                                        conjunto=campos_novos.get('conjunto', conjunto_pendente),
                                        data_contrato=campos_novos.get('data_contrato'),
                                        ocorrencia=campos_novos.get('ocorrencia', ''),
                                        cod_imovel=campos_novos.get('cod_imovel', ''),
                                        chave=campos_novos.get('chave', ''),
                                        lote=campos_novos.get('lote', ''),
                                        sinal=campos_novos.get('sinal', ''),
                                    )
                                    criados += 1

                                # Criar/atualizar mutuário correspondente para evitar "(sem cadastro)".
                                campos_mut = _linha_para_campos_mutuario(
                                    row=row,
                                    mapping=mapping,
                                    conjunto_destino=conjunto_pendente,
                                    codigo_contrato=codigo,
                                )
                                mutuario = Mutuario.objects.filter(codigo=codigo).first()
                                if not mutuario:
                                    mutuario = Mutuario.objects.create(
                                        codigo=campos_mut['codigo'],
                                        codimovel=campos_mut.get('codimovel', ''),
                                        conjunto=campos_mut.get('conjunto', ''),
                                        conjseg=campos_mut.get('conjseg', ''),
                                        nome=campos_mut.get('nome', ''),
                                        ident='',
                                        orgao='',
                                        dtnasc=None,
                                        cpf=campos_mut.get('cpf', ''),
                                        renda=0,
                                        crenda=0,
                                        endereco=campos_mut.get('endereco', ''),
                                        numero='',
                                        compl='',
                                        tipoimovel='',
                                        bairro='',
                                        cidade=campos_mut.get('cidade', ''),
                                        cep='',
                                        uf=campos_mut.get('uf', ''),
                                    )
                                    mutuarios_criados += 1
                                else:
                                    alterado_mut = _aplicar_modo_mutuario(
                                        mutuario=mutuario,
                                        campos_novos=campos_mut,
                                        modo_importacao=modo_importacao,
                                        campos_sobrescrever=campos_sobrescrever,
                                    )
                                    if alterado_mut:
                                        mutuario.save()
                                        mutuarios_atualizados += 1

                                _vincular_contrato_mutuario(contrato.id, mutuario.id)
                                vinculados += 1

                            import_resultado = {
                                'arquivo': arquivo_pendente,
                                'conjunto': conjunto_pendente,
                                'modo_importacao': modo_importacao,
                                'criados': criados,
                                'atualizados': atualizados,
                                'ignorados': ignorados,
                                'mutuarios_criados': mutuarios_criados,
                                'mutuarios_atualizados': mutuarios_atualizados,
                                'vinculados': vinculados,
                                'erros': erros[:10],
                            }

                            request.session.pop('cadmut_pending_file', None)
                            request.session.pop('cadmut_pending_conjunto', None)
                            request.session.pop('cadmut_pending_mapping', None)
                    except Exception as exc:
                        upload_error = f'Falha ao confirmar gravação: {exc}'

        elif acao == 'ocr_extrair_cadastro':
            arquivo_pdf = request.FILES.get('arquivo_ocr')
            conjunto_ocr = str(request.POST.get('conjunto_ocr', '')).strip()
            codigo_manual_ocr = str(request.POST.get('codigo_manual_ocr', '')).strip()

            if not conjunto_ocr:
                ocr_error = 'Selecione o conjunto de destino do OCR.'
            elif not arquivo_pdf:
                ocr_error = 'Envie um arquivo PDF para OCR.'
            elif not arquivo_pdf.name.lower().endswith('.pdf'):
                ocr_error = 'O OCR CADMUT aceita apenas PDF nesta versão inicial.'
            else:
                temp_path = None
                try:
                    from ocr_contrato_processor import ContratoOCRExtractor
                    from principal.ocr_hibrido import analisar_ocr_hibrido

                    sufixo = Path(arquivo_pdf.name).suffix or '.pdf'
                    fd, temp_str = tempfile.mkstemp(suffix=sufixo, prefix='cadmut_ocr_')
                    os.close(fd)
                    temp_path = Path(temp_str)

                    with open(temp_path, 'wb') as f:
                        for chunk in arquivo_pdf.chunks():
                            f.write(chunk)

                    lote = _extrair_lote_ocr_por_pagina(temp_path)
                    chaves_vistas = {}
                    for item in lote:
                        _enriquecer_item_com_campos_raw_azure(item)
                        item['dados']['conjunto'] = conjunto_ocr
                        for campo_norm in [
                            'nome', 'cpf', 'endereco', 'numero', 'compl', 'bairro', 'cidade',
                            'uf', 'cep', 'tipoimovel', 'vlfinanc', 'vlprop', 'prestacao_inicial',
                            'prazo', 'ident', 'orgao', 'dtnasc', 'data_primeiro_venc',
                            'sa', 'tx_juros', 'cat_prof', 'pr'
                        ]:
                            item['dados'][campo_norm] = _normalizar_campo_ocr(campo_norm, item['dados'].get(campo_norm, ''))
                        if not _valor_para_texto(item['dados'].get('numero', '')):
                            item['dados']['numero'] = _inferir_numero_do_endereco(item['dados'].get('endereco', ''))
                        item['sem_dados'] = _ocr_item_sem_dados(item.get('dados', {}))

                        codigo_item = _valor_para_texto(item['dados'].get('codigo', ''))
                        cpf_item = _limpar_cpf(item['dados'].get('cpf', ''))
                        chave_item = (codigo_item, cpf_item)
                        if codigo_item and cpf_item and chave_item in chaves_vistas:
                            _consolidar_dados_ocr_duplicado(chaves_vistas[chave_item], item)
                            item['duplicado'] = True
                        else:
                            item['duplicado'] = False
                            if codigo_item and cpf_item:
                                chaves_vistas[chave_item] = item

                        if codigo_manual_ocr and len(lote) == 1:
                            item['dados']['codigo'] = codigo_manual_ocr
                            item['codigo'] = codigo_manual_ocr
                            item['codigo_valido'] = _codigo_valido_ocr(codigo_manual_ocr)
                            item['sem_dados'] = False
                            item['duplicado'] = False

                        _aplicar_complemento_manual_em_dados_ocr(item['dados'])

                    validos = [x for x in lote if x['codigo_valido'] and not x.get('duplicado')]
                    sem_dados = [x for x in lote if x.get('sem_dados')]
                    duplicados = [x for x in lote if x.get('duplicado')]
                    invalidos = [x for x in lote if (not x['codigo_valido']) and (not x.get('sem_dados')) and (not x.get('duplicado'))]

                    # Gera CSV de prévia do lote para auditoria
                    preview_dir = Path(__file__).resolve().parents[1] / 'exports' / 'cadmut_bases'
                    preview_dir.mkdir(parents=True, exist_ok=True)
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    preview_csv = preview_dir / f'ocr_lote_preview_{ts}_conj_{conjunto_ocr}.csv'
                    diagnostico_csv = preview_dir / f'ocr_lote_diag_campos_{ts}_conj_{conjunto_ocr}.csv'
                    with preview_csv.open('w', encoding='utf-8', newline='') as fcsv:
                        writer = csv.writer(fcsv, delimiter=';')
                        writer.writerow([
                            'pagina', 'metodo', 'score', 'codigo', 'codigo_provisorio', 'nome', 'cpf',
                            'endereco', 'numero', 'compl', 'bairro', 'cidade', 'uf', 'cep',
                            'codimovel', 'tipoimovel', 'ident', 'orgao', 'dtnasc',
                            'vlfinanc', 'vlprop', 'prestacao_inicial', 'prazo',
                            'data_primeiro_venc', 'sa', 'tx_juros', 'cat_prof', 'pr',
                            'codigo_valido', 'duplicado'
                        ])
                        for item in lote:
                            d = item['dados']
                            writer.writerow([
                                item['pagina'], _valor_para_texto(item.get('metodo', '')),
                                item['score'], _valor_para_texto(d.get('codigo', '')),
                                bool(item.get('codigo_provisorio', False)),
                                _valor_para_texto(d.get('nome', '')), _formatar_cpf(d.get('cpf', '')),
                                _valor_para_texto(d.get('endereco', '')), _valor_para_texto(d.get('numero', '')),
                                _valor_para_texto(d.get('compl', '')), _valor_para_texto(d.get('bairro', '')),
                                _valor_para_texto(d.get('cidade', '')),
                                _valor_para_texto(d.get('uf', '')).upper(), _valor_para_texto(d.get('cep', '')),
                                _valor_para_texto(d.get('codimovel', '')), _valor_para_texto(d.get('tipoimovel', '')),
                                _valor_para_texto(d.get('ident', '')), _valor_para_texto(d.get('orgao', '')),
                                _valor_para_texto(d.get('dtnasc', '')),
                                _formatar_decimal_ocr(d.get('vlfinanc', ''), casas=2),
                                _formatar_decimal_ocr(d.get('vlprop', ''), casas=2),
                                _formatar_decimal_ocr(d.get('prestacao_inicial', ''), casas=6),
                                _valor_para_texto(d.get('prazo', '')),
                                _valor_para_texto(d.get('data_primeiro_venc', '')),
                                _valor_para_texto(d.get('sa', '')),
                                _valor_para_texto(d.get('tx_juros', '')),
                                _valor_para_texto(d.get('cat_prof', '')),
                                _valor_para_texto(d.get('pr', '')),
                                item['codigo_valido'], bool(item.get('duplicado', False))
                            ])

                    with diagnostico_csv.open('w', encoding='utf-8', newline='') as fdiag:
                        writer_diag = csv.writer(fdiag, delimiter=';')
                        writer_diag.writerow(['pagina', 'metodo', 'field_name', 'field_name_norm', 'value', 'confidence'])
                        for item in lote:
                            campos_raw = item.get('azure_fields_raw', []) or []
                            for raw in campos_raw:
                                writer_diag.writerow([
                                    item.get('pagina', ''),
                                    _valor_para_texto(item.get('metodo', '')),
                                    _valor_para_texto(raw.get('field_name', '')),
                                    _valor_para_texto(raw.get('field_name_norm', '')),
                                    _valor_para_texto(raw.get('value', '')),
                                    raw.get('confidence', 0),
                                ])

                    if len(validos) == 1 and len(lote) == 1:
                        dados = validos[0]['dados']
                        ocr_preview = {
                            'arquivo': arquivo_pdf.name,
                            'metodo': validos[0]['metodo'],
                            'score': validos[0]['score'],
                            'codigo': _valor_para_texto(dados.get('codigo', '')),
                            'nome': _valor_para_texto(dados.get('nome', '')),
                            'cpf': _valor_para_texto(dados.get('cpf', '')),
                            'endereco': _valor_para_texto(dados.get('endereco', '')),
                            'numero': _valor_para_texto(dados.get('numero', '')),
                            'compl': _valor_para_texto(dados.get('compl', '')),
                            'bairro': _valor_para_texto(dados.get('bairro', '')),
                            'cidade': _valor_para_texto(dados.get('cidade', '')),
                            'uf': _valor_para_texto(dados.get('uf', '')).upper(),
                            'cep': _valor_para_texto(dados.get('cep', '')),
                            'codimovel': _valor_para_texto(dados.get('codimovel', '')),
                            'tipoimovel': _valor_para_texto(dados.get('tipoimovel', '')),
                            'ident': _valor_para_texto(dados.get('ident', '')),
                            'orgao': _valor_para_texto(dados.get('orgao', '')),
                            'dtnasc': _valor_para_texto(dados.get('dtnasc', '')),
                            'vlfinanc': _formatar_decimal_ocr(dados.get('vlfinanc', ''), casas=2),
                            'vlprop': _formatar_decimal_ocr(dados.get('vlprop', ''), casas=2),
                            'prestacao_inicial': _formatar_decimal_ocr(dados.get('prestacao_inicial', ''), casas=6),
                            'prazo': _valor_para_texto(dados.get('prazo', '')),
                            'data_primeiro_venc': _valor_para_texto(dados.get('data_primeiro_venc', '')),
                            'sa': _valor_para_texto(dados.get('sa', '')),
                            'tx_juros': _valor_para_texto(dados.get('tx_juros', '')),
                            'cat_prof': _valor_para_texto(dados.get('cat_prof', '')),
                            'pr': _valor_para_texto(dados.get('pr', '')),
                            'conjunto': conjunto_ocr,
                            'faltando_criticos': validos[0].get('faltando_criticos', []),
                        }

                        request.session['cadmut_pending_ocr'] = {
                            'dados': {
                                'codigo': ocr_preview['codigo'],
                                'nome': ocr_preview['nome'],
                                'cpf': ocr_preview['cpf'],
                                'endereco': ocr_preview['endereco'],
                                'numero': ocr_preview['numero'],
                                'compl': ocr_preview['compl'],
                                'bairro': ocr_preview['bairro'],
                                'cidade': ocr_preview['cidade'],
                                'uf': ocr_preview['uf'],
                                'cep': ocr_preview['cep'],
                                'codimovel': ocr_preview['codimovel'],
                                'tipoimovel': ocr_preview['tipoimovel'],
                                'ident': ocr_preview['ident'],
                                'orgao': ocr_preview['orgao'],
                                'dtnasc': ocr_preview['dtnasc'],
                                'vlfinanc': ocr_preview['vlfinanc'],
                                'vlprop': ocr_preview['vlprop'],
                                'prestacao_inicial': ocr_preview['prestacao_inicial'],
                                'prazo': ocr_preview['prazo'],
                                'data_primeiro_venc': ocr_preview['data_primeiro_venc'],
                                'sa': ocr_preview['sa'],
                                'tx_juros': ocr_preview['tx_juros'],
                                'cat_prof': ocr_preview['cat_prof'],
                                'pr': ocr_preview['pr'],
                                'conjunto': ocr_preview['conjunto'],
                            },
                            'arquivo': ocr_preview['arquivo'],
                            'score': ocr_preview['score'],
                        }
                    else:
                        ocr_lote_preview = {
                            'arquivo': arquivo_pdf.name,
                            'conjunto': conjunto_ocr,
                            'total_paginas': len(lote),
                            'total_validos': len(validos),
                            'total_invalidos': len(invalidos),
                            'total_sem_dados': len(sem_dados),
                            'total_duplicados': len(duplicados),
                            'total_provisorios': sum(1 for x in lote if x.get('codigo_provisorio')),
                            'metodos_usados': sorted({
                                _valor_para_texto(x.get('metodo', 'desconhecido')) for x in lote
                            }),
                            'preview_csv': preview_csv.name,
                            'diagnostico_csv': diagnostico_csv.name,
                            'itens': [
                                {
                                    'pagina': item['pagina'],
                                    'metodo': _valor_para_texto(item.get('metodo', '')),
                                    'score': item['score'],
                                    'codigo': _valor_para_texto(item['dados'].get('codigo', '')),
                                    'nome': _valor_para_texto(item['dados'].get('nome', '')),
                                    'cpf': _formatar_cpf(item['dados'].get('cpf', '')),
                                    'endereco': _valor_para_texto(item['dados'].get('endereco', '')),
                                    'numero': _valor_para_texto(item['dados'].get('numero', '')),
                                    'bairro': _valor_para_texto(item['dados'].get('bairro', '')),
                                    'cidade': _valor_para_texto(item['dados'].get('cidade', '')),
                                    'uf': _valor_para_texto(item['dados'].get('uf', '')).upper(),
                                    'vlfinanc': _formatar_decimal_ocr(item['dados'].get('vlfinanc', ''), casas=2),
                                    'prestacao_inicial': _formatar_decimal_ocr(item['dados'].get('prestacao_inicial', ''), casas=6),
                                    'codigo_valido': item['codigo_valido'],
                                    'codigo_provisorio': bool(item.get('codigo_provisorio', False)),
                                    'sem_dados': bool(item.get('sem_dados', False)),
                                    'duplicado': bool(item.get('duplicado', False)),
                                }
                                for item in lote[:80]
                            ],
                        }
                        request.session['cadmut_pending_ocr_lote'] = {
                            'arquivo': arquivo_pdf.name,
                            'conjunto': conjunto_ocr,
                            'itens': [
                                {
                                    'codigo': _valor_para_texto(item['dados'].get('codigo', '')),
                                    'nome': _valor_para_texto(item['dados'].get('nome', '')),
                                    'cpf': _formatar_cpf(item['dados'].get('cpf', '')),
                                    'endereco': _valor_para_texto(item['dados'].get('endereco', '')),
                                    'numero': _valor_para_texto(item['dados'].get('numero', '')),
                                    'compl': _valor_para_texto(item['dados'].get('compl', '')),
                                    'bairro': _valor_para_texto(item['dados'].get('bairro', '')),
                                    'cidade': _valor_para_texto(item['dados'].get('cidade', '')),
                                    'uf': _valor_para_texto(item['dados'].get('uf', '')).upper(),
                                    'cep': _valor_para_texto(item['dados'].get('cep', '')),
                                    'codimovel': _valor_para_texto(item['dados'].get('codimovel', '')),
                                    'tipoimovel': _valor_para_texto(item['dados'].get('tipoimovel', '')),
                                    'ident': _valor_para_texto(item['dados'].get('ident', '')),
                                    'orgao': _valor_para_texto(item['dados'].get('orgao', '')),
                                    'dtnasc': _valor_para_texto(item['dados'].get('dtnasc', '')),
                                    'vlfinanc': _formatar_decimal_ocr(item['dados'].get('vlfinanc', ''), casas=2),
                                    'vlprop': _formatar_decimal_ocr(item['dados'].get('vlprop', ''), casas=2),
                                    'prestacao_inicial': _formatar_decimal_ocr(item['dados'].get('prestacao_inicial', ''), casas=6),
                                    'prazo': _valor_para_texto(item['dados'].get('prazo', '')),
                                    'data_primeiro_venc': _valor_para_texto(item['dados'].get('data_primeiro_venc', '')),
                                    'sa': _valor_para_texto(item['dados'].get('sa', '')),
                                    'tx_juros': _valor_para_texto(item['dados'].get('tx_juros', '')),
                                    'cat_prof': _valor_para_texto(item['dados'].get('cat_prof', '')),
                                    'pr': _valor_para_texto(item['dados'].get('pr', '')),
                                    'conjunto': conjunto_ocr,
                                    'codigo_valido': item['codigo_valido'],
                                    'metodo': _valor_para_texto(item.get('metodo', '')),
                                    'sem_dados': bool(item.get('sem_dados', False)),
                                    'duplicado': bool(item.get('duplicado', False)),
                                }
                                for item in lote
                            ],
                            'preview_csv': preview_csv.name,
                            'diagnostico_csv': diagnostico_csv.name,
                        }
                        ocr_success = (
                            f'OCR em lote concluído: {len(validos)} contrato(s) válidos em {len(lote)} página(s). '
                            f'Prévia CSV gerada: {preview_csv.name}'
                        )
                except Exception as exc:
                    ocr_error = f'Falha no OCR: {exc}'
                finally:
                    if temp_path and temp_path.exists():
                        try:
                            temp_path.unlink()
                        except Exception:
                            pass

        elif acao == 'ocr_salvar_complemento_manual':
            codigo_manual = re.sub(r'\s+', '', _valor_para_texto(request.POST.get('codigo_manual_complemento', '')))
            if not codigo_manual:
                ocr_error = 'Informe um código para salvar o complemento manual.'
            else:
                dados_override = {
                    'data_primeiro_venc': _valor_para_texto(request.POST.get('manual_data_primeiro_venc', '')),
                    'sa': _normalizar_campo_ocr('sa', request.POST.get('manual_sa', '')),
                    'tx_juros': _normalizar_campo_ocr('tx_juros', request.POST.get('manual_tx_juros', '')),
                    'cat_prof': _normalizar_campo_ocr('cat_prof', request.POST.get('manual_cat_prof', '')),
                    'pr': _normalizar_campo_ocr('pr', request.POST.get('manual_pr', '')),
                    'ident': _normalizar_campo_ocr('ident', request.POST.get('manual_ident', '')),
                    'orgao': _normalizar_campo_ocr('orgao', request.POST.get('manual_orgao', '')),
                    'dtnasc': _valor_para_texto(request.POST.get('manual_dtnasc', '')),
                    'updated_at': datetime.now().isoformat(timespec='seconds'),
                }

                data_1_venc = _parse_data_flexivel(dados_override.get('data_primeiro_venc'))
                if data_1_venc is not None:
                    dados_override['data_primeiro_venc'] = data_1_venc.strftime('%Y-%m-%d')
                else:
                    dados_override['data_primeiro_venc'] = ''

                dt_nasc = _parse_data_flexivel(dados_override.get('dtnasc'))
                if dt_nasc is not None:
                    dados_override['dtnasc'] = dt_nasc.strftime('%Y-%m-%d')
                else:
                    dados_override['dtnasc'] = ''

                if not any(_valor_para_texto(dados_override.get(c, '')) for c in ['data_primeiro_venc', 'sa', 'tx_juros', 'cat_prof', 'pr', 'ident', 'orgao', 'dtnasc']):
                    ocr_error = 'Preencha ao menos um campo para salvar o complemento manual.'
                else:
                    overrides = _carregar_complementos_manuais_ocr()
                    overrides[codigo_manual] = dados_override
                    _salvar_complementos_manuais_ocr(overrides)

                    pending_ocr = request.session.get('cadmut_pending_ocr', {}) or {}
                    dados_ocr = pending_ocr.get('dados', {}) or {}
                    if re.sub(r'\s+', '', _valor_para_texto(dados_ocr.get('codigo', ''))) == codigo_manual:
                        _aplicar_complemento_manual_em_dados_ocr(dados_ocr)
                        pending_ocr['dados'] = dados_ocr
                        request.session['cadmut_pending_ocr'] = pending_ocr

                    ocr_success = f'Complemento manual salvo para o contrato {codigo_manual}.'

        elif acao == 'ocr_confirmar_cadastro':
            pending_ocr = request.session.get('cadmut_pending_ocr', {}) or {}
            dados_ocr = pending_ocr.get('dados', {}) or {}
            _aplicar_complemento_manual_em_dados_ocr(dados_ocr)
            modo_importacao = request.POST.get('modo_importacao_ocr', 'complementar_vazios').strip() or 'complementar_vazios'
            campos_sobrescrever = request.POST.getlist('campos_sobrescrever_ocr')

            if not dados_ocr or not dados_ocr.get('codigo'):
                ocr_error = 'Nenhum OCR pendente para confirmar.'
            elif modo_importacao == 'sobrescrever_especificos' and not campos_sobrescrever:
                ocr_error = 'Selecione ao menos um campo para sobrescrever no modo 3.'
            else:
                try:
                    conjunto_destino = _valor_para_texto(dados_ocr.get('conjunto', ''))
                    campos_contrato = _campos_contrato_por_dados_ocr(dados_ocr, conjunto_destino)
                    if not campos_contrato:
                        ocr_error = 'OCR sem código de contrato válido para gravação.'
                    else:
                        codigo = campos_contrato['codigo']
                        contrato = Contrato.objects.filter(codigo=codigo).first()
                        if contrato:
                            alterado = _aplicar_modo_importacao(
                                contrato=contrato,
                                campos_novos=campos_contrato,
                                modo_importacao=modo_importacao,
                                campos_sobrescrever=campos_sobrescrever,
                            )
                            if alterado:
                                contrato.save()
                        else:
                            contrato = Contrato.objects.create(
                                codigo=codigo,
                                conjunto=campos_contrato.get('conjunto', ''),
                                data_contrato=campos_contrato.get('data_contrato'),
                                ocorrencia=campos_contrato.get('ocorrencia', ''),
                                cod_imovel=campos_contrato.get('cod_imovel', ''),
                                chave=campos_contrato.get('chave', ''),
                                lote=campos_contrato.get('lote', ''),
                                sinal=campos_contrato.get('sinal', ''),
                                vlfinanc=campos_contrato.get('vlfinanc'),
                                vlprop=campos_contrato.get('vlprop'),
                                prestacao_inicial=campos_contrato.get('prestacao_inicial'),
                                prazo=campos_contrato.get('prazo'),
                                data_primeiro_venc=campos_contrato.get('data_primeiro_venc'),
                                sa=campos_contrato.get('sa', ''),
                                tx_juros=campos_contrato.get('tx_juros'),
                                cat_prof=campos_contrato.get('cat_prof', ''),
                                pr=campos_contrato.get('pr', ''),
                            )

                        campos_mut = _campos_mutuario_por_dados_ocr(dados_ocr, codigo, conjunto_destino)
                        mutuario = Mutuario.objects.filter(codigo=codigo).first()
                        if not mutuario:
                            mutuario = Mutuario.objects.create(
                                codigo=campos_mut['codigo'],
                                codimovel=campos_mut.get('codimovel', ''),
                                conjunto=campos_mut.get('conjunto', ''),
                                conjseg=campos_mut.get('conjseg', ''),
                                nome=campos_mut.get('nome', ''),
                                ident=campos_mut.get('ident', ''),
                                orgao=campos_mut.get('orgao', ''),
                                dtnasc=campos_mut.get('dtnasc'),
                                cpf=campos_mut.get('cpf', ''),
                                renda=0,
                                crenda=0,
                                endereco=campos_mut.get('endereco', ''),
                                numero=campos_mut.get('numero', ''),
                                compl=campos_mut.get('compl', ''),
                                tipoimovel=campos_mut.get('tipoimovel', ''),
                                bairro=campos_mut.get('bairro', ''),
                                cidade=campos_mut.get('cidade', ''),
                                cep=campos_mut.get('cep', ''),
                                uf=campos_mut.get('uf', ''),
                            )
                        else:
                            alterado_mut = _aplicar_modo_mutuario(
                                mutuario=mutuario,
                                campos_novos=campos_mut,
                                modo_importacao=modo_importacao,
                                campos_sobrescrever=campos_sobrescrever,
                            )
                            if alterado_mut:
                                mutuario.save()

                        _vincular_contrato_mutuario(contrato.id, mutuario.id)
                        ocr_success = f'OCR aplicado com sucesso no contrato {codigo}.'
                        request.session.pop('cadmut_pending_ocr', None)
                except Exception as exc:
                    ocr_error = f'Falha ao confirmar OCR: {exc}'

        elif acao == 'ocr_confirmar_lote':
            pending_lote = request.session.get('cadmut_pending_ocr_lote', {}) or {}
            itens = pending_lote.get('itens', []) or []
            modo_importacao = request.POST.get('modo_importacao_ocr_lote', 'complementar_vazios').strip() or 'complementar_vazios'
            campos_sobrescrever = request.POST.getlist('campos_sobrescrever_ocr_lote')

            if not itens:
                ocr_error = 'Nenhum lote OCR pendente para confirmar.'
            elif modo_importacao == 'sobrescrever_especificos' and not campos_sobrescrever:
                ocr_error = 'Selecione ao menos um campo para sobrescrever no modo 3.'
            else:
                criados = 0
                atualizados = 0
                ignorados = 0
                vinculados = 0
                for d in itens:
                    _aplicar_complemento_manual_em_dados_ocr(d)
                    if d.get('duplicado'):
                        ignorados += 1
                        continue

                    if not d.get('codigo_valido'):
                        ignorados += 1
                        continue

                    conjunto_destino = _valor_para_texto(d.get('conjunto', ''))
                    campos_contrato = _campos_contrato_por_dados_ocr(d, conjunto_destino)
                    if not campos_contrato:
                        ignorados += 1
                        continue

                    codigo = campos_contrato['codigo']
                    contrato = Contrato.objects.filter(codigo=codigo).first()
                    if contrato:
                        alterado = _aplicar_modo_importacao(
                            contrato=contrato,
                            campos_novos=campos_contrato,
                            modo_importacao=modo_importacao,
                            campos_sobrescrever=campos_sobrescrever,
                        )
                        if alterado:
                            contrato.save()
                            atualizados += 1
                        else:
                            ignorados += 1
                    else:
                        contrato = Contrato.objects.create(
                            codigo=codigo,
                            conjunto=campos_contrato.get('conjunto', ''),
                            data_contrato=campos_contrato.get('data_contrato'),
                            ocorrencia=campos_contrato.get('ocorrencia', ''),
                            cod_imovel=campos_contrato.get('cod_imovel', ''),
                            chave=campos_contrato.get('chave', ''),
                            lote=campos_contrato.get('lote', ''),
                            sinal=campos_contrato.get('sinal', ''),
                            vlfinanc=campos_contrato.get('vlfinanc'),
                            vlprop=campos_contrato.get('vlprop'),
                            prestacao_inicial=campos_contrato.get('prestacao_inicial'),
                            prazo=campos_contrato.get('prazo'),
                            data_primeiro_venc=campos_contrato.get('data_primeiro_venc'),
                            sa=campos_contrato.get('sa', ''),
                            tx_juros=campos_contrato.get('tx_juros'),
                            cat_prof=campos_contrato.get('cat_prof', ''),
                            pr=campos_contrato.get('pr', ''),
                        )
                        criados += 1

                    campos_mut = _campos_mutuario_por_dados_ocr(d, codigo, conjunto_destino)
                    mutuario = Mutuario.objects.filter(codigo=codigo).first()
                    if not mutuario:
                        mutuario = Mutuario.objects.create(
                            codigo=campos_mut['codigo'],
                            codimovel=campos_mut.get('codimovel', ''),
                            conjunto=campos_mut.get('conjunto', ''),
                            conjseg=campos_mut.get('conjseg', ''),
                            nome=campos_mut.get('nome', ''),
                            ident=campos_mut.get('ident', ''),
                            orgao=campos_mut.get('orgao', ''),
                            dtnasc=campos_mut.get('dtnasc'),
                            cpf=campos_mut.get('cpf', ''),
                            renda=0,
                            crenda=0,
                            endereco=campos_mut.get('endereco', ''),
                            numero=campos_mut.get('numero', ''),
                            compl=campos_mut.get('compl', ''),
                            tipoimovel=campos_mut.get('tipoimovel', ''),
                            bairro=campos_mut.get('bairro', ''),
                            cidade=campos_mut.get('cidade', ''),
                            cep=campos_mut.get('cep', ''),
                            uf=campos_mut.get('uf', ''),
                        )
                    else:
                        alterado_mut = _aplicar_modo_mutuario(
                            mutuario=mutuario,
                            campos_novos=campos_mut,
                            modo_importacao=modo_importacao,
                            campos_sobrescrever=campos_sobrescrever,
                        )
                        if alterado_mut:
                            mutuario.save()

                    _vincular_contrato_mutuario(contrato.id, mutuario.id)
                    vinculados += 1

                ocr_success = (
                    f'Lote OCR aplicado: criados={criados}, atualizados={atualizados}, '
                    f'ignorados={ignorados}, vinculados={vinculados}.'
                )
                request.session.pop('cadmut_pending_ocr_lote', None)

    if preview_base is None:
        arquivo_pendente = request.session.get('cadmut_pending_file', '')
        if arquivo_pendente:
            caminho_pendente = base_dir / arquivo_pendente
            if caminho_pendente.exists():
                preview_base = _ler_preview_base(caminho_pendente)

    if ocr_preview is None:
        pending_ocr = request.session.get('cadmut_pending_ocr', {}) or {}
        if pending_ocr:
            d = pending_ocr.get('dados', {}) or {}
            ocr_preview = {
                'arquivo': pending_ocr.get('arquivo', ''),
                'metodo': 'pending',
                'score': pending_ocr.get('score', 0),
                'codigo': _valor_para_texto(d.get('codigo', '')),
                'nome': _valor_para_texto(d.get('nome', '')),
                'cpf': _valor_para_texto(d.get('cpf', '')),
                'endereco': _valor_para_texto(d.get('endereco', '')),
                'numero': _valor_para_texto(d.get('numero', '')),
                'compl': _valor_para_texto(d.get('compl', '')),
                'bairro': _valor_para_texto(d.get('bairro', '')),
                'cidade': _valor_para_texto(d.get('cidade', '')),
                'uf': _valor_para_texto(d.get('uf', '')),
                'cep': _valor_para_texto(d.get('cep', '')),
                'codimovel': _valor_para_texto(d.get('codimovel', '')),
                'tipoimovel': _valor_para_texto(d.get('tipoimovel', '')),
                'ident': _valor_para_texto(d.get('ident', '')),
                'orgao': _valor_para_texto(d.get('orgao', '')),
                'dtnasc': _valor_para_texto(d.get('dtnasc', '')),
                'vlfinanc': _formatar_decimal_ocr(d.get('vlfinanc', ''), casas=2),
                'vlprop': _formatar_decimal_ocr(d.get('vlprop', ''), casas=2),
                'prestacao_inicial': _formatar_decimal_ocr(d.get('prestacao_inicial', ''), casas=6),
                'prazo': _valor_para_texto(d.get('prazo', '')),
                'data_primeiro_venc': _valor_para_texto(d.get('data_primeiro_venc', '')),
                'sa': _valor_para_texto(d.get('sa', '')),
                'tx_juros': _valor_para_texto(d.get('tx_juros', '')),
                'cat_prof': _valor_para_texto(d.get('cat_prof', '')),
                'pr': _valor_para_texto(d.get('pr', '')),
                'conjunto': _valor_para_texto(d.get('conjunto', '')),
                'faltando_criticos': [],
            }

    ocr_manual_override = {}
    if ocr_preview:
        codigo_preview = re.sub(r'\s+', '', _valor_para_texto(ocr_preview.get('codigo', '')))
        if codigo_preview:
            ocr_manual_override = _carregar_complementos_manuais_ocr().get(codigo_preview, {}) or {}

    if ocr_lote_preview is None:
        pending_lote = request.session.get('cadmut_pending_ocr_lote', {}) or {}
        if pending_lote:
            itens = pending_lote.get('itens', [])
            ocr_lote_preview = {
                'arquivo': pending_lote.get('arquivo', ''),
                'conjunto': pending_lote.get('conjunto', ''),
                'total_paginas': len(itens),
                'total_validos': sum(1 for x in itens if x.get('codigo_valido') and not x.get('duplicado')),
                'total_invalidos': sum(1 for x in itens if (not x.get('codigo_valido')) and (not x.get('sem_dados'))),
                'total_sem_dados': sum(1 for x in itens if x.get('sem_dados')),
                'total_duplicados': sum(1 for x in itens if x.get('duplicado')),
                'metodos_usados': sorted({
                    _valor_para_texto(x.get('metodo', 'desconhecido')) for x in itens
                }),
                'preview_csv': pending_lote.get('preview_csv', ''),
                'diagnostico_csv': pending_lote.get('diagnostico_csv', ''),
                'itens': [
                    {
                        'pagina': i + 1,
                        'metodo': _valor_para_texto(x.get('metodo', '')),
                        'score': '-',
                        'codigo': _valor_para_texto(x.get('codigo', '')),
                        'nome': _valor_para_texto(x.get('nome', '')),
                        'cpf': _formatar_cpf(x.get('cpf', '')),
                        'endereco': _valor_para_texto(x.get('endereco', '')),
                        'numero': _valor_para_texto(x.get('numero', '')),
                        'bairro': _valor_para_texto(x.get('bairro', '')),
                        'cidade': _valor_para_texto(x.get('cidade', '')),
                        'uf': _valor_para_texto(x.get('uf', '')).upper(),
                        'vlfinanc': _formatar_decimal_ocr(x.get('vlfinanc', ''), casas=2),
                        'prestacao_inicial': _formatar_decimal_ocr(x.get('prestacao_inicial', ''), casas=6),
                        'codigo_valido': x.get('codigo_valido', False),
                        'sem_dados': x.get('sem_dados', False),
                        'duplicado': x.get('duplicado', False),
                    }
                    for i, x in enumerate(itens[:80])
                ],
            }

    contratos_com_conjunto = Contrato.objects.exclude(conjunto__isnull=True).exclude(conjunto='').count()
    contratos_sem_conjunto = Contrato.objects.filter(Q(conjunto__isnull=True) | Q(conjunto='')).count()

    arquivos_base = []
    for f in sorted(base_dir.glob('*'), key=lambda p: p.stat().st_mtime, reverse=True):
        if f.is_file():
            arquivos_base.append({
                'nome': f.name,
                'tamanho_kb': round(f.stat().st_size / 1024, 1),
                'modificado_em': datetime.fromtimestamp(f.stat().st_mtime).strftime('%d/%m/%Y %H:%M:%S')
            })

    context = {
        'total_conjuntos': ConjuntoHabitacional.objects.count(),
        'total_mutuarios': Mutuario.objects.count(),
        'total_contratos': Contrato.objects.count(),
        'contratos_com_conjunto': contratos_com_conjunto,
        'contratos_sem_conjunto': contratos_sem_conjunto,
        'total_parcelas': ParcelaContrato.objects.count(),
        'upload_success': upload_success,
        'upload_error': upload_error,
        'conjunto_success': conjunto_success,
        'conjunto_error': conjunto_error,
        'conjuntos_cadmut': ConjuntoHabitacional.objects.all().order_by('conjunto'),
        'arquivos_base': arquivos_base[:20],
        'preview_base': preview_base,
        'import_resultado': import_resultado,
        'tem_pendente_confirmacao': bool(request.session.get('cadmut_pending_file', '')),
        'opcoes_modo_importacao': opcoes_modo_importacao,
        'campos_sobrescrita': campos_sobrescrita,
        'ocr_preview': ocr_preview,
        'ocr_lote_preview': ocr_lote_preview,
        'ocr_error': ocr_error,
        'ocr_success': ocr_success,
        'tem_pendente_ocr': bool(request.session.get('cadmut_pending_ocr')),
        'tem_pendente_ocr_lote': bool(request.session.get('cadmut_pending_ocr_lote')),
        'ocr_manual_override': ocr_manual_override,
        'arquivo_overrides_ocr': _arquivo_complementos_manuais_ocr().name,
    }
    return render(request, 'principal/cadmut.html', context)

def clientes(request):
    lista_clientes = Cliente.objects.all()
    return render(request, 'principal/clientes.html', {'clientes': lista_clientes})

def conjuntos(request):
    conjuntos_list = []
    
    for conj in ConjuntoHabitacional.objects.all().order_by('conjunto'):
        # Contar contratos pelo campo conjunto (string)
        qtd_contratos = Contrato.objects.filter(conjunto=conj.conjunto).count()
        
        conjuntos_list.append({
            'conjunto': conj.conjunto,
            'nome': conj.nome,
            'contrato': conj.contrato,
            'qtd_contratos': qtd_contratos,
            'qtd_mutuarios': qtd_contratos  # Sempre igual ao número de contratos (1 para 1)
        })
    
    return render(request, 'principal/conjuntos.html', {'conjuntos': conjuntos_list})

def mutuarios(request):
    busca = request.GET.get('q', '')
    busca_nome = request.GET.get('nome', '')
    busca_codigo = request.GET.get('codigo', '')
    busca_cpf = request.GET.get('cpf', '')
    pagina = request.GET.get('pagina', 1)
    
    # Buscar contratos (não mutuários diretamente)
    contratos_qs = Contrato.objects.all()
    
    # Aplicar filtros no conjunto se fornecido
    if busca_codigo:
        conjunto_normalizado = busca_codigo.strip().zfill(3)
        contratos_qs = contratos_qs.filter(conjunto=conjunto_normalizado)
    
    # Buscar IDs dos contratos filtrados
    contratos_ids = list(contratos_qs.values_list('id', flat=True))
    
    if not contratos_ids:
        # Sem contratos, retornar vazio
        qs = Mutuario.objects.none()
    else:
        # Buscar mutuários vinculados aos contratos filtrados
        conn = sqlite3.connect(r'C:\Users\fabri\cofluhab\cofluhab\db.sqlite3')
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT mutuario_id 
            FROM contrato_mutuario_map 
            WHERE contrato_id IN ({})
        """.format(','.join('?' * len(contratos_ids))), contratos_ids)
        mutuario_ids_com_contrato = [row[0] for row in cur.fetchall()]
        conn.close()
        
        qs = Mutuario.objects.filter(id__in=mutuario_ids_com_contrato).select_related('conjunto_fk', 'endereco_fk')
    
    # Filtro por nome (já filtrou contratos por conjunto acima)
    if busca_nome:
        qs = qs.filter(nome__icontains=busca_nome)
    
    # Filtro por CPF
    if busca_cpf:
        qs = qs.filter(cpf__icontains=busca_cpf)
    
    # Busca geral (nome ou CPF) - conjunto já filtrado
    if busca:
        qs = qs.filter(Q(nome__icontains=busca) | Q(cpf__icontains=busca))
    
    # Ordenar e contar total antes da paginação
    qs = qs.order_by('nome')
    
    # Contar contratos correspondentes (não mutuários)
    if busca_codigo:
        total_mutuarios = contratos_qs.count()  # Mostrar total de contratos do conjunto
    else:
        total_mutuarios = qs.count()  # Mostrar total de mutuários únicos
    
    # Paginação: 200 mutuários por página
    paginator = Paginator(qs, 200)
    page_obj = paginator.get_page(pagina)
    
    return render(request, 'principal/mutuarios.html', {
        'mutuarios': page_obj, 
        'page_obj': page_obj,
        'busca': busca,
        'busca_nome': busca_nome,
        'busca_codigo': busca_codigo,
        'busca_cpf': busca_cpf,
        'total': Mutuario.objects.count(),
        'total_filtrado': total_mutuarios
    })

def mutuario_detail(request, codigo):
    mutuario = get_object_or_404(Mutuario, codigo=codigo)
    
    # Buscar contratos vinculados via mapeamento
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
    contratos_ids = []
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT contrato_id FROM contrato_mutuario_map WHERE mutuario_id = ?", (mutuario.id,))
        contratos_ids = [row[0] for row in cur.fetchall()]
        conn.close()
    except Exception as e:
        print(f"Erro ao buscar contratos para mutuário {codigo}: {e}")
    
    contratos = Contrato.objects.filter(id__in=contratos_ids) if contratos_ids else []
    
    # Buscar parcelas dos contratos vinculados (últimas 50)
    parcelas_recentes = []
    if contratos:
        parcelas_recentes = ParcelaContrato.objects.filter(
            contrato__in=contratos
        ).select_related('contrato').order_by('-dtvenc')[:50]
    
    return render(request, 'principal/mutuario_detail.html', {
        'mutuario': mutuario, 
        'contratos': contratos,
        'parcelas_recentes': parcelas_recentes
    })

def mutuario_editar(request, codigo):
    from django.shortcuts import redirect
    from django.contrib import messages
    
    mutuario = get_object_or_404(Mutuario, codigo=codigo)
    
    if request.method == 'POST':
        # Atualizar campos
        mutuario.nome = request.POST.get('nome', mutuario.nome)
        mutuario.cpf = request.POST.get('cpf', mutuario.cpf)
        mutuario.ident = request.POST.get('ident', mutuario.ident)
        mutuario.orgao = request.POST.get('orgao', mutuario.orgao)
        
        # Data de nascimento
        dtnasc = request.POST.get('dtnasc', '')
        if dtnasc:
            try:
                mutuario.dtnasc = datetime.strptime(dtnasc, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Data de nascimento inválida.')
        
        mutuario.renda = request.POST.get('renda', mutuario.renda) or None
        mutuario.crenda = request.POST.get('crenda', mutuario.crenda) or None
        mutuario.endereco = request.POST.get('endereco', mutuario.endereco)
        mutuario.numero = request.POST.get('numero', mutuario.numero)
        mutuario.compl = request.POST.get('compl', mutuario.compl)
        mutuario.bairro = request.POST.get('bairro', mutuario.bairro)
        mutuario.cidade = request.POST.get('cidade', mutuario.cidade)
        mutuario.uf = request.POST.get('uf', mutuario.uf)
        mutuario.cep = request.POST.get('cep', mutuario.cep)
        mutuario.tipoimovel = request.POST.get('tipoimovel', mutuario.tipoimovel)
        
        mutuario.save()
        messages.success(request, 'Mutuário atualizado com sucesso.')
        
        return redirect('mutuario_detail', codigo=codigo)
    
    return render(request, 'principal/mutuario_editar.html', {'mutuario': mutuario})

def enderecos(request):
    conjunto_filtro = request.GET.get('conjunto', '')
    pagina = request.GET.get('pagina', 1)
    
    if conjunto_filtro:
        import sqlite3
        
        # Buscar contratos do conjunto
        contratos = Contrato.objects.filter(conjunto=conjunto_filtro).order_by('codigo')
        contratos_ids = [c.id for c in contratos]
        
        # Buscar mapeamento contrato -> mutuário
        conn = sqlite3.connect(r'C:\Users\fabri\cofluhab\cofluhab\db.sqlite3')
        cur = conn.cursor()
        
        cur.execute("""
            SELECT contrato_id, mutuario_id 
            FROM contrato_mutuario_map 
            WHERE contrato_id IN ({})
        """.format(','.join('?' * len(contratos_ids))), contratos_ids)
        
        mapeamento = {}
        for contrato_id, mutuario_id in cur.fetchall():
            if contrato_id not in mapeamento:
                mapeamento[contrato_id] = []
            mapeamento[contrato_id].append(mutuario_id)
        
        conn.close()
        
        # Buscar todos os mutuários necessários
        mutuario_ids_unicos = []
        for ids in mapeamento.values():
            if ids:
                mutuario_ids_unicos.append(ids[0])  # Pegar apenas o primeiro
        
        mutuarios_dict = {}
        if mutuario_ids_unicos:
            mutuarios = Mutuario.objects.filter(id__in=mutuario_ids_unicos).select_related('endereco_fk')
            mutuarios_dict = {m.id: m for m in mutuarios}
        
        # Montar lista de endereços com contratos
        enderecos_list = []
        for contrato in contratos:
            mutuario_ids_contrato = mapeamento.get(contrato.id, [])
            mutuario = mutuarios_dict.get(mutuario_ids_contrato[0]) if mutuario_ids_contrato else None
            
            if mutuario:
                # Buscar endereço do FK
                if mutuario.endereco_fk:
                    endereco = mutuario.endereco_fk.endereco
                    numero = mutuario.endereco_fk.numero
                    compl = mutuario.endereco_fk.compl
                    bairro = mutuario.endereco_fk.bairro
                    cidade = mutuario.endereco_fk.cidade
                    uf = mutuario.endereco_fk.uf
                    cep = mutuario.endereco_fk.cep
                else:
                    endereco = numero = compl = bairro = cidade = uf = cep = ''
                
                # Extrair número para ordenação numérica
                try:
                    numero_int = int(''.join(filter(str.isdigit, numero))) if numero else 0
                except ValueError:
                    numero_int = 0
                
                enderecos_list.append({
                    'mutuario': mutuario.nome,
                    'codigo_mutuario': mutuario.codigo,
                    'endereco': endereco,
                    'numero': numero,
                    'numero_int': numero_int,
                    'compl': compl,
                    'bairro': bairro,
                    'cidade': cidade,
                    'uf': uf,
                    'cep': cep,
                    'cod_imovel': mutuario.codimovel,
                    'contrato': contrato.codigo,
                })
        
        # Ordenar por endereço (alfabético) e depois por número (numérico)
        enderecos_list.sort(key=lambda x: (x['endereco'] or '', x['numero_int'], x['compl'] or ''))
        
        # Paginar resultados
        paginator = Paginator(enderecos_list, 200)
        page_obj = paginator.get_page(pagina)
        
        return render(request, 'principal/enderecos.html', {
            'page_obj': page_obj,
            'conjunto_selecionado': conjunto_filtro,
            'total_enderecos': len(enderecos_list)
        })
    else:
        # Listar conjuntos
        conjuntos_list = []
        for conj in ConjuntoHabitacional.objects.all().order_by('conjunto'):
            qtd_contratos = Contrato.objects.filter(conjunto=conj.conjunto).count()
            conjuntos_list.append({
                'conjunto': conj.conjunto,
                'nome': conj.nome,
                'qtd_contratos': qtd_contratos,
                'qtd_mutuarios': qtd_contratos  # Sempre igual ao número de contratos
            })
        
        return render(request, 'principal/enderecos.html', {
            'conjuntos': conjuntos_list
        })

def movimentacoes(request):
    busca = request.GET.get('q', '')
    qs = Movimentacao.objects.all()
    if busca:
        qs = qs.filter(Q(codigo__icontains=busca) | Q(tipo__icontains=busca) | Q(descricao__icontains=busca))
    qs = qs.order_by('-data')[:200]
    return render(request, 'principal/movimentacoes.html', {'movimentacoes': qs, 'busca': busca, 'total': Movimentacao.objects.count()})

def contratos(request):
    busca = request.GET.get('q', '')
    busca_codigo = request.GET.get('codigo', '')
    busca_conjunto = request.GET.get('conjunto', '')
    pagina = request.GET.get('pagina', 1)
    
    qs = Contrato.objects.all()
    
    # Filtro por código do contrato
    if busca_codigo:
        qs = qs.filter(codigo__icontains=busca_codigo)
    
    # Filtro por conjunto
    if busca_conjunto:
        qs = qs.filter(conjunto__icontains=busca_conjunto)
    
    # Busca geral (código, conjunto ou nome do mutuário)
    if busca:
        # Buscar contratos por código ou conjunto
        contratos_codigo = qs.filter(Q(codigo__icontains=busca) | Q(conjunto__icontains=busca))
        
        # Buscar mutuários por nome ou código
        mutuarios = Mutuario.objects.filter(Q(nome__icontains=busca) | Q(codigo__icontains=busca))
        
        # Pegar IDs dos contratos vinculados a esses mutuários
        db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
        contrato_ids = []
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            for mut in mutuarios:
                cur.execute("SELECT contrato_id FROM contrato_mutuario_map WHERE mutuario_id=?", (mut.id,))
                contrato_ids.extend([row[0] for row in cur.fetchall()])
        except Exception as e:
            print(f"Erro ao buscar contratos por mutuário: {e}")
        finally:
            if 'conn' in locals():
                conn.close()
        
        # Combinar resultados
        if contrato_ids:
            contratos_mutuario = Contrato.objects.filter(id__in=contrato_ids)
            qs = (contratos_codigo | contratos_mutuario).distinct()
        else:
            qs = contratos_codigo
    
    # Ordenar e contar total antes da paginação
    qs = qs.order_by('codigo')
    total_contratos = qs.count()
    
    # Paginação: 200 contratos por página
    paginator = Paginator(qs, 200)
    page_obj = paginator.get_page(pagina)
    
    # Adicionar total de parcelas, saldo atual e dados do mutuário
    contratos_list = []
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
    
    # Buscar mutuários vinculados aos contratos via mapeamento
    contratos_ids = [c.id for c in page_obj]
    
    # Buscar mapeamento contrato -> mutuário
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        SELECT contrato_id, mutuario_id 
        FROM contrato_mutuario_map 
        WHERE contrato_id IN ({})
    """.format(','.join('?' * len(contratos_ids))), contratos_ids)
    
    mapeamento = {}
    mutuario_ids = []
    for contrato_id, mutuario_id in cur.fetchall():
        if contrato_id not in mapeamento:
            mapeamento[contrato_id] = []
        mapeamento[contrato_id].append(mutuario_id)
        mutuario_ids.append(mutuario_id)
    conn.close()
    
    # Buscar mutuários de uma vez
    mutuarios_dict = {}
    if mutuario_ids:
        mutuarios = Mutuario.objects.filter(id__in=mutuario_ids).select_related('endereco_fk')
        mutuarios_dict = {m.id: m for m in mutuarios}
    
    # Buscar estatísticas de parcelas de uma vez
    parcelas_stats = ParcelaContrato.objects.filter(contrato_id__in=contratos_ids).values('contrato_id').annotate(
        total=Count('id'),
        max_nmens=Max('nmens')
    )
    parcelas_map = {p['contrato_id']: p for p in parcelas_stats}
    
    # Buscar últimas parcelas para pegar saldo devedor
    ultimas_parcelas = {}
    for c_id in contratos_ids:
        stats = parcelas_map.get(c_id)
        if stats:
            ultima = ParcelaContrato.objects.filter(contrato_id=c_id, nmens=stats['max_nmens']).only('sddev').first()
            if ultima:
                ultimas_parcelas[c_id] = ultima.sddev
    
    for c in page_obj:
        # Usar estatísticas já calculadas
        stats = parcelas_map.get(c.id, {})
        c.total_parcelas = stats.get('total', 0)
        c.saldo_atual = ultimas_parcelas.get(c.id, 0)
        
        # Buscar mutuário principal vinculado ao contrato
        mutuario_ids_contrato = mapeamento.get(c.id, [])
        mutuario = mutuarios_dict.get(mutuario_ids_contrato[0]) if mutuario_ids_contrato else None
        
        c.mutuario_nome = mutuario.nome if mutuario else None
        c.mutuario_cpf = mutuario.cpf if mutuario else None
        
        # Buscar endereço do FK
        if mutuario and mutuario.endereco_fk:
            c.mutuario_endereco = mutuario.endereco_fk.endereco or mutuario.endereco
            c.mutuario_numero = mutuario.endereco_fk.numero or mutuario.numero
            c.mutuario_cidade = mutuario.endereco_fk.cidade or mutuario.cidade
        else:
            c.mutuario_endereco = mutuario.endereco if mutuario else None
            c.mutuario_numero = mutuario.numero if mutuario else None
            c.mutuario_cidade = mutuario.cidade if mutuario else None
        
        contratos_list.append(c)
    
    return render(request, 'principal/contratos.html', {
        'contratos': contratos_list, 
        'page_obj': page_obj,
        'busca': busca,
        'busca_codigo': busca_codigo,
        'busca_conjunto': busca_conjunto,
        'total': Contrato.objects.count(),
        'total_filtrado': total_contratos
    })

def contrato_detail(request, pk):
    contrato = get_object_or_404(Contrato, pk=pk)
    parcelas_qs = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')
    total_parcelas = parcelas_qs.count()
    
    # Calcular saldo atual e total pago (original)
    ultima_parcela = parcelas_qs.last()
    saldo_atual = ultima_parcela.sddev if ultima_parcela and ultima_parcela.sddev else Decimal('0')
    
    # Total pago = soma de todas as parcelas pagas (que têm dtpgto)
    parcelas_pagas = parcelas_qs.filter(dtpgto__isnull=False)
    total_pago = sum([p.vlautent for p in parcelas_pagas if p.vlautent]) if parcelas_pagas.exists() else Decimal('0')
    
    # Buscar mutuário vinculado pelo código (original com SQLite)
    mutuario = None
    try:
        # Primeiro tenta pelo código (relação direta)
        mutuario = Mutuario.objects.filter(codigo=contrato.codigo).first()
        
        # Se não encontrar, tenta pela tabela de mapeamento
        if not mutuario:
            db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
            result = cur.fetchone()
            if result:
                mutuario = Mutuario.objects.get(id=result[0])
            conn.close()
    except Exception as e:
        print(f"Erro ao buscar mutuário: {e}")

    mutuario_nome_display = None
    if mutuario:
        mutuario_nome_display = None if _nome_mutuario_suspeito(mutuario.nome) else mutuario.nome

    cod_imovel_display = contrato.cod_imovel
    if _parece_endereco(cod_imovel_display):
        cod_imovel_display = (mutuario.codimovel if mutuario and mutuario.codimovel else '')
    
    # Verificar se há parcelas em aberto (sem pagamento)
    parcelas_em_aberto = parcelas_qs.filter(dtpgto__isnull=True).count()
    atendimentos_crm = AtendimentoCRM.objects.filter(contrato=contrato).order_by('-data_atendimento')[:100]
    crm_importados = int(request.GET.get('crm_importados', 0) or 0)
    crm_ignorados = int(request.GET.get('crm_ignorados', 0) or 0)
    crm_erros = int(request.GET.get('crm_erros', 0) or 0)
    
    # Calcular campos derivados para cada parcela (formato PRINTEVO) - original
    parcelas_list = []
    parcelas_data = list(parcelas_qs)
    
    for i, p in enumerate(parcelas_data):
        # Correção Monetária (CORR MONET) = CM da parcela
        corr_monet = p.cm if p.cm else Decimal('0')
        
        # Razão Progressiva (RZ PROGR) = (Encargo Mensal Total) / Saldo Devedor
        # Regra de consistência: usar vlautent quando disponível.
        # EM costuma representar encargo/prestação e, se somado aos componentes,
        # pode duplicar valores em várias parcelas.
        if p.vlautent and p.vlautent > 0:
            enc_mensal_total = Decimal(str(p.vlautent))
        else:
            enc_mensal_total = Decimal('0')
            if p.juros: enc_mensal_total += p.juros
            if p.amort: enc_mensal_total += p.amort
            if p.seguro: enc_mensal_total += p.seguro
            if p.tca: enc_mensal_total += p.tca
            if p.fcvs: enc_mensal_total += p.fcvs
            if p.rp: enc_mensal_total += p.rp
            if p.em and enc_mensal_total == 0:
                enc_mensal_total = Decimal(str(p.em))
        
        rz_progr = Decimal('0')
        if p.sddev and p.sddev > 0:
            rz_progr = enc_mensal_total / p.sddev
        
        # Reajuste (REAJ) = detectar quando há mudança significativa no encargo
        reaj = None
        if i > 0:
            p_ant = parcelas_data[i-1]
            if p_ant.vlautent and p_ant.vlautent > 0:
                enc_ant = Decimal(str(p_ant.vlautent))
            else:
                enc_ant = Decimal('0')
                if p_ant.juros: enc_ant += p_ant.juros
                if p_ant.amort: enc_ant += p_ant.amort
                if p_ant.seguro: enc_ant += p_ant.seguro
                if p_ant.tca: enc_ant += p_ant.tca
                if p_ant.fcvs: enc_ant += p_ant.fcvs
                if p_ant.rp: enc_ant += p_ant.rp
                if p_ant.em and enc_ant == 0:
                    enc_ant = Decimal(str(p_ant.em))
            
            if enc_ant > 0 and enc_mensal_total > 0:
                variacao = (enc_mensal_total / enc_ant) - 1
                # Se variação for > 1%, considerar reajuste
                if abs(variacao) > Decimal('0.01'):
                    reaj = variacao
        
        # Adicionar campos calculados
        p.corr_monet = corr_monet
        p.rz_progr = rz_progr
        p.enc_mensal_total = enc_mensal_total
        p.reaj = reaj
        
        parcelas_list.append(p)

    # Saldo inicial exibido deve respeitar o valor financiado do contrato (OCR)
    # quando disponível; fallback para o primeiro saldo de parcela.
    saldo_inicial_exibicao = Decimal(str(contrato.vlfinanc or '0'))
    if saldo_inicial_exibicao <= 0 and parcelas_data:
        saldo_inicial_exibicao = parcelas_data[0].sddev or Decimal('0')
    
    # NOVA INTEGRAÇÃO: Simulação de Evolução Histórica (inner function para escopo)
    @lru_cache(maxsize=1)
    def simular_evolucao_historica(contrato_id):
        from decimal import ROUND_HALF_UP
        
        # 1. Pegar todas as parcelas REAIS do banco de dados primeiro
        parcelas_reais = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')
        dict_parcelas = {p.nmens: p for p in parcelas_reais}
        
        prazo_meses = contrato.prazo if contrato.prazo else 360
        evolucao = []
        fcvs_acum = Decimal('0')
        anomalias = 0
        
        # Início da simulação
        data_simulacao_atual = contrato.data_contrato or date(1984, 10, 30)
        saldo_simulado = Decimal('0')
        current_moeda = get_moeda_vigente(data_simulacao_atual)

        for mes_num in range(1, prazo_meses + 1):
            # --- REGRA DE OURO: Se a parcela existe no banco, use o dado REAL ---
            if mes_num in dict_parcelas:
                p_real = dict_parcelas[mes_num]
                saldo_ant = saldo_simulado
                saldo_simulado = p_real.sddev
                data_simulacao_atual = p_real.dtvenc or data_simulacao_atual
                
                # Calcula a correção teórica apenas para fins de exibição
                chave_mes = data_simulacao_atual.strftime('%Y-%m')
                indice = INDICES_HISTORICOS.get(chave_mes, Decimal('0'))
                correcao_teorica = saldo_ant * indice
                
                # Se o saldo cresceu mesmo com pagamento, houve resíduo FCVS
                if saldo_simulado > saldo_ant and saldo_ant > 0:
                    fcvs_acum += (saldo_simulado - saldo_ant)
                    anomalias += 1

                evolucao.append({
                    'mes': mes_num,
                    'data': data_simulacao_atual.strftime('%Y-%m'),
                    'saldo_ant': float(saldo_ant),
                    'correcao': float(p_real.cm or 0),
                    'saldo_novo': float(saldo_simulado),
                    'indice': f"{float(indice * 100):.2f}%",
                    'anomalia': 'DADO REAL' if mes_num > 1 else 'INÍCIO',
                    'fcvs_excedente': float(fcvs_acum),
                    'moeda': get_moeda_vigente(data_simulacao_atual)
                })
            else:
                # --- Se não existe parcela, aí sim nós simulamos ---
                # (Aqui entra a lógica de conversão de moeda que corrigimos antes)
                # ... lógica de simulação para meses futuros ...
                pass

            # Avança a data para o próximo loop
            data_simulacao_atual = (data_simulacao_atual.replace(day=1) + timedelta(days=32)).replace(day=1)
    @lru_cache(maxsize=1)
    def simular_evolucao_historica(contrato_id):
        from decimal import ROUND_HALF_UP
        
        # 1. Obter contrato e parcelas reais
        contrato = Contrato.objects.get(id=contrato_id)
        parcelas_reais = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')
        dict_parcelas = {p.nmens: p for p in parcelas_reais}
        
        prazo_meses = contrato.prazo if contrato.prazo else 360
        evolucao = []
        fcvs_acum = Decimal('0')
        anomalias = 0
        
        # Configuração inicial
        data_simulacao_atual = contrato.data_contrato or date(1984, 10, 30)
        saldo_simulado = Decimal('0')
        current_moeda = get_moeda_vigente(data_simulacao_atual)

        for mes_num in range(1, prazo_meses + 1):
            saldo_ant = saldo_simulado
            
            # --- CORREÇÃO DO FCVS: Redenominação do Acumulado ---
            moeda_mes_atual = get_moeda_vigente(data_simulacao_atual)
            if mes_num > 1 and moeda_mes_atual != current_moeda:
                for data_limite, moeda_anterior, fator, moeda_nova in NOMINAL_CONVERSION_FACTORS:
                    if data_simulacao_atual >= data_limite and current_moeda != moeda_nova:
                        # Divide o FCVS acumulado pelo fator da época (Corte de zeros)
                        fcvs_acum = (fcvs_acum / fator).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        break
                current_moeda = moeda_mes_atual

            # --- REGRA DE OURO: Usar dado real do banco ---
            if mes_num in dict_parcelas:
                p_real = dict_parcelas[mes_num]
                saldo_simulado = p_real.sddev if p_real.sddev is not None else Decimal('0')
                data_simulacao_atual = p_real.dtvenc or data_simulacao_atual
                
                # Cálculo do índice para exibição
                chave_mes = data_simulacao_atual.strftime('%Y-%m')
                indice = INDICES_HISTORICOS.get(chave_mes, Decimal('0'))
                
                # Cálculo do FCVS do mês: Se o saldo subiu, a diferença é resíduo
                anomalia_detectada = ''
                if saldo_simulado > saldo_ant and saldo_ant > 0:
                    excedente = saldo_simulado - saldo_ant
                    fcvs_acum += excedente
                    anomalias += 1
                    anomalia_detectada = 'RESÍDUO FCVS'

                evolucao.append({
                    'mes': mes_num,
                    'data': data_simulacao_atual.strftime('%Y-%m'),
                    'saldo_ant': float(saldo_ant),
                    'correcao': float(p_real.cm or 0),
                    'saldo_novo': float(saldo_simulado),
                    'indice': f"{float(indice * 100):.2f}%",
                    'anomalia': anomalia_detectada or 'OK',
                    'fcvs_excedente': float(fcvs_acum),
                    'moeda': current_moeda
                })
            else:
                # Caso não existam parcelas futuras no banco
                evolucao.append({
                    'mes': mes_num, 'data': data_simulacao_atual.strftime('%Y-%m'),
                    'saldo_ant': float(saldo_simulado), 'correcao': 0.0, 'saldo_novo': float(saldo_simulado),
                    'indice': "0.00%", 'anomalia': 'FIM CONTRATO',
                    'fcvs_excedente': float(fcvs_acum), 'moeda': current_moeda
                })

            # Avança para o próximo mês
            data_simulacao_atual = (data_simulacao_atual.replace(day=1) + timedelta(days=32)).replace(day=1)

        return evolucao[:12] + evolucao[-12:], anomalias, float(fcvs_acum)

    # Usa calcular_fcvs_residual_global (simulação completa com índices reais + PES)
    # em vez de simular_evolucao_historica (que depende de parcelas no banco)
    try:
        evolucao_historica, resumo_anomalias, fcvs_residual = calcular_fcvs_residual_global(contrato.id)
        residual_fcvs_final = fcvs_residual
    except Exception as e:
        print(f"[contrato_detail] Erro em calcular_fcvs_residual_global: {e}")
        evolucao_historica, resumo_anomalias, residual_fcvs_final, fcvs_residual = [], 0, 0.0, 0.0

    # Se não há parcelas reais, popula a tabela principal com parcelas simuladas
    if total_parcelas == 0 and evolucao_historica:
        from types import SimpleNamespace
        sim_parcelas = []
        for e in evolucao_historica:
            try:
                ano, mes_n = map(int, e['data'].split('-'))
                dtvenc_sim = date(ano, mes_n, 25)
            except Exception:
                dtvenc_sim = None
            sddev_val  = Decimal(str(e['saldo_novo']))  if e['saldo_novo']  else Decimal('0')
            encargo_v  = Decimal(str(e['encargo']))     if e.get('encargo') else Decimal('0')
            amort_v    = Decimal(str(e['amort']))       if e.get('amort')   else Decimal('0')
            juros_v    = Decimal(str(e['juros']))       if e.get('juros')   else Decimal('0')
            corr_v     = Decimal(str(e['correcao']))    if e.get('correcao')else Decimal('0')
            fcvs_v     = Decimal(str(e['fcvs_mes']))    if e.get('fcvs_mes')else Decimal('0')
            prest_v    = Decimal(str(e['prest_pes']))   if e.get('prest_pes')else Decimal('0')
            rz         = (encargo_v / sddev_val).quantize(Decimal('0.000001')) if sddev_val > 0 and encargo_v > 0 else None
            sim_parcelas.append(SimpleNamespace(
                dtvenc         = dtvenc_sim,
                dtpgto         = None,
                nmens          = e['mes'],
                juros          = juros_v,
                amort          = amort_v,
                corr_monet     = corr_v,
                sddev          = sddev_val,
                sddev_original = Decimal(str(e['saldo_ant'])) if e.get('mes') == 1 else None,
                seguro         = Decimal('0'),
                tca            = Decimal('0'),
                fcvs           = fcvs_v,
                rz_progr       = rz,
                enc_mensal_total = encargo_v,
                reaj           = None,
                moeda          = e.get('moeda', ''),
                simulado       = True,
            ))
        parcelas_list     = sim_parcelas
        total_parcelas    = len(sim_parcelas)
        saldo_atual       = sim_parcelas[-1].sddev if sim_parcelas else Decimal('0')
        parcelas_simuladas = True
    
    # Gerar HTML da tabela (para template)
    table_html = """
    <table class="printevo-table" style="font-size: 11px; margin-top: 10px;">
        <thead>
            <tr>
                <th>Mês</th>
                <th>Data</th>
                <th>Saldo Anterior (R$)</th>
                <th>Correção (R$)</th>
                <th>Novo Saldo (R$)</th>
                <th>Índice (%)</th>
                <th>Anomalia</th>
                <th>FCVS Residual (R$)</th>
            </tr>
        </thead>
        <tbody>
    """
    for ev in evolucao_historica:
        cor = "style='color: red; font-weight: bold;'" if ev['anomalia'] else ""
        table_html += f"""
            <tr>
                                    'codigo_provisorio': bool(item.get('codigo_provisorio', False)),
                <td>{ev['mes']}</td>
                <td>{ev['data']}</td>
                <td>{ev['saldo_ant']:.2f}</td>
                <td>{ev['correcao']:.2f}</td>
                <td>{ev['saldo_novo']:.2f}</td>
                <td>{ev['indice']}</td>
                <td {cor}>{ev['anomalia'] or 'OK'}</td>
                <td style="color: {'red' if ev['fcvs_excedente'] > 0 else 'green'}; font-weight: bold;">{ev['fcvs_excedente']:.2f}</td>
            </tr>
        """
    table_html += "</tbody></table>"
    
    return render(request, 'principal/contrato_detail.html', {
        'contrato': contrato, 
        'parcelas': parcelas_list, 
        'mutuario': mutuario, 
        'cod_imovel_display': cod_imovel_display,
        'mutuario_nome_display': mutuario_nome_display,
        'total_parcelas': total_parcelas,
        'saldo_atual': saldo_atual,
        'saldo_inicial_exibicao': saldo_inicial_exibicao,
        'total_pago': total_pago,
        'parcelas_pagas': parcelas_pagas.count(),
        'parcelas_em_aberto': parcelas_em_aberto,
        'atendimentos_crm': atendimentos_crm,
        'crm_importados': crm_importados,
        'crm_ignorados': crm_ignorados,
        'crm_erros': crm_erros,
        # Novos para visualização
        'evolucao_historica': evolucao_historica,
        'resumo_anomalias': resumo_anomalias,
        'fcvs_residual': round(residual_fcvs_final, 2),
        'parcelas_simuladas': parcelas_simuladas if 'parcelas_simuladas' in dir() else False,
        'table_evolucao_html': table_html,
    })


def adicionar_atendimento_crm(request, pk):
    """Adiciona um registro de atendimento/follow-up para um contrato."""
    contrato = get_object_or_404(Contrato, pk=pk)

    if request.method != 'POST':
        return redirect('contrato_detail', pk=contrato.id)

    assunto = (request.POST.get('assunto') or '').strip()
    observacoes = (request.POST.get('observacoes') or '').strip()

    if not assunto or not observacoes:
        return redirect(f"/contrato/{contrato.id}/#crm-atendimento")

    data_retorno_str = (request.POST.get('data_retorno') or '').strip()
    data_retorno = None
    if data_retorno_str:
        try:
            data_retorno = datetime.strptime(data_retorno_str, '%Y-%m-%d').date()
        except ValueError:
            data_retorno = None

    responsavel = (request.POST.get('responsavel') or '').strip()
    if not responsavel and hasattr(request, 'user') and getattr(request.user, 'is_authenticated', False):
        responsavel = request.user.username

    AtendimentoCRM.objects.create(
        contrato=contrato,
        tipo_contato=(request.POST.get('tipo_contato') or 'OUTRO').strip() or 'OUTRO',
        assunto=assunto,
        observacoes=observacoes,
        acordo_novo=(request.POST.get('acordo_novo') or '').strip(),
        proximo_passo=(request.POST.get('proximo_passo') or '').strip(),
        data_retorno=data_retorno,
        status=(request.POST.get('status') or 'ABERTO').strip() or 'ABERTO',
        responsavel=responsavel,
    )

    return redirect(f"/contrato/{contrato.id}/#crm-atendimento")


def importar_atendimentos_crm(request, pk):
    """Importa atendimentos CRM em lote a partir de planilha Excel."""
    contrato_origem = get_object_or_404(Contrato, pk=pk)

    if request.method != 'POST':
        return redirect(f"/contrato/{contrato_origem.id}/#crm-atendimento")

    arquivo = request.FILES.get('arquivo_crm')
    if not arquivo:
        return redirect(f"/contrato/{contrato_origem.id}/?crm_erros=1#crm-atendimento")

    try:
        from openpyxl import load_workbook
    except ImportError:
        return redirect(f"/contrato/{contrato_origem.id}/?crm_erros=1#crm-atendimento")

    def normalizar_texto(valor):
        if valor is None:
            return ''
        txt = str(valor).strip()
        txt = unicodedata.normalize('NFKD', txt).encode('ascii', 'ignore').decode('ascii')
        return txt.lower().strip()

    def normalizar_codigo(valor):
        if valor is None:
            return ''
        s = ''.join(ch for ch in str(valor).strip() if ch.isdigit())
        if not s:
            s = str(valor).strip()
        return s.lstrip('0') or s

    def parse_data(valor):
        if valor is None or valor == '':
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        txt = str(valor).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(txt, fmt).date()
            except ValueError:
                continue
        return None

    wb = load_workbook(filename=arquivo, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return redirect(f"/contrato/{contrato_origem.id}/?crm_erros=1#crm-atendimento")

    headers = [normalizar_texto(c) for c in rows[0]]
    idx_map = {h: i for i, h in enumerate(headers) if h}

    contrato_col = None
    for chave in ('contrato', 'numero contrato', 'num contrato', 'codigo contrato', 'codigo'):
        if chave in idx_map:
            contrato_col = idx_map[chave]
            break

    if contrato_col is None:
        return redirect(f"/contrato/{contrato_origem.id}/?crm_erros=1#crm-atendimento")

    def idx(*aliases):
        for a in aliases:
            if a in idx_map:
                return idx_map[a]
        return None

    i_assunto = idx('assunto', 'titulo', 'tema')
    i_observ = idx('observacoes', 'observacao', 'descricao', 'atendimento')
    i_acordo = idx('acordo', 'acordo novo', 'novo acordo')
    i_prox = idx('proximo passo', 'proximo_passo', 'pendencia', 'pendencias')
    i_retorno = idx('data retorno', 'data_retorno', 'retorno')
    i_tipo = idx('tipo contato', 'tipo_contato', 'canal')
    i_status = idx('status', 'situacao')
    i_resp = idx('responsavel', 'atendente')

    contratos = Contrato.objects.all().only('id', 'codigo')
    mapa_contrato = {}
    for c in contratos:
        mapa_contrato[normalizar_codigo(c.codigo)] = c

    importados = 0
    ignorados = 0
    erros = 0

    for row in rows[1:]:
        try:
            if row is None:
                continue
            cod = normalizar_codigo(row[contrato_col] if contrato_col < len(row) else '')
            if not cod or cod not in mapa_contrato:
                ignorados += 1
                continue

            assunto = ''
            if i_assunto is not None and i_assunto < len(row) and row[i_assunto] is not None:
                assunto = str(row[i_assunto]).strip()
            if not assunto:
                assunto = 'Importacao em lote'

            observ = ''
            if i_observ is not None and i_observ < len(row) and row[i_observ] is not None:
                observ = str(row[i_observ]).strip()
            if not observ:
                ignorados += 1
                continue

            tipo_contato = 'OUTRO'
            if i_tipo is not None and i_tipo < len(row) and row[i_tipo] is not None:
                t = normalizar_texto(row[i_tipo])
                if 'telefone' in t:
                    tipo_contato = 'TELEFONE'
                elif 'whats' in t:
                    tipo_contato = 'WHATSAPP'
                elif 'mail' in t:
                    tipo_contato = 'EMAIL'
                elif 'presen' in t:
                    tipo_contato = 'PRESENCIAL'

            status = 'ABERTO'
            if i_status is not None and i_status < len(row) and row[i_status] is not None:
                s = normalizar_texto(row[i_status])
                if 'andamento' in s:
                    status = 'EM_ANDAMENTO'
                elif 'conclu' in s or 'final' in s:
                    status = 'CONCLUIDO'

            acordo_novo = ''
            if i_acordo is not None and i_acordo < len(row) and row[i_acordo] is not None:
                acordo_novo = str(row[i_acordo]).strip()

            proximo_passo = ''
            if i_prox is not None and i_prox < len(row) and row[i_prox] is not None:
                proximo_passo = str(row[i_prox]).strip()

            responsavel = ''
            if i_resp is not None and i_resp < len(row) and row[i_resp] is not None:
                responsavel = str(row[i_resp]).strip()

            data_retorno = None
            if i_retorno is not None and i_retorno < len(row):
                data_retorno = parse_data(row[i_retorno])

            AtendimentoCRM.objects.create(
                contrato=mapa_contrato[cod],
                tipo_contato=tipo_contato,
                assunto=assunto,
                observacoes=observ,
                acordo_novo=acordo_novo,
                proximo_passo=proximo_passo,
                data_retorno=data_retorno,
                status=status,
                responsavel=responsavel,
            )
            importados += 1
        except Exception:
            erros += 1

    return redirect(
        f"/contrato/{contrato_origem.id}/?crm_importados={importados}&crm_ignorados={ignorados}&crm_erros={erros}#crm-atendimento"
    )


def editar_atendimento_crm(request, pk, atendimento_id):
    """Edita um registro de atendimento/follow-up de um contrato."""
    contrato = get_object_or_404(Contrato, pk=pk)
    atendimento = get_object_or_404(AtendimentoCRM, pk=atendimento_id, contrato=contrato)

    if request.method != 'POST':
        return redirect(f"/contrato/{contrato.id}/#crm-atendimento")

    assunto = (request.POST.get('assunto') or '').strip()
    observacoes = (request.POST.get('observacoes') or '').strip()
    if not assunto or not observacoes:
        return redirect(f"/contrato/{contrato.id}/#crm-atendimento")

    data_retorno_str = (request.POST.get('data_retorno') or '').strip()
    data_retorno = None
    if data_retorno_str:
        try:
            data_retorno = datetime.strptime(data_retorno_str, '%Y-%m-%d').date()
        except ValueError:
            data_retorno = None

    atendimento.tipo_contato = (request.POST.get('tipo_contato') or 'OUTRO').strip() or 'OUTRO'
    atendimento.status = (request.POST.get('status') or 'ABERTO').strip() or 'ABERTO'
    atendimento.assunto = assunto
    atendimento.observacoes = observacoes
    atendimento.acordo_novo = (request.POST.get('acordo_novo') or '').strip()
    atendimento.proximo_passo = (request.POST.get('proximo_passo') or '').strip()
    atendimento.data_retorno = data_retorno
    atendimento.responsavel = (request.POST.get('responsavel') or '').strip()
    atendimento.save()

    return redirect(f"/contrato/{contrato.id}/#crm-atendimento")


def excluir_atendimento_crm(request, pk, atendimento_id):
    """Exclui um registro de atendimento/follow-up de um contrato."""
    contrato = get_object_or_404(Contrato, pk=pk)
    atendimento = get_object_or_404(AtendimentoCRM, pk=atendimento_id, contrato=contrato)

    if request.method == 'POST':
        atendimento.delete()

    return redirect(f"/contrato/{contrato.id}/#crm-atendimento")

# ===== FUNÇÕES AUXILIARES DE FORMATAÇÃO FH1 =====
def fmt_num(valor, tamanho, decimais=0):
    """Formata número com zeros à esquerda"""
    if decimais > 0:
        valor_int = int(float(valor or 0) * (10 ** decimais))
        return str(valor_int).zfill(tamanho)
    return str(int(valor or 0)).zfill(tamanho)

def fmt_alfa(texto, tamanho):
    """Formata texto com espaços à direita"""
    if not texto:
        return ' ' * tamanho
    return str(texto)[:tamanho].ljust(tamanho)

def fmt_data(data):
    """Formata data como DDMMAA"""
    if not data:
        return '000000'
    return data.strftime('%d%m%y')

def fmt_data_ddmmaaaa(data):
    """Formata data como DDMMAAAA (8 dígitos)"""
    if not data:
        return '00000000'
    return data.strftime('%d%m%Y')

def pad_right(s, total):
    """Completa string com espaços à direita até tamanho especificado"""
    if len(s) >= total:
        return s[:total]
    return s + (' ' * (total - len(s)))


# ===== FUNÇÕES RCV (REGISTRO DE COMPROVAÇÃO DE VALORES) =====
def gerar_linha_rcv_simplificado(matricula_sem_dv_5, contrato_13, hipoteca_1, data_termino_ddmmaaaa):
    """
    Retorna a linha detalhe do RCV simplificado (27 bytes base).
    Campos:
    - matricula_sem_dv_5: Matrícula sem DV (5 bytes)
    - contrato_13: Número do contrato (13 bytes)
    - hipoteca_1: Indicador de hipoteca (1 byte)
    - data_termino_ddmmaaaa: Data término DDMMAAAA (8 bytes)
    Total: 27 bytes
    """
    linha = ""
    linha += matricula_sem_dv_5
    linha += contrato_13
    linha += hipoteca_1
    linha += data_termino_ddmmaaaa
    return linha

def validar_rcv(header, detalhes, trailer):
    """
    Validações bloqueantes para arquivo RCV.
    Mesma filosofia do Gatekeeper FH1.
    """
    erros = []
    
    # Validar HEADER
    if len(header) != 430:
        erros.append(f"HEADER inválido: {len(header)} bytes (esperado 430)")
    if header[0:2] != "33":
        erros.append(f"HEADER posição 1-2 (UFS): esperado '33', encontrado '{header[0:2]}'")
    if header[2:8] != "000442":
        erros.append(f"HEADER posição 3-8 (Matrícula): esperado '000442', encontrado '{header[2:8]}'")
    if header[8] != "0":
        erros.append(f"HEADER posição 9 (Tipo): esperado '0', encontrado '{header[8]}'")
    
    # Validar DETALHES
    for i, detalhe in enumerate(detalhes, start=1):
        if len(detalhe) != 430:
            erros.append(f"Detalhe {i} inválido: {len(detalhe)} bytes (esperado 430)")
    
    # Validar TRAILER
    if len(trailer) != 430:
        erros.append(f"TRAILER inválido: {len(trailer)} bytes (esperado 430)")
    if trailer[0:2] != "33":
        erros.append(f"TRAILER posição 1-2 (UFS): esperado '33', encontrado '{trailer[0:2]}'")
    if trailer[2:8] != "000442":
        erros.append(f"TRAILER posição 3-8 (Matrícula): esperado '000442', encontrado '{trailer[2:8]}'")
    if trailer[8] != "9":
        erros.append(f"TRAILER posição 9 (Tipo): esperado '9', encontrado '{trailer[8]}'")
    
    if erros:
        msg = "🚨 FALHA CRÍTICA DE VALIDAÇÃO RCV:\n" + "\n".join(erros)
        raise ValueError(msg)
    
    return True


def validar_pacote_fh1(header, linha_i, trailer, fcvs_esperado, saldo_esperado):
    """
    Realiza auditoria técnica rigorosa no pacote FH1 antes da liberação.
    Levanta erro se houver qualquer desalinhamento.
    """
    erros = []
    
    # 1. Validação de Comprimento (430 bytes fixos)
    if len(header) != 430: 
        erros.append(f"HEADER inválido: {len(header)} bytes (esperado 430)")
    if len(linha_i) != 430: 
        erros.append(f"REGISTRO I inválido: {len(linha_i)} bytes (esperado 430)")
    if len(trailer) != 430: 
        erros.append(f"TRAILER inválido: {len(trailer)} bytes (esperado 430)")
    
    # 2. Validação do Campo Crítico 60 (VAF3) no Registro I
    # Posição 375-388 (1-based) -> [374:388] (0-based)
    vaf3_no_arquivo = linha_i[374:388]
    vaf3_calculado = fmt_num(fcvs_esperado, 14, 2)
    if vaf3_no_arquivo != vaf3_calculado:
        erros.append(f"Divergência VAF3: No arquivo [{vaf3_no_arquivo}] vs Calculado [{vaf3_calculado}]")
        
    # 3. Validação de Tipos de Registro
    if header[8] != '0': 
        erros.append("HEADER: Identificador de tipo '0' não encontrado na posição 9")
    if linha_i[8:21].strip() == "": 
        erros.append("REGISTRO I: Número do contrato não identificado")
    if trailer[8] != '9': 
        erros.append("TRAILER: Identificador de tipo '9' não encontrado na posição 9")

    # 4. Validação de Posições Fixas (protege contra alteração de ordem)
    # HEADER: validar posições críticas
    if header[0:2] != "33":
        erros.append(f"HEADER posição 1-2 (UFS): esperado '33', encontrado '{header[0:2]}'")
    if header[2:8] != "000442":
        erros.append(f"HEADER posição 3-8 (Matrícula): esperado '000442', encontrado '{header[2:8]}'")
    if header[8] != "0":
        erros.append(f"HEADER posição 9 (Tipo): esperado '0', encontrado '{header[8]}'")
    
    # TRAILER: validar posições críticas
    if trailer[0:2] != "33":
        erros.append(f"TRAILER posição 1-2 (UFS): esperado '33', encontrado '{trailer[0:2]}'")
    if trailer[2:8] != "000442":
        erros.append(f"TRAILER posição 3-8 (Matrícula): esperado '000442', encontrado '{trailer[2:8]}'")
    if trailer[8] != "9":
        erros.append(f"TRAILER posição 9 (Tipo): esperado '9', encontrado '{trailer[8]}'")

    # 5. Validação de Totais no Trailer
    # Posição 16-29 no trailer contém o total FCVS
    total_fcvs_trailer = trailer[15:29]
    if vaf3_calculado != total_fcvs_trailer:
        erros.append(f"TRAILER: Total FCVS [{total_fcvs_trailer}] difere do calculado [{vaf3_calculado}]")

    if erros:
        msg = "🚨 FALHA CRÍTICA DE VALIDAÇÃO FH1:\n" + "\n".join(erros)
        raise ValueError(msg)
    
    return True


def exportar_evolucao_txt(request, pk):
    """Exporta arquivo FH1 no formato oficial da CEF com HEADER + registros + TRAILER"""
    from django.http import HttpResponse
    import logging
    logger = logging.getLogger(__name__)
    
    contrato = get_object_or_404(Contrato, pk=pk)
    
    # ===== OPÇÃO 3: VALIDAÇÃO PREVENTIVA =====
    if request.GET.get('validar') != '0':  # Por padrão valida, exceto se ?validar=0
        from .validators import validar_antes_exportar, pode_exportar
        
        # Pre-flight check
        validacao = validar_antes_exportar(contrato)
        
        # Se houver erros críticos, bloquear exportação
        pode_exp, motivo = pode_exportar(contrato)
        if not pode_exp:
            # Criar registro no histórico ANTES de bloquear
            try:
                from .models import ValidacaoAI
                ValidacaoAI.objects.create(
                    tipo_arquivo='FH1',
                    contrato=contrato,
                    status='REPROVADO',
                    relatorio_completo=f"❌ Exportação bloqueada pela validação preventiva\n\n{motivo}\n\n📋 Erros:\n" + "\n".join([f"• {e['campo']}: {e['mensagem']}" for e in validacao['erros']]),
                    erros_encontrados="\n".join([f"{e['campo']}: {e['mensagem']}" for e in validacao['erros']]),
                    tempo_execucao=0.1,
                    tamanho_arquivo=0,
                    agentes_utilizados='Validação Preventiva (Opção 3)',
                    correcao_automatica=False,
                    correcoes_aplicadas=''
                )
                print(f"💾 Validação REPROVADA registrada no histórico")
            except Exception as e:
                print(f"⚠️ Erro ao salvar histórico de bloqueio: {e}")
            
            html_erros = "<h1>❌ Exportação bloqueada</h1>"
            html_erros += f"<p><strong>{motivo}</strong></p>"
            html_erros += "<h2>Erros encontrados:</h2><ul>"
            for erro in validacao['erros']:
                html_erros += f"<li><strong>{erro['campo']}</strong>: {erro['mensagem']}<br>"
                html_erros += f"<em>Sugestão: {erro['sugestao']}</em></li>"
            html_erros += "</ul>"
            html_erros += f"<p><a href='/contrato/{pk}/'>← Voltar para o contrato</a></p>"
            html_erros += f"<p><a href='?validar=0'>⚠️ Forçar exportação (não recomendado)</a></p>"
            return HttpResponse(html_erros, status=400)
        
        # Se houver warnings, registrar mas não bloquear
        if validacao['warnings']:
            print("\n⚠️ VALIDAÇÃO PREVENTIVA - WARNINGS:")
            for warning in validacao['warnings']:
                print(f"  • {warning['campo']}: {warning['mensagem']}")
    
    # Buscar mutuário
    mutuario = None
    try:
        db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
        result = cur.fetchone()
        if result:
            mutuario = Mutuario.objects.get(id=result[0])
        conn.close()
    except:
        pass
    
    # Buscar parcelas
    parcelas = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')
    primeira = parcelas.first()
    ultima = parcelas.last()
    
    # Inicializar logger
    import logging
    logger = logging.getLogger(__name__)
    
    # Calcular FCVS residual usando calcular_fcvs_residual_global (função global disponível)
    try:
        evolucao_completa, anomalias, fcvs_residual = calcular_fcvs_residual_global(contrato.id)
        fcvs_residual = Decimal(str(fcvs_residual))
        print(f"\n{'='*80}")
        print(f">>> FCVS Residual calculado: R$ {fcvs_residual:,.2f}")
        print(f">>> Tipo: {type(fcvs_residual)}")
        print(f">>> Valor bruto: {fcvs_residual}")
        print(f"{'='*80}\n")
        logger.info(f">>> FCVS Residual calculado: R$ {fcvs_residual:,.2f}")
    except Exception as e:
        print(f"\n!!! ERRO ao calcular FCVS: {e}\n")
        logger.error(f"Erro ao calcular FCVS residual para contrato {contrato.id}: {e}")
        fcvs_residual = Decimal(0)
    
    # Buscar conjunto
    conjunto = ConjuntoHabitacional.objects.filter(conjunto=contrato.conjunto).first()
    
    # Construir registro FH1 (430 caracteres) conforme manual CEF
    # COM RASTREAMENTO DE POSIÇÃO
    linha = ''
    pos = 0  # Contador de posição (1-based no manual, mas 0-based no código)
    
    def add_campo(valor, nome_campo, pos_inicio_manual):
        nonlocal linha, pos
        antes = len(linha)
        linha += valor
        depois = len(linha)
        tamanho = depois - antes
        logger.info(f"  Campo: {nome_campo:30s} | Pos manual: {pos_inicio_manual:3d}-{pos_inicio_manual+tamanho-1:3d} | Real: {antes:3d}-{depois-1:3d} | Val: [{valor[:20]}...]" if len(valor) > 20 else f"  Campo: {nome_campo:30s} | Pos manual: {pos_inicio_manual:3d}-{pos_inicio_manual+tamanho-1:3d} | Real: {antes:3d}-{depois-1:3d} | Val: [{valor}]")
        pos = depois
    
    logger.info("=" * 80)
    logger.info("=== CONSTRUÇÃO DA LINHA I (COM RASTREAMENTO) ===")
    
    add_campo('33', 'UFS', 1)
    add_campo('000442', 'MAT. AG. FINANC./DV', 3)
    add_campo(fmt_alfa(contrato.codigo, 13), 'N.º CONTRATO', 9)
    add_campo('1', 'HIPOTECA', 22)
    add_campo('1', 'TIPO DE REGISTRO', 23)
    add_campo('00', 'SEQUENCIAL', 24)
    add_campo('0', 'CONSTANTE', 26)
    add_campo(fmt_alfa(mutuario.nome if mutuario else '', 40), 'NOME MUTUÁRIO', 27)
    add_campo('1', 'TIPO', 67)
    add_campo(fmt_alfa(mutuario.cpf if mutuario else '', 17), 'CPF/CI', 68)
    add_campo(fmt_data(mutuario.dtnasc if mutuario and mutuario.dtnasc else None), 'DATA NASCIMENTO', 85)
    add_campo('00000', 'COD MUNICÍPIO', 91)
    add_campo('RJ', 'UF', 96)
    add_campo(fmt_alfa(conjunto.nome if conjunto else '', 38), 'ENDEREÇO', 98)
    add_campo(fmt_data(contrato.data_contrato), 'DATA CONTRATO', 136)
    add_campo(fmt_num(0, 12, 2), 'VALOR GARANTIA', 142)
    add_campo('00', 'IM', 154)
    add_campo(fmt_data(contrato.data_contrato), 'DATA LEGISLAÇÃO', 156)
    # Usar sddev_original se disponível (valor antes das conversões monetárias)
    valor_fin = primeira.sddev_original if primeira and primeira.sddev_original else (primeira.sddev if primeira and primeira.sddev else Decimal(0))
    add_campo(fmt_num(valor_fin, 12, 2), 'VALOR FINANCIAMENTO', 162)
    add_campo(fmt_num(0, 12, 2), 'VALOR FCVS', 174)
    add_campo('00000', 'COD CATEGORIA', 186)
    add_campo('S', 'SEGURO CRÉDITO', 191)
    add_campo('N', 'CARÊNCIA', 192)
    add_campo('S', 'SEGURO DFI', 193)
    add_campo('N', 'PROER', 194)
    add_campo(' ', 'VAGO', 195)
    prazo = contrato.prazo if contrato.prazo else parcelas.count()
    add_campo(fmt_num(prazo, 3), 'PRAZO CONTRATADO', 196)
    add_campo('060000', 'TAXA JUROS', 199)
    add_campo('0000', 'CES', 205)
    add_campo('SFH', 'PLANO', 209)
    add_campo('0', 'ST', 212)
    add_campo('N', 'RJ', 213)
    add_campo('00', 'RR', 214)
    add_campo('SAL', 'INDEX', 216)
    add_campo(fmt_num(prazo, 3), 'PRAZO FCVS', 219)
    add_campo('060000', 'TAXA FCVS', 222)
    add_campo('0000', 'CES FCVS', 228)
    add_campo('SFH', 'PLANO', 232)
    add_campo('0', 'ST', 235)
    add_campo('N', 'RJ', 236)
    add_campo('00', 'RR', 237)
    add_campo('SAL', 'INDEX', 239)
    add_campo(fmt_data(primeira.dtvenc if primeira else None), 'DATA SALDO', 242)
    # Usar valor original se disponível (antes da conversão monetária)
    saldo = ultima.sddev_original if ultima and ultima.sddev_original else (ultima.sddev if ultima and ultima.sddev else Decimal(0))
    add_campo(fmt_num(saldo, 12, 2), 'SALDO DEVEDOR', 248)
    add_campo(fmt_data(primeira.dtvenc if primeira else None), '1º VENCIMENTO', 260)
    
    # Valores da primeira parcela
    seguro = primeira.seguro if primeira and primeira.seguro else Decimal(0)
    add_campo(fmt_num(seguro, 8, 2), 'SEGURO', 266)
    
    # Prestação = Encargo Mensal Total (juros + amort + seguro + tca + fcvs + em + rp)
    # O campo 'em' contém valores adicionais que completam o encargo mensal total
    prestacao_total = Decimal('0')
    if primeira:
        if primeira.juros: prestacao_total += primeira.juros
        if primeira.amort: prestacao_total += primeira.amort
        if primeira.seguro: prestacao_total += primeira.seguro
        if primeira.tca: prestacao_total += primeira.tca
        if primeira.fcvs: prestacao_total += primeira.fcvs
        if primeira.em: prestacao_total += primeira.em
        if primeira.rp: prestacao_total += primeira.rp
    add_campo(fmt_num(prestacao_total, 10, 2), 'PRESTAÇÃO', 274)
    
    tca_val = primeira.tca if primeira and primeira.tca else Decimal(0)
    add_campo(fmt_num(tca_val, 8, 2), 'TCA/TAC', 284)
    fcvs = primeira.fcvs if primeira and primeira.fcvs else Decimal(0)
    add_campo(fmt_num(fcvs, 8, 2), 'FCVS MENSAL', 292)
    add_campo(fmt_num(0, 8, 2), 'RAZÃO', 300)
    add_campo('   ', 'TIPO EVENTO', 308)
    add_campo('000000', 'DATA EVENTO', 311)
    add_campo('00', 'OR/CO', 317)
    add_campo('0100', '% CAIXA', 319)
    add_campo(fmt_num(0, 18), 'Nº CONTR EMPR', 323)
    add_campo('000000', 'TAXA JUROS EVENTO', 341)
    add_campo(fmt_num(0, 14, 2), 'VAF1 - VAF Básico', 347)
    add_campo(fmt_num(0, 14, 2), 'VAF2 - VAF Complementar', 361)
    
    # *** CAMPO CRÍTICO: FCVS RESIDUAL ***
    print(f"\n{'='*80}")
    print(f">>> ANTES de adicionar VAF3: len(linha) = {len(linha)}")
    print(f">>> FCVS Residual disponível: {fcvs_residual}")
    print(f">>> fmt_num(fcvs_residual, 14, 2) = [{fmt_num(fcvs_residual, 14, 2)}]")
    print(f"{'='*80}\n")
    
    logger.info("=" * 80)
    logger.info(f">>> ANTES de adicionar VAF3: len(linha) = {len(linha)}")
    logger.info(f">>> FCVS Residual: R$ {fcvs_residual:,.2f}")
    logger.info(f">>> fmt_num(fcvs_residual, 14, 2) = [{fmt_num(fcvs_residual, 14, 2)}]")
    add_campo(fmt_num(fcvs_residual, 14, 2), '*** VAF3 - VAF Residual/FCVS ***', 375)
    logger.info(f">>> DEPOIS de adicionar VAF3: len(linha) = {len(linha)}")
    logger.info(f">>> Slice [374:388]: [{linha[374:388]}]")
    logger.info("=" * 80)
    
    print(f"\n{'='*80}")
    print(f">>> DEPOIS de adicionar VAF3: len(linha) = {len(linha)}")
    print(f">>> Slice [374:388]: [{linha[374:388]}]")
    print(f"{'='*80}\n")
    
    add_campo(fmt_num(0, 14), 'JUROS', 389)
    add_campo('D', 'DEBITO/CRÉDITO', 403)
    add_campo('00', 'QTD ALTERAÇÕES', 404)
    add_campo('33', 'UFS', 406)
    add_campo('000442', 'MAT AG', 408)
    add_campo(date.today().strftime('%d%m%y'), 'DATA GERAÇÃO', 414)
    add_campo('001', 'NÚMERO LOTE', 420)
    add_campo('S', 'FORMA ENVIO', 423)
    add_campo('I', 'TIPO MOVIMENTO', 424)
    add_campo('      ', 'FILLER', 425)
    
    logger.info("=" * 80)
    logger.info(f">>> LINHA COMPLETA: {len(linha)} bytes")
    logger.info("=" * 80)
    
    # ===== TESTE DEFINITIVO DO SLICE (EXIGIDO PELO USUÁRIO) =====
    print(f"\n{'='*80}")
    print(f"===== TESTE DEFINITIVO DE VALIDAÇÃO CEF =====")
    print(f"LEN I: {len(linha)}")
    print(f"VAF3 375-388: [{linha[374:388]}]")
    print(f"VAF3 esperado: [00000002206651]")
    print(f"Match: {linha[374:388] == '00000002206651'}")
    print(f"Posição encontrada (find): {linha.find('00000002206651')}")
    print(f"{'='*80}\n")
    
    # ====================
    # VALIDAÇÃO BLOQUEANTE - NÃO GERA ARQUIVO SE HOUVER ERRO
    # ====================
    import logging
    logger = logging.getLogger(__name__)
    
    erros_criticos = []
    
    # 1. Validar tamanho EXATO
    tamanho_atual = len(linha)
    if tamanho_atual != 430:
        erros_criticos.append(f"Registro I: tamanho {tamanho_atual} (esperado 430)")
        logger.error(f"❌ ERRO CRÍTICO: Registro I com {tamanho_atual} bytes!")
    else:
        logger.info(f"✓ Registro I: 430 bytes")
    
    # 2. Validar FCVS nas posições 375-388
    fcvs_slice = linha[374:388]
    fcvs_esperado = fmt_num(fcvs_residual, 14, 2)
    if fcvs_slice != fcvs_esperado:
        erros_criticos.append(f"FCVS posições 375-388: '{fcvs_slice}' ≠ '{fcvs_esperado}'")
        logger.error(f"❌ ERRO CRÍTICO: FCVS={fcvs_slice}, Esperado={fcvs_esperado}")
    else:
        logger.info(f"✓ FCVS (375-388): {fcvs_slice} = R$ {fcvs_residual:,.2f}")
    
    # 3. Validar encoding
    try:
        linha.encode('latin-1')
        logger.info(f"✓ Encoding: latin-1")
    except UnicodeEncodeError as e:
        erros_criticos.append(f"Caracteres inválidos: {e}")
        logger.error(f"❌ ERRO CRÍTICO: {e}")
    
    # 4. Validar campos chave
    validacoes = {
        'UFS': {'pos': (0, 2), 'esperado': '33'},
        'Matrícula': {'pos': (2, 8), 'esperado': '000442'},
        'Tipo Registro': {'pos': (22, 23), 'esperado': '1'},
    }
    
    for campo, dados in validacoes.items():
        valor = linha[dados['pos'][0]:dados['pos'][1]]
        if valor != dados['esperado']:
            erros_criticos.append(f"{campo}: '{valor}' ≠ '{dados['esperado']}'")
            logger.error(f"❌ {campo}: {valor}")
        else:
            logger.info(f"✓ {campo}: {valor}")
    
    # SE HOUVER ERROS, NÃO GERA ARQUIVO
    if erros_criticos:
        from django.http import HttpResponse
        erro_html = "<h1>❌ ERRO: Arquivo FH1 NÃO PODE SER GERADO</h1>"
        erro_html += "<h2>Erros Críticos Encontrados:</h2><ul>"
        for erro in erros_criticos:
            erro_html += f"<li>{erro}</li>"
        erro_html += "</ul>"
        erro_html += f"<p><strong>FCVS Calculado:</strong> R$ {fcvs_residual:,.2f}</p>"
        erro_html += f"<p><strong>Tamanho linha:</strong> {tamanho_atual} bytes</p>"
        erro_html += "<p>Corrija os erros antes de enviar à CEF.</p>"
        return HttpResponse(erro_html, status=400)
    
    logger.info("=" * 60)
    
    # VALIDAÇÃO ADICIONAL (não bloqueante, apenas informativa)
    logger.info(f"=== VALIDAÇÃO FH1 - Contrato {contrato.codigo} ===")
    logger.info(f"Tamanho: {tamanho_atual} bytes ✓")
    
    campos_info = {
        'Contrato': {'slice': linha[8:21], 'posicao': '9-21'},
        'Data Contrato': {'slice': linha[135:141], 'posicao': '136-141'},
        'Saldo Devedor': {'slice': linha[247:259], 'posicao': '248-259'},
        'FCVS Residual': {'slice': linha[374:388], 'posicao': '375-388'}
    }
    
    for campo, dados in campos_info.items():
        logger.info(f"{campo} (pos {dados['posicao']}): [{dados['slice']}]")
    
    logger.info("=" * 50)
    
    # ====================
    # MONTAR ARQUIVO COMPLETO: HEADER + REGISTROS + TRAILER
    # ====================
    
    linhas_arquivo = []
    data_hoje = date.today()
    
    # ===== HEADER (tipo "0") - 430 bytes =====
    header = ''
    header += '33'  # 01. UFS (1-2)
    header += '000442'  # 02. MAT. AG. FINANC./DV (3-8)
    header += '0'  # 03. TIPO REGISTRO = "0" HEADER (9)
    header += fmt_data(data_hoje)  # 04. DATA GERAÇÃO DDMMAA (10-15)
    header += '001'  # 05. SEQUENCIAL DO ARQUIVO (16-18)
    header += 'COFLUHAB'.ljust(30)  # 06. NOME DA COHAB (19-48)
    header += fmt_num(1, 6)  # 07. QTD REGISTROS TIPO I (49-54)
    header += fmt_num(fcvs_residual, 14, 2)  # 08. TOTAL FCVS RESIDUAL (55-68)
    header += ' ' * (430 - len(header))  # FILLER até 430 bytes
    
    linhas_arquivo.append(header)
    
    # ===== REGISTRO TIPO "I" (Habilitação) - 430 bytes =====
    linhas_arquivo.append(linha)
    
    # ===== TRAILER (tipo "9") - 430 bytes =====
    trailer = ''
    trailer += '33'  # 01. UFS (1-2)
    trailer += '000442'  # 02. MAT. AG. FINANC./DV (3-8)
    trailer += '9'  # 03. TIPO REGISTRO = "9" TRAILER (9)
    trailer += fmt_num(1, 6)  # 04. TOTAL REGISTROS TIPO I (10-15)
    trailer += fmt_num(fcvs_residual, 14, 2)  # 05. SOMA FCVS RESIDUAL (16-29)
    trailer += fmt_num(saldo, 14, 2)  # 06. SOMA SALDO DEVEDOR (30-43)
    
    # 🧪 MODO DE TESTE: Forçar erro se parâmetro ?teste_erro=1 presente
    if request.GET.get('teste_erro') == '1':
        print("\n" + "="*80)
        print("🧪 MODO DE TESTE ATIVADO - Gerando TRAILER com ERRO PROPOSITAL")
        print("="*80 + "\n")
        # NÃO adiciona padding - deixa com 43 bytes para testar Gatekeeper
        # (trailer fica com ~43 bytes propositalmente)
    else:
        # GARANTIR 430 bytes SEMPRE: completar com espaços até 430
        tamanho_trailer = len(trailer)
        if tamanho_trailer < 430:
            trailer += ' ' * (430 - tamanho_trailer)
        elif tamanho_trailer > 430:
            trailer = trailer[:430]  # Truncar se ultrapassar (nunca deveria acontecer)
    
    linhas_arquivo.append(trailer)
    
    # ====================
    # 5 LOGS DEFINITIVOS SOLICITADOS PELO USUÁRIO
    # ====================
    linha_i = linhas_arquivo[1]  # Registro tipo I
    
    print("\n" + "="*80)
    print("===== 5 LOGS DEFINITIVOS PARA CEF =====")
    print(f"LEN HEADER: {len(header)}")
    print(f"LEN I: {len(linha_i)}")
    print(f"LEN TRAILER: {len(trailer)}")
    print(f"VAF3 375-388: {linha_i[374:388]}")
    print(f"POS VAF3: {linha_i.find('00000002206651')}")
    print("="*80 + "\n")
    
    # ====================
    # DEBUG DETALHADO: Verificar cada linha
    # ====================
    logger.info("=" * 60)
    logger.info("=== DEBUG DETALHADO - ESTRUTURA DO ARQUIVO ===")
    
    for idx, l in enumerate(linhas_arquivo, start=1):
        nome_linha = ['HEADER', 'REGISTRO I', 'TRAILER'][idx-1]
        logger.info(f"{nome_linha}: len={len(l)} | inicio={repr(l[:10])} | fim={repr(l[-10:])}")
    
    # Verificar FCVS no registro I especificamente
    linha_i = linhas_arquivo[1]  # Registro tipo I
    fcvs_slice = linha_i[374:388]  # Posições 375-388 (0-indexed)
    logger.info(f"")
    logger.info(f">>> FCVS Residual (posições 375-388): [{fcvs_slice}]")
    logger.info(f">>> Valor esperado: [{fmt_num(fcvs_residual, 14, 2)}]")
    logger.info(f">>> Valor calculado: R$ {fcvs_residual}")
    logger.info(f">>> Match: {'✓ CORRETO' if fcvs_slice == fmt_num(fcvs_residual, 14, 2) else '✗ ERRO'}")
    logger.info("=" * 60)
    
    # ====================
    # GATEKEEPER - CAMADA DE SEGURANÇA DEFINITIVA
    # ====================
    try:
        # EXECUTA O GATEKEEPER
        validar_pacote_fh1(header, linha_i, trailer, fcvs_residual, saldo)
        
        print(f"\n{'='*80}")
        print("✅ VALIDAÇÃO GATEKEEPER APROVADA - Arquivo seguro para CEF")
        print(f"{'='*80}\n")
        
        # Se chegou aqui, o arquivo está perfeito.
        # Monta o conteúdo final com quebra de linha Windows (CRLF)
        conteudo_str = f"{header}\r\n{linha_i}\r\n{trailer}\r\n"
        
        # Garantir encoding latin-1 estrito (padrão bancário CEF)
        # Usar .encode() garante que o Django não altere o encoding
        try:
            conteudo_bytes = conteudo_str.encode("latin-1", errors="strict")
        except UnicodeEncodeError as e:
            raise ValueError(f"Erro de encoding: caractere inválido para latin-1: {e}")
        
        logger.info(f"=== ARQUIVO FH1 COMPLETO ===")
        logger.info(f"HEADER: {len(header)} bytes")
        logger.info(f"Registros tipo I: 1 registro ({len(linha_i)} bytes)")
        logger.info(f"TRAILER: {len(trailer)} bytes")
        logger.info(f"Total linhas: 3")
        logger.info(f"Total bytes (com CRLF): {len(conteudo_bytes)} bytes")
        logger.info(f"Encoding: latin-1 (strict)")
        logger.info("=" * 50)
        
        # ===== SEMPRE CRIAR REGISTRO NO HISTÓRICO =====
        import time
        from .models import ValidacaoAI
        from .validators import validar_antes_exportar
        
        tempo_inicio = time.time()
        registro_criado = False
        
        try:
            validacao_preventiva = validar_antes_exportar(contrato)
            
            # Tentar validação AI se não for explicitamente desabilitada
            if request.GET.get('validar_ai') != '0':
                try:
                    from .ai_agents import (
                        validar_arquivo_com_ai, 
                        corrigir_com_agente_autofix,
                        CREWAI_DISPONIVEL
                    )
                    
                    if CREWAI_DISPONIVEL:
                        print(f"\n{'='*80}")
                        print("🤖 INICIANDO VALIDAÇÃO AI AUTOMÁTICA")
                        print(f"{'='*80}\n")
                        
                        resultado_ai = validar_arquivo_com_ai(conteudo_str, "FH1")
                        tempo_execucao = time.time() - tempo_inicio
                        
                        # AUTO-CORREÇÃO INTELIGENTE
                        correcao_info = None
                        if not resultado_ai.get('aprovado'):
                            print("\n🧠 INICIANDO AUTO-FIX INTELIGENTE...")
                            correcao_info = corrigir_com_agente_autofix(conteudo_str, resultado_ai.get('erros', []), "FH1")
                            
                            if correcao_info and correcao_info['sucesso']:
                                conteudo_str = correcao_info['conteudo_corrigido']
                                conteudo_bytes = conteudo_str.encode('latin-1')
                                print(f"\n📦 Arquivo corrigido ({len(conteudo_bytes)} bytes)")
                        
                        # Salvar no histórico
                        ValidacaoAI.objects.create(
                            tipo_arquivo='FH1',
                            contrato=contrato,
                            status='APROVADO' if resultado_ai.get('aprovado') else 'REPROVADO',
                            relatorio_completo=resultado_ai.get('resultado', ''),
                            erros_encontrados=resultado_ai.get('erro', '') if not resultado_ai.get('sucesso') else '',
                            tempo_execucao=tempo_execucao,
                            tamanho_arquivo=len(conteudo_bytes),
                            agentes_utilizados='QA Engineer, Auto-Fix Engineer' if correcao_info else 'QA Engineer',
                            correcao_automatica=correcao_info['sucesso'] if correcao_info else False,
                            correcoes_aplicadas='\n'.join(correcao_info['correcoes_aplicadas']) if correcao_info else ''
                        )
                        registro_criado = True
                        print(f"✅ Validação AI concluída em {tempo_execucao:.2f}s\n")
                    
                except Exception as e:
                    print(f"⚠️ Erro na validação AI: {e}")
            
            # Se não criou registro ainda (CrewAI indisponível ou erro), criar agora
            if not registro_criado:
                tempo_execucao = time.time() - tempo_inicio
                ValidacaoAI.objects.create(
                    tipo_arquivo='FH1',
                    contrato=contrato,
                    status='APROVADO',
                    relatorio_completo=f"✅ Arquivo exportado com sucesso\n\n📊 Estatísticas:\n- Tamanho: {len(conteudo_bytes)} bytes\n- Validação Preventiva: {len(validacao_preventiva['erros'])} erros, {len(validacao_preventiva['warnings'])} avisos\n- Mutuário: {'✓' if mutuario else '✗'}\n- Parcelas: {parcelas.count()}\n\n⚠️ CrewAI não disponível (Python 3.14)",
                    tempo_execucao=tempo_execucao,
                    tamanho_arquivo=len(conteudo_bytes),
                    agentes_utilizados='Gatekeeper',
                    correcao_automatica=False,
                    correcoes_aplicadas=''
                )
                print(f"💾 Registro de exportação salvo no histórico\n")
                
        except Exception as e:
            print(f"⚠️ Erro ao criar registro: {e}")
            import traceback
            traceback.print_exc()
        
        # Retorna bytes diretamente para evitar conversão automática do Django
        response = HttpResponse(conteudo_bytes, content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="FH1_{contrato.codigo}.txt"'
        
        print(f"✅ Arquivo FH1 gerado e validado com sucesso: {contrato.codigo}")
        return response

    except ValueError as e:
        # Se houver erro, impede o download e mostra o erro no log/tela
        print(f"❌ Erro na geração do FH1: {str(e)}")
        logger.error(f"❌ Erro na validação FH1: {str(e)}")
        
        erro_html = f"<h1>❌ ARQUIVO FH1 BLOQUEADO</h1>"
        erro_html += f"<h2>Validação Gatekeeper falhou - Não é seguro enviar à CEF</h2>"
        erro_html += f"<pre style='background:#fee;padding:20px;border:2px solid red'>{str(e)}</pre>"
        erro_html += f"<p><strong>Contrato:</strong> {contrato.codigo}</p>"
        
        return HttpResponse(erro_html, status=400)


def gerar_arquivo_rcv(request):
    """
    Gera arquivo RCV (Registro de Comprovação de Valores) para múltiplos contratos.
    Endpoint: /rcv/gerar/?contratos=1,2,3
    """
    from django.http import HttpResponse
    import logging
    logger = logging.getLogger(__name__)
    
    # Parâmetros
    contratos_ids = request.GET.get('contratos', '')
    teste_erro = request.GET.get('teste_erro') == '1'
    
    if not contratos_ids:
        return HttpResponse("<h1>❌ ERRO</h1><p>Informe os IDs dos contratos: ?contratos=1,2,3</p>", status=400)
    
    try:
        ids = [int(x.strip()) for x in contratos_ids.split(',')]
        contratos = Contrato.objects.filter(id__in=ids)
        
        if not contratos.exists():
            return HttpResponse("<h1>❌ ERRO</h1><p>Nenhum contrato encontrado com os IDs fornecidos</p>", status=404)
        
    except ValueError:
        return HttpResponse("<h1>❌ ERRO</h1><p>IDs de contratos inválidos</p>", status=400)
    
    # Dados fixos
    UFS = "33"
    MATRICULA_6 = "000442"
    MATRICULA_SEM_DV_5 = "00044"  # Primeiros 5 dígitos
    data_hoje = date.today()
    data_ddmmaa = fmt_data(data_hoje)
    data_ddmmaaaa = fmt_data_ddmmaaaa(data_hoje)
    sequencial_lote = "001"
    
    print(f"\n{'='*80}")
    print(f">>> GERANDO ARQUIVO RCV - {len(contratos)} contratos")
    print(f"{'='*80}\n")
    
    # ===== HEADER (tipo "0") - 430 bytes =====
    header = ""
    header += UFS  # 1-2
    header += MATRICULA_6  # 3-8
    header += "0"  # 9: Tipo HEADER
    header += data_ddmmaa  # 10-15
    header += sequencial_lote  # 16-18
    header += "COFLUHAB".ljust(30)  # 19-48
    header += fmt_num(len(contratos), 6)  # 49-54: Qtd registros
    header += "0".zfill(14)  # 55-68: Totalizador (não usado no RCV)
    header = pad_right(header, 430)
    
    linhas = [header]
    
    # ===== DETALHES (um por contrato) =====
    for contrato in contratos:
        contrato_13 = str(contrato.codigo).ljust(13)[:13]
        hipoteca_1 = "1"  # Padrão: com hipoteca
        
        # Gerar linha base (27 bytes)
        linha_det = gerar_linha_rcv_simplificado(
            MATRICULA_SEM_DV_5,
            contrato_13,
            hipoteca_1,
            data_ddmmaaaa
        )
        
        # Completar para 430 bytes (modo seguro para layouts fixos CEF)
        if teste_erro:
            # Modo teste: deixa com 27 bytes para testar validação
            pass
        else:
            linha_det = pad_right(linha_det, 430)
        
        linhas.append(linha_det)
        
        print(f"  • Contrato {contrato.codigo}: {len(linha_det)} bytes")
    
    # ===== TRAILER (tipo "9") - 430 bytes =====
    trailer = ""
    trailer += UFS  # 1-2
    trailer += MATRICULA_6  # 3-8
    trailer += "9"  # 9: Tipo TRAILER
    trailer += fmt_num(len(contratos), 6)  # 10-15: Total registros
    trailer += "0".zfill(14)  # 16-29: Totalizador (não usado)
    
    if teste_erro:
        # Modo teste: deixa incompleto
        pass
    else:
        trailer = pad_right(trailer, 430)
    
    linhas.append(trailer)
    
    print(f"\n>>> HEADER: {len(header)} bytes")
    print(f">>> DETALHES: {len(contratos)} linhas")
    print(f">>> TRAILER: {len(trailer)} bytes\n")
    
    # ===== GATEKEEPER RCV - CAMADA DE SEGURANÇA =====
    detalhes = linhas[1:-1]  # Pegar só os detalhes (entre header e trailer)
    
    try:
        # EXECUTA O GATEKEEPER RCV
        validar_rcv(header, detalhes, trailer)
        
        print(f"{'='*80}")
        print("✅ VALIDAÇÃO GATEKEEPER RCV APROVADA - Arquivo seguro para CEF")
        print(f"{'='*80}\n")
        
        # Monta conteúdo final com CRLF
        conteudo_str = "\r\n".join(linhas) + "\r\n"
        
        # Encoding latin-1 strict
        try:
            conteudo_bytes = conteudo_str.encode("latin-1", errors="strict")
        except UnicodeEncodeError as e:
            raise ValueError(f"Erro de encoding: caractere inválido para latin-1: {e}")
        
        logger.info(f"=== ARQUIVO RCV COMPLETO ===")
        logger.info(f"HEADER: {len(header)} bytes")
        logger.info(f"Detalhes: {len(detalhes)} registros")
        logger.info(f"TRAILER: {len(trailer)} bytes")
        logger.info(f"Total linhas: {len(linhas)}")
        logger.info(f"Total bytes: {len(conteudo_bytes)}")
        logger.info(f"Encoding: latin-1 (strict)")
        logger.info("=" * 50)
        
        # ===== VALIDAÇÃO AI AUTOMÁTICA RCV =====
        if request.GET.get('validar_ai') != '0':  # Por padrão validar, exceto se ?validar_ai=0
            try:
                import time
                from .ai_agents import validar_arquivo_com_ai, corrigir_com_agente_autofix, CREWAI_DISPONIVEL
                from .models import ValidacaoAI
                
                if CREWAI_DISPONIVEL:
                    print(f"\n{'='*80}")
                    print("🤖 INICIANDO VALIDAÇÃO AI AUTOMÁTICA - RCV")
                    print(f"{'='*80}\n")
                    
                    inicio = time.time()
                    resultado_ai = validar_arquivo_com_ai(conteudo_str, "RCV")
                    tempo_execucao = time.time() - inicio
                    
                    # ===== AUTO-CORREÇÃO INTELIGENTE (OPÇÃO 2) =====
                    correcao_info = None
                    analise_ia = ""
                    if not resultado_ai.get('aprovado'):
                        print("\n🧠 INICIANDO AUTO-FIX INTELIGENTE (Agente AI)...")
                        correcao_info = corrigir_com_agente_autofix(
                            conteudo_str, 
                            resultado_ai.get('resultado', ''),
                            "RCV"
                        )
                        
                        if 'analise_ia' in correcao_info:
                            analise_ia = correcao_info['analise_ia']
                            print(f"\n💡 ANÁLISE DO AGENTE AUTO-FIX:")
                            print(analise_ia[:500] + "...")
                        
                        if correcao_info['sucesso']:
                            print(f"✅ Auto-correção inteligente aplicada: {correcao_info['total_correcoes']} correções")
                            for correcao in correcao_info['correcoes_aplicadas']:
                                print(f"   {correcao}")
                            
                            # Substituir conteúdo pelo corrigido
                            conteudo_str = correcao_info['conteudo_corrigido']
                            conteudo_bytes = conteudo_str.encode('latin-1')
                            
                            print(f"\n📦 Arquivo corrigido ({len(conteudo_bytes)} bytes)")
                        else:
                            print("ℹ️ Nenhuma correção automática aplicada")
                    
                    # Salvar no histórico (sem contrato específico, pois são múltiplos)
                    validacao = ValidacaoAI.objects.create(
                        tipo_arquivo='RCV',
                        contrato=None,  # RCV tem múltiplos contratos
                        status='APROVADO' if resultado_ai.get('aprovado') else 'REPROVADO',
                        relatorio_completo=resultado_ai.get('resultado', '') + ('\n\n' + analise_ia if analise_ia else ''),
                        erros_encontrados=resultado_ai.get('erro', '') if not resultado_ai.get('sucesso') else '',
                        tempo_execucao=tempo_execucao,
                        tamanho_arquivo=len(conteudo_bytes),
                        agentes_utilizados='QA Engineer, Auto-Fix Engineer',
                        correcao_automatica=correcao_info['sucesso'] if correcao_info else False,
                        correcoes_aplicadas='\n'.join(correcao_info['correcoes_aplicadas']) if correcao_info else ''
                    )
                    
                    print(f"✅ Validação AI concluída em {tempo_execucao:.2f}s")
                    print(f"📊 Status: {validacao.status}")
                    print(f"💾 Histórico salvo (ID: {validacao.id})\n")
            except Exception as e:
                print(f"⚠️ Erro na validação AI (não bloqueia export): {e}")
        else:
            # Criar registro básico para histórico
            try:
                validacao = ValidacaoAI.objects.create(
                    tipo_arquivo='RCV',
                    contrato=None,  # RCV tem múltiplos contratos
                    status='APROVADO',
                    relatorio_completo=f"✅ Arquivo RCV exportado com sucesso\n\n📊 Estatísticas:\n- Tamanho: {len(conteudo_bytes)} bytes\n- Total de contratos: {len(contratos)}\n- Data de geração: {data_hoje.strftime('%d/%m/%Y')}",
                    tempo_execucao=0.1,
                    tamanho_arquivo=len(conteudo_bytes),
                    agentes_utilizados='Gatekeeper (validação manual)',
                    correcao_automatica=False,
                    correcoes_aplicadas=''
                )
                print(f"💾 Registro de exportação RCV salvo (ID: {validacao.id})")
            except Exception as e:
                print(f"⚠️ Erro ao salvar histórico: {e}")
        
        # Retorna bytes
        data_fmt = data_hoje.strftime("%Y%m%d")
        filename = f"RCV_{data_fmt}_{len(contratos)}contratos.txt"
        response = HttpResponse(conteudo_bytes, content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print(f"✅ Arquivo RCV gerado: {len(contratos)} contratos")
        return response
        
    except ValueError as e:
        # BLOQUEIO: validação falhou
        print(f"❌ Erro na geração do RCV: {str(e)}")
        logger.error(f"❌ Erro na validação RCV: {str(e)}")
        
        erro_html = f"<h1>❌ ARQUIVO RCV BLOQUEADO</h1>"
        erro_html += f"<h2>Validação Gatekeeper falhou - Não é seguro enviar à CEF</h2>"
        erro_html += f"<pre style='background:#fee;padding:20px;border:2px solid red'>{str(e)}</pre>"
        erro_html += f"<p><strong>Contratos:</strong> {len(contratos)}</p>"
        
        return HttpResponse(erro_html, status=400)


def integracao_cef(request):
    """
    Página principal de integração com a CEF.
    Centraliza todas as funcionalidades de envio/recebimento de arquivos.
    """
    return render(request, 'principal/integracao_cef.html')


def validacoes_ai(request):
    """
    Dashboard de histórico de validações AI
    """
    from .models import ValidacaoAI
    
    validacoes = ValidacaoAI.objects.all().order_by('-data_validacao')
    
    # Estatísticas
    total = validacoes.count()
    aprovadas = validacoes.filter(status='APROVADO').count()
    reprovadas = validacoes.filter(status='REPROVADO').count()
    
    context = {
        'validacoes': validacoes,
        'total': total,
        'aprovadas': aprovadas,
        'reprovadas': reprovadas,
    }
    
    return render(request, 'principal/validacoes_ai.html', context)


def validacao_ai_detail(request, validacao_id):
    """
    Detalhes de uma validação AI específica
    """
    from .models import ValidacaoAI
    
    validacao = get_object_or_404(ValidacaoAI, id=validacao_id)
    
    context = {
        'validacao': validacao
    }
    
    return render(request, 'principal/validacao_ai_detail.html', context)


def aprendizados_ai(request):
    """
    OPÇÃO 2: Dashboard de aprendizados do Auto-Fix Engineer
    """
    from .models import AprendizadoAI, ValidacaoAI
    
    aprendizados = AprendizadoAI.objects.all()
    
    # Estatísticas
    total = aprendizados.count()
    implementados = aprendizados.filter(implementado=True).count()
    pendentes = total - implementados
    
    # Erros mais comuns
    erros_top = aprendizados.order_by('-ocorrencias')[:5]
    
    context = {
        'aprendizados': aprendizados,
        'total': total,
        'implementados': implementados,
        'pendentes': pendentes,
        'erros_top': erros_top,
    }
    
    return render(request, 'principal/aprendizados_ai.html', context)


def analisar_padroes_ai(request):
    """
    OPÇÃO 2: Executa análise de padrões pelo Auto-Fix Engineer
    Fallback: Se CrewAI não disponível, usa análise básica de padrões
    """
    from .models import ValidacaoAI, AprendizadoAI
    from .ai_agents import analisar_padroes_erros_com_ai, CREWAI_DISPONIVEL
    from django.http import JsonResponse
    import re
    
    if request.method == 'POST':
        try:
            # Buscar últimas 50 validações
            validacoes = ValidacaoAI.objects.filter(status='REPROVADO')[:50]
            
            if not validacoes.exists():
                return JsonResponse({
                    'sucesso': False,
                    'erro': 'Nenhuma validação reprovada encontrada para análise'
                })
            
            # Se CrewAI disponível, usar análise inteligente
            if CREWAI_DISPONIVEL:
                historico = []
                for val in validacoes:
                    historico.append({
                        'erros': val.erros_encontrados or val.relatorio_completo,
                        'correcoes': val.correcoes_aplicadas,
                        'tipo_arquivo': val.tipo_arquivo,
                        'data': val.data_validacao
                    })
                
                resultado = analisar_padroes_erros_com_ai(historico)
                
                if resultado.get('sucesso'):
                    for erro, count in resultado.get('erros_mais_comuns', {}).items():
                        aprendizado, created = AprendizadoAI.objects.get_or_create(
                            tipo_erro=erro,
                            defaults={
                                'ocorrencias': count,
                                'causa_raiz': 'Analisando...',
                                'sugestao_codigo': resultado.get('analise_completa', ''),
                                'prevencao': 'Em análise',
                                'prioridade': 8
                            }
                        )
                        if not created:
                            aprendizado.ocorrencias += count
                            aprendizado.save()
                    
                    return JsonResponse({
                        'sucesso': True,
                        'mensagem': f'Análise AI concluída! {len(resultado.get("erros_mais_comuns", {}))} padrões identificados.',
                        'analise': resultado.get('analise_completa'),
                        'modo': 'AI'
                    })
            
            # FALLBACK: Análise básica sem CrewAI
            print("\n⚠️ CrewAI não disponível. Usando análise básica de padrões...")
            
            padroes = {}
            for val in validacoes:
                texto = val.erros_encontrados or val.relatorio_completo
                
                # Detectar padrões comuns
                if 'HEADER' in texto and '431' in texto:
                    padroes['HEADER_431_BYTES'] = padroes.get('HEADER_431_BYTES', 0) + 1
                
                if 'TRAILER' in texto and '71' in texto:
                    padroes['TRAILER_71_BYTES'] = padroes.get('TRAILER_71_BYTES', 0) + 1
                
                if '430 bytes' in texto or '430bytes' in texto:
                    padroes['TAMANHO_LINHA_INCORRETO'] = padroes.get('TAMANHO_LINHA_INCORRETO', 0) + 1
                
                if 'mutuário' in texto.lower() or 'mutuario' in texto.lower():
                    padroes['PROBLEMA_MUTUARIO'] = padroes.get('PROBLEMA_MUTUARIO', 0) + 1
            
            # Criar/atualizar aprendizados
            aprendizados_criados = 0
            
            if 'HEADER_431_BYTES' in padroes:
                apr, created = AprendizadoAI.objects.get_or_create(
                    tipo_erro='HEADER_431_BYTES',
                    defaults={
                        'ocorrencias': padroes['HEADER_431_BYTES'],
                        'causa_raiz': 'HEADER sendo gerado com 431 bytes ao invés de 430 bytes. Causa: preenchimento incorreto na função de geração.',
                        'sugestao_codigo': 'Verificar função pad_right() em views.py. Ajustar para garantir exatamente 430 bytes no HEADER.',
                        'prevencao': 'Validar tamanho da linha antes de adicionar à lista. Adicionar assert len(header) == 430.',
                        'arquivo_afetado': 'views.py (exportar_evolucao_txt)',
                        'prioridade': 9
                    }
                )
                if not created:
                    apr.ocorrencias += padroes['HEADER_431_BYTES']
                    apr.save()
                aprendizados_criados += 1
            
            if 'TRAILER_71_BYTES' in padroes:
                apr, created = AprendizadoAI.objects.get_or_create(
                    tipo_erro='TRAILER_71_BYTES',
                    defaults={
                        'ocorrencias': padroes['TRAILER_71_BYTES'],
                        'causa_raiz': 'TRAILER sendo gerado com apenas 71 bytes ao invés de 430 bytes. Falta preenchimento.',
                        'sugestao_codigo': 'Adicionar preenchimento com espaços: trailer = pad_right(trailer, 430)',
                        'prevencao': 'Sempre aplicar pad_right() antes de adicionar qualquer linha ao arquivo.',
                        'arquivo_afetado': 'views.py (exportar_evolucao_txt)',
                        'prioridade': 9
                    }
                )
                if not created:
                    apr.ocorrencias += padroes['TRAILER_71_BYTES']
                    apr.save()
                aprendizados_criados += 1
            
            if 'PROBLEMA_MUTUARIO' in padroes:
                apr, created = AprendizadoAI.objects.get_or_create(
                    tipo_erro='PROBLEMA_MUTUARIO',
                    defaults={
                        'ocorrencias': padroes['PROBLEMA_MUTUARIO'],
                        'causa_raiz': 'Contratos sem mutuário vinculado ou com dados incompletos.',
                        'sugestao_codigo': 'Adicionar validação preventiva antes de exportar. Verificar se contrato tem mutuário.',
                        'prevencao': 'Usar validators.py (Opção 3) para bloquear exportação de contratos sem mutuário.',
                        'arquivo_afetado': 'validators.py, views.py',
                        'prioridade': 7
                    }
                )
                if not created:
                    apr.ocorrencias += padroes['PROBLEMA_MUTUARIO']
                    apr.save()
                aprendizados_criados += 1
            
            return JsonResponse({
                'sucesso': True,
                'mensagem': f'✅ Análise básica concluída! {aprendizados_criados} padrões identificados (total: {len(validacoes)} validações).',
                'analise': f'Análise sem AI: {len(padroes)} tipos de problemas detectados',
                'modo': 'BASICO',
                'padroes': padroes,
                'nota': '⚠️ CrewAI não disponível. Usando análise básica de padrões.'
            })
            
        except Exception as e:
            return JsonResponse({
                'sucesso': False,
                'erro': str(e)
            })
    
    return render(request, 'principal/analisar_padroes.html')


def implementar_aprendizado(request, aprendizado_id):
    """
    OPÇÃO 2: Marca um aprendizado como implementado
    POST /aprendizados-ai/<id>/implementar/
    """
    from .models import AprendizadoAI
    from django.http import JsonResponse
    from django.utils import timezone
    import json
    
    if request.method == 'POST':
        try:
            aprendizado = get_object_or_404(AprendizadoAI, id=aprendizado_id)
            
            # Dados do POST (opcional)
            try:
                data = json.loads(request.body)
                arquivo = data.get('arquivo_modificado', '')
                linha = data.get('linha', '')
                comentario = data.get('comentario', '')
            except:
                arquivo = linha = comentario = ''
            
            # Marcar como implementado
            aprendizado.implementado = True
            aprendizado.data_implementacao = timezone.now()
            
            # Atualizar arquivo afetado
            if arquivo:
                aprendizado.arquivo_afetado = f"{arquivo}" + (f" linha {linha}" if linha else "")
            
            # Adicionar comentário
            if comentario:
                aprendizado.causa_raiz += f"\n\n[IMPLEMENTADO em {timezone.now().strftime('%Y-%m-%d %H:%M')}]\n{comentario}"
            
            aprendizado.save()
            
            return JsonResponse({
                'sucesso': True,
                'mensagem': f'✅ Aprendizado marcado como implementado!',
                'id': aprendizado.id
            })
            
        except Exception as e:
            return JsonResponse({
                'sucesso': False,
                'erro': str(e)
            }, status=500)
    
    return JsonResponse({'erro': 'Método não permitido'}, status=405)


def detalhes_aprendizado(request, aprendizado_id):
    """
    Retorna detalhes completos de um aprendizado AI em JSON
    """
    try:
        aprendizado = AprendizadoAI.objects.get(id=aprendizado_id)
        
        data = {
            'sucesso': True,
            'aprendizado': {
                'id': aprendizado.id,
                'tipo_erro': aprendizado.tipo_erro,
                'ocorrencias': aprendizado.ocorrencias,
                'causa_raiz': aprendizado.causa_raiz,
                'sugestao_codigo': aprendizado.sugestao_codigo,
                'prevencao': aprendizado.prevencao,
                'arquivo_afetado': aprendizado.arquivo_afetado or 'N/A',
                'prioridade': aprendizado.prioridade,
                'implementado': aprendizado.implementado,
                'data_analise': aprendizado.data_analise.strftime('%d/%m/%Y %H:%M'),
                'data_implementacao': aprendizado.data_implementacao.strftime('%d/%m/%Y %H:%M') if aprendizado.data_implementacao else None,
                'arquivo_modificado': getattr(aprendizado, 'arquivo_modificado', None),
                'linha': getattr(aprendizado, 'linha', None),
                'comentario': getattr(aprendizado, 'comentario', None)
            }
        }
        
        return JsonResponse(data)
        
    except AprendizadoAI.DoesNotExist:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Aprendizado não encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


def testar_ai_agents(request):
    """
    Página de testes dos agentes AI
    """
    from django.http import JsonResponse
    from .ai_agents import exemplo_validacao_simples, validar_arquivo_com_ai, CREWAI_DISPONIVEL
    
    if request.method == 'POST':
        # Verificar se CrewAI está disponível
        if not CREWAI_DISPONIVEL:
            return JsonResponse({
                'sucesso': False,
                'erro': '⚠️ CrewAI ainda não está disponível. A instalação está em andamento. Por favor, aguarde alguns minutos e recarregue a página.'
            })
        
        # Exemplo de validação
        tipo_teste = request.POST.get('tipo', 'simples')
        
        try:
            if tipo_teste == 'simples':
                resultado = exemplo_validacao_simples()
                return JsonResponse({
                    'sucesso': True,
                    'resultado': str(resultado)
                })
            
            elif tipo_teste == 'validar_arquivo':
                conteudo = request.POST.get('conteudo', '')
                tipo_arquivo = request.POST.get('tipo_arquivo', 'FH1')
                resultado = validar_arquivo_com_ai(conteudo, tipo_arquivo)
                return JsonResponse(resultado)
        except Exception as e:
            return JsonResponse({
                'sucesso': False,
                'erro': f'Erro ao executar teste: {str(e)}'
            })
    
    # Página de interface
    context = {
        'crewai_disponivel': CREWAI_DISPONIVEL
    }
    return render(request, 'principal/test_ai_agents.html', context)


def contrato_editar(request, pk):
    """Editar dados do contrato e, opcionalmente, dados básicos do mutuário."""
    from django.contrib import messages

    contrato = get_object_or_404(Contrato, pk=pk)

    mutuario = Mutuario.objects.filter(codigo=contrato.codigo).first()
    if not mutuario:
        db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
            result = cur.fetchone()
            if result:
                mutuario = Mutuario.objects.filter(id=result[0]).first()
            conn.close()
        except Exception:
            mutuario = None

    field_errors = {}
    atualizar_mutuario_checked = False

    if request.method == 'POST':
        erros = []
        field_errors = defaultdict(list)
        atualizar_mutuario = request.POST.get('atualizar_mutuario') == 'sim'
        atualizar_mutuario_checked = atualizar_mutuario

        def add_error(field_name, message):
            erros.append(message)
            field_errors[field_name].append(message)

        # Campos texto do contrato
        contrato.conjunto = (request.POST.get('conjunto', '') or '').strip()
        contrato.cod_imovel = (request.POST.get('cod_imovel', '') or '').strip()
        contrato.ocorrencia = (request.POST.get('ocorrencia', '') or '').strip().upper()
        contrato.chave = (request.POST.get('chave', '') or '').strip()
        contrato.lote = (request.POST.get('lote', '') or '').strip()
        contrato.sinal = (request.POST.get('sinal', '') or '').strip()
        contrato.sa = (request.POST.get('sa', '') or '').strip().upper()
        contrato.cat_prof = (request.POST.get('cat_prof', '') or '').strip()
        contrato.pr = (request.POST.get('pr', '') or '').strip().upper()

        # Datas
        data_contrato_raw = (request.POST.get('data_contrato', '') or '').strip()
        if data_contrato_raw:
            data_contrato = _parse_data_flexivel(data_contrato_raw)
            if data_contrato is None:
                add_error('data_contrato', 'Data do contrato inválida.')
            else:
                contrato.data_contrato = data_contrato
        else:
            contrato.data_contrato = None

        data_primeiro_venc_raw = (request.POST.get('data_primeiro_venc', '') or '').strip()
        if data_primeiro_venc_raw:
            data_primeiro_venc = _parse_data_flexivel(data_primeiro_venc_raw)
            if data_primeiro_venc is None:
                add_error('data_primeiro_venc', 'Data do 1º vencimento inválida.')
            else:
                contrato.data_primeiro_venc = data_primeiro_venc
        else:
            contrato.data_primeiro_venc = None

        # Números
        tx_juros_raw = (request.POST.get('tx_juros', '') or '').strip()
        if tx_juros_raw:
            tx_juros = _parse_decimal_ocr(tx_juros_raw)
            if tx_juros is None:
                add_error('tx_juros', 'Taxa de juros inválida.')
            else:
                contrato.tx_juros = tx_juros
        else:
            contrato.tx_juros = None

        prazo_raw = (request.POST.get('prazo', '') or '').strip()
        if prazo_raw:
            prazo = _parse_int_ocr(prazo_raw)
            if prazo is None:
                add_error('prazo', 'Prazo inválido.')
            else:
                contrato.prazo = prazo
        else:
            contrato.prazo = None

        vlfinanc_raw = (request.POST.get('vlfinanc', '') or '').strip()
        if vlfinanc_raw:
            vlfinanc = _parse_decimal_ocr(vlfinanc_raw)
            if vlfinanc is None:
                add_error('vlfinanc', 'Valor do financiamento inválido.')
            else:
                contrato.vlfinanc = vlfinanc
        else:
            contrato.vlfinanc = None

        vlprop_raw = (request.POST.get('vlprop', '') or '').strip()
        if vlprop_raw:
            vlprop = _parse_decimal_ocr(vlprop_raw)
            if vlprop is None:
                add_error('vlprop', 'Valor do imóvel inválido.')
            else:
                contrato.vlprop = vlprop
        else:
            contrato.vlprop = None

        prestacao_inicial_raw = (request.POST.get('prestacao_inicial', '') or '').strip()
        if prestacao_inicial_raw:
            prestacao_inicial = _parse_decimal_ocr(prestacao_inicial_raw)
            if prestacao_inicial is None:
                add_error('prestacao_inicial', 'Prestação inicial inválida.')
            else:
                contrato.prestacao_inicial = prestacao_inicial
        else:
            contrato.prestacao_inicial = None

        # Regras mínimas para evitar salvar cadastro incompleto
        if not contrato.conjunto:
            add_error('conjunto', 'Conjunto é obrigatório.')
        if not contrato.sa:
            add_error('sa', 'Sistema de amortização (SA) é obrigatório.')
        if contrato.tx_juros is None:
            add_error('tx_juros', 'Taxa de juros é obrigatória.')
        elif contrato.tx_juros <= 0:
            add_error('tx_juros', 'Taxa de juros deve ser maior que zero.')
        if contrato.prazo is None:
            add_error('prazo', 'Prazo é obrigatório.')
        elif contrato.prazo <= 0:
            add_error('prazo', 'Prazo deve ser maior que zero.')
        if contrato.vlfinanc is None:
            add_error('vlfinanc', 'Valor do financiamento é obrigatório.')
        elif contrato.vlfinanc <= 0:
            add_error('vlfinanc', 'Valor do financiamento deve ser maior que zero.')

        if atualizar_mutuario and not mutuario:
            add_error('atualizar_mutuario', 'Não foi possível localizar o mutuário vinculado para atualizar.')

        if atualizar_mutuario and mutuario:
            mutuario.nome = (request.POST.get('mutuario_nome', mutuario.nome) or '').strip()
            cpf_mutuario = request.POST.get('mutuario_cpf', mutuario.cpf)
            mutuario.cpf = _limpar_cpf(cpf_mutuario)
            mutuario.ident = (request.POST.get('mutuario_ident', mutuario.ident) or '').strip()
            mutuario.orgao = (request.POST.get('mutuario_orgao', mutuario.orgao) or '').strip()
            mutuario.endereco = (request.POST.get('mutuario_endereco', mutuario.endereco) or '').strip()
            mutuario.numero = (request.POST.get('mutuario_numero', mutuario.numero) or '').strip()
            mutuario.compl = (request.POST.get('mutuario_compl', mutuario.compl) or '').strip()
            mutuario.bairro = (request.POST.get('mutuario_bairro', mutuario.bairro) or '').strip()
            mutuario.cidade = (request.POST.get('mutuario_cidade', mutuario.cidade) or '').strip()
            mutuario.uf = (request.POST.get('mutuario_uf', mutuario.uf) or '').strip().upper()
            mutuario.cep = (request.POST.get('mutuario_cep', mutuario.cep) or '').strip()
            mutuario.telefone = (request.POST.get('mutuario_telefone', mutuario.telefone) or '').strip()
            mutuario.email = (request.POST.get('mutuario_email', mutuario.email) or '').strip()
            mutuario.tipoimovel = (request.POST.get('mutuario_tipoimovel', mutuario.tipoimovel) or '').strip()

            mutuario_dtnasc_raw = (request.POST.get('mutuario_dtnasc', '') or '').strip()
            if mutuario_dtnasc_raw:
                mutuario_dtnasc = _parse_data_flexivel(mutuario_dtnasc_raw)
                if mutuario_dtnasc is None:
                    add_error('mutuario_dtnasc', 'Data de nascimento do mutuário inválida.')
                else:
                    mutuario.dtnasc = mutuario_dtnasc
            else:
                mutuario.dtnasc = None

            mutuario_renda_raw = (request.POST.get('mutuario_renda', '') or '').strip()
            if mutuario_renda_raw:
                renda = _parse_decimal_ocr(mutuario_renda_raw)
                if renda is None:
                    add_error('mutuario_renda', 'Renda do mutuário inválida.')
                else:
                    mutuario.renda = float(renda)
            else:
                mutuario.renda = None

            mutuario_crenda_raw = (request.POST.get('mutuario_crenda', '') or '').strip()
            if mutuario_crenda_raw:
                crenda = _parse_decimal_ocr(mutuario_crenda_raw)
                if crenda is None:
                    add_error('mutuario_crenda', 'Complemento de renda do mutuário inválido.')
                else:
                    mutuario.crenda = float(crenda)
            else:
                mutuario.crenda = None

            if not mutuario.nome:
                add_error('mutuario_nome', 'Nome do mutuário é obrigatório quando a atualização do mutuário está marcada.')
            if len(mutuario.cpf or '') != 11:
                add_error('mutuario_cpf', 'CPF do mutuário inválido. Informe 11 dígitos.')
            if not mutuario.ident:
                add_error('mutuario_ident', 'Identidade do mutuário é obrigatória quando a atualização do mutuário está marcada.')
            if not mutuario.orgao:
                add_error('mutuario_orgao', 'Órgão expedidor é obrigatório quando a atualização do mutuário está marcada.')
            if not mutuario.endereco:
                add_error('mutuario_endereco', 'Logradouro do mutuário é obrigatório quando a atualização do mutuário está marcada.')
            if not mutuario.numero:
                add_error('mutuario_numero', 'Número do imóvel é obrigatório quando a atualização do mutuário está marcada.')
            if not mutuario.bairro:
                add_error('mutuario_bairro', 'Bairro do mutuário é obrigatório quando a atualização do mutuário está marcada.')
            if not mutuario.cidade:
                add_error('mutuario_cidade', 'Cidade do mutuário é obrigatória quando a atualização do mutuário está marcada.')
            if len(mutuario.uf or '') != 2:
                add_error('mutuario_uf', 'UF do mutuário inválida. Informe 2 letras.')

            cep_digitos = re.sub(r'\D+', '', mutuario.cep or '')
            if len(cep_digitos) != 8:
                add_error('mutuario_cep', 'CEP do mutuário inválido. Informe 8 dígitos.')

            telefone_digitos = re.sub(r'\D+', '', mutuario.telefone or '')
            if len(telefone_digitos) not in (10, 11):
                add_error('mutuario_telefone', 'Telefone inválido. Informe DDD + número.')

            email_txt = (mutuario.email or '').strip()
            if not email_txt:
                add_error('mutuario_email', 'E-mail é obrigatório quando a atualização do mutuário está marcada.')
            elif not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email_txt):
                add_error('mutuario_email', 'E-mail inválido.')

        if erros:
            messages.error(request, 'Corrija os campos destacados para salvar.')
        else:
            contrato.save()
            if atualizar_mutuario and mutuario:
                mutuario.save()

            messages.success(request, 'Contrato atualizado com sucesso.')
            return redirect('contrato_detail', pk=contrato.id)

    return render(request, 'principal/contrato_editar.html', {
        'contrato': contrato,
        'mutuario': mutuario,
        'field_errors': field_errors,
        'atualizar_mutuario_checked': atualizar_mutuario_checked,
    })


def debito_prestacoes(request, codigo):
    """Mostra débito de prestações em aberto de um contrato"""
    from datetime import date
    from decimal import Decimal
    
    contrato = get_object_or_404(Contrato, codigo=codigo)
    
    # Data de referência para cálculo (hoje ou data fornecida)
    # Compatibilidade com template antigo: usa ate_data.
    data_calculo_str = request.GET.get('ate_data', '') or request.GET.get('data_calculo', '')
    if data_calculo_str:
        try:
            data_calculo = datetime.strptime(data_calculo_str, '%Y-%m-%d').date()
        except ValueError:
            data_calculo = date.today()
    else:
        data_calculo = date.today()
    
    parcelas = ParcelaContrato.objects.filter(contrato=contrato, dtpgto__isnull=True).order_by('dtvenc')
    
    debito_total = Decimal('0')
    mora_total_geral = Decimal('0')
    encargo_total_geral = Decimal('0')
    parcelas_list = []
    acumulado = Decimal('0')
    
    # Taxa de mora: 0,0333% ao dia (1% ao mês / 30 dias)
    taxa_mora_dia = Decimal('0.000333')
    
    for p in parcelas:
        # Calcular encargo mensal (usar vlautent se disponível, senão somar componentes)
        if p.vlautent and p.vlautent > 0:
            encargo_mensal = Decimal(str(p.vlautent))
        else:
            encargo_mensal = Decimal('0')
            if p.juros: encargo_mensal += Decimal(str(p.juros))
            if p.amort: encargo_mensal += Decimal(str(p.amort))
            if p.seguro: encargo_mensal += Decimal(str(p.seguro))
            if p.tca: encargo_mensal += Decimal(str(p.tca))
            if p.fcvs: encargo_mensal += Decimal(str(p.fcvs))
            if p.em: encargo_mensal += Decimal(str(p.em))
            if p.rp: encargo_mensal += Decimal(str(p.rp))
        
        # Calcular dias de atraso
        dias_atraso = 0
        mora_total = Decimal('0')
        mora_dia = Decimal('0')
        
        if p.dtvenc and p.dtvenc < data_calculo:
            dias_atraso = (data_calculo - p.dtvenc).days
            mora_dia = encargo_mensal * taxa_mora_dia
            mora_total = mora_dia * dias_atraso
        
        # Total = encargo + mora
        total = encargo_mensal + mora_total
        acumulado += total
        
        debito_total += total
        mora_total_geral += mora_total
        encargo_total_geral += encargo_mensal
        
        parcelas_list.append({
            'parcela': p,
            'nmens': p.nmens,
            'dtvenc': p.dtvenc,
            'dias_atraso': dias_atraso,
            'mora_dia': mora_dia,
            'encargo_mensal': encargo_mensal,
            'mora_total': mora_total,
            'total': total,
            'acumulado': acumulado
        })

    if (request.GET.get('exportar', '') or '').lower() == 'excel':
        try:
            from openpyxl import Workbook
        except ImportError:
            return HttpResponse('openpyxl nao esta instalado no ambiente.', status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Debito Prestacoes'

        ws.append([
            'Contrato', 'Conjunto', 'Parcela', 'Vencimento',
            'Dias Atraso', 'Encargo Mensal',
            'Mora Total', 'Total', 'Acumulado'
        ])

        for item in parcelas_list:
            ws.append([
                contrato.codigo,
                contrato.conjunto,
                item['nmens'],
                item['dtvenc'].strftime('%d/%m/%Y') if item['dtvenc'] else '',
                item['dias_atraso'],
                float(item['encargo_mensal']),
                float(item['mora_total']),
                float(item['total']),
                float(item['acumulado']),
            ])

        ws.append(['', '', '', 'TOTAIS', '', '', float(encargo_total_geral), float(mora_total_geral), float(debito_total), ''])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="debito_prestacoes_{contrato.codigo}_{data_calculo.strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response
    
    context = {
        'contrato': contrato,
        'parcelas': parcelas_list,
        'debito_total': debito_total,
        'mora_total_geral': mora_total_geral,
        'encargo_total_geral': encargo_total_geral,
        'data_calculo': data_calculo,
        'ate_data': data_calculo,
        'taxa_mora_dia': taxa_mora_dia * 100,  # Para exibir em %
        'taxa_mora': taxa_mora_dia * 100,
        'total_parcelas': len(parcelas_list),
        'total_parcelas_aberto': len(parcelas_list),
    }
    return render(request, 'principal/debito_prestacoes.html', context)


def relatorio_divida_seguro(request):
    """Tela para gerar e listar relatórios de dívida de seguro por ano."""
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

    def _resolver_exports_dir():
        exports_candidates = [
            os.path.abspath(os.path.join(base_dir, '..', 'exports')),
            os.path.join(base_dir, 'exports'),
        ]
        return next((p for p in exports_candidates if os.path.isdir(p)), exports_candidates[0])

    exports_dir = _resolver_exports_dir()
    script_path = os.path.join(base_dir, 'scripts', 'gerar_relatorio_divida_seguro_v2.py')
    preferred_python = os.path.abspath(os.path.join(base_dir, '..', 'venv_django', 'Scripts', 'python.exe'))
    python_exec = preferred_python if os.path.isfile(preferred_python) else sys.executable

    ano_inicial_estudo = 1992
    ano_final_estudo = 2019
    anos_disponiveis = list(range(ano_inicial_estudo, ano_final_estudo + 1))
    exec_result = None
    ano_selecionado = request.GET.get('ano', '')
    ano_inicio_selecionado = request.GET.get('ano_inicio', str(ano_inicial_estudo))
    ano_fim_selecionado = request.GET.get('ano_fim', str(ano_final_estudo))
    contrato_selecionado = (request.GET.get('contrato') or '').strip()

    try:
        if ano_selecionado and (int(ano_selecionado) < ano_inicial_estudo or int(ano_selecionado) > ano_final_estudo):
            ano_selecionado = ''
    except Exception:
        ano_selecionado = ''

    # Se o período informado for diferente (ex.: 2010-2019), o intervalo prevalece sobre Ano único.
    try:
        if ano_inicio_selecionado and ano_fim_selecionado and int(ano_inicio_selecionado) != int(ano_fim_selecionado):
            ano_selecionado = ''
    except Exception:
        pass

    # Quando vier por link anual (?ano=YYYY), sincroniza os campos de intervalo para o mesmo ano.
    if ano_selecionado:
        ano_inicio_selecionado = ano_selecionado
        ano_fim_selecionado = ano_selecionado

    if request.method == 'POST':
        ano_form = (request.POST.get('ano') or '').strip()
        ano_inicio_form = (request.POST.get('ano_inicio') or '').strip()
        ano_fim_form = (request.POST.get('ano_fim') or '').strip()

        ano_inicio_selecionado = ano_inicio_form or str(ano_inicial_estudo)
        ano_fim_selecionado = ano_fim_form or str(ano_final_estudo)

        anos_para_gerar = []
        try:
            ano_inicio = int(ano_inicio_form)
            ano_fim = int(ano_fim_form)
            if ano_inicio < ano_inicial_estudo or ano_fim > ano_final_estudo or ano_inicio > ano_fim:
                raise ValueError()
            anos_para_gerar = list(range(ano_inicio, ano_fim + 1))
        except Exception:
            exec_result = {
                'ok': False,
                'returncode': -2,
                'output': (
                    f'Informe um intervalo válido entre {ano_inicial_estudo} e {ano_final_estudo} '
                    f'(ano inicial <= ano final).'
                ),
            }

        if not ano_form and not ano_inicio_form and not ano_fim_form:
            exec_result = {
                'ok': False,
                'returncode': -2,
                'output': (
                    'Geracao consolidada (1992-2019) desativada na tela para evitar travamentos. '\
                    'Use um ano especifico ou um intervalo de anos aqui, ou rode no terminal: '\
                    'python manage.py gerar_relatorio_divida_seguro'
                ),
            }
        elif exec_result is None and anos_para_gerar:
            logs_execucao = []
            sucesso_total = True
            ultimo_returncode = 0

            logs_execucao.append(
                f"Gerando período {anos_para_gerar[0]}-{anos_para_gerar[-1]} "
                f"({len(anos_para_gerar)} ano(s))"
            )

            for ano_exec in anos_para_gerar:
                cmd = [python_exec, script_path, '--ano', str(ano_exec)]
                try:
                    run = subprocess.run(
                        cmd,
                        cwd=base_dir,
                        capture_output=True,
                        text=True,
                        timeout=240,
                        check=False,
                    )
                    saida = (run.stdout or '') + ('\n' + run.stderr if run.stderr else '')
                    trecho_saida = saida[-700:] if saida else '(sem saída)'
                    logs_execucao.append(f"[{ano_exec}] returncode={run.returncode}\n{trecho_saida}")
                    if run.returncode != 0:
                        sucesso_total = False
                    ultimo_returncode = run.returncode
                except subprocess.TimeoutExpired:
                    sucesso_total = False
                    ultimo_returncode = -1
                    logs_execucao.append(f"[{ano_exec}] timeout de 240s")

            # Para período multi-ano, não fixa no último ano (evita impressão de "gerou só 2019").
            if len(anos_para_gerar) == 1:
                ano_selecionado = str(anos_para_gerar[0])
            else:
                ano_selecionado = ''
            exec_result = {
                'ok': sucesso_total,
                'returncode': ultimo_returncode,
                'output': "\n\n".join(logs_execucao)[-5000:],
            }

    # Anos de referência para os quadros de resumo, respeitando o filtro do menu.
    anos_resumo = anos_disponiveis
    try:
        ano_ini_ref = int(ano_inicio_selecionado)
        ano_fim_ref = int(ano_fim_selecionado)
        intervalo_valido = (
            ano_inicial_estudo <= ano_ini_ref <= ano_final_estudo
            and ano_inicial_estudo <= ano_fim_ref <= ano_final_estudo
            and ano_ini_ref <= ano_fim_ref
        )

        if intervalo_valido:
            anos_resumo = list(range(ano_ini_ref, ano_fim_ref + 1))
            if ano_ini_ref == ano_fim_ref and ano_selecionado:
                ano_ref = int(ano_selecionado)
                if ano_inicial_estudo <= ano_ref <= ano_final_estudo:
                    anos_resumo = [ano_ref]
    except Exception:
        anos_resumo = anos_disponiveis

    arquivos = []
    arquivos_por_periodo = defaultdict(list)
    resumo_base = None
    resumo_global = None
    arquivos_fixados = []

    def _resumir_csv(path_csv):
        total = 0
        encontrados = 0
        nao_encontrados = 0
        tipos = set()
        total_seguro = Decimal('0')

        try:
            with open(path_csv, 'r', encoding='utf-8-sig', newline='') as fp:
                reader = csv.DictReader(fp, delimiter=';')
                for row in reader:
                    total += 1
                    status = (row.get('status_confronto') or row.get('status') or '').strip().upper()
                    if status == 'ENCONTRADO':
                        encontrados += 1
                    elif status == 'NAO_ENCONTRADO_NO_BANCO':
                        nao_encontrados += 1

                    motivo = (row.get('motivo_finalizacao') or '').strip()
                    if motivo:
                        tipos.add(motivo)

                    total_seguro += _parse_decimal_br(row.get('valor_seguro_pdf', '0'))

            periodo_txt = '-'

            return {
                'arquivo_csv': os.path.basename(path_csv),
                'total_registros': total,
                'encontrados': encontrados,
                'nao_encontrados': nao_encontrados,
                'periodo': periodo_txt,
                'tipos': sorted(tipos),
                'total_seguro': f"{total_seguro:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                'total_seguro_decimal': total_seguro,
            }
        except Exception:
            return None

    def _resumir_ultimos_csvs_anuais() -> dict | None:
        total = 0
        encontrados = 0
        nao_encontrados = 0
        tipos = set()
        total_seguro = Decimal('0')
        anos_com_csv = []
        anos_detalhe = []

        for ano in anos_resumo:
            ano_csv = sorted(Path(exports_dir).glob(f'divida_seguro_{ano}_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
            if not ano_csv:
                anos_detalhe.append({
                    'ano': ano,
                    'registros': 0,
                    'total_seguro': '0,00',
                    'zerado': False,
                    'sem_csv': True,
                })
                continue
            anos_com_csv.append(ano)
            total_seguro_ano = Decimal('0')
            registros_ano = 0
            try:
                with open(ano_csv[0], 'r', encoding='utf-8-sig', newline='') as fp:
                    reader = csv.DictReader(fp, delimiter=';')
                    for row in reader:
                        registros_ano += 1
                        total += 1
                        status = (row.get('status_confronto') or row.get('status') or '').strip().upper()
                        if status == 'ENCONTRADO':
                            encontrados += 1
                        elif status == 'NAO_ENCONTRADO_NO_BANCO':
                            nao_encontrados += 1

                        motivo = (row.get('motivo_finalizacao') or '').strip()
                        if motivo:
                            tipos.add(motivo)

                        valor_seguro = _parse_decimal_br(row.get('valor_seguro_pdf', '0'))
                        total_seguro += valor_seguro
                        total_seguro_ano += valor_seguro

                anos_detalhe.append({
                    'ano': ano,
                    'registros': registros_ano,
                    'total_seguro': f"{total_seguro_ano:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'zerado': total_seguro_ano == 0,
                    'sem_csv': False,
                })
            except Exception:
                anos_detalhe.append({
                    'ano': ano,
                    'registros': 0,
                    'total_seguro': '0,00',
                    'zerado': False,
                    'sem_csv': True,
                })
                continue

        if not anos_com_csv:
            return None

        return {
            'arquivo_csv': f"Agregado dinâmico ({len(anos_com_csv)} ano(s))",
            'total_registros': total,
            'encontrados': encontrados,
            'nao_encontrados': nao_encontrados,
            'periodo': f"{min(anos_com_csv)}-{max(anos_com_csv)}",
            'tipos': sorted(tipos),
            'total_seguro': f"{total_seguro:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
            'total_seguro_decimal': total_seguro,
            'anos_com_csv': anos_com_csv,
            'anos_detalhe': anos_detalhe,
        }
    if os.path.isdir(exports_dir):
        # Arquivos fixados (sem necessidade de novo upload): último consolidado + último por ano.
        consolidado_csv = sorted(Path(exports_dir).glob('divida_seguro_consolidado_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
        consolidado_md = sorted(Path(exports_dir).glob('laudo_divida_seguro_consolidado_*.md'), key=lambda p: p.stat().st_mtime, reverse=True)
        resumo_base_anos = _resumir_ultimos_csvs_anuais()
        if resumo_base_anos:
            resumo_base = resumo_base_anos
        elif consolidado_csv:
            arquivos_fixados.append({
                'nome': consolidado_csv[0].name,
                'tipo': 'CSV Consolidado',
                'modificado_em': datetime.fromtimestamp(consolidado_csv[0].stat().st_mtime),
            })
            resumo_base = _resumir_csv(str(consolidado_csv[0]))
        if consolidado_md:
            arquivos_fixados.append({
                'nome': consolidado_md[0].name,
                'tipo': 'Laudo Consolidado',
                'modificado_em': datetime.fromtimestamp(consolidado_md[0].stat().st_mtime),
            })

        for ano in anos_disponiveis:
            ano_csv = sorted(Path(exports_dir).glob(f'divida_seguro_{ano}_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
            ano_md = sorted(Path(exports_dir).glob(f'laudo_divida_seguro_{ano}_*.md'), key=lambda p: p.stat().st_mtime, reverse=True)
            if ano_csv:
                arquivos_fixados.append({
                    'nome': ano_csv[0].name,
                    'tipo': f'CSV {ano}',
                    'modificado_em': datetime.fromtimestamp(ano_csv[0].stat().st_mtime),
                })
            if ano_md:
                arquivos_fixados.append({
                    'nome': ano_md[0].name,
                    'tipo': f'Laudo {ano}',
                    'modificado_em': datetime.fromtimestamp(ano_md[0].stat().st_mtime),
                })

        comparativo_csv = sorted(Path(exports_dir).glob('confronto_divida_seguro_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
        if comparativo_csv:
            arquivos_fixados.append({
                'nome': comparativo_csv[0].name,
                'tipo': 'CSV Comparativo',
                'modificado_em': datetime.fromtimestamp(comparativo_csv[0].stat().st_mtime),
            })

        for path in Path(exports_dir).glob('divida_seguro_*.csv'):
            nome = path.name
            if ano_selecionado and f'_{ano_selecionado}_' not in nome:
                continue
            item = {
                'nome': nome,
                'tipo': 'CSV',
                'modificado_em': datetime.fromtimestamp(path.stat().st_mtime),
            }
            arquivos.append(item)

            periodo = 'Consolidado'
            match_ano = re.search(r'divida_seguro_(\d{4})_', nome)
            if match_ano:
                periodo = match_ano.group(1)
            elif 'consolidado' in nome:
                periodo = 'Consolidado'
            arquivos_por_periodo[periodo].append(item)

        for path in Path(exports_dir).glob('laudo_divida_seguro_*.md'):
            nome = path.name
            if ano_selecionado and f'_{ano_selecionado}_' not in nome and 'consolidado' not in nome:
                continue
            item = {
                'nome': nome,
                'tipo': 'Laudo',
                'modificado_em': datetime.fromtimestamp(path.stat().st_mtime),
            }
            arquivos.append(item)

            periodo = 'Consolidado'
            match_ano = re.search(r'laudo_divida_seguro_(\d{4})_', nome)
            if match_ano:
                periodo = match_ano.group(1)
            elif 'consolidado' in nome:
                periodo = 'Consolidado'
            arquivos_por_periodo[periodo].append(item)

        for path in Path(exports_dir).glob('confronto_divida_seguro_*.csv'):
            nome = path.name
            if ano_selecionado and ano_selecionado not in nome:
                continue
            item = {
                'nome': nome,
                'tipo': 'CSV Comparativo',
                'modificado_em': datetime.fromtimestamp(path.stat().st_mtime),
            }
            arquivos.append(item)
            arquivos_por_periodo['Comparativo'].append(item)

    arquivos.sort(key=lambda x: x['modificado_em'], reverse=True)

    # Ordena os blocos de periodo: anos desc + consolidado por ultimo
    periodos_ordenados = []
    for periodo, itens in arquivos_por_periodo.items():
        itens.sort(key=lambda x: x['modificado_em'], reverse=True)
        periodos_ordenados.append((periodo, itens))

    def _period_key(item):
        periodo = item[0]
        if periodo == 'Consolidado':
            return -1
        try:
            return int(periodo)
        except ValueError:
            return 0

    periodos_ordenados.sort(key=_period_key, reverse=True)

    preview_relatorio = None

    def _fmt_decimal_br(valor: Decimal) -> str:
        return f"{valor:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')

    def _ipca_mensal(ano: int, mes: int) -> Decimal:
        chave = f"{ano}-{mes:02d}"
        try:
            return Decimal(str(INDICES_HISTORICOS.get(chave, Decimal('0'))))
        except Exception:
            return Decimal('0')

    def _ipca_fator_acumulado_ate_data(ano: int, mes: int, data_ref: date | None = None) -> Decimal:
        """Retorna fator acumulado de IPCA do mês-base até a data de referência (inclusive por competência)."""
        if not data_ref:
            data_ref = date.today()

        ano_mes_inicio = (ano, mes)
        ano_mes_fim = (data_ref.year, data_ref.month)
        if ano_mes_inicio > ano_mes_fim:
            return Decimal('1')

        fator = Decimal('1')
        a = ano
        m = mes
        while (a, m) <= ano_mes_fim:
            fator *= (Decimal('1') + _ipca_mensal(a, m))
            if m == 12:
                a += 1
                m = 1
            else:
                m += 1
        return fator

    _bcb_month_cache: dict[tuple[int, int, int], Decimal | None] = {}

    def _iterar_ano_mes(inicio_ano: int, inicio_mes: int, fim_ano: int, fim_mes: int):
        a = inicio_ano
        m = inicio_mes
        while (a, m) <= (fim_ano, fim_mes):
            yield a, m
            if m == 12:
                a += 1
                m = 1
            else:
                m += 1

    def _ultimo_dia_mes(ano: int, mes: int) -> int:
        if mes == 12:
            return 31
        prox = date(ano, mes + 1, 1)
        return (prox - timedelta(days=1)).day

    def _bcb_serie_valor_mes(codigo_serie: int, ano: int, mes: int) -> Decimal | None:
        """
        Busca no SGS/BCB o último valor disponível no mês para uma série.
        Retorna percentual (ex: 0.38 para 0,38%).
        """
        chave = (codigo_serie, ano, mes)
        if chave in _bcb_month_cache:
            return _bcb_month_cache[chave]

        inicio = f"01/{mes:02d}/{ano}"
        fim = f"{_ultimo_dia_mes(ano, mes):02d}/{mes:02d}/{ano}"
        url = (
            f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo_serie}/dados"
            f"?formato=json&dataInicial={inicio}&dataFinal={fim}"
        )

        try:
            req = urllib_request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib_request.urlopen(req, timeout=8) as resp:
                payload = resp.read().decode('utf-8', errors='ignore')
            dados = json.loads(payload)
            if not isinstance(dados, list) or not dados:
                _bcb_month_cache[chave] = None
                return None
            valor_txt = str(dados[-1].get('valor', '')).strip().replace(',', '.')
            valor = Decimal(valor_txt) if valor_txt else None
            _bcb_month_cache[chave] = valor
            return valor
        except Exception:
            _bcb_month_cache[chave] = None
            return None

    def _taxa_moratoria_mensal(ano: int, mes: int) -> Decimal:
        # Regra histórica adotada no projeto: 1% a.m. até 08/2024.
        if (ano, mes) <= (2024, 8):
            return Decimal('0.01')

        # Pós Lei 14.905/24 (proxy operacional): taxa legal ~= SELIC mensal - IPCA mensal, piso 0.
        selic_aa = _bcb_serie_valor_mes(1178, ano, mes)  # % a.a.
        ipca_mes = _bcb_serie_valor_mes(433, ano, mes)   # % no mês
        if selic_aa is None or ipca_mes is None:
            return Decimal('0.01')

        try:
            selic_mensal = Decimal(str((1.0 + float(selic_aa) / 100.0) ** (1.0 / 12.0) - 1.0))
            ipca_mensal = Decimal(str(float(ipca_mes) / 100.0))
            taxa = selic_mensal - ipca_mensal
            return taxa if taxa > 0 else Decimal('0')
        except Exception:
            return Decimal('0.01')

    def _taxa_juros_anual_cef(ano: int) -> Decimal:
        """Calcula taxa de juros aplicada pela CEF em um ano (juros / principal)."""
        try:
            total_principal = Decimal('0')
            total_juros = Decimal('0')
            
            for mes in range(1, 13):
                entry = cef_relatorio.get((ano, mes), {})
                total_principal += Decimal(str(entry.get('principal', 0)))
                total_juros += Decimal(str(entry.get('juros', 0)))
            
            if total_principal == 0:
                return Decimal('0')
            
            return total_juros / total_principal
        except Exception as e:
            import sys
            print(f"[ERROR _taxa_juros_anual_cef] ano={ano}: {e}", file=sys.stderr)
            return Decimal('0')

    def _fator_juros_mora_ate_data(ano: int, mes: int, data_ref: date | None = None) -> Decimal:
        """Fator de juros moratórios baseado na taxa anual da CEF para o ano do contrato."""
        try:
            # Usa a taxa de juros que a CEF aplicou naquele ano
            taxa_juros_ano = _taxa_juros_anual_cef(ano)
            return Decimal('1') + taxa_juros_ano
        except Exception as e:
            import sys
            print(f"[ERROR _fator_juros_mora_ate_data] ano={ano}, mes={mes}: {e}", file=sys.stderr)
            return Decimal('1')

    def _load_cef_relatorio() -> dict:
        """Carrega o relatório CEF do Excel. Retorna dict keyed (ano, mes) com valores."""
        xlsx_path = os.path.join(base_dir, 'manual', 'divida_seguro', 'relatorio cef divida seguro total.xlsx')
        if not os.path.isfile(xlsx_path):
            return {}
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            data: dict = {}
            for row in ws.iter_rows(min_row=3, values_only=True):
                competencia = row[0]
                if competencia is None:
                    continue
                if not hasattr(competencia, 'year'):
                    continue
                ano = competencia.year
                mes = competencia.month
                principal = float(row[1] or 0)
                atualizacao = float(row[2] or 0)
                sca = float(row[3] or 0)
                sca_atu = float(row[4] or 0)
                multa = float(row[5] or 0)
                juros = float(row[6] or 0)
                total = float(row[7] or 0)
                if total == 0:
                    total = principal + atualizacao + sca + sca_atu + multa + juros
                data[(ano, mes)] = {
                    'principal': principal,
                    'atualizacao': atualizacao,
                    'multa': multa,
                    'juros': juros,
                    'total': total,
                }
            return data
        except Exception:
            return {}

    cef_relatorio = _load_cef_relatorio()

    def _extract_mes_from_fonte(fonte: str) -> str:
        txt = (fonte or '').strip()
        m = re.search(r'_(\d{4})_(\d{2})\.pdf$', txt, re.IGNORECASE)
        if m:
            return m.group(2)
        m2 = re.match(r'CAD APOLICE (\d{2})_', txt, re.IGNORECASE)
        if m2:
            return m2.group(1)
        return ''

    def _build_resumo_global() -> dict | None:
        if not resumo_base:
            return None

        total_pdf_global = Decimal('0')
        total_cef_principal_global = 0.0
        total_cef_corrigido_global = 0.0
        total_cofluhab_atualizado_juros_global = Decimal('0')
        for ano in anos_resumo:
            fator_juros_ano = _fator_juros_mora_ate_data(ano, 1)
            for mes_num in range(1, 13):
                entry = cef_relatorio.get((ano, mes_num), {})
                total_cef_principal_global += entry.get('principal', 0.0)
                total_cef_corrigido_global += entry.get('total', 0.0)

                seguro_mes = (
                    ParcelaContrato.objects
                    .filter(dtvenc__year=ano, dtvenc__month=mes_num)
                    .aggregate(total=Sum('seguro'))
                    .get('total')
                ) or Decimal('0')
                seguro_mes_dec = Decimal(str(seguro_mes))
                fator_corr = _ipca_fator_acumulado_ate_data(ano, mes_num)
                total_cofluhab_atualizado_juros_global += seguro_mes_dec * fator_corr * fator_juros_ano

        meses_pdf_total = 0
        anos_com_csv = 0
        for ano in anos_resumo:
            ano_csv = sorted(Path(exports_dir).glob(f'divida_seguro_{ano}_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
            if not ano_csv:
                continue
            anos_com_csv += 1
            meses_set = set()
            try:
                with open(ano_csv[0], 'r', encoding='utf-8-sig', newline='') as fp:
                    reader = csv.DictReader(fp, delimiter=';')
                    for row in reader:
                        total_pdf_global += _parse_decimal_br(row.get('valor_seguro_pdf', '0'))
                        mes = _extract_mes_from_fonte(row.get('fonte', ''))
                        if mes:
                            meses_set.add(mes)
            except Exception:
                pass
            meses_pdf_total += len(meses_set)

        dif_global = total_cef_corrigido_global - float(total_pdf_global)
        dif_global_cef_cofluhab_juros = Decimal(str(total_cef_corrigido_global)) - total_cofluhab_atualizado_juros_global
        fmt_float = lambda v: f"{v:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')

        # Soma total histórico COFLUHAB: soma de ParcelaContrato.seguro para todos os anos disponíveis
        from django.db.models import Sum as _Sum
        qs_historico = (
            ParcelaContrato.objects
            .filter(dtvenc__year__in=anos_resumo)
            .aggregate(total=_Sum('seguro'))
        )
        historico_cofluhab_total = Decimal(str(qs_historico['total'] or 0))

        periodo_resumo_txt = f"{min(anos_resumo)}-{max(anos_resumo)}" if anos_resumo else '-'

        return {
            'total_pdf_global': _fmt_decimal_br(total_pdf_global),
            'total_cef_principal_global': fmt_float(total_cef_principal_global),
            'total_cef_corrigido_global': fmt_float(total_cef_corrigido_global),
            'diferenca_global': fmt_float(dif_global),
            'diferenca_global_negativa': dif_global < 0,
            'cofluhab_atualizado_juros_global': _fmt_decimal_br(total_cofluhab_atualizado_juros_global),
            'diferenca_cef_cofluhab_juros_global': _fmt_decimal_br(dif_global_cef_cofluhab_juros),
            'diferenca_cef_cofluhab_juros_global_negativa': dif_global_cef_cofluhab_juros < 0,
            'meses_pdf_total': meses_pdf_total,
            'meses_esperados_total': len(anos_resumo) * 12,
            'anos_com_csv': anos_com_csv,
            'historico_cofluhab_total': _fmt_decimal_br(historico_cofluhab_total),
            'periodo_resumo': periodo_resumo_txt,
        }

    resumo_global = _build_resumo_global()

    def _build_confronto_global_por_ano() -> list[dict]:
        linhas = []
        fmt_float = lambda v: f"{v:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')

        for ano in anos_resumo:
            total_pdf_ano = Decimal('0')
            meses_pdf = set()

            ano_csv = sorted(Path(exports_dir).glob(f'divida_seguro_{ano}_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
            if ano_csv:
                try:
                    with open(ano_csv[0], 'r', encoding='utf-8-sig', newline='') as fp:
                        reader = csv.DictReader(fp, delimiter=';')
                        for row in reader:
                            total_pdf_ano += _parse_decimal_br(row.get('valor_seguro_pdf', '0'))
                            mes = _extract_mes_from_fonte(row.get('fonte', ''))
                            if mes:
                                meses_pdf.add(mes)
                except Exception:
                    pass

            total_cef_principal_ano = 0.0
            total_cef_corrigido_ano = 0.0
            for mes_num in range(1, 13):
                entry = cef_relatorio.get((ano, mes_num), {})
                total_cef_principal_ano += entry.get('principal', 0.0)
                total_cef_corrigido_ano += entry.get('total', 0.0)

            total_historico_cofluhab_ano = Decimal('0')
            total_cofluhab_atualizado_ano = Decimal('0')
            total_cofluhab_atualizado_juros_ano = Decimal('0')
            
            # Calcular fator de juros apenas uma vez por ano
            fator_juros_ano = _fator_juros_mora_ate_data(ano, 1)
            
            for mes_num in range(1, 13):
                seguro_mes = (
                    ParcelaContrato.objects
                    .filter(dtvenc__year=ano, dtvenc__month=mes_num)
                    .aggregate(total=Sum('seguro'))
                    .get('total')
                ) or Decimal('0')
                seguro_mes_dec = Decimal(str(seguro_mes))
                total_historico_cofluhab_ano += seguro_mes_dec
                fator_corr = _ipca_fator_acumulado_ate_data(ano, mes_num)
                total_cofluhab_atualizado_ano += seguro_mes_dec * fator_corr
                total_cofluhab_atualizado_juros_ano += seguro_mes_dec * fator_corr * fator_juros_ano

            dif_cef_pdf = Decimal(str(total_cef_corrigido_ano)) - total_pdf_ano
            dif_cef_cofluhab = Decimal(str(total_cef_corrigido_ano)) - total_cofluhab_atualizado_ano
            dif_cef_cofluhab_juros = Decimal(str(total_cef_corrigido_ano)) - total_cofluhab_atualizado_juros_ano

            linhas.append({
                'ano': ano,
                'cef_historico': fmt_float(total_cef_principal_ano),
                'cef_total_corrigido': fmt_float(total_cef_corrigido_ano),
                'encontrado_pdfs': _fmt_decimal_br(total_pdf_ano),
                'historico_cofluhab': _fmt_decimal_br(total_historico_cofluhab_ano),
                'cofluhab_atualizado': _fmt_decimal_br(total_cofluhab_atualizado_ano),
                'cofluhab_atualizado_juros': _fmt_decimal_br(total_cofluhab_atualizado_juros_ano),
                'diferenca_cef_pdf': _fmt_decimal_br(dif_cef_pdf),
                'diferenca_cef_cofluhab': _fmt_decimal_br(dif_cef_cofluhab),
                'diferenca_cef_cofluhab_juros': _fmt_decimal_br(dif_cef_cofluhab_juros),
                'meses_pdf_count': len(meses_pdf),
            })

        return linhas

    confronto_global_por_ano = _build_confronto_global_por_ano()

    def _load_cef_relatorio() -> dict:
        """Carrega o relatório CEF do Excel. Retorna dict keyed (ano, mes) com valores."""
        xlsx_path = os.path.join(base_dir, 'manual', 'divida_seguro', 'relatorio cef divida seguro total.xlsx')
        if not os.path.isfile(xlsx_path):
            return {}
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            data: dict = {}
            for row in ws.iter_rows(min_row=3, values_only=True):
                competencia = row[0]
                if competencia is None:
                    continue
                if not hasattr(competencia, 'year'):
                    continue
                ano = competencia.year
                mes = competencia.month
                principal = float(row[1] or 0)
                atualizacao = float(row[2] or 0)
                sca = float(row[3] or 0)
                sca_atu = float(row[4] or 0)
                multa = float(row[5] or 0)
                juros = float(row[6] or 0)
                total = float(row[7] or 0)
                if total == 0:
                    total = principal + atualizacao + sca + sca_atu + multa + juros
                data[(ano, mes)] = {
                    'principal': principal,
                    'atualizacao': atualizacao,
                    'multa': multa,
                    'juros': juros,
                    'total': total,
                }
            return data
        except Exception:
            return {}

    cef_relatorio = _load_cef_relatorio()

    def _normalize_contract_code(value):
        digits = ''.join(ch for ch in str(value or '') if ch.isdigit())
        if not digits:
            return ''
        return str(int(digits))

    def _map_valor_base_contrato(codigos_contrato):
        codigos_set = {c for c in codigos_contrato if c}
        if not codigos_set:
            return {}

        contratos = []
        for c in Contrato.objects.all().only('id', 'codigo', 'vlfinanc'):
            code = _normalize_contract_code(c.codigo)
            if code in codigos_set:
                contratos.append((code, c.id, c.vlfinanc))

        if not contratos:
            return {}

        id_para_code = {cid: code for code, cid, _ in contratos}
        base = {}
        for code, _, vlfinanc in contratos:
            if vlfinanc is not None:
                base[code] = Decimal(vlfinanc)

        ultimo_sddev = Subquery(
            ParcelaContrato.objects.filter(contrato_id=OuterRef('pk')).order_by('-nmens').values('sddev')[:1]
        )
        contratos_com_sddev = Contrato.objects.filter(id__in=list(id_para_code.keys())).annotate(ultimo_sddev=ultimo_sddev).only('id')
        for c in contratos_com_sddev:
            code = id_para_code.get(c.id)
            if code and c.ultimo_sddev is not None:
                base[code] = Decimal(c.ultimo_sddev)

        return base

    try:
        # Pré-computa o seguro base da carteira vigente por mês diretamente do banco
        # (ParcelaContrato.seguro, agrupado por mês da dtvenc)
        seguro_base_db: dict = {}  # {mes_num (int): Decimal}
        if ano_selecionado:
            from django.db.models import Sum as _Sum
            qs_seg = (
                ParcelaContrato.objects
                .filter(dtvenc__year=int(ano_selecionado))
                .values('dtvenc__month')
                .annotate(total_seguro=_Sum('seguro'))
            )
            for row in qs_seg:
                m = row['dtvenc__month']
                seguro_base_db[m] = Decimal(str(row['total_seguro'] or 0))

        # Verifica se há arquivos de confronto mês a mês para o ano selecionado.
        # Esses arquivos têm o padrão: confronto_divida_seguro_MMAAAA_timestamp.csv
        confronto_mensal_paths = []
        if ano_selecionado:
            confronto_mensal_paths = sorted(
                Path(exports_dir).glob(f'confronto_divida_seguro_*{ano_selecionado}_*.csv'),
                key=lambda p: p.name
            )

        if confronto_mensal_paths:
            # --- MODO COMPARATIVO MENSAL: mês a mês com confronto de finalização ---
            periodo_files: dict = {}
            for p in confronto_mensal_paths:
                m_match = re.match(r'confronto_divida_seguro_(\d{2})(\d{4})_', p.name)
                if m_match:
                    periodo_key = f"{m_match.group(1)}/{m_match.group(2)}"
                    if periodo_key not in periodo_files or p.stat().st_mtime > periodo_files[periodo_key].stat().st_mtime:
                        periodo_files[periodo_key] = p

            meses_nomes = {
                '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março',
                '04': 'Abril', '05': 'Maio', '06': 'Junho',
                '07': 'Julho', '08': 'Agosto', '09': 'Setembro',
                '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro',
            }

            meses_data = []
            total_anual_cef = Decimal('0')
            total_anual_fora_prazo = Decimal('0')
            total_anual_linhas = 0
            total_anual_meses_pos = 0
            qtd_anual_meses_pos = 0
            all_rows_index: dict = {}

            for periodo_key in sorted(periodo_files.keys()):
                csv_mes_path = periodo_files[periodo_key]
                mm = periodo_key[:2]
                nome_mes = meses_nomes.get(mm, mm)

                with open(csv_mes_path, 'r', encoding='utf-8-sig', newline='') as fp:
                    reader = csv.DictReader(fp, delimiter=';')
                    rows_mes = list(reader)

                linhas_mes = []
                total_cef_mes = Decimal('0')
                total_dentro_prazo_mes = Decimal('0')
                total_fora_prazo_mes = Decimal('0')
                total_meses_pos_mes = 0
                qtd_meses_pos_mes = 0

                for row in rows_mes:
                    contrato = (row.get('contrato_pdf') or '').strip()
                    valor_cef_txt = row.get('valor_seguro_pdf', '')
                    valor_cef = _parse_decimal_br(valor_cef_txt)
                    meses_txt = (row.get('meses_apos_finalizacao') or '').strip()
                    try:
                        meses_pos = int(meses_txt) if meses_txt else 0
                    except ValueError:
                        meses_pos = 0

                    entrada = {
                        'contrato': contrato,
                        'nome': row.get('nome_mutuario', ''),
                        'status': row.get('status_confronto', ''),
                        'motivo_finalizacao': row.get('motivo_finalizacao', ''),
                        'data_finalizacao': row.get('data_finalizacao', ''),
                        'meses_apos_finalizacao': meses_txt,
                        'valor_cef_cobrando': valor_cef_txt,
                    }
                    linhas_mes.append(entrada)

                    if contrato and contrato not in all_rows_index:
                        all_rows_index[contrato] = entrada.copy()
                        all_rows_index[contrato]['periodo'] = periodo_key

                    total_cef_mes += valor_cef
                    total_anual_cef += valor_cef
                    if meses_pos > 0:
                        total_fora_prazo_mes += valor_cef
                        total_anual_fora_prazo += valor_cef
                        total_meses_pos_mes += meses_pos
                        qtd_meses_pos_mes += 1
                        total_anual_meses_pos += meses_pos
                        qtd_anual_meses_pos += 1
                    else:
                        total_dentro_prazo_mes += valor_cef

                total_anual_linhas += len(rows_mes)
                # Busca dados CEF para este mês
                cef_mes = cef_relatorio.get((int(ano_selecionado), int(mm)), {}) if ano_selecionado else {}
                cef_principal = cef_mes.get('principal', 0.0)
                cef_atualizacao = cef_mes.get('atualizacao', 0.0)
                cef_multa = cef_mes.get('multa', 0.0)
                cef_juros = cef_mes.get('juros', 0.0)
                cef_total = cef_mes.get('total', 0.0)
                seguro_base_mes = seguro_base_db.get(int(mm), Decimal('0')) if ano_selecionado else Decimal('0')
                fator_ipca_acumulado = _ipca_fator_acumulado_ate_data(int(ano_selecionado), int(mm)) if ano_selecionado else Decimal('1')
                cofluhab_atualizado_mes = seguro_base_mes * fator_ipca_acumulado
                diferenca = cef_total - float(cofluhab_atualizado_mes)
                meses_data.append({
                    'periodo': periodo_key,
                    'nome_mes': nome_mes,
                    'arquivo': csv_mes_path.name,
                    'total_linhas': len(rows_mes),
                    'total_cef_cobrando': _fmt_decimal_br(total_cef_mes),
                    'cobranca_fora_prazo': _fmt_decimal_br(total_fora_prazo_mes),
                    'seguro_base_cofluhab': _fmt_decimal_br(seguro_base_mes),
                    'seguro_base_cofluhab_valor': seguro_base_mes,
                    'fator_ipca_acumulado': f"{fator_ipca_acumulado:,.6f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'media_meses_pos': f"{(total_meses_pos_mes / qtd_meses_pos_mes):.1f}" if qtd_meses_pos_mes else '0',
                    'linhas': linhas_mes,
                    'atualizacao_cofluhab': _fmt_decimal_br(cofluhab_atualizado_mes),
                    'atualizacao_cofluhab_valor': cofluhab_atualizado_mes,
                    'cef_principal': f"{cef_principal:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'cef_atualizacao': f"{cef_atualizacao:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'cef_multa': f"{cef_multa:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'cef_juros': f"{cef_juros:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'cef_total': f"{cef_total:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'cef_tem_dados': bool(cef_mes),
                    'diferenca': f"{diferenca:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                    'diferenca_negativa': diferenca < 0,
                    'diferenca_zero': abs(diferenca) < 0.01,
                })

            contrato_ativo = contrato_selecionado
            if not contrato_ativo and meses_data and meses_data[0]['linhas']:
                contrato_ativo = meses_data[0]['linhas'][0]['contrato']
            contrato_detalhe = all_rows_index.get(contrato_ativo)

            # Totais anuais CEF do relatório Excel — soma apenas meses com dados de PDF
            total_anual_cef_principal = 0.0
            total_anual_seguro_base_cofluhab = Decimal('0')
            total_anual_atualizacao_cofluhab = Decimal('0')
            total_anual_cef_total = 0.0
            for mes in meses_data:
                try:
                    mm_str, yyyy_str = mes['periodo'].split('/')
                    cef_entry = cef_relatorio.get((int(yyyy_str), int(mm_str)), {})
                    total_anual_cef_principal += cef_entry.get('principal', 0.0)
                    total_anual_cef_total += cef_entry.get('total', 0.0)
                except Exception:
                    pass
                total_anual_seguro_base_cofluhab += Decimal(str(mes.get('seguro_base_cofluhab_valor', Decimal('0'))))
                total_anual_atualizacao_cofluhab += Decimal(str(mes.get('atualizacao_cofluhab_valor', Decimal('0'))))
            # Total CEF do ano completo (todos os 12 meses)
            cef_ano_completo_total = 0.0
            if ano_selecionado:
                for mes_num in range(1, 13):
                    cef_entry = cef_relatorio.get((int(ano_selecionado), mes_num), {})
                    cef_ano_completo_total += cef_entry.get('total', 0.0)
            dif_anual = total_anual_cef_total - float(total_anual_atualizacao_cofluhab)
            fmt = lambda v: f"{v:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')

            preview_relatorio = {
                'modo': 'comparativo_mensal',
                'meses': meses_data,
                'total_anual_linhas': total_anual_linhas,
                'total_anual_cef': _fmt_decimal_br(total_anual_cef),
                'media_meses_pos_finalizacao': f"{(total_anual_meses_pos / qtd_anual_meses_pos):.1f}" if qtd_anual_meses_pos else '0',
                'contrato_ativo': contrato_ativo,
                'contrato_detalhe': contrato_detalhe,
                'total_anual_cef_principal': fmt(total_anual_cef_principal),
                'total_anual_seguro_base_cofluhab': _fmt_decimal_br(total_anual_seguro_base_cofluhab),
                'total_anual_atualizacao_cofluhab': _fmt_decimal_br(total_anual_atualizacao_cofluhab),
                'total_anual_cef_total': fmt(total_anual_cef_total),
                'total_anual_cobranca_fora_prazo': _fmt_decimal_br(total_anual_fora_prazo),
                'total_anual_diferenca': fmt(dif_anual),
                'total_anual_diferenca_negativa': dif_anual < 0,
                'cef_relatorio_disponivel': bool(cef_relatorio),
                'cef_ano_completo_total': fmt(cef_ano_completo_total),
                'meses_pdf_count': len(meses_data),
            }
        else:
            # --- MODO CONSOLIDADO: CSV de dívida por ano ou geral ---
            csv_paths = []
            if ano_selecionado:
                csv_paths = sorted(Path(exports_dir).glob(f'divida_seguro_{ano_selecionado}_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)
            elif not ano_selecionado:
                csv_paths = sorted(Path(exports_dir).glob('divida_seguro_consolidado_*.csv'), key=lambda p: p.stat().st_mtime, reverse=True)

            if csv_paths or ano_selecionado:
                rows_raw = []
                csv_preview_name = f'divida_seguro_{ano_selecionado}_*.csv (não encontrado)' if ano_selecionado else ''
                if csv_paths:
                    csv_preview_path = csv_paths[0]
                    csv_preview_name = csv_preview_path.name
                    with open(csv_preview_path, 'r', encoding='utf-8-sig', newline='') as fp:
                        reader = csv.DictReader(fp, delimiter=';')
                        rows_raw = list(reader)

                total_cef = Decimal('0')
                total_fora_prazo = Decimal('0')
                total_linhas = 0

                # Agrupa por mês extraído da coluna `fonte` (ex: RIE_OFI_2010_03.pdf → "03")
                meses_nomes = {
                    '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março',
                    '04': 'Abril', '05': 'Maio', '06': 'Junho',
                    '07': 'Julho', '08': 'Agosto', '09': 'Setembro',
                    '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro',
                }
                from collections import OrderedDict
                meses_dict: dict = OrderedDict()

                # Para ano selecionado, garante os 12 meses mesmo sem PDF
                if ano_selecionado:
                    for mes_key in [f"{m:02d}" for m in range(1, 13)]:
                        meses_dict[mes_key] = {
                            'mes_key': mes_key,
                            'nome_mes': meses_nomes.get(mes_key, mes_key),
                            'fonte': 'PDF não encontrado',
                            'linhas': [],
                            'total_cef': Decimal('0'),
                            'tem_pdf': False,
                        }

                for row in rows_raw:
                    total_linhas += 1
                    contrato = (row.get('contrato_pdf') or '').strip()
                    fonte = (row.get('fonte') or '').strip()
                    valor_cef_txt = row.get('valor_seguro_pdf', row.get('valor_cef_cobrando', ''))
                    valor_cef = _parse_decimal_br(valor_cef_txt)
                    total_cef += valor_cef
                    meses_txt = (row.get('meses_apos_finalizacao') or '').strip()
                    try:
                        meses_pos = int(meses_txt) if meses_txt else 0
                    except ValueError:
                        meses_pos = 0
                    if meses_pos > 0:
                        total_fora_prazo += valor_cef

                    # Tenta extrair mês da fonte (padrão: TIPO_AAAA_MM.pdf ou CAD APOLICE MM_AAAA.pdf)
                    mes_key = ''
                    m = re.search(r'_(\d{4})_(\d{2})\.pdf$', fonte, re.IGNORECASE)
                    if m:
                        mes_key = m.group(2)
                    else:
                        m2 = re.match(r'CAD APOLICE (\d{2})_', fonte, re.IGNORECASE)
                        if m2:
                            mes_key = m2.group(1)
                    if not mes_key:
                        mes_key = '00'

                    if mes_key not in meses_dict:
                        meses_dict[mes_key] = {
                            'mes_key': mes_key,
                            'nome_mes': meses_nomes.get(mes_key, f'Mês {mes_key}' if mes_key != '00' else 'Sem mês'),
                            'fonte': fonte or 'PDF não encontrado',
                            'linhas': [],
                            'total_cef': Decimal('0'),
                            'tem_pdf': False,
                        }

                    if fonte:
                        meses_dict[mes_key]['fonte'] = fonte
                    meses_dict[mes_key]['tem_pdf'] = True
                    meses_dict[mes_key]['linhas'].append({
                        'contrato': contrato,
                        'nome': row.get('nome_mutuario', ''),
                        'status': row.get('status_confronto', row.get('status', '')),
                        'motivo_finalizacao': row.get('motivo_finalizacao', ''),
                        'data_finalizacao': row.get('data_finalizacao', ''),
                        'meses_apos_finalizacao': row.get('meses_apos_finalizacao', ''),
                        'valor_cef_cobrando': valor_cef_txt,
                    })
                    meses_dict[mes_key]['total_cef'] += valor_cef

                meses_data = []
                ordered_keys = [f"{m:02d}" for m in range(1, 13)] if ano_selecionado else sorted(meses_dict.keys())
                if '00' in meses_dict and '00' not in ordered_keys:
                    ordered_keys.append('00')

                for mes_key in ordered_keys:
                    d = meses_dict.get(mes_key)
                    if not d:
                        continue

                    total_pdf_mes = float(d['total_cef'])
                    cef_mes = {}
                    if ano_selecionado and mes_key.isdigit() and len(mes_key) == 2:
                        cef_mes = cef_relatorio.get((int(ano_selecionado), int(mes_key)), {})

                    cef_principal = cef_mes.get('principal', 0.0)
                    cef_atualizacao = cef_mes.get('atualizacao', 0.0)
                    cef_multa = cef_mes.get('multa', 0.0)
                    cef_juros = cef_mes.get('juros', 0.0)
                    cef_total = cef_mes.get('total', 0.0)
                    total_fora_prazo_mes = Decimal('0')
                    total_dentro_prazo_mes = Decimal('0')
                    for l in d['linhas']:
                        valor_linha = _parse_decimal_br(l.get('valor_cef_cobrando', '0'))
                        meses_txt_l = str(l.get('meses_apos_finalizacao', '')).strip()
                        try:
                            meses_pos_l = int(meses_txt_l) if meses_txt_l else 0
                        except ValueError:
                            meses_pos_l = 0
                        if meses_pos_l > 0:
                            total_fora_prazo_mes += valor_linha
                        else:
                            total_dentro_prazo_mes += valor_linha

                    seguro_base_mes = seguro_base_db.get(int(mes_key), Decimal('0')) if (ano_selecionado and mes_key.isdigit() and len(mes_key) == 2) else Decimal('0')
                    fator_ipca_acumulado = _ipca_fator_acumulado_ate_data(int(ano_selecionado), int(mes_key)) if (ano_selecionado and mes_key.isdigit() and len(mes_key) == 2) else Decimal('1')
                    cofluhab_atualizado_mes = seguro_base_mes * fator_ipca_acumulado
                    diferenca = cef_total - float(cofluhab_atualizado_mes)
                    fmt = lambda v: f"{v:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')

                    meses_data.append({
                        'mes_key': mes_key,
                        'nome_mes': d['nome_mes'],
                        'fonte': d['fonte'],
                        'tem_pdf': d.get('tem_pdf', False),
                        'total_linhas': len(d['linhas']),
                        'total_cef_cobrando': _fmt_decimal_br(d['total_cef']),
                        'cobranca_fora_prazo': _fmt_decimal_br(total_fora_prazo_mes),
                        'seguro_base_cofluhab': _fmt_decimal_br(seguro_base_mes),
                        'seguro_base_cofluhab_valor': seguro_base_mes,
                        'fator_ipca_acumulado': f"{fator_ipca_acumulado:,.6f}".replace('.', '_').replace(',', '.').replace('_', ','),
                        'linhas': d['linhas'],
                        'atualizacao_cofluhab': _fmt_decimal_br(cofluhab_atualizado_mes),
                        'atualizacao_cofluhab_valor': cofluhab_atualizado_mes,
                        'cef_principal': fmt(cef_principal),
                        'cef_atualizacao': fmt(cef_atualizacao),
                        'cef_multa': fmt(cef_multa),
                        'cef_juros': fmt(cef_juros),
                        'cef_total': fmt(cef_total),
                        'cef_tem_dados': bool(cef_mes),
                        'diferenca': fmt(diferenca),
                        'diferenca_negativa': diferenca < 0,
                        'diferenca_zero': abs(diferenca) < 0.01,
                    })

                # Detalhe do contrato clicado (se houver)
                contrato_detalhe = None
                if contrato_selecionado:
                    for row in rows_raw:
                        if (row.get('contrato_pdf') or '').strip() == contrato_selecionado:
                            valor_cef_txt = row.get('valor_seguro_pdf', row.get('valor_cef_cobrando', ''))
                            contrato_detalhe = {
                                'contrato': contrato_selecionado,
                                'nome': row.get('nome_mutuario', ''),
                                'status': row.get('status_confronto', row.get('status', '')),
                                'motivo_finalizacao': row.get('motivo_finalizacao', ''),
                                'data_finalizacao': row.get('data_finalizacao', ''),
                                'meses_apos_finalizacao': row.get('meses_apos_finalizacao', ''),
                                'fonte': row.get('fonte', ''),
                                'valor_cef_cobrando': valor_cef_txt,
                            }
                            break

                # Totais anuais CEF do relatório Excel — sempre 12 meses quando ano é selecionado
                total_anual_cef_principal_c = 0.0
                total_anual_seguro_base_cofluhab_c = Decimal('0')
                total_anual_atualizacao_cofluhab_c = Decimal('0')
                total_anual_cef_total_c = 0.0
                if ano_selecionado:
                    for mes_num in range(1, 13):
                        cef_entry = cef_relatorio.get((int(ano_selecionado), mes_num), {})
                        total_anual_cef_principal_c += cef_entry.get('principal', 0.0)
                        total_anual_cef_total_c += cef_entry.get('total', 0.0)

                for mes in meses_data:
                    total_anual_seguro_base_cofluhab_c += Decimal(str(mes.get('seguro_base_cofluhab_valor', Decimal('0'))))
                    total_anual_atualizacao_cofluhab_c += Decimal(str(mes.get('atualizacao_cofluhab_valor', Decimal('0'))))

                cef_ano_completo_total_c = total_anual_cef_total_c
                dif_anual_c = total_anual_cef_total_c - float(total_anual_atualizacao_cofluhab_c)
                fmt = lambda v: f"{v:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')
                meses_pdf_count_real = sum(1 for mes in meses_data if mes.get('tem_pdf'))

                preview_relatorio = {
                    'modo': 'consolidado_mensal',
                    'arquivo': csv_preview_name,
                    'total_linhas': total_linhas,
                    'total_cef_cobrando': _fmt_decimal_br(total_cef),
                    'meses': meses_data,
                    'contrato_ativo': contrato_selecionado or '',
                    'contrato_detalhe': contrato_detalhe,
                    'total_anual_cef_principal': fmt(total_anual_cef_principal_c),
                    'total_anual_seguro_base_cofluhab': _fmt_decimal_br(total_anual_seguro_base_cofluhab_c),
                    'total_anual_atualizacao_cofluhab': _fmt_decimal_br(total_anual_atualizacao_cofluhab_c),
                    'total_anual_cef_total': fmt(total_anual_cef_total_c),
                    'total_anual_cobranca_fora_prazo': _fmt_decimal_br(total_fora_prazo),
                    'total_anual_diferenca': fmt(dif_anual_c),
                    'total_anual_diferenca_negativa': dif_anual_c < 0,
                    'cef_relatorio_disponivel': bool(cef_relatorio),
                    'cef_ano_completo_total': fmt(cef_ano_completo_total_c),
                    'meses_pdf_count': meses_pdf_count_real,
                }

    except Exception:
        preview_relatorio = None

    mostrar_detalhes = request.GET.get('detalhes') == '1'
    aba_relatorio = request.GET.get('aba_relatorio', 'geral')
    if aba_relatorio not in {'geral', 'anual'}:
        aba_relatorio = 'geral'

    if request.GET.get('exportar_excel') == '1':
        wb = Workbook()

        ws_resumo = wb.active
        ws_resumo.title = 'Resumo_Global'
        ws_resumo.append(['Campo', 'Valor'])

        resumo_global = resumo_global or {}
        ws_resumo.append(['Periodo', f"{ano_inicio_selecionado} a {ano_fim_selecionado}"])
        ws_resumo.append(['CEF historico (principal)', float(resumo_global.get('cef_historico', 0.0) or 0.0)])
        ws_resumo.append(['CEF total atualizado', float(resumo_global.get('cef_total_corrigido', 0.0) or 0.0)])
        ws_resumo.append(['Encontrado nos PDFs', float(resumo_global.get('encontrado_pdfs', 0.0) or 0.0)])
        ws_resumo.append(['COFLUHAB atualizado', float(resumo_global.get('cofluhab_atualizado', 0.0) or 0.0)])
        ws_resumo.append(['Diferenca CEF total - PDFs', float(resumo_global.get('diferenca_cef_pdf', 0.0) or 0.0)])
        ws_resumo.append(['Diferenca CEF total - COFLUHAB', float(resumo_global.get('diferenca_cef_cofluhab', 0.0) or 0.0)])
        ws_resumo.append(['Meses com PDF', int(resumo_global.get('meses_pdf_count', 0) or 0)])

        ws_ano = wb.create_sheet('Por_Ano')
        ws_ano.append([
            'Ano',
            'CEF historico (principal)',
            'CEF total atualizado',
            'Encontrado nos PDFs',
            'Historico COFLUHAB',
            'COFLUHAB atualizado',
            'COFLUHAB atualizado + juros',
            'Diferenca CEF total - PDFs',
            'Diferenca CEF total - COFLUHAB total com juros',
            'Meses com PDF',
        ])

        for item in (confronto_global_por_ano or []):
            ws_ano.append([
                int(item.get('ano') or 0),
                float(item.get('cef_historico', 0.0) or 0.0),
                float(item.get('cef_total_corrigido', 0.0) or 0.0),
                float(item.get('encontrado_pdfs', 0.0) or 0.0),
                float(item.get('historico_cofluhab', 0.0) or 0.0),
                float(item.get('cofluhab_atualizado', 0.0) or 0.0),
                float(item.get('cofluhab_atualizado_juros', 0.0) or 0.0),
                float(item.get('diferenca_cef_pdf', 0.0) or 0.0),
                float(item.get('diferenca_cef_cofluhab_juros', 0.0) or 0.0),
                int(item.get('meses_pdf_count', 0) or 0),
            ])

        for ws in (ws_resumo, ws_ano):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

        for row in ws_resumo.iter_rows(min_row=3, max_row=8, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '#,##0.00'

        for row in ws_ano.iter_rows(min_row=2, min_col=2, max_col=9):
            for cell in row:
                cell.number_format = '#,##0.00'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"resumo_dividas_seguro_{ano_inicio_selecionado}_{ano_fim_selecionado}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'principal/relatorio_divida_seguro.html', {
        'anos_disponiveis': anos_disponiveis,
        'ano_selecionado': ano_selecionado,
        'ano_inicio_selecionado': ano_inicio_selecionado,
        'ano_fim_selecionado': ano_fim_selecionado,
        'ano_inicial_estudo': ano_inicial_estudo,
        'ano_final_estudo': ano_final_estudo,
        'mostrar_detalhes': mostrar_detalhes,
        'exec_result': exec_result,
        'arquivos': arquivos,
        'arquivos_exibicao': arquivos[:30],
        'arquivos_por_periodo': periodos_ordenados,
        'resumo_base': resumo_base,
        'resumo_global': resumo_global,
        'confronto_global_por_ano': confronto_global_por_ano,
        'aba_relatorio': aba_relatorio,
        'arquivos_fixados': sorted(arquivos_fixados, key=lambda x: x['modificado_em'], reverse=True),
        'preview_relatorio': preview_relatorio,
    })


def download_relatorio_divida_seguro(request, filename):
    """Download seguro dos arquivos de relatório de dívida de seguro."""
    if '/' in filename or '\\' in filename:
        raise Http404('Arquivo inválido')

    permitido = (
        filename.startswith('divida_seguro_')
        or filename.startswith('laudo_divida_seguro_')
        or filename.startswith('confronto_divida_seguro_')
    )
    if not permitido:
        raise Http404('Arquivo não permitido')

    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    exports_candidates = [
        os.path.abspath(os.path.join(base_dir, '..', 'exports')),
        os.path.join(base_dir, 'exports'),
    ]
    exports_dir = next((p for p in exports_candidates if os.path.isdir(p)), exports_candidates[0])
    full_path = os.path.join(exports_dir, filename)

    if not os.path.isfile(full_path):
        raise Http404('Arquivo não encontrado')

    return FileResponse(open(full_path, 'rb'), as_attachment=True, filename=filename)


def visualizar_relatorio_divida_seguro(request, filename):
    """Visualização amigável de CSV/Markdown dos relatórios de dívida de seguro."""
    if '/' in filename or '\\' in filename:
        raise Http404('Arquivo inválido')

    permitido = (
        filename.startswith('divida_seguro_')
        or filename.startswith('laudo_divida_seguro_')
        or filename.startswith('confronto_divida_seguro_')
    )
    if not permitido:
        raise Http404('Arquivo não permitido')

    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    exports_candidates = [
        os.path.abspath(os.path.join(base_dir, '..', 'exports')),
        os.path.join(base_dir, 'exports'),
    ]
    exports_dir = next((p for p in exports_candidates if os.path.isdir(p)), exports_candidates[0])
    full_path = os.path.join(exports_dir, filename)

    if not os.path.isfile(full_path):
        raise Http404('Arquivo não encontrado')

    contexto = {
        'filename': filename,
        'is_csv': filename.lower().endswith('.csv'),
        'is_md': filename.lower().endswith('.md'),
        'headers': [],
        'rows': [],
        'markdown_content': '',
        'total_rows': 0,
        'sumario': {},
    }

    if contexto['is_csv']:
        with open(full_path, 'r', encoding='utf-8-sig', newline='') as fp:
            reader = csv.DictReader(fp, delimiter=';')
            headers = reader.fieldnames or []
            rows = []
            total_rows = 0
            encontrados = 0
            nao_encontrados = 0
            total_cef = Decimal('0')
            total_divida = Decimal('0')
            contratos_unicos = set()

            for row in reader:
                total_rows += 1
                contrato = (row.get('contrato_pdf') or row.get('contrato') or '').strip()
                if contrato:
                    contratos_unicos.add(contrato)
                status = (row.get('status') or row.get('status_confronto') or '').strip().upper()
                if status == 'ENCONTRADO':
                    encontrados += 1
                elif 'NAO_ENCONTRADO' in status:
                    nao_encontrados += 1

                total_cef += _parse_decimal_br(row.get('valor_cef_cobrando', row.get('valor_seguro_pdf', '0')))
                total_divida += _parse_decimal_br(row.get('valor_divida_banco', '0'))

                if len(rows) < 300:
                    rows.append(row)

        # Enriquecimento do resumo para arquivos anuais: confronto CEF x PDF x COFLUHAB.
        ano_relatorio = None
        match_ano = re.search(r'divida_seguro_(\d{4})_', filename)
        if match_ano:
            try:
                ano_relatorio = int(match_ano.group(1))
            except Exception:
                ano_relatorio = None

        total_cef_historico_ano = None
        total_cef_corrigido_ano = None
        total_cofluhab_atualizado_ano = None
        diferenca_cef_pdf_ano = None
        diferenca_cef_cofluhab_ano = None

        if ano_relatorio:
            # 1) Totais CEF do Excel (principal e total corrigido do ano completo)
            try:
                xlsx_path = os.path.join(base_dir, 'manual', 'divida_seguro', 'relatorio cef divida seguro total.xlsx')
                if os.path.isfile(xlsx_path):
                    import openpyxl
                    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                    ws = wb.active
                    soma_principal = 0.0
                    soma_total = 0.0
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        competencia = row[0]
                        if not hasattr(competencia, 'year') or competencia.year != ano_relatorio:
                            continue
                        principal = float(row[1] or 0)
                        atualizacao = float(row[2] or 0)
                        sca = float(row[3] or 0)
                        sca_atu = float(row[4] or 0)
                        multa = float(row[5] or 0)
                        juros = float(row[6] or 0)
                        total_linha = float(row[7] or 0)
                        if total_linha == 0:
                            total_linha = principal + atualizacao + sca + sca_atu + multa + juros
                        soma_principal += principal
                        soma_total += total_linha
                    total_cef_historico_ano = Decimal(str(soma_principal))
                    total_cef_corrigido_ano = Decimal(str(soma_total))
            except Exception:
                pass

            # 2) Total COFLUHAB atualizado (soma mensal de seguro * fator IPCA acumulado)
            try:
                def _ipca_mensal(ano: int, mes: int) -> Decimal:
                    chave = f"{ano}-{mes:02d}"
                    try:
                        return Decimal(str(INDICES_HISTORICOS.get(chave, Decimal('0'))))
                    except Exception:
                        return Decimal('0')

                def _ipca_fator_acumulado(ano: int, mes: int) -> Decimal:
                    fim = date.today()
                    fator = Decimal('1')
                    y, m = ano, mes
                    while (y, m) <= (fim.year, fim.month):
                        fator *= (Decimal('1') + _ipca_mensal(y, m))
                        if m == 12:
                            y += 1
                            m = 1
                        else:
                            m += 1
                    return fator

                total_cofluhab_atualizado = Decimal('0')
                for mes_num in range(1, 13):
                    seguro_mes = (
                        ParcelaContrato.objects
                        .filter(dtvenc__year=ano_relatorio, dtvenc__month=mes_num)
                        .aggregate(total=Sum('seguro'))
                        .get('total')
                    ) or Decimal('0')
                    seguro_mes_dec = Decimal(str(seguro_mes))
                    total_cofluhab_atualizado += seguro_mes_dec * _ipca_fator_acumulado(ano_relatorio, mes_num)

                total_cofluhab_atualizado_ano = total_cofluhab_atualizado
            except Exception:
                pass

        if total_cef_corrigido_ano is not None:
            diferenca_cef_pdf_ano = total_cef_corrigido_ano - total_cef
            if total_cofluhab_atualizado_ano is not None:
                diferenca_cef_cofluhab_ano = total_cef_corrigido_ano - total_cofluhab_atualizado_ano

        def _fmt_dec(v: Decimal | None):
            if v is None:
                return '-'
            return f"{v:,.2f}".replace('.', '_').replace(',', '.').replace('_', ',')

        contexto.update({
            'headers': headers,
            'rows': rows,
            'total_rows': total_rows,
            'sumario': {
                'ano_relatorio': ano_relatorio,
                'contratos_unicos': len(contratos_unicos),
                'encontrados': encontrados,
                'nao_encontrados': nao_encontrados,
                'total_cef_cobrando': f"{total_cef:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                'total_divida_banco': f"{total_divida:,.2f}".replace('.', '_').replace(',', '.').replace('_', ','),
                'total_cef_historico_ano': _fmt_dec(total_cef_historico_ano),
                'total_cef_corrigido_ano': _fmt_dec(total_cef_corrigido_ano),
                'total_cofluhab_atualizado_ano': _fmt_dec(total_cofluhab_atualizado_ano),
                'diferenca_cef_pdf_ano': _fmt_dec(diferenca_cef_pdf_ano),
                'diferenca_cef_cofluhab_ano': _fmt_dec(diferenca_cef_cofluhab_ano),
            },
        })
    elif contexto['is_md']:
        with open(full_path, 'r', encoding='utf-8') as fp:
            contexto['markdown_content'] = fp.read()

    return render(request, 'principal/visualizar_relatorio_divida_seguro.html', contexto)


def parcelas_pagas(request, codigo):
    """Mostra relatório de parcelas pagas de um contrato, com data e valor de pagamento."""
    contrato = get_object_or_404(Contrato, codigo=codigo)
    data_sentinela_pagamento = date(1980, 1, 1)

    def normalizar_data_pagamento(dtpgto):
        if dtpgto == data_sentinela_pagamento:
            return None
        # Dados importados com ano 190x representam 200x na base legada.
        if dtpgto and dtpgto.year < 1950:
            try:
                return dtpgto.replace(year=dtpgto.year + 100)
            except ValueError:
                return dtpgto
        return dtpgto

    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')

    data_inicio = None
    data_fim = None

    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        except ValueError:
            data_inicio = None

    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            data_fim = None

    parcelas_qs = ParcelaContrato.objects.filter(
        contrato=contrato,
        dtpgto__isnull=False
    ).order_by('dtpgto', 'nmens')

    if data_inicio:
        parcelas_qs = parcelas_qs.filter(dtpgto__gte=data_inicio)
    if data_fim:
        parcelas_qs = parcelas_qs.filter(dtpgto__lte=data_fim)

    parcelas_list = []
    total_pago = Decimal('0')
    parcelas_sem_data = 0

    for p in parcelas_qs:
        if p.vlautent and p.vlautent > 0:
            valor_pago = Decimal(str(p.vlautent))
        else:
            valor_pago = Decimal('0')
            if p.juros:
                valor_pago += Decimal(str(p.juros))
            if p.amort:
                valor_pago += Decimal(str(p.amort))
            if p.seguro:
                valor_pago += Decimal(str(p.seguro))
            if p.tca:
                valor_pago += Decimal(str(p.tca))
            if p.fcvs:
                valor_pago += Decimal(str(p.fcvs))
            if p.em:
                valor_pago += Decimal(str(p.em))
            if p.rp:
                valor_pago += Decimal(str(p.rp))

        dtpgto_corrigida = normalizar_data_pagamento(p.dtpgto)
        if not dtpgto_corrigida:
            parcelas_sem_data += 1

        dias_para_pagar = None
        if p.dtvenc and dtpgto_corrigida:
            dias_para_pagar = (dtpgto_corrigida - p.dtvenc).days

        total_pago += valor_pago
        parcelas_list.append({
            'parcela': p,
            'nmens': p.nmens,
            'dtvenc': p.dtvenc,
            'dtpgto': dtpgto_corrigida,
            'dias_para_pagar': dias_para_pagar,
            'valor_pago': valor_pago,
        })

    parcelas_list.sort(key=lambda item: (
        item['dtvenc'] is None,
        item['dtvenc'] or date.max,
        item['nmens'],
        item['dtpgto'] or date.max,
    ))

    exportar = (request.GET.get('exportar', '') or '').lower()
    if exportar == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="parcelas_pagas_{contrato.codigo}.csv"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'Contrato', 'Conjunto', 'Parcela', 'Vencimento', 'Pagamento',
            'Diferenca Dias', 'Valor Pago'
        ])

        for item in parcelas_list:
            writer.writerow([
                contrato.codigo,
                contrato.conjunto,
                item['nmens'],
                item['dtvenc'].strftime('%d/%m/%Y') if item['dtvenc'] else '',
                item['dtpgto'].strftime('%d/%m/%Y') if item['dtpgto'] else '',
                item['dias_para_pagar'] if item['dias_para_pagar'] is not None else '',
                f"{item['valor_pago']:.2f}",
            ])

        writer.writerow([])
        writer.writerow(['TOTAL PAGO', '', '', '', '', '', f"{total_pago:.2f}"])
        return response

    if exportar == 'excel':
        try:
            from openpyxl import Workbook
        except ImportError:
            return HttpResponse('openpyxl nao esta instalado no ambiente.', status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Parcelas Pagas'

        ws.append([
            'Contrato', 'Conjunto', 'Parcela', 'Vencimento', 'Pagamento',
            'Diferenca Dias', 'Valor Pago'
        ])

        for item in parcelas_list:
            ws.append([
                contrato.codigo,
                contrato.conjunto,
                item['nmens'],
                item['dtvenc'].strftime('%d/%m/%Y') if item['dtvenc'] else '',
                item['dtpgto'].strftime('%d/%m/%Y') if item['dtpgto'] else '',
                item['dias_para_pagar'] if item['dias_para_pagar'] is not None else '',
                float(item['valor_pago']),
            ])

        ws.append(['', '', '', '', '', 'TOTAL PAGO', float(total_pago)])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="parcelas_pagas_{contrato.codigo}.xlsx"'
        )
        wb.save(response)
        return response

    context = {
        'contrato': contrato,
        'parcelas': parcelas_list,
        'total_pago': total_pago,
        'total_parcelas_pagas': len(parcelas_list),
        'parcelas_sem_data': parcelas_sem_data,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }
    return render(request, 'principal/parcelas_pagas.html', context)


def cadastrar_mutuario(request):
    """Função placeholder para cadastro de mutuário"""
    from django.http import HttpResponse
    return HttpResponse("Função de cadastro em desenvolvimento")


def contratos_sem_mutuario(request):
    """Lista contratos que não possuem mutuário vinculado"""
    # Buscar todos os IDs de contratos que TÊM mutuário
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT contrato_id FROM contrato_mutuario_map")
    contratos_com_mutuario = [row[0] for row in cur.fetchall()]
    conn.close()
    
    # Buscar contratos SEM mutuário
    contratos_sem = Contrato.objects.exclude(id__in=contratos_com_mutuario).order_by('codigo')
    
    # Paginação
    paginator = Paginator(contratos_sem, 100)
    page_obj = paginator.get_page(request.GET.get('pagina', 1))
    
    context = {
        'contratos': page_obj,
        'page_obj': page_obj,
        'total': contratos_sem.count()
    }
    return render(request, 'principal/contratos_sem_mutuario.html', context)


def atualizacao_monetaria(request):
    """Relatório de atualizações monetárias e conversões"""
    # Informações sobre as conversões monetárias
    conversoes = []
    for data_limite, moeda_anterior, fator, moeda_nova in NOMINAL_CONVERSION_FACTORS:
        conversoes.append({
            'data': data_limite,
            'moeda_de': moeda_anterior.replace('_', ' ').title(),
            'moeda_para': moeda_nova.replace('_', ' ').title(),
            'fator': fator,
            'descricao': f'1 {moeda_nova.replace("_", " ").title()} = {fator} {moeda_anterior.replace("_", " ").title()}'
        })
    
    # Buscar alguns contratos para exemplificar
    contratos_exemplo = []
    for contrato in Contrato.objects.all()[:10]:
        primeira = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens').first()
        ultima = ParcelaContrato.objects.filter(contrato=contrato).order_by('-nmens').first()
        
        if primeira and ultima:
            moeda_inicial = get_moeda_vigente(primeira.dtvenc) if primeira.dtvenc else 'Cr$'
            moeda_final = get_moeda_vigente(ultima.dtvenc) if ultima.dtvenc else 'R$'
            
            contratos_exemplo.append({
                'contrato': contrato,
                'saldo_inicial': primeira.sddev,
                'moeda_inicial': moeda_inicial,
                'data_inicial': primeira.dtvenc,
                'saldo_final': ultima.sddev,
                'moeda_final': moeda_final,
                'data_final': ultima.dtvenc
            })
    
    context = {
        'conversoes': conversoes,
        'contratos_exemplo': contratos_exemplo,
        'moedas': MOEDA_POR_PERIODO
    }
    return render(request, 'principal/atualizacao_monetaria.html', context)


def amortizacao_negativa(request):
    """Detecta contratos com amortização negativa (saldo crescente)"""
    contratos_com_anomalia = []
    
    # Buscar contratos com parcelas
    contratos = Contrato.objects.all()[:200]  # Limitar para performance
    
    for contrato in contratos:
        parcelas = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')[:50]
        
        if parcelas.count() < 2:
            continue
            
        parcelas_list = list(parcelas)
        anomalias = 0
        
        for i in range(1, len(parcelas_list)):
            p_ant = parcelas_list[i-1]
            p_atual = parcelas_list[i]
            
            if p_ant.sddev and p_atual.sddev:
                # Detectar crescimento do saldo (amortização negativa)
                if p_atual.sddev > p_ant.sddev * Decimal('1.02'):  # Cresceu mais de 2%
                    anomalias += 1
        
        if anomalias > 0:
            contratos_com_anomalia.append({
                'contrato': contrato,
                'anomalias': anomalias,
                'primeira_parcela': parcelas_list[0] if parcelas_list else None,
                'ultima_parcela': parcelas_list[-1] if parcelas_list else None
            })
    
    # Paginação
    paginator = Paginator(contratos_com_anomalia, 50)
    page_obj = paginator.get_page(request.GET.get('pagina', 1))
    
    context = {
        'contratos': page_obj,
        'page_obj': page_obj,
        'total': len(contratos_com_anomalia)
    }
    return render(request, 'principal/amortizacao_negativa.html', context)


def relatorio_debitos(request):
    """Relatório de débitos por conjunto habitacional"""
    conjunto_filtro = request.GET.get('conjunto', '')
    ordenar_por = request.GET.get('ordenar', 'parcelas')  # 'parcelas' ou 'valor'
    
    contratos_debito = []
    

def _parse_decimal_br(texto):
    texto = (texto or '').strip()
    if not texto:
        return Decimal('0')
    try:
        return Decimal(texto.replace('.', '').replace(',', '.'))
    except Exception:
        return Decimal('0')

    # Filtrar contratos por conjunto se fornecido
    if conjunto_filtro:
        contratos_qs = Contrato.objects.filter(conjunto=conjunto_filtro)
    else:
        contratos_qs = Contrato.objects.all()
    
    # Limitar para performance (processar em lotes)
    contratos_qs = contratos_qs[:500]
    
    for contrato in contratos_qs:
        # Buscar parcelas em aberto (não pagas)
        parcelas_abertas = ParcelaContrato.objects.filter(
            contrato=contrato, 
            dtpgto__isnull=True
        ).order_by('dtvenc')
        
        if parcelas_abertas.count() > 0:
            # Calcular total do débito
            total_debito = Decimal('0')
            for p in parcelas_abertas:
                if p.vlautent:
                    total_debito += p.vlautent
                else:
                    # Somar componentes se vlautent não existir
                    if p.juros: total_debito += p.juros
                    if p.amort: total_debito += p.amort
                    if p.seguro: total_debito += p.seguro
                    if p.tca: total_debito += p.tca
            
            # Buscar mutuário
            db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
            mutuario = None
            try:
                conn = sqlite3.connect(db_path)
                cur = conn.cursor()
                cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
                result = cur.fetchone()
                if result:
                    mutuario = Mutuario.objects.get(id=result[0])
                conn.close()
            except:
                pass
            
            contratos_debito.append({
                'contrato': contrato,
                'mutuario': mutuario,
                'parcelas_abertas': parcelas_abertas.count(),
                'total_debito': total_debito,
                'primeira_vencida': parcelas_abertas.first()
            })
    
    # Ordenar
    if ordenar_por == 'valor':
        contratos_debito.sort(key=lambda x: x['total_debito'], reverse=True)
    else:
        contratos_debito.sort(key=lambda x: x['parcelas_abertas'], reverse=True)
    
    # Calcular estatísticas
    total_contratos_debito = len(contratos_debito)
    total_geral_debito = sum([c['total_debito'] for c in contratos_debito])
    
    # Paginação
    paginator = Paginator(contratos_debito, 50)
    page_obj = paginator.get_page(request.GET.get('pagina', 1))
    
    # Buscar lista de conjuntos para filtro
    conjuntos = ConjuntoHabitacional.objects.all().order_by('conjunto')
    
    context = {
        'contratos_debito': page_obj,
        'page_obj': page_obj,
        'total_contratos_debito': total_contratos_debito,
        'total_geral_debito': total_geral_debito,
        'conjuntos': conjuntos,
        'conjunto_filtro': conjunto_filtro,
        'ordenar_por': ordenar_por,
        'stats': {
            'total_contratos': total_contratos_debito,
            'total_debito': total_geral_debito,
            'media_debito': total_geral_debito / total_contratos_debito if total_contratos_debito > 0 else 0
        }
    }
    return render(request, 'principal/relatorio_debitos.html', context)


def fcvs(request):
    """Análise do FCVS (Fundo de Compensação de Variações Salariais)"""
    # Filtros
    conjunto_filtro = request.GET.get('conjunto', '')
    codigo_contrato_filtro = request.GET.get('codigo_contrato', '')
    tipo_analise = request.GET.get('tipo_analise', 'todos')  # 'todos', 'com_fcvs', 'anomalias'
    analisar_conjunto = request.GET.get('analisar_conjunto', '')  # Analisa um conjunto específico
    
    contratos_fcvs = []
    
    # Filtrar contratos
    contratos_qs = Contrato.objects.all()
    
    if conjunto_filtro:
        contratos_qs = contratos_qs.filter(conjunto=conjunto_filtro)
    
    if codigo_contrato_filtro:
        contratos_qs = contratos_qs.filter(codigo__icontains=codigo_contrato_filtro)
    
    # Se está analisando um conjunto específico, busca todos os contratos desse conjunto
    if analisar_conjunto:
        contratos_qs = contratos_qs.filter(conjunto=analisar_conjunto)
        contratos_qs = contratos_qs[:2000]  # Limite para análise de conjunto
    elif conjunto_filtro or codigo_contrato_filtro:
        contratos_qs = contratos_qs[:2000]
    else:
        contratos_qs = contratos_qs[:500]  # Limite padrão sem filtros
    
    for contrato in contratos_qs:
        parcelas = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')
        
        if parcelas.count() < 2:
            continue
        
        primeira = parcelas.first()
        ultima = parcelas.last()
        
        # Usar a função de cálculo correto de FCVS com conversões monetárias
        evolucao_completa, anomalias, fcvs_residual = calcular_fcvs_residual_global(contrato.id)
        
        # Método 1: Contribuição mensal de FCVS nas parcelas
        total_contribuicao_fcvs = Decimal('0')
        parcelas_com_fcvs = 0
        for p in parcelas:
            if p.fcvs and p.fcvs > 0:
                total_contribuicao_fcvs += p.fcvs
                parcelas_com_fcvs += 1
        
        # Método 2: FCVS calculado com conversões monetárias (valor correto)
        fcvs_estimado = Decimal(str(fcvs_residual))
        
        # Informações sobre período de alta inflação
        periodo_critico = False
        if primeira.dtvenc:
            # Período crítico: 1986-1994 (hiperinflação)
            if primeira.dtvenc.year >= 1986 and primeira.dtvenc.year <= 1994:
                periodo_critico = True
        
        # Incluir baseado no tipo de análise
        incluir = False
        if tipo_analise == 'todos':
            incluir = True
        elif tipo_analise == 'com_fcvs' and (total_contribuicao_fcvs > 0 or fcvs_estimado > 0):
            incluir = True
        elif tipo_analise == 'anomalias' and anomalias > 0:
            incluir = True
        
        if incluir:
            # Buscar mutuário usando SQLite direto
            mutuario = None
            try:
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
                conn = sqlite3.connect(db_path)
                cur = conn.cursor()
                cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
                result = cur.fetchone()
                if result:
                    mut = Mutuario.objects.get(id=result[0])
                    mutuario = mut.nome
                conn.close()
            except:
                pass
            
            contratos_fcvs.append({
                'contrato_id': contrato.id,
                'codigo_contrato': contrato.codigo,
                'conjunto': contrato.conjunto,
                'mutuario': mutuario,
                'contribuicao_fcvs': total_contribuicao_fcvs,
                'fcvs_estimado': fcvs_estimado,
                'parcelas_com_fcvs': parcelas_com_fcvs,
                'anomalias': anomalias,
                'primeira_parcela': primeira,
                'ultima_parcela': ultima,
                'periodo': f"{primeira.dtvenc.strftime('%m/%Y') if primeira.dtvenc else '?'} - {ultima.dtvenc.strftime('%m/%Y') if ultima.dtvenc else '?'}",
                'periodo_critico': periodo_critico,
                'total_parcelas': parcelas.count()
            })
    
    # Ordenar por FCVS total (contribuição + estimado)
    contratos_fcvs.sort(key=lambda x: x['contribuicao_fcvs'] + x['fcvs_estimado'], reverse=True)
    
    # Calcular totais
    total_contribuicao_fcvs = sum([c['contribuicao_fcvs'] for c in contratos_fcvs])
    total_fcvs_estimado = sum([c['fcvs_estimado'] for c in contratos_fcvs])
    total_contratos_periodo_critico = sum([1 for c in contratos_fcvs if c['periodo_critico']])
    
    # Resumo do conjunto analisado
    resumo_conjunto_analisado = None
    if analisar_conjunto and contratos_fcvs:
        resumo_conjunto_analisado = {
            'conjunto': analisar_conjunto,
            'total_contribuicao_fcvs': Decimal('0'),
            'total_fcvs_estimado': Decimal('0'),
            'total_fcvs': Decimal('0'),
            'contratos_com_fcvs': 0,
            'contratos_analisados': 0,
            'anomalias': 0
        }
        
        for item in contratos_fcvs:
            resumo_conjunto_analisado['total_contribuicao_fcvs'] += item['contribuicao_fcvs']
            resumo_conjunto_analisado['total_fcvs_estimado'] += item['fcvs_estimado']
            resumo_conjunto_analisado['total_fcvs'] += item['contribuicao_fcvs'] + item['fcvs_estimado']
            resumo_conjunto_analisado['contratos_analisados'] += 1
            resumo_conjunto_analisado['anomalias'] += item['anomalias']
            
            if item['contribuicao_fcvs'] > 0 or item['fcvs_estimado'] > 0:
                resumo_conjunto_analisado['contratos_com_fcvs'] += 1
        
        # Calcular média e formatar
        if resumo_conjunto_analisado['contratos_com_fcvs'] > 0:
            resumo_conjunto_analisado['media_fcvs'] = float(resumo_conjunto_analisado['total_fcvs']) / resumo_conjunto_analisado['contratos_com_fcvs']
        else:
            resumo_conjunto_analisado['media_fcvs'] = 0
        
        resumo_conjunto_analisado['total_fcvs'] = float(resumo_conjunto_analisado['total_fcvs'])
        resumo_conjunto_analisado['total_contribuicao_fcvs'] = float(resumo_conjunto_analisado['total_contribuicao_fcvs'])
        resumo_conjunto_analisado['total_fcvs_estimado'] = float(resumo_conjunto_analisado['total_fcvs_estimado'])
    
    # Paginação
    paginator = Paginator(contratos_fcvs, 50)
    page_obj = paginator.get_page(request.GET.get('pagina', 1))
    
    # Buscar conjuntos para filtro
    conjuntos = ConjuntoHabitacional.objects.all().order_by('conjunto')
    
    context = {
        'contratos': page_obj,
        'page_obj': page_obj,
        'total': len(contratos_fcvs),
        'total_contribuicao_fcvs': total_contribuicao_fcvs,
        'total_fcvs_estimado': total_fcvs_estimado,
        'total_fcvs_geral': total_contribuicao_fcvs + total_fcvs_estimado,
        'total_contratos_periodo_critico': total_contratos_periodo_critico,
        'conjuntos': conjuntos,
        'conjunto_filtro': conjunto_filtro,
        'codigo_contrato_filtro': codigo_contrato_filtro,
        'tipo_analise': tipo_analise,
        'resumo_conjunto_analisado': resumo_conjunto_analisado,
        'analisar_conjunto': analisar_conjunto
    }
    return render(request, 'principal/fcvs.html', context)


def fcvs_contribuicao(request):
    """Relatório de contribuição FCVS por mês/ano com exportação."""
    hoje = date.today()
    ano_atual = hoje.year

    periodicidade = (request.GET.get('periodicidade', 'mensal') or 'mensal').strip().lower()
    if periodicidade not in ('mensal', 'trimestral'):
        periodicidade = 'mensal'

    meses_nome = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Marco', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro',
    }

    try:
        ano = int(request.GET.get('ano', ano_atual))
    except Exception:
        ano = ano_atual

    try:
        mes = int(request.GET.get('mes', hoje.month))
    except Exception:
        mes = hoje.month

    if ano < 2000:
        ano = 2000
    if ano > ano_atual:
        ano = ano_atual
    if mes < 1 or mes > 12:
        mes = hoje.month

    modo = (request.GET.get('modo', 'resumo') or 'resumo').strip().lower()
    # Tela principal otimizada: aceita somente resumo e detalhe mensal sob demanda.
    if modo not in ('resumo', 'mensal'):
        modo = 'resumo'

    try:
        trimestre = int(request.GET.get('trimestre', ((hoje.month - 1) // 3) + 1))
    except Exception:
        trimestre = ((hoje.month - 1) // 3) + 1
    if trimestre < 1 or trimestre > 4:
        trimestre = ((hoje.month - 1) // 3) + 1

    exportar = (request.GET.get('exportar', '') or '').strip().lower()

    base_qs = ParcelaContrato.objects.filter(
        dtvenc__isnull=False,
        dtvenc__year__gte=2000,
        fcvs__gt=0,
    ).select_related('contrato')

    anos_com_dados = list(
        base_qs.values_list('dtvenc__year', flat=True).distinct().order_by('dtvenc__year')
    )
    if anos_com_dados:
        # Exibe somente anos que realmente possuem FCVS>0 no banco carregado.
        anos_disponiveis = anos_com_dados
        if ano not in anos_disponiveis:
            anos_ate_selecao = [y for y in anos_disponiveis if y <= ano]
            ano = (anos_ate_selecao[-1] if anos_ate_selecao else anos_disponiveis[-1])
    else:
        anos_disponiveis = list(range(2000, ano_atual + 1))

    compute_mensal = modo == 'mensal' or (exportar in ('csv', 'excel') and modo == 'mensal')
    compute_trimestral = False
    compute_anual = False

    compute_grupos = False

    need_total_mensal = compute_mensal or modo == 'resumo' or modo == 'grupo_mensal'
    need_total_trimestral = False
    need_total_anual = False

    mensal_qs = base_qs.filter(dtvenc__year=ano, dtvenc__month=mes)
    mensal_por_contrato = []
    if compute_mensal:
        mensal_por_contrato = list(
            mensal_qs.values('contrato__codigo', 'contrato__conjunto')
            .annotate(
                total_fcvs=Sum('fcvs'),
                total_parcelas=Count('id'),
            )
            .order_by('-total_fcvs', 'contrato__codigo')
        )
    if need_total_mensal:
        total_mensal = mensal_qs.aggregate(total=Sum('fcvs'))['total'] or Decimal('0')
    else:
        total_mensal = Decimal('0')

    trimestre_meses = {
        1: [1, 2, 3],
        2: [4, 5, 6],
        3: [7, 8, 9],
        4: [10, 11, 12],
    }
    meses_trimestre = trimestre_meses[trimestre]
    trimestral_qs = base_qs.filter(dtvenc__year=ano, dtvenc__month__in=meses_trimestre)
    trimestral_por_contrato = []
    if compute_trimestral:
        trimestral_por_contrato = list(
            trimestral_qs.values('contrato__codigo', 'contrato__conjunto')
            .annotate(
                total_fcvs=Sum('fcvs'),
                total_parcelas=Count('id'),
            )
            .order_by('-total_fcvs', 'contrato__codigo')
        )
    if need_total_trimestral:
        total_trimestral = trimestral_qs.aggregate(total=Sum('fcvs'))['total'] or Decimal('0')
    else:
        total_trimestral = Decimal('0')

    trimestral_por_mes = []
    if compute_trimestral:
        for mes_q in meses_trimestre:
            mes_q_qs = base_qs.filter(dtvenc__year=ano, dtvenc__month=mes_q)
            trimestral_por_mes.append({
                'mes_numero': mes_q,
                'mes_nome': meses_nome[mes_q],
                'total_fcvs': mes_q_qs.aggregate(total=Sum('fcvs'))['total'] or Decimal('0'),
                'total_contratos': mes_q_qs.values('contrato').distinct().count(),
                'total_parcelas': mes_q_qs.count(),
            })

    anual_qs = base_qs.filter(dtvenc__year=ano)
    anual_por_mes_raw = {
        item['dtvenc__month']: item
        for item in anual_qs.values('dtvenc__month').annotate(
            total_fcvs=Sum('fcvs'),
            total_parcelas=Count('id'),
            total_contratos=Count('contrato', distinct=True),
        )
    }

    resumo_meses = []
    for numero_mes in range(1, 13):
        item = anual_por_mes_raw.get(numero_mes)
        resumo_meses.append({
            'mes_numero': numero_mes,
            'mes_nome': meses_nome[numero_mes],
            'total_fcvs': (item['total_fcvs'] if item else Decimal('0')),
            'total_parcelas': (item['total_parcelas'] if item else 0),
            'total_contratos': (item['total_contratos'] if item else 0),
            'is_mes_selecionado': numero_mes == mes,
        })

    total_anual_resumo = sum((item['total_fcvs'] for item in resumo_meses), Decimal('0'))

    resumo_trimestres = []
    for numero_trimestre, meses_t in ((1, (1, 2, 3)), (2, (4, 5, 6)), (3, (7, 8, 9)), (4, (10, 11, 12))):
        linhas_t = [item for item in resumo_meses if item['mes_numero'] in meses_t]
        resumo_trimestres.append({
            'trimestre': numero_trimestre,
            'label': f"{numero_trimestre}o Trimestre",
            'total_fcvs': sum((item['total_fcvs'] for item in linhas_t), Decimal('0')),
            'total_contratos': sum((item['total_contratos'] for item in linhas_t), 0),
            'total_parcelas': sum((item['total_parcelas'] for item in linhas_t), 0),
        })

    contratos_trimestrais_ano = []
    total_fcvs_contratos_trimestrais_ano = Decimal('0')
    if periodicidade == 'trimestral':
        contrato_ids_ano = list(
            base_qs.filter(dtvenc__year=ano)
            .values_list('contrato_id', flat=True)
            .distinct()
        )

        if contrato_ids_ano:
            linhas_historicas = list(
                base_qs.filter(contrato_id__in=contrato_ids_ano)
                .values('contrato_id', 'contrato__codigo', 'contrato__conjunto', 'dtvenc')
                .order_by('contrato_id', 'dtvenc')
            )

            totais_ano_por_contrato = {
                item['contrato_id']: {
                    'total_fcvs_ano': Decimal(str(item['total_fcvs'] or 0)),
                    'total_parcelas_ano': item['total_parcelas'] or 0,
                }
                for item in base_qs.filter(dtvenc__year=ano, contrato_id__in=contrato_ids_ano)
                .values('contrato_id')
                .annotate(total_fcvs=Sum('fcvs'), total_parcelas=Count('id'))
            }

            contratos_hist = {}
            for row in linhas_historicas:
                cid = row['contrato_id']
                item = contratos_hist.setdefault(cid, {
                    'contrato': row['contrato__codigo'],
                    'conjunto': row['contrato__conjunto'],
                    'meses_set': set(),
                })
                dt = row['dtvenc']
                item['meses_set'].add((dt.year, dt.month))

            for cid, item in contratos_hist.items():
                meses = sorted(item['meses_set'])
                if len(meses) < 2:
                    continue

                idx = [(ano_m * 12 + mes_m) for ano_m, mes_m in meses]
                intervalos = [idx[i] - idx[i - 1] for i in range(1, len(idx))]

                # Trimestral puro: todas as cobrancas seguem salto de 3 meses.
                if any(it != 3 for it in intervalos):
                    continue

                totais_ano = totais_ano_por_contrato.get(cid)
                if not totais_ano:
                    continue

                contratos_trimestrais_ano.append({
                    'contrato': item['contrato'],
                    'conjunto': item['conjunto'],
                    'primeiro_mes_historico': f"{meses[0][1]:02d}/{meses[0][0]}",
                    'ultimo_mes_historico': f"{meses[-1][1]:02d}/{meses[-1][0]}",
                    'total_fcvs_ano': totais_ano['total_fcvs_ano'],
                    'total_parcelas_ano': totais_ano['total_parcelas_ano'],
                })

            contratos_trimestrais_ano.sort(
                key=lambda x: (x['total_fcvs_ano'], x['total_parcelas_ano']),
                reverse=True,
            )
            total_fcvs_contratos_trimestrais_ano = sum(
                (item['total_fcvs_ano'] for item in contratos_trimestrais_ano),
                Decimal('0'),
            )

    anual_por_mes = []
    if compute_anual:
        anual_por_mes = resumo_meses
        total_anual = total_anual_resumo
    elif need_total_anual:
        total_anual = total_anual_resumo
    else:
        total_anual = Decimal('0')

    periodicidade_stats = {
        'total_contratos': 0,
        'trimestral_puro': 0,
        'mensal_puro': 0,
        'misto_regular': 0,
        'irregular': 0,
        'pontual': 0,
        'diferente_3m': 0,
    }
    contratos_mensal_puro = []
    contratos_trimestral_puro = []
    contratos_diferente_3m = []
    total_grupo_mensal_mes = Decimal('0')
    total_grupo_mensal_historico = Decimal('0')
    total_grupo_trimestral_trimestre = Decimal('0')
    total_grupo_trimestral_historico = Decimal('0')

    if compute_grupos:
        # Analise de cadencia de contribuicao por contrato (mensal x trimestral x irregular)
        linhas_cadencia = list(
            base_qs.values('contrato_id', 'contrato__codigo', 'contrato__conjunto', 'dtvenc', 'fcvs')
            .order_by('contrato_id', 'dtvenc')
        )

        contratos_cadencia = {}
        for row in linhas_cadencia:
            cid = row['contrato_id']
            item = contratos_cadencia.setdefault(cid, {
                'contrato': row['contrato__codigo'],
                'conjunto': row['contrato__conjunto'],
                'meses_set': set(),
                'total_fcvs': Decimal('0'),
                'total_parcelas': 0,
            })
            dt = row['dtvenc']
            item['meses_set'].add((dt.year, dt.month))
            item['total_fcvs'] += Decimal(str(row['fcvs'] or 0))
            item['total_parcelas'] += 1

        total_mes_por_contrato = {
            item['contrato_id']: item['total_fcvs']
            for item in base_qs.filter(dtvenc__year=ano, dtvenc__month=mes)
            .values('contrato_id')
            .annotate(total_fcvs=Sum('fcvs'))
        }
        total_trimestre_por_contrato = {
            item['contrato_id']: item['total_fcvs']
            for item in base_qs.filter(dtvenc__year=ano, dtvenc__month__in=meses_trimestre)
            .values('contrato_id')
            .annotate(total_fcvs=Sum('fcvs'))
        }

        for cid, item in contratos_cadencia.items():
            meses = sorted(item['meses_set'])
            idx = [(ano_m * 12 + mes_m) for ano_m, mes_m in meses]
            intervalos = [idx[i] - idx[i - 1] for i in range(1, len(idx))]

            count_1 = sum(1 for it in intervalos if it == 1)
            count_3 = sum(1 for it in intervalos if it == 3)
            count_outros = sum(1 for it in intervalos if it not in (1, 3))

            if not intervalos:
                perfil = 'PONTUAL'
                periodicidade_stats['pontual'] += 1
            elif count_outros == 0 and count_3 > 0 and count_1 == 0:
                perfil = 'TRIMESTRAL_PURO'
                periodicidade_stats['trimestral_puro'] += 1
            elif count_outros == 0 and count_1 > 0 and count_3 == 0:
                perfil = 'MENSAL_PURO'
                periodicidade_stats['mensal_puro'] += 1
            elif count_outros == 0 and count_1 > 0 and count_3 > 0:
                perfil = 'MISTO_REGULAR'
                periodicidade_stats['misto_regular'] += 1
            else:
                perfil = 'IRREGULAR'
                periodicidade_stats['irregular'] += 1

            periodicidade_stats['total_contratos'] += 1

            dados_base = {
                'contrato_id': cid,
                'contrato': item['contrato'],
                'conjunto': item['conjunto'],
                'primeiro_mes': f"{meses[0][1]:02d}/{meses[0][0]}",
                'ultimo_mes': f"{meses[-1][1]:02d}/{meses[-1][0]}",
                'qtd_meses': len(meses),
                'total_parcelas': item['total_parcelas'],
                'total_fcvs_historico': item['total_fcvs'],
            }

            if perfil == 'MENSAL_PURO':
                contratos_mensal_puro.append({
                    **dados_base,
                    'total_fcvs_mes': Decimal(str(total_mes_por_contrato.get(cid) or 0)),
                })
            elif perfil == 'TRIMESTRAL_PURO':
                contratos_trimestral_puro.append({
                    **dados_base,
                    'total_fcvs_trimestre': Decimal(str(total_trimestre_por_contrato.get(cid) or 0)),
                })

            diferente_3m = len(intervalos) > 0 and any(it != 3 for it in intervalos)
            if diferente_3m:
                periodicidade_stats['diferente_3m'] += 1
                contratos_diferente_3m.append({
                    'contrato': item['contrato'],
                    'conjunto': item['conjunto'],
                    'perfil': perfil,
                    'total_fcvs': item['total_fcvs'],
                    'total_parcelas': item['total_parcelas'],
                    'qtd_meses': len(meses),
                    'primeiro_mes': f"{meses[0][1]:02d}/{meses[0][0]}",
                    'ultimo_mes': f"{meses[-1][1]:02d}/{meses[-1][0]}",
                    'intervalos': ', '.join(str(x) for x in intervalos[:12]) + ('...' if len(intervalos) > 12 else ''),
                    'qtd_intervalos_fora_3': sum(1 for it in intervalos if it != 3),
                })

        contratos_diferente_3m.sort(key=lambda x: (x['qtd_intervalos_fora_3'], x['total_fcvs']), reverse=True)
        contratos_mensal_puro.sort(key=lambda x: (x['total_fcvs_mes'], x['total_fcvs_historico']), reverse=True)
        contratos_trimestral_puro.sort(key=lambda x: (x['total_fcvs_trimestre'], x['total_fcvs_historico']), reverse=True)

        total_grupo_mensal_mes = sum((item['total_fcvs_mes'] for item in contratos_mensal_puro), Decimal('0'))
        total_grupo_mensal_historico = sum((item['total_fcvs_historico'] for item in contratos_mensal_puro), Decimal('0'))
        total_grupo_trimestral_trimestre = sum((item['total_fcvs_trimestre'] for item in contratos_trimestral_puro), Decimal('0'))
        total_grupo_trimestral_historico = sum((item['total_fcvs_historico'] for item in contratos_trimestral_puro), Decimal('0'))

    if exportar in ('csv', 'excel'):
        if modo == 'grupo_mensal':
            if exportar == 'csv':
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="fcvs_grupo_mensal_{ano}_{mes:02d}.csv"'
                writer = csv.writer(response, delimiter=';')
                writer.writerow([
                    'Contrato', 'Conjunto', 'Primeiro Mes', 'Ultimo Mes',
                    'Meses com FCVS', 'Parcelas FCVS',
                    f'FCVS Mes {mes:02d}/{ano}', 'FCVS Historico'
                ])
                for item in contratos_mensal_puro:
                    writer.writerow([
                        item['contrato'],
                        item['conjunto'],
                        item['primeiro_mes'],
                        item['ultimo_mes'],
                        item['qtd_meses'],
                        item['total_parcelas'],
                        f"{Decimal(str(item['total_fcvs_mes'])):.2f}",
                        f"{Decimal(str(item['total_fcvs_historico'])):.2f}",
                    ])
                writer.writerow([
                    'TOTAL GRUPO MENSAL', '', '', '', '', '',
                    f"{total_grupo_mensal_mes:.2f}",
                    f"{total_grupo_mensal_historico:.2f}",
                ])
                return response

            try:
                from openpyxl import Workbook
            except ImportError:
                return HttpResponse('openpyxl nao esta instalado no ambiente.', status=500)

            wb = Workbook()
            ws = wb.active
            ws.title = 'Grupo Mensal'
            ws.append([
                'Contrato', 'Conjunto', 'Primeiro Mes', 'Ultimo Mes',
                'Meses com FCVS', 'Parcelas FCVS',
                f'FCVS Mes {mes:02d}/{ano}', 'FCVS Historico'
            ])
            for item in contratos_mensal_puro:
                ws.append([
                    item['contrato'],
                    item['conjunto'],
                    item['primeiro_mes'],
                    item['ultimo_mes'],
                    item['qtd_meses'],
                    item['total_parcelas'],
                    float(item['total_fcvs_mes']),
                    float(item['total_fcvs_historico']),
                ])
            ws.append([
                'TOTAL GRUPO MENSAL', '', '', '', '', '',
                float(total_grupo_mensal_mes),
                float(total_grupo_mensal_historico),
            ])

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="fcvs_grupo_mensal_{ano}_{mes:02d}.xlsx"'
            wb.save(response)
            return response

        if modo == 'grupo_trimestral':
            nome_trimestre = f'T{trimestre}'
            if exportar == 'csv':
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="fcvs_grupo_trimestral_{ano}_{nome_trimestre}.csv"'
                writer = csv.writer(response, delimiter=';')
                writer.writerow([
                    'Contrato', 'Conjunto', 'Primeiro Mes', 'Ultimo Mes',
                    'Meses com FCVS', 'Parcelas FCVS',
                    f'FCVS Trimestre {nome_trimestre}/{ano}', 'FCVS Historico'
                ])
                for item in contratos_trimestral_puro:
                    writer.writerow([
                        item['contrato'],
                        item['conjunto'],
                        item['primeiro_mes'],
                        item['ultimo_mes'],
                        item['qtd_meses'],
                        item['total_parcelas'],
                        f"{Decimal(str(item['total_fcvs_trimestre'])):.2f}",
                        f"{Decimal(str(item['total_fcvs_historico'])):.2f}",
                    ])
                writer.writerow([
                    'TOTAL GRUPO TRIMESTRAL', '', '', '', '', '',
                    f"{total_grupo_trimestral_trimestre:.2f}",
                    f"{total_grupo_trimestral_historico:.2f}",
                ])
                return response

            try:
                from openpyxl import Workbook
            except ImportError:
                return HttpResponse('openpyxl nao esta instalado no ambiente.', status=500)

            wb = Workbook()
            ws = wb.active
            ws.title = 'Grupo Trimestral'
            ws.append([
                'Contrato', 'Conjunto', 'Primeiro Mes', 'Ultimo Mes',
                'Meses com FCVS', 'Parcelas FCVS',
                f'FCVS Trimestre {nome_trimestre}/{ano}', 'FCVS Historico'
            ])
            for item in contratos_trimestral_puro:
                ws.append([
                    item['contrato'],
                    item['conjunto'],
                    item['primeiro_mes'],
                    item['ultimo_mes'],
                    item['qtd_meses'],
                    item['total_parcelas'],
                    float(item['total_fcvs_trimestre']),
                    float(item['total_fcvs_historico']),
                ])
            ws.append([
                'TOTAL GRUPO TRIMESTRAL', '', '', '', '', '',
                float(total_grupo_trimestral_trimestre),
                float(total_grupo_trimestral_historico),
            ])

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="fcvs_grupo_trimestral_{ano}_{nome_trimestre}.xlsx"'
            wb.save(response)
            return response

        if modo == 'trimestral':
            linhas_tri = trimestral_por_contrato
            nome_trimestre = f'T{trimestre}'

            if exportar == 'csv':
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="fcvs_contribuicao_trimestral_{ano}_{nome_trimestre}.csv"'
                writer = csv.writer(response, delimiter=';')
                writer.writerow(['Ano', 'Trimestre', 'Contrato', 'Conjunto', 'Contribuicao FCVS', 'Parcelas'])
                for item in linhas_tri:
                    writer.writerow([
                        ano,
                        nome_trimestre,
                        item['contrato__codigo'],
                        item['contrato__conjunto'],
                        f"{Decimal(str(item['total_fcvs'])):.2f}",
                        item['total_parcelas'],
                    ])
                writer.writerow(['', '', 'TOTAL TRIMESTRAL', '', f"{total_trimestral:.2f}", ''])
                return response

            try:
                from openpyxl import Workbook
            except ImportError:
                return HttpResponse('openpyxl nao esta instalado no ambiente.', status=500)

            wb = Workbook()
            ws = wb.active
            ws.title = f'{ano}-T{trimestre}'
            ws.append(['Ano', 'Trimestre', 'Contrato', 'Conjunto', 'Contribuicao FCVS', 'Parcelas'])
            for item in linhas_tri:
                ws.append([
                    ano,
                    nome_trimestre,
                    item['contrato__codigo'],
                    item['contrato__conjunto'],
                    float(item['total_fcvs']),
                    item['total_parcelas'],
                ])
            ws.append(['', '', 'TOTAL TRIMESTRAL', '', float(total_trimestral), ''])

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="fcvs_contribuicao_trimestral_{ano}_{nome_trimestre}.xlsx"'
            wb.save(response)
            return response

        if modo == 'anual':
            linhas_anual = anual_por_mes

            if exportar == 'csv':
                response = HttpResponse(content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="fcvs_contribuicao_anual_{ano}.csv"'
                writer = csv.writer(response, delimiter=';')
                writer.writerow(['Ano', 'Mes', 'Contribuicao FCVS', 'Contratos', 'Parcelas'])
                for item in linhas_anual:
                    writer.writerow([
                        ano,
                        item['mes_nome'],
                        f"{Decimal(str(item['total_fcvs'])):.2f}",
                        item['total_contratos'],
                        item['total_parcelas'],
                    ])
                writer.writerow(['', 'TOTAL ANUAL', f"{total_anual:.2f}", '', ''])
                return response

            try:
                from openpyxl import Workbook
            except ImportError:
                return HttpResponse('openpyxl nao esta instalado no ambiente.', status=500)

            wb = Workbook()
            ws = wb.active
            ws.title = f'FCVS {ano}'
            ws.append(['Ano', 'Mes', 'Contribuicao FCVS', 'Contratos', 'Parcelas'])
            for item in linhas_anual:
                ws.append([
                    ano,
                    item['mes_nome'],
                    float(item['total_fcvs']),
                    item['total_contratos'],
                    item['total_parcelas'],
                ])
            ws.append(['', 'TOTAL ANUAL', float(total_anual), '', ''])

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="fcvs_contribuicao_anual_{ano}.xlsx"'
            wb.save(response)
            return response

        linhas_mensal = mensal_por_contrato

        if exportar == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="fcvs_contribuicao_mensal_{ano}_{mes:02d}.csv"'
            writer = csv.writer(response, delimiter=';')
            writer.writerow(['Ano', 'Mes', 'Contrato', 'Conjunto', 'Contribuicao FCVS', 'Parcelas'])
            for item in linhas_mensal:
                writer.writerow([
                    ano,
                    meses_nome[mes],
                    item['contrato__codigo'],
                    item['contrato__conjunto'],
                    f"{Decimal(str(item['total_fcvs'])):.2f}",
                    item['total_parcelas'],
                ])
            writer.writerow(['', '', 'TOTAL MENSAL', '', f"{total_mensal:.2f}", ''])
            return response

        try:
            from openpyxl import Workbook
        except ImportError:
            return HttpResponse('openpyxl nao esta instalado no ambiente.', status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = f'{ano}-{mes:02d}'
        ws.append(['Ano', 'Mes', 'Contrato', 'Conjunto', 'Contribuicao FCVS', 'Parcelas'])
        for item in linhas_mensal:
            ws.append([
                ano,
                meses_nome[mes],
                item['contrato__codigo'],
                item['contrato__conjunto'],
                float(item['total_fcvs']),
                item['total_parcelas'],
            ])
        ws.append(['', '', 'TOTAL MENSAL', '', float(total_mensal), ''])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="fcvs_contribuicao_mensal_{ano}_{mes:02d}.xlsx"'
        wb.save(response)
        return response

    context = {
        'ano': ano,
        'periodicidade': periodicidade,
        'mes': mes,
        'mes_nome_selecionado': meses_nome.get(mes, ''),
        'trimestre': trimestre,
        'modo': modo,
        'meses_nome': meses_nome,
        'anos_disponiveis': anos_disponiveis,
        'compute_mensal': compute_mensal,
        'compute_trimestral': compute_trimestral,
        'compute_anual': compute_anual,
        'need_total_mensal': need_total_mensal,
        'need_total_trimestral': need_total_trimestral,
        'need_total_anual': need_total_anual,
        'mensal_por_contrato': mensal_por_contrato,
        'total_mensal': total_mensal,
        'trimestral_por_contrato': trimestral_por_contrato,
        'trimestral_por_mes': trimestral_por_mes,
        'total_trimestral': total_trimestral,
        'anual_por_mes': anual_por_mes,
        'total_anual': total_anual,
        'resumo_meses': resumo_meses,
        'resumo_trimestres': resumo_trimestres,
        'contratos_trimestrais_ano': contratos_trimestrais_ano,
        'total_fcvs_contratos_trimestrais_ano': total_fcvs_contratos_trimestrais_ano,
        'total_anual_resumo': total_anual_resumo,
        'periodicidade_stats': periodicidade_stats,
        'compute_grupos': compute_grupos,
        'contratos_mensal_puro': contratos_mensal_puro,
        'contratos_trimestral_puro': contratos_trimestral_puro,
        'total_grupo_mensal_mes': total_grupo_mensal_mes,
        'total_grupo_mensal_historico': total_grupo_mensal_historico,
        'total_grupo_trimestral_trimestre': total_grupo_trimestral_trimestre,
        'total_grupo_trimestral_historico': total_grupo_trimestral_historico,
        'contratos_diferente_3m': contratos_diferente_3m,
    }
    return render(request, 'principal/fcvs_contribuicao.html', context)


def relatorio_caixa(request):
    """Relatório para Caixa Econômica Federal"""
    # Filtros
    conjunto_filtro = request.GET.get('conjunto', '')
    ano_filtro = request.GET.get('ano', '')
    active_tab = request.GET.get('aba', 'geral').strip().lower()
    if active_tab not in ('geral', 'seguro'):
        active_tab = 'geral'
    
    # Aplicar filtro de conjunto
    contratos_query = Contrato.objects.all()
    if conjunto_filtro:
        contratos_query = contratos_query.filter(conjunto=conjunto_filtro)
    
    # Estatísticas gerais (considerando filtro)
    total_contratos = contratos_query.count()
    
    # Mutuários e conjuntos totais (sem filtro)
    total_mutuarios = Mutuario.objects.count()
    total_conjuntos = ConjuntoHabitacional.objects.count()
    
    # Contratos por conjunto
    contratos_por_conjunto = []
    conjuntos_a_listar = ConjuntoHabitacional.objects.all().order_by('conjunto')
    if conjunto_filtro:
        conjuntos_a_listar = conjuntos_a_listar.filter(conjunto=conjunto_filtro)
    
    for conjunto in conjuntos_a_listar:
        qtd = Contrato.objects.filter(conjunto=conjunto.conjunto).count()
        if qtd > 0:
            contratos_por_conjunto.append({
                'conjunto': conjunto,
                'quantidade': qtd
            })
    
    # Movimentações recentes (filtradas por conjunto se aplicável)
    movimentacoes_query = Movimentacao.objects.all()
    if conjunto_filtro:
        # Filtrar movimentações pelo conjunto
        movimentacoes_query = movimentacoes_query.filter(conjunto=conjunto_filtro)
    movimentacoes = movimentacoes_query.order_by('-data')[:100]
    
    # Contratos ativos (com parcelas) - aplicar filtro
    contratos_ativos = []
    contratos_seguro = []
    total_seguro_carteira = Decimal('0')
    total_seguro_pago = Decimal('0')
    total_parcelas_com_seguro = 0
    total_parcelas_sem_seguro = 0
    limite_contratos = 100 if conjunto_filtro else 50  # Mais contratos quando filtrado
    for contrato in contratos_query[:limite_contratos]:
        parcelas = ParcelaContrato.objects.filter(contrato=contrato)
        if parcelas.exists():
            parcelas_pagas = parcelas.filter(dtpgto__isnull=False).count()
            parcelas_total = parcelas.count()
            ultima_parcela = parcelas.order_by('-nmens').first()
            primeira_parcela = parcelas.order_by('nmens').first()

            seguro_total_contrato = Decimal('0')
            seguro_pago_contrato = Decimal('0')
            parcelas_com_seguro_contrato = 0
            parcelas_sem_seguro_contrato = 0
            for p in parcelas.only('seguro', 'dtpgto'):
                valor_seguro = Decimal(str(p.seguro)) if p.seguro else Decimal('0')
                if valor_seguro > 0:
                    parcelas_com_seguro_contrato += 1
                    seguro_total_contrato += valor_seguro
                    if p.dtpgto:
                        seguro_pago_contrato += valor_seguro
                else:
                    parcelas_sem_seguro_contrato += 1

            seguro_aberto_contrato = seguro_total_contrato - seguro_pago_contrato
            total_seguro_carteira += seguro_total_contrato
            total_seguro_pago += seguro_pago_contrato
            total_parcelas_com_seguro += parcelas_com_seguro_contrato
            total_parcelas_sem_seguro += parcelas_sem_seguro_contrato
            
            # Buscar mutuário
            mutuario_obj = None
            try:
                db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
                conn = sqlite3.connect(db_path)
                cur = conn.cursor()
                cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
                result = cur.fetchone()
                if result:
                    mutuario_obj = Mutuario.objects.get(id=result[0])
                conn.close()
            except Exception as e:
                pass
            
            contratos_ativos.append({
                'contrato': contrato,
                'conjunto': contrato.conjunto,
                'codigo': contrato.codigo,
                'mutuario': mutuario_obj,  # Passar o objeto completo
                'data_contrato': primeira_parcela.dtvenc if primeira_parcela and primeira_parcela.dtvenc else None,
                'prazo': parcelas_total,  # Nome esperado pelo template
                'saldo_atual': float(ultima_parcela.sddev) if ultima_parcela and ultima_parcela.sddev else 0.0,  # Nome esperado pelo template
                'parcelas_pagas': parcelas_pagas,
                'parcelas_total': parcelas_total,
                'percentual': (parcelas_pagas / parcelas_total * 100) if parcelas_total > 0 else 0,
                'saldo_devedor': float(ultima_parcela.sddev) if ultima_parcela and ultima_parcela.sddev else 0.0,
                'seguro_total': float(seguro_total_contrato),
                'seguro_pago': float(seguro_pago_contrato),
                'seguro_aberto': float(seguro_aberto_contrato),
                'parcelas_com_seguro': parcelas_com_seguro_contrato,
                'parcelas_sem_seguro': parcelas_sem_seguro_contrato,
            })

            if parcelas_com_seguro_contrato > 0:
                contratos_seguro.append({
                    'contrato': contrato,
                    'conjunto': contrato.conjunto,
                    'codigo': contrato.codigo,
                    'mutuario': mutuario_obj,
                    'parcelas_com_seguro': parcelas_com_seguro_contrato,
                    'parcelas_sem_seguro': parcelas_sem_seguro_contrato,
                    'seguro_total': float(seguro_total_contrato),
                    'seguro_pago': float(seguro_pago_contrato),
                    'seguro_aberto': float(seguro_aberto_contrato),
                })
    
    # Saldo total devedor
    saldo_total_devedor = sum([c['saldo_devedor'] for c in contratos_ativos])
    total_seguro_aberto = total_seguro_carteira - total_seguro_pago
    contratos_com_seguro = len(contratos_seguro)
    contratos_sem_seguro = max(0, len(contratos_ativos) - contratos_com_seguro)

    contratos_seguro.sort(key=lambda item: item['seguro_total'], reverse=True)
    
    # Conjuntos para filtro
    conjuntos = ConjuntoHabitacional.objects.all().order_by('conjunto')
    
    # Contar conjuntos filtrados
    if conjunto_filtro:
        total_conjuntos_filtrado = 1
    else:
        total_conjuntos_filtrado = total_conjuntos
    
    context = {
        'total_contratos': total_contratos,
        'total_mutuarios': total_mutuarios,
        'total_conjuntos': total_conjuntos_filtrado,
        'contratos_por_conjunto': contratos_por_conjunto,
        'movimentacoes': movimentacoes,
        'contratos': contratos_ativos,  # Nome esperado pelo template
        'contratos_ativos': contratos_ativos,  # Manter para compatibilidade
        'saldo_total_devedor': saldo_total_devedor,
        'total_seguro_carteira': float(total_seguro_carteira),
        'total_seguro_pago': float(total_seguro_pago),
        'total_seguro_aberto': float(total_seguro_aberto),
        'total_parcelas_com_seguro': total_parcelas_com_seguro,
        'total_parcelas_sem_seguro': total_parcelas_sem_seguro,
        'contratos_com_seguro': contratos_com_seguro,
        'contratos_sem_seguro': contratos_sem_seguro,
        'contratos_seguro': contratos_seguro,
        'conjuntos': conjuntos,
        'conjunto_filtro': conjunto_filtro,
        'ano_filtro': ano_filtro,
        'active_tab': active_tab,
        'data_relatorio': date.today()
    }
    return render(request, 'principal/relatorio_caixa.html', context)


def calcular_fcvs_residual_global(contrato_id):
    """
    Calcula o FCVS residual de um contrato.

    Modo 1 — Parcelas reais (MOVMUT): quando o banco tem ParcelaContrato importadas,
             lê saldo devedor real e detecta resíduos FCVS.
    Modo 2 — Simulação BNH (fallback): quando não há parcelas reais, simula a evolução
             mês a mês a partir dos parâmetros do contrato (vlfinanc, SA, tx_juros, prazo)
             usando correção monetária histórica (ORTN/OTN/BTNF/TR) e PES anual.
    """
    from decimal import ROUND_HALF_UP
    from .simulador_sfh import simular_evolucao_sfh, carregar_indices_sfh, carregar_indices_pes

    # Obter contrato e parcelas reais
    contrato = Contrato.objects.get(id=contrato_id)
    parcelas_reais = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')
    dict_parcelas = {p.nmens: p for p in parcelas_reais}

    prazo_meses = contrato.prazo if contrato.prazo else 360
    # Para contratos OCR sem parcelas reais, a simulação deve iniciar no 1º vencimento
    # (quando informado) para refletir a base de cálculo da evolução financeira.
    data_inicio = contrato.data_primeiro_venc or contrato.data_contrato or date(1984, 10, 30)

    # ---------------------------------------------------------------
    # MODO 2: Simulação quando não há parcelas reais no banco
    # ---------------------------------------------------------------
    if not dict_parcelas:
        # Busca vlfinanc e prestação inicial no próprio Contrato (novos campos)
        vlfinanc = Decimal(str(contrato.vlfinanc or '0'))
        prestacao_inicial = Decimal(str(contrato.prestacao_inicial or '0')) or None

        # Fallback: parcela 1 se existir (caso partial import)
        if vlfinanc <= 0:
            p1 = ParcelaContrato.objects.filter(contrato=contrato, nmens=1).first()
            if p1 and p1.sddev:
                vlfinanc = Decimal(str(p1.sddev))

        tx_juros = float(contrato.tx_juros or 10)
        sa = contrato.sa or 'SAC'

        # Índices CM: usa INDICES_HISTORICOS já carregado
        indices_cm = {k: Decimal(str(v)) for k, v in INDICES_HISTORICOS.items()}

        csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'indices_historicos.csv')
        indices_full = carregar_indices_sfh(csv_path)
        indices_full.update(indices_cm)

        csv_pes = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'indices_pes.csv')
        indices_pes = carregar_indices_pes(csv_pes) if os.path.exists(csv_pes) else None

        evolucao_sim, fcvs_final = simular_evolucao_sfh(
            vlfinanc=vlfinanc if vlfinanc > 0 else Decimal('1'),
            sa=sa,
            tx_juros_aa=Decimal(str(tx_juros)),
            prazo=prazo_meses,
            data_contrato=data_inicio,
            prestacao_inicial=prestacao_inicial,
            indices_cm=indices_full,
            indices_pes=indices_pes,
        )

        # Converte formato para compatibilidade com o template existente
        evolucao_fmt = []
        for e in evolucao_sim:
            evolucao_fmt.append({
                'mes':          e['mes'],
                'data':         e['data'],
                'saldo_ant':    e['saldo_ant'],
                'correcao':     e['cm_valor'],
                'saldo_novo':   e['saldo_novo'],
                'indice':       f"{e['cm_pct']:.2f}%",
                'anomalia':     'SIMULADO',
                'fcvs_excedente': e['fcvs_acum'],
                'moeda':        e['moeda'],
                # campos extras da simulação
                'amort':        e.get('amort', 0),
                'juros':        e.get('juros', 0),
                'encargo':      e.get('encargo', 0),
                'prest_pes':    e.get('prest_pes', 0),
                'fcvs_mes':     e.get('fcvs_mes', 0),
            })

        anomalias_sim = sum(1 for e in evolucao_sim if e['fcvs_mes'] > 0)
        return evolucao_fmt, anomalias_sim, float(fcvs_final)

    # ---------------------------------------------------------------
    # MODO 1: Parcelas reais (comportamento original)
    # ---------------------------------------------------------------
    evolucao = []
    fcvs_acum = Decimal('0')
    anomalias = 0

    data_simulacao_atual = data_inicio
    saldo_simulado = Decimal('0')
    current_moeda = get_moeda_vigente(data_simulacao_atual)

    for mes_num in range(1, prazo_meses + 1):
        saldo_ant = saldo_simulado

        # Correção do FCVS: Redenominação do Acumulado
        moeda_mes_atual = get_moeda_vigente(data_simulacao_atual)
        if mes_num > 1 and moeda_mes_atual != current_moeda:
            for data_limite, moeda_anterior, fator, moeda_nova in NOMINAL_CONVERSION_FACTORS:
                if data_simulacao_atual >= data_limite and current_moeda != moeda_nova:
                    fcvs_acum = (fcvs_acum / fator).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    break
            current_moeda = moeda_mes_atual

        if mes_num in dict_parcelas:
            p_real = dict_parcelas[mes_num]
            saldo_simulado = p_real.sddev if p_real.sddev is not None else Decimal('0')
            data_simulacao_atual = p_real.dtvenc or data_simulacao_atual

            chave_mes = data_simulacao_atual.strftime('%Y-%m')
            indice = INDICES_HISTORICOS.get(chave_mes, Decimal('0'))

            anomalia_detectada = ''
            if saldo_simulado > saldo_ant and saldo_ant > 0:
                excedente = saldo_simulado - saldo_ant
                fcvs_acum += excedente
                anomalias += 1
                anomalia_detectada = 'RESIDUO FCVS'

            evolucao.append({
                'mes':          mes_num,
                'data':         data_simulacao_atual.strftime('%Y-%m'),
                'saldo_ant':    float(saldo_ant),
                'correcao':     float(p_real.cm or 0),
                'saldo_novo':   float(saldo_simulado),
                'indice':       f"{float(indice * 100):.2f}%",
                'anomalia':     anomalia_detectada or 'OK',
                'fcvs_excedente': float(fcvs_acum),
                'moeda':        current_moeda,
            })
        else:
            evolucao.append({
                'mes': mes_num, 'data': data_simulacao_atual.strftime('%Y-%m'),
                'saldo_ant': float(saldo_simulado), 'correcao': 0.0,
                'saldo_novo': float(saldo_simulado), 'indice': "0.00%",
                'anomalia': 'FIM CONTRATO',
                'fcvs_excedente': float(fcvs_acum), 'moeda': current_moeda,
            })

        data_simulacao_atual = (data_simulacao_atual.replace(day=1) + timedelta(days=32)).replace(day=1)

    return evolucao[:12] + evolucao[-12:], anomalias, float(fcvs_acum)


def relatorio_fh1(request, pk):
    """Gera relatório FH1 para habilitação ao FCVS"""
    contrato = get_object_or_404(Contrato, pk=pk)
    parcelas = ParcelaContrato.objects.filter(contrato=contrato).order_by('nmens')
    primeira = parcelas.first()
    ultima = parcelas.last()
    
    # Buscar conjunto
    conjunto = ConjuntoHabitacional.objects.filter(conjunto=contrato.conjunto).first()
    
    # Calcular FCVS residual usando função correta com conversões monetárias
    evolucao, anomalias_count, fcvs_residual = calcular_fcvs_residual_global(contrato.id)
    anomalias = anomalias_count
    
    # Buscar Mutuário
    mutuario = None
    try:
        mutuario = Mutuario.objects.filter(codigo=contrato.codigo).first()
        if not mutuario:
            db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
            result = cur.fetchone()
            if result:
                mutuario = Mutuario.objects.get(id=result[0])
            conn.close()
    except Exception as e:
        print(f"Erro ao buscar mutuário: {e}")
    
    # Saldo atual
    saldo_atual = ultima.sddev if ultima and ultima.sddev else Decimal('0')

    context = {
        'contrato': contrato,
        'mutuario': mutuario,
        'conjunto': conjunto,
        'fcvs_residual': fcvs_residual,
        'saldo_atual': saldo_atual,
        'total_parcelas': parcelas.count(),
        'anomalias': anomalias,
        'data_emissao': date.today(),
        'primeira_parcela': primeira.dtvenc if primeira else None,
        'ultima_parcela': ultima.dtvenc if ultima else None,
        'agente_financeiro': "COFLUHAB - CIA FLUMINENSE DE HABITAÇÃO",
        'codigo_agente': "33.000.000/0001-00",
    }
    
    return render(request, 'principal/relatorio_fh1.html', context)


def fh1_completo(request, pk):
    """Gera relatório FH1 completo corrigido para COHAB"""
    contrato = get_object_or_404(Contrato, pk=pk)
    
    # 1. Chamar a simulação oficial (agora global)
    evolucao_completa, total_anomalias, fcvs_residual_total = calcular_fcvs_residual_global(contrato.id)
    
    # 2. Filtrar e ordenar o TOP 10 de anomalias (usando valores já convertidos)
    anomalias_reais = [
        e for e in evolucao_completa 
        if e['anomalia'] == 'RESÍDUO FCVS'
    ]
    # Ordena pelo valor do crescimento (saldo_novo - saldo_ant)
    anomalias_reais.sort(key=lambda x: x['saldo_novo'] - x['saldo_ant'], reverse=True)
    top_10 = anomalias_reais[:10]
    
    # 3. Buscar Mutuário e Conjunto
    conjunto = ConjuntoHabitacional.objects.filter(conjunto=contrato.conjunto).first()
    mutuario = None
    try:
        mutuario = Mutuario.objects.filter(codigo=contrato.codigo).first()
        if not mutuario:
            db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT mutuario_id FROM contrato_mutuario_map WHERE contrato_id = ?", (contrato.id,))
            result = cur.fetchone()
            if result:
                mutuario = Mutuario.objects.get(id=result[0])
            conn.close()
    except Exception as e:
        print(f"Erro ao buscar mutuário: {e}")
    
    # Buscar endereço
    endereco = None
    try:
        from principal.models import EnderecoContrato
        endereco = EnderecoContrato.objects.filter(contrato=contrato).first()
    except:
        pass
    
    # 4. Dados finais
    primeira = evolucao_completa[0] if evolucao_completa else None
    ultima = evolucao_completa[-1] if evolucao_completa else None

    context = {
        'contrato': contrato,
        'mutuario': mutuario,
        'conjunto': conjunto,
        'endereco': endereco,
        'fcvs_residual': fcvs_residual_total,
        'saldo_atual': ultima['saldo_novo'] if ultima else 0,
        'saldo_inicial': primeira['saldo_ant'] if primeira else 0,
        'total_parcelas': len(evolucao_completa),
        'anomalias': total_anomalias,
        'anomalias_total': total_anomalias,
        'anomalias_top10': top_10,
        'primeira_parcela': primeira['data'] if primeira else None,
        'ultima_parcela': ultima['data'] if ultima else None,
        'periodo_inicio': primeira['data'] if primeira else '',
        'periodo_fim': ultima['data'] if ultima else '',
        'data_emissao': date.today(),
    }
    
    return render(request, 'principal/fh1_completo.html', context)


def carteira_fcvs(request):
    """Dashboard consolidado de FCVS para toda a carteira"""
    conjunto_filtro = request.GET.get('conjunto', '').strip()
    contrato_filtro = request.GET.get('contrato', '').strip()

    cache_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fcvs_dashboard_cache.json')

    def _load_cache():
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return None

    cache = _load_cache()

    # Sem cache: calcula ao vivo SOMENTE o recorte filtrado (evita tela zerada)
    if not cache:
        contratos_query = Contrato.objects.all()
        if conjunto_filtro:
            contratos_query = contratos_query.filter(conjunto__icontains=conjunto_filtro)
        if contrato_filtro:
            contratos_query = contratos_query.filter(codigo__icontains=contrato_filtro)

        # Sem filtro, evita cálculo global pesado sem cache
        if not conjunto_filtro and not contrato_filtro:
            context = {
                'carteira_fcvs': [],
                'resumo_por_conjunto': [],
                'total_contratos_analisados': 0,
                'contratos_com_fcvs': 0,
                'total_fcvs_acumulado': 0,
                'total_fcvs_filtro': 0,
                'total_fcvs_geral': 0,
                'contratos_com_fcvs_filtro': 0,
                'contratos_com_fcvs_geral': 0,
                'media_fcvs_por_contrato': 0,
                'data_analise': date.today(),
                'conjunto_filtro': conjunto_filtro,
                'contrato_filtro': contrato_filtro,
                'conjuntos': ConjuntoHabitacional.objects.all().values_list('conjunto', 'nome').distinct(),
                'cache_ausente': True,
            }
            return render(request, 'principal/carteira_fcvs.html', context)

        carteira_fcvs = []
        resumo_por_conjunto = {}
        total_fcvs_filtro = 0.0
        contratos_com_fcvs_filtro = 0

        for contrato in contratos_query.only('id', 'codigo', 'conjunto').iterator():
            try:
                evolucao, anomalias, fcvs_residual = calcular_fcvs_residual_global(contrato.id)
                fcvs_residual = float(fcvs_residual)
                saldo_atual_calc = evolucao[-1]['saldo_novo'] if evolucao else 0
            except Exception:
                continue

            cj = contrato.conjunto
            if cj not in resumo_por_conjunto:
                resumo_por_conjunto[cj] = {
                    'conjunto': cj,
                    'total_fcvs': 0.0,
                    'contratos_com_fcvs': 0,
                    'total_analisados': 0,
                    'media_fcvs': 0.0,
                }
            resumo_por_conjunto[cj]['total_analisados'] += 1

            if fcvs_residual > 100:
                total_fcvs_filtro += fcvs_residual
                contratos_com_fcvs_filtro += 1
                resumo_por_conjunto[cj]['total_fcvs'] += fcvs_residual
                resumo_por_conjunto[cj]['contratos_com_fcvs'] += 1

                carteira_fcvs.append({
                    'contrato': contrato,
                    'fcvs_residual': fcvs_residual,
                    'anomalias': anomalias,
                    'saldo_atual': saldo_atual_calc,
                    'prioridade': 'ALTA' if fcvs_residual > 10000 else 'MÉDIA'
                })

        carteira_fcvs.sort(key=lambda x: x['fcvs_residual'], reverse=True)
        carteira_fcvs = carteira_fcvs[:300]

        resumo_conjuntos_list = []
        for _, stats in resumo_por_conjunto.items():
            stats['media_fcvs'] = stats['total_fcvs'] / stats['contratos_com_fcvs'] if stats['contratos_com_fcvs'] > 0 else 0
            resumo_conjuntos_list.append(stats)
        resumo_conjuntos_list.sort(key=lambda x: x['total_fcvs'], reverse=True)

        context = {
            'carteira_fcvs': carteira_fcvs,
            'resumo_por_conjunto': resumo_conjuntos_list,
            'total_contratos_analisados': contratos_query.count(),
            'contratos_com_fcvs': contratos_com_fcvs_filtro,
            'total_fcvs_acumulado': total_fcvs_filtro,
            'total_fcvs_filtro': total_fcvs_filtro,
            'total_fcvs_geral': 0,
            'contratos_com_fcvs_filtro': contratos_com_fcvs_filtro,
            'contratos_com_fcvs_geral': 0,
            'media_fcvs_por_contrato': total_fcvs_filtro / contratos_com_fcvs_filtro if contratos_com_fcvs_filtro > 0 else 0,
            'data_analise': date.today(),
            'conjunto_filtro': conjunto_filtro,
            'contrato_filtro': contrato_filtro,
            'conjuntos': ConjuntoHabitacional.objects.all().values_list('conjunto', 'nome').distinct(),
            'cache_ausente': True,
        }
        return render(request, 'principal/carteira_fcvs.html', context)

    contratos_cache = cache.get('contratos', [])

    # aplica filtro sobre os dados já calculados
    filtrados = []
    for item in contratos_cache:
        if conjunto_filtro and conjunto_filtro.lower() not in str(item.get('conjunto', '')).lower():
            continue
        if contrato_filtro and contrato_filtro.lower() not in str(item.get('codigo', '')).lower():
            continue
        filtrados.append(item)

    total_contratos_analisados = len(filtrados)
    contratos_com_fcvs_filtro = sum(1 for x in filtrados if float(x.get('fcvs_residual', 0)) > 100)
    total_fcvs_filtro = sum(float(x.get('fcvs_residual', 0)) for x in filtrados if float(x.get('fcvs_residual', 0)) > 100)

    # resumo por conjunto do recorte filtrado
    resumo_por_conjunto = {}
    for item in filtrados:
        cj = item.get('conjunto', '')
        if cj not in resumo_por_conjunto:
            resumo_por_conjunto[cj] = {
                'conjunto': cj,
                'total_fcvs': 0.0,
                'contratos_com_fcvs': 0,
                'total_analisados': 0,
                'media_fcvs': 0.0,
            }
        resumo_por_conjunto[cj]['total_analisados'] += 1
        if float(item.get('fcvs_residual', 0)) > 100:
            resumo_por_conjunto[cj]['total_fcvs'] += float(item.get('fcvs_residual', 0))
            resumo_por_conjunto[cj]['contratos_com_fcvs'] += 1

    resumo_conjuntos_list = []
    for _, stats in resumo_por_conjunto.items():
        stats['media_fcvs'] = stats['total_fcvs'] / stats['contratos_com_fcvs'] if stats['contratos_com_fcvs'] > 0 else 0
        resumo_conjuntos_list.append(stats)
    resumo_conjuntos_list.sort(key=lambda x: x['total_fcvs'], reverse=True)

    # tabela (top do recorte)
    filtrados_fcvs = [x for x in filtrados if float(x.get('fcvs_residual', 0)) > 100]
    filtrados_fcvs.sort(key=lambda x: float(x.get('fcvs_residual', 0)), reverse=True)
    top = filtrados_fcvs[:300]

    mapa_contratos = {c.id: c for c in Contrato.objects.filter(id__in=[int(x.get('id')) for x in top if x.get('id')]).only('id', 'codigo', 'conjunto')}

    carteira_fcvs = []
    for item in top:
        contrato_obj = mapa_contratos.get(int(item.get('id')))
        if not contrato_obj:
            continue
        carteira_fcvs.append({
            'contrato': contrato_obj,
            'fcvs_residual': float(item.get('fcvs_residual', 0)),
            'anomalias': int(item.get('anomalias', 0)),
            'saldo_atual': float(item.get('saldo_atual', 0)),
            'prioridade': 'ALTA' if float(item.get('fcvs_residual', 0)) > 10000 else 'MÉDIA',
        })

    conjuntos = ConjuntoHabitacional.objects.all().values_list('conjunto', 'nome').distinct()

    total_fcvs_geral = float(cache.get('total_fcvs_geral', 0))
    contratos_com_fcvs_geral = int(cache.get('contratos_com_fcvs_geral', 0))

    context = {
        'carteira_fcvs': carteira_fcvs,
        'resumo_por_conjunto': resumo_conjuntos_list,
        'total_contratos_analisados': total_contratos_analisados,
        'contratos_com_fcvs': contratos_com_fcvs_filtro,
        'total_fcvs_acumulado': total_fcvs_filtro,
        'total_fcvs_filtro': total_fcvs_filtro,
        'total_fcvs_geral': total_fcvs_geral,
        'contratos_com_fcvs_filtro': contratos_com_fcvs_filtro,
        'contratos_com_fcvs_geral': contratos_com_fcvs_geral,
        'media_fcvs_por_contrato': total_fcvs_filtro / contratos_com_fcvs_filtro if contratos_com_fcvs_filtro > 0 else 0,
        'data_analise': date.today(),
        'conjunto_filtro': conjunto_filtro,
        'contrato_filtro': contrato_filtro,
        'conjuntos': conjuntos,
        'cache_ausente': False,
    }

    return render(request, 'principal/carteira_fcvs.html', context)
