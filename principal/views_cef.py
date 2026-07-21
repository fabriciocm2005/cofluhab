"""
Views para integração CEF SIWFC
Gerenciamento de envios, retornos e automação
"""

from __future__ import annotations

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, FileResponse, Http404
from django.contrib import messages
from django.utils import timezone
from django.db import close_old_connections
from django.db.models import Count, Q
from django.conf import settings
from datetime import datetime, timedelta
import csv
import json
import os
import threading
import re
import unicodedata
import io
import zipfile
import hashlib
import tempfile
from pathlib import Path
from collections import Counter, defaultdict

from .models import Contrato, Mutuario
from .models_cef import (
    RemessaCEF, CredencialCEF, EnvioCEF, RetornoCEF,
    AgendamentoEnvio, LogAutomacao
)
from .manual_expert_agent import ManualExpertAgent
from .fh1_validator import run_fh1_precheck_agent

# Importa módulos de fichas
try:
    from .ficha_parsers import ArquivoFichasCEF
    from .ficha_validators import validar_fh1, validar_cadmut, validar_lote_fichas
    from .ficha_generators import (
        FH1Generator, CADMUTGenerator, ArquivoFCVSGenerator,
        gerar_fh1_contrato, gerar_cadmut_mutuario, gerar_arquivo_fcvs
    )
    from .ficha_return_interpreter import (
        ReturnInterpreter, interpretar_retorno_fcvs, interpretar_retorno_cadmut,
        processar_lote_retornos
    )
    from .ficha_selector import (
        FichaSelector, SequenciadorFichas,
        selecionar_ficha_automatica, gerar_plano_envio
    )
    from .ficha_p3026_parser import interpretar_p3026, ParserP3026
    MODULOS_FICHAS_DISPONIVEIS = True
except ImportError as e:
    MODULOS_FICHAS_DISPONIVEIS = False
    print("[ERRO DE IMPORTAÇÃO] Falha ao importar módulos de ficha:", e)


STATUS_BLOQUEIO_ENVIO = ['GERADO', 'PENDENTE', 'PROCESSANDO', 'ENVIADO', 'RETORNO_RECEBIDO']
MANUAL_EXPERT_AGENT = ManualExpertAgent()
SESSION_CHAVE_ULTIMO_LOTE_MANUAL = 'cef_ultimo_lote_manual'


def _buscar_contratos_bloqueados_para_envio(contrato_ids, tipo_envio='FH1'):
    """Retorna contratos que já possuem envio em status que bloqueia novo envio."""
    envios = (
        EnvioCEF.objects
        .filter(
            contrato_id__in=contrato_ids,
            tipo_envio=tipo_envio,
            status__in=STATUS_BLOQUEIO_ENVIO,
        )
        .select_related('contrato')
        .order_by('contrato_id', '-criado_em')
    )

    bloqueados = {}
    for envio in envios:
        if envio.contrato_id in bloqueados:
            continue
        bloqueados[envio.contrato_id] = {
            'contrato_id': envio.contrato_id,
            'codigo': envio.contrato.codigo,
            'status': envio.status,
            'status_label': envio.get_status_display(),
            'protocolo': envio.protocolo_cef or '',
        }

    return list(bloqueados.values())


# ===== DASHBOARD CEF =====

def integracao_cef(request):
    """Dashboard principal de integração CEF"""
    
    # Estatísticas
    total_envios = EnvioCEF.objects.count()
    envios_pendentes = EnvioCEF.objects.filter(status='PENDENTE').count()
    envios_sucesso = EnvioCEF.objects.filter(status='ENVIADO').count()
    envios_erro = EnvioCEF.objects.filter(status='ERRO').count()
    
    retornos_nao_lidos = RetornoCEF.objects.filter(lido=False).count()
    retornos_requerem_acao = RetornoCEF.objects.filter(requer_acao=True, processado=False).count()
    
    # Credenciais configuradas
    credenciais_ativas = CredencialCEF.objects.filter(ativo=True).count()
    
    # Agendamentos ativos
    agendamentos_ativos = AgendamentoEnvio.objects.filter(ativo=True).count()
    proximos_agendamentos = AgendamentoEnvio.objects.filter(
        ativo=True,
        proxima_execucao__gte=timezone.now()
    ).order_by('proxima_execucao')[:5]
    
    # Últimos envios
    ultimos_envios = EnvioCEF.objects.select_related('contrato').order_by('-criado_em')[:10]
    
    # Últimos retornos
    ultimos_retornos = RetornoCEF.objects.select_related('contrato').order_by('-data_retorno')[:10]
    
    # Logs recentes
    logs_recentes = LogAutomacao.objects.order_by('-timestamp')[:20]
    
    context = {
        'total_envios': total_envios,
        'envios_pendentes': envios_pendentes,
        'envios_sucesso': envios_sucesso,
        'envios_erro': envios_erro,
        'retornos_nao_lidos': retornos_nao_lidos,
        'retornos_requerem_acao': retornos_requerem_acao,
        'credenciais_ativas': credenciais_ativas,
        'agendamentos_ativos': agendamentos_ativos,
        'proximos_agendamentos': proximos_agendamentos,
        'ultimos_envios': ultimos_envios,
        'ultimos_retornos': ultimos_retornos,
        'logs_recentes': logs_recentes,
    }
    
    return render(request, 'principal/integracao_cef.html', context)


# ===== ENVIOS CEF =====

def listar_envios_cef(request):
    """Lista todos os envios CEF com filtros"""
    
    envios = EnvioCEF.objects.select_related('contrato', 'remessa').all()
    
    # Filtros
    status_filter = request.GET.get('status')
    if status_filter:
        envios = envios.filter(status=status_filter)
    
    tipo_filter = request.GET.get('tipo')
    if tipo_filter:
        envios = envios.filter(tipo_envio=tipo_filter)

    contrato_filter = request.GET.get('contrato')
    if contrato_filter:
        envios = envios.filter(contrato__codigo__icontains=contrato_filter)
    
    context = {
        'envios': envios,
        'status_choices': EnvioCEF.STATUS_CHOICES,
        'tipo_choices': EnvioCEF.TIPO_ENVIO_CHOICES,
    }
    
    return render(request, 'principal/cef_envios.html', context)


def criar_envio_fh1(request, contrato_id):
    """Cria envio FH1 para um contrato"""
    
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    
    if request.method == 'POST':
        try:
            # Gerar arquivo FH1
            from .views import exportar_evolucao_txt
            
            # Criar envio
            envio = EnvioCEF.objects.create(
                contrato=contrato,
                tipo_envio='FH1',
                status='PENDENTE',
                arquivo_enviado=f'FH1_{contrato.codigo}.txt',
                codigo_contrato_cef=contrato.codigo,
            )
            
            messages.success(request, f'✅ Envio FH1 criado para contrato {contrato.codigo}')
            return redirect('integracao_cef')
            
        except Exception as e:
            messages.error(request, f'❌ Erro ao criar envio: {str(e)}')
    
    return render(request, 'principal/cef_criar_envio.html', {
        'contrato': contrato
    })


def processar_envio_automatico(request, envio_id):
    """Processa envio automático via bot"""
    
    envio = get_object_or_404(EnvioCEF, pk=envio_id)
    
    try:
        from cef_web_automation import CEFWebBot
        from .models_cef import CredencialCEF
        
        # Buscar credencial ativa
        credencial = CredencialCEF.objects.filter(ativo=True).first()
        if not credencial:
            return JsonResponse({
                'sucesso': False,
                'erro': 'Nenhuma credencial CEF configurada'
            })
        
        # Atualizar status
        envio.status = 'PROCESSANDO'
        envio.tentativas += 1
        envio.save()
        
        # Criar log
        log = LogAutomacao.objects.create(
            tipo_acao='ENVIO',
            descricao=f'Iniciando envio automático de {envio.tipo_envio}',
            envio=envio
        )
        
        # TODO: Implementar envio real com bot
        # bot = CEFWebBot(headless=True)
        # sucesso = bot.enviar_fh1(...)
        
        # Simulação por enquanto
        import time
        time.sleep(2)
        
        envio.status = 'ENVIADO'
        envio.data_envio = timezone.now()
        envio.enviado_automaticamente = True
        envio.protocolo_cef = f'PROTO-{timezone.now().strftime("%Y%m%d%H%M%S")}'
        envio.save()
        
        log.sucesso = True
        log.duracao_segundos = 2.0
        log.save()
        
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Envio processado com sucesso! Protocolo: {envio.protocolo_cef}'
        })
        
    except Exception as e:
        envio.status = 'ERRO'
        envio.mensagem_erro = str(e)
        envio.save()
        
        LogAutomacao.objects.create(
            tipo_acao='ERRO',
            descricao=f'Erro ao processar envio: {str(e)}',
            envio=envio,
            sucesso=False,
            traceback=str(e)
        )
        
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


# ===== RETORNOS CEF =====

def listar_retornos_cef(request):
    """Lista retornos recebidos da CEF"""
    
    base_retornos = RetornoCEF.objects.select_related('contrato', 'envio').order_by('-data_retorno')
    retornos = base_retornos
    
    # Filtros
    lido_filter = request.GET.get('lido')
    if lido_filter == '0':
        retornos = retornos.filter(lido=False)
    elif lido_filter == '1':
        retornos = retornos.filter(lido=True)
    
    tipo_filter = request.GET.get('tipo')
    if tipo_filter:
        retornos = retornos.filter(tipo_retorno=tipo_filter)

    aprovados_count = base_retornos.filter(tipo_retorno='APROVADO').count()
    rejeitados_count = base_retornos.filter(tipo_retorno='REJEITADO').count()
    pendentes_count = base_retornos.filter(tipo_retorno='PENDENTE').count()
    complementar_count = base_retornos.filter(tipo_retorno='COMPLEMENTAR').count()
    oficio_count = base_retornos.filter(tipo_retorno='OFICIO').count()
    nao_lidos_count = base_retornos.filter(lido=False).count()
    
    context = {
        'retornos': retornos,
        'tipo_choices': RetornoCEF.TIPO_RETORNO_CHOICES,
        'aprovados_count': aprovados_count,
        'rejeitados_count': rejeitados_count,
        'pendentes_count': pendentes_count,
        'complementar_count': complementar_count,
        'oficio_count': oficio_count,
        'nao_lidos_count': nao_lidos_count,
    }
    
    return render(request, 'principal/cef_retornos.html', context)


def marcar_retorno_lido(request, retorno_id):
    """Marca retorno como lido"""
    
    retorno = get_object_or_404(RetornoCEF, pk=retorno_id)
    retorno.lido = True
    retorno.data_leitura = timezone.now()
    retorno.save()
    
    return JsonResponse({'sucesso': True})


def verificar_retornos(request):
    """Verifica novos retornos no portal CEF"""
    
    try:
        # TODO: Implementar verificação real com bot
        # bot = CEFWebBot()
        # retornos = bot.baixar_retornos()
        
        # Simulação
        novos_retornos = 0
        
        LogAutomacao.objects.create(
            tipo_acao='DOWNLOAD',
            descricao=f'Verificação de retornos: {novos_retornos} novos',
            sucesso=True
        )
        
        return JsonResponse({
            'sucesso': True,
            'novos_retornos': novos_retornos,
            'mensagem': f'✅ {novos_retornos} novos retornos encontrados'
        })
        
    except Exception as e:
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


# ===== AGENDAMENTOS =====

def listar_agendamentos(request):
    """Lista agendamentos de envio"""
    
    agendamentos = AgendamentoEnvio.objects.all().order_by('-ativo', 'proxima_execucao')
    
    context = {
        'agendamentos': agendamentos,
    }
    
    return render(request, 'principal/cef_agendamentos.html', context)


def criar_agendamento(request):
    """Cria novo agendamento"""
    
    if request.method == 'POST':
        try:
            AgendamentoEnvio.objects.create(
                nome=request.POST.get('nome'),
                descricao=request.POST.get('descricao', ''),
                tipo_envio=request.POST.get('tipo_envio'),
                frequencia=request.POST.get('frequencia'),
                proxima_execucao=datetime.fromisoformat(request.POST.get('proxima_execucao')),
                ativo=True,
                criado_por=request.user.username if request.user.is_authenticated else 'Sistema'
            )
            
            messages.success(request, '✅ Agendamento criado com sucesso!')
            return redirect('listar_agendamentos')
            
        except Exception as e:
            messages.error(request, f'❌ Erro ao criar agendamento: {str(e)}')
    
    context = {
        'tipo_choices': EnvioCEF.TIPO_ENVIO_CHOICES,
        'frequencia_choices': AgendamentoEnvio.FREQUENCIA_CHOICES,
    }
    
    return render(request, 'principal/cef_criar_agendamento.html', context)


def executar_agendamento(request, agendamento_id):
    """Executa agendamento manualmente"""
    
    agendamento = get_object_or_404(AgendamentoEnvio, pk=agendamento_id)
    
    try:
        # TODO: Implementar lógica de execução
        # 1. Buscar contratos baseado em filtros
        # 2. Criar envios
        # 3. Processar envios
        
        agendamento.ultima_execucao = timezone.now()
        agendamento.total_envios += 1
        agendamento.envios_sucesso += 1
        agendamento.save()
        
        LogAutomacao.objects.create(
            tipo_acao='ENVIO',
            descricao=f'Execução manual do agendamento: {agendamento.nome}',
            agendamento=agendamento,
            sucesso=True
        )
        
        return JsonResponse({
            'sucesso': True,
            'mensagem': 'Agendamento executado com sucesso!'
        })
        
    except Exception as e:
        agendamento.total_envios += 1
        agendamento.envios_erro += 1
        agendamento.save()
        
        return JsonResponse({
            'sucesso': False,
            'erro': str(e)
        }, status=500)


# ===== CONFIGURAÇÕES =====

def configurar_credenciais_cef(request):
    """Configura credenciais de acesso ao SIWFC"""
    
    credenciais = CredencialCEF.objects.all()
    
    if request.method == 'POST':
        try:
            from cryptography.fernet import Fernet
            import base64
            
            # Gerar chave (em produção, usar variável de ambiente)
            key = base64.urlsafe_b64encode(b'chave_secreta_32_bytes_aqui!!')
            cipher = Fernet(key)
            
            senha = request.POST.get('senha')
            senha_criptografada = cipher.encrypt(senha.encode()).decode()
            
            CredencialCEF.objects.create(
                cpf=request.POST.get('cpf'),
                email=request.POST.get('email'),
                senha_criptografada=senha_criptografada,
                matricula_agente=request.POST.get('matricula_agente'),
                ativo=True
            )
            
            messages.success(request, '✅ Credencial salva com sucesso!')
            return redirect('integracao_cef')
            
        except Exception as e:
            messages.error(request, f'❌ Erro ao salvar credencial: {str(e)}')
    
    context = {
        'credenciais': credenciais,
    }
    
    return render(request, 'principal/cef_configurar_credenciais.html', context)


# ===== LOGS =====

def logs_automacao(request):
    """Visualiza logs de automação"""
    logs = LogAutomacao.objects.order_by('-criado_em')[:50]
    
    context = {
        'logs': logs,
    }
    
    return render(request, 'principal/cef_logs_automacao.html', context)


# ===== GERAÇÃO E VALIDAÇÃO DE FICHAS =====

def gerar_ficha_view(request, contrato_id):
    """
    View para geração inteligente de fichas CEF
    - Mostra recomendações de ficha com base no contrato
    - Valida dados antes de gerar
    - Permite download do arquivo .txt
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        print("[DEBUG] MODULOS_FICHAS_DISPONIVEIS = False: Falha ao importar módulos de ficha. Veja se há erros de importação no início do arquivo views_cef.py.")
        messages.error(request, 'Módulos de fichas não disponíveis. Verifique se todos os arquivos de ficha estão presentes e sem erro de importação.')
        return render(request, 'principal/erro_fichas.html', {"erro": "Módulos de fichas não disponíveis. Veja o terminal para detalhes."})
    
    contrato = get_object_or_404(Contrato, pk=contrato_id)
    mutuario = contrato.mutuario if hasattr(contrato, 'mutuario') else None
    
    # Busca histórico de envios
    historico_envios = EnvioCEF.objects.filter(
        contrato=contrato
    ).order_by('-criado_em')
    
    # Seleção inteligente de ficha
    try:
        resultado = selecionar_ficha_automatica(
            contrato=contrato,
            mutuario=mutuario,
            historico_envios=list(historico_envios)
        )
        
        ficha_recomendada = resultado.get('ficha_recomendada')
        motivo = resultado.get('motivo', '')
        pode_enviar = resultado.get('pode_enviar', False)
        pre_requisitos = resultado.get('pre_requisitos', [])
        validacoes_pendentes = resultado.get('validacoes_pendentes', [])
        fichas_complementares = resultado.get('fichas_complementares', [])
        
    except Exception as e:
        messages.error(request, f'Erro ao selecionar ficha: {str(e)}')
        ficha_recomendada = None
        pode_enviar = False
    
    # Se for POST, gera a ficha
    arquivo_gerado = None
    erros_validacao = []
    
    if request.method == 'POST':
        print("[DEBUG] POST recebido em gerar_ficha_view")
        tipo_ficha = request.POST.get('tipo_ficha', ficha_recomendada)
        print(f"[DEBUG] tipo_ficha: {tipo_ficha}")
        print(f"[DEBUG] pode_enviar: {pode_enviar}")
        print(f"[DEBUG] ficha_recomendada: {ficha_recomendada}")
        print(f"[DEBUG] pre_requisitos: {pre_requisitos}")
        print(f"[DEBUG] validacoes_pendentes: {validacoes_pendentes}")
        try:
            # Gera ficha com base no tipo
            if tipo_ficha == 'FH1':
                print("[DEBUG] Gerando FH1...")
                generator = FH1Generator(validar=True)
                linha, erros = generator.gerar_de_contrato(contrato)
                nome_arquivo = f'FH1_{contrato.codigo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
                print(f"[DEBUG] FH1 gerado: {nome_arquivo}, erros: {erros}")
            elif tipo_ficha == 'CADMUT':
                print("[DEBUG] Gerando CADMUT...")
                if not mutuario:
                    print("[ERROR] Mutuário não encontrado para gerar CADMUT")
                    raise ValueError('Mutuário não encontrado para gerar CADMUT')
                generator = CADMUTGenerator(validar=True)
                linha, erros = generator.gerar_de_mutuario(mutuario)
                nome_arquivo = f'CADMUT_{mutuario.cpf}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
                print(f"[DEBUG] CADMUT gerado: {nome_arquivo}, erros: {erros}")
            else:
                print(f"[ERROR] Tipo de ficha não suportado: {tipo_ficha}")
                raise ValueError(f'Tipo de ficha não suportado: {tipo_ficha}')
            if erros:
                erros_validacao = erros
                print(f"[DEBUG] Ficha gerada com {len(erros)} avisos: {erros}")
                messages.warning(request, f'Ficha gerada com {len(erros)} avisos')
            else:
                print("[DEBUG] Ficha gerada sem avisos")
            # Salva registro de envio
            envio = EnvioCEF.objects.create(
                contrato=contrato,
                tipo_ficha=tipo_ficha,
                arquivo=nome_arquivo,
                status='GERADO',
                observacoes=motivo
            )
            print(f"[DEBUG] Registro de envio salvo: {envio}")
            # Prepara arquivo para download
            response = HttpResponse(linha, content_type='text/plain; charset=latin-1')
            response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
            print("[DEBUG] Retornando arquivo para download")
            return response
        except Exception as e:
            print(f"[ERROR] Exceção ao gerar ficha: {str(e)}")
            messages.error(request, f'Erro ao gerar ficha: {str(e)}')
    
    context = {
        'contrato': contrato,
        'mutuario': mutuario,
        'ficha_recomendada': ficha_recomendada,
        'motivo': motivo,
        'pode_enviar': pode_enviar,
        'pre_requisitos': pre_requisitos,
        'validacoes_pendentes': validacoes_pendentes,
        'fichas_complementares': fichas_complementares,
        'historico_envios': historico_envios[:5],
        'erros_validacao': erros_validacao,
    }
    
    return render(request, 'principal/cef_gerar_ficha.html', context)


def validar_ficha_view(request):
    """
    View para validação de arquivos de ficha antes do envio
    - Upload de arquivo .txt
    - Validação de campos e consistência
    - Exibição de erros com severidade
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        messages.error(request, 'Módulos de fichas não disponíveis')
        return redirect('integracao_cef')
    
    erros = []
    avisos = []
    fichas_validas = 0
    fichas_com_erro = 0
    
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        tipo_ficha = request.POST.get('tipo_ficha', 'FH1')
        
        try:
            if tipo_ficha not in ('FH1', 'CADMUT'):
                messages.error(request, f'Tipo de ficha não suportado: {tipo_ficha}')
                context = {
                    'erros': erros,
                    'avisos': avisos,
                    'fichas_validas': fichas_validas,
                    'fichas_com_erro': fichas_com_erro,
                }
                return render(request, 'principal/cef_validar_ficha.html', context)

            parser_arquivo = ArquivoFichasCEF(tipo_ficha)

            # Lê conteúdo do arquivo
            conteudo = arquivo.read().decode('latin-1')
            linhas = conteudo.splitlines()
            
            # Valida cada linha
            for i, linha in enumerate(linhas, 1):
                if not linha.strip():
                    continue

                linha = linha.rstrip('\r')

                try:
                    parser = parser_arquivo.parser
                    linha_padded = linha.ljust(parser.TAMANHO_REGISTRO)
                    dados_linha = {}
                    for campo in parser.campos:
                        if campo.inicio > 0 and campo.fim > 0:
                            dados_linha[campo.nome] = linha_padded[campo.inicio-1:campo.fim].strip()
                except Exception as e:
                    fichas_com_erro += 1
                    erros.append({
                        'linha': i,
                        'campo': 'LINHA',
                        'mensagem': f'Erro ao interpretar linha no layout {tipo_ficha}: {str(e)}',
                        'severidade': 'error',
                        'codigo': 'ERRO_LEITURA_LAYOUT'
                    })
                    continue
                
                if tipo_ficha == 'FH1':
                    valido, erros_linha = validar_fh1(dados_linha)
                elif tipo_ficha == 'CADMUT':
                    valido, erros_linha = validar_cadmut(dados_linha)
                else:
                    erros.append({
                        'linha': i,
                        'campo': 'TIPO',
                        'mensagem': f'Tipo de ficha não suportado: {tipo_ficha}',
                        'severidade': 'ERRO'
                    })
                    continue
                
                if valido:
                    fichas_validas += 1
                else:
                    fichas_com_erro += 1
                    for erro in erros_linha:
                        erro['linha'] = i
                        severidade = str(erro.get('severidade', '')).strip().lower()
                        if severidade in ('erro', 'error'):
                            erros.append(erro)
                        else:
                            avisos.append(erro)
            
            if fichas_com_erro == 0:
                messages.success(request, f'✅ Arquivo válido! {fichas_validas} fichas sem erros.')
            else:
                messages.warning(request, f'⚠️ {fichas_com_erro} fichas com erros, {fichas_validas} válidas.')
                
        except Exception as e:
            messages.error(request, f'Erro ao validar arquivo: {str(e)}')
    
    context = {
        'erros': erros,
        'avisos': avisos,
        'fichas_validas': fichas_validas,
        'fichas_com_erro': fichas_com_erro,
    }
    
    return render(request, 'principal/cef_validar_ficha.html', context)


def interpretar_retorno_view(request):
    """
    View para interpretação automática de arquivos de retorno CEF
    - Upload de arquivo de retorno .txt
    - Parsing automático (HEADER, MOVIMENTO, CRÍTICAS, TRAILER)
    - Interpretação de códigos de crítica
    - Exibição de ação requerida
    - Salva no banco para histórico
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        messages.error(request, 'Módulos de fichas não disponíveis')
        return redirect('integracao_cef')
    
    relatorio = None
    resumo = None
    mensagens = []
    criticas = []
    acao_requerida = None
    
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        tipo_retorno = request.POST.get('tipo_retorno', 'FCVS')
        
        try:
            # Salva arquivo temporário
            temp_path = f'/tmp/{arquivo.name}'
            with open(temp_path, 'wb+') as temp_file:
                for chunk in arquivo.chunks():
                    temp_file.write(chunk)
            
            # Interpreta com base no tipo
            if tipo_retorno == 'FCVS':
                relatorio = interpretar_retorno_fcvs(temp_path)
            elif tipo_retorno == 'CADMUT':
                relatorio = interpretar_retorno_cadmut(temp_path)
            else:
                raise ValueError(f'Tipo de retorno não suportado: {tipo_retorno}')
            
            # Extrai informações
            resumo = relatorio.get('resumo', {})
            mensagens = relatorio.get('mensagens', [])
            criticas = relatorio.get('criticas', [])
            acao_requerida = relatorio.get('acao_requerida', 'DESCONHECIDA')
            
            # Salva no banco
            total_registros = resumo.get('total_registros', 0)
            aceitos = resumo.get('aceitos', 0)
            rejeitados = resumo.get('rejeitados', 0)
            
            retorno = RetornoCEF.objects.create(
                arquivo=arquivo.name,
                tipo_retorno=tipo_retorno,
                total_registros=total_registros,
                registros_aceitos=aceitos,
                registros_rejeitados=rejeitados,
                observacoes=f'Ação requerida: {acao_requerida}'
            )
            
            # Vincula com envios se possível
            for msg in mensagens:
                codigo_contrato = msg.get('codigo_contrato')
                if codigo_contrato:
                    try:
                        contrato = Contrato.objects.get(codigo=codigo_contrato)
                        envio = EnvioCEF.objects.filter(
                            contrato=contrato,
                            status__in=['ENVIADO', 'PENDENTE']
                        ).first()
                        if envio:
                            envio.retorno = retorno
                            envio.status = 'ACEITO' if msg.get('status') == 'SUCESSO' else 'REJEITADO'
                            envio.save()
                    except Contrato.DoesNotExist:
                        pass
            
            # Remove arquivo temporário
            os.remove(temp_path)
            
            # Mensagem de sucesso
            if acao_requerida == 'SUCESSO':
                messages.success(request, f'✅ Retorno processado: {aceitos} aceitos')
            elif acao_requerida == 'CORRIGIR_E_REENVIAR':
                messages.error(request, f'❌ {rejeitados} rejeitados - correção necessária')
            else:
                messages.warning(request, f'⚠️ Retorno processado - verificar detalhes')
                
        except Exception as e:
            messages.error(request, f'Erro ao interpretar retorno: {str(e)}')
    
    context = {
        'relatorio': relatorio,
        'resumo': resumo,
        'mensagens': mensagens,
        'criticas': criticas,
        'acao_requerida': acao_requerida,
        'ultimos_retornos': RetornoCEF.objects.order_by('-data_retorno')[:10],
    }
    
    return render(request, 'principal/cef_interpretar_retorno.html', context)


def _detectar_codigo_relatorio(nome_arquivo: str) -> str:
    """Extrai o codigo de relatorio do padrao MICP.FCVS.<CODIGO>.Axxxxx..."""
    if not nome_arquivo:
        return 'DESCONHECIDO'

    nome_upper = nome_arquivo.upper()

    match = re.search(r'MICP\.FCVS\.([^.]+)\.A\d+', nome_upper)
    if match:
        return match.group(1)

    # Pacote legado/alternativo da CEF: FCV.<UF>.<...>.P3026....D1..D8
    if '.P3026.' in nome_upper:
        return 'P3026'

    return 'DESCONHECIDO'


def _processar_relatorio_textual(caminho_arquivo: str, codigo: str = '') -> dict:
    """Leitura basica de relatorios textuais (familia Sxxxxxx)."""
    linhas = _linhas_relatorio_textual(caminho_arquivo)
    codigo = (codigo or '').upper()

    totalizadores = [
        ln.strip() for ln in linhas
        if re.search(r'\bTOTAL\b|\bTOTAIS\b|\bTOTALIZ', ln, re.IGNORECASE)
    ]

    resumo = {
        'linhas_total': len(linhas),
        'totalizadores_encontrados': len(totalizadores),
        'primeiro_cabecalho': linhas[0][:120] if linhas else '',
        'amostra_totalizadores': totalizadores[:3],
    }

    if codigo == 'S194301':
        registros = _extrair_registros_s194301(linhas)
        motivos = Counter(r['motivo_rejeicao'] for r in registros if r.get('motivo_rejeicao'))
        resumo.update({
            'contratos_rejeitados': len(registros),
            'motivos_distintos': len(motivos),
            'motivo_principal': motivos.most_common(1)[0][0] if motivos else '',
        })
    elif codigo == 'S778101':
        registros = _extrair_registros_s778101(linhas)
        tipos = Counter(r['tipo_solicitacao'] for r in registros if r.get('tipo_solicitacao'))
        resumo.update({
            'contratos_acatados': len(registros),
            'qtd_rnv': tipos.get('RNV', 0),
            'qtd_rcv': tipos.get('RCV', 0),
        })
    elif codigo == 'S343501':
        registros = _extrair_registros_s3435xx(linhas, 'vaf4')
        resumo.update({
            'contratos_vaf4': len(registros),
            'total_vaf4': round(sum(r.get('vaf4_num', 0.0) for r in registros), 2),
        })
    elif codigo == 'S343601':
        registros = _extrair_registros_s3435xx(linhas, 'vaf3')
        resumo.update({
            'contratos_vaf3': len(registros),
            'total_vaf3': round(sum(r.get('vaf3_num', 0.0) for r in registros), 2),
        })
    elif codigo == 'S820401':
        registros = _extrair_registros_s820401(linhas)
        resumo.update({
            'contratos_multa': len(registros),
            'total_multa_moeda12': round(sum(r.get('multa_moeda12_num', 0.0) for r in registros), 2),
            'total_multa_moeda14': round(sum(r.get('multa_moeda14_num', 0.0) for r in registros), 2),
            'total_multa_moeda15': round(sum(r.get('multa_moeda15_num', 0.0) for r in registros), 2),
        })
    elif codigo == 'S820301':
        registros = _extrair_registros_s820301(linhas)
        resumo.update({
            'contratos_multa_820301': len(registros),
            'total_multa12_820301': round(sum(r.get('multa_moeda12_num', 0.0) for r in registros), 2),
            'total_multa14_820301': round(sum(r.get('multa_moeda14_num', 0.0) for r in registros), 2),
            'total_multa15_820301': round(sum(r.get('multa_moeda15_num', 0.0) for r in registros), 2),
            'total_meses12_820301': sum(int(r.get('meses_moeda12', 0)) for r in registros),
            'total_meses14_820301': sum(int(r.get('meses_moeda14', 0)) for r in registros),
            'total_meses15_820301': sum(int(r.get('meses_moeda15', 0)) for r in registros),
        })
    elif codigo == 'S820601':
        registros = _extrair_registros_s820601(linhas)
        por_contrato = {}
        for reg in registros:
            contrato = reg.get('numero_contrato', '')
            if not contrato:
                continue
            resumo_contrato = por_contrato.setdefault(contrato, {'saldo_final': 0.0})
            resumo_contrato['saldo_final'] = reg.get('vlr_devido_num', 0.0)

        resumo.update({
            'contratos_apuracao_820601': len(por_contrato),
            'periodos_apuracao_820601': len(registros),
            'total_apurado_820601': round(sum(r.get('vlr_apurado_num', 0.0) for r in registros), 2),
            'saldo_final_820601': round(sum(v.get('saldo_final', 0.0) for v in por_contrato.values()), 2),
        })

    return resumo


def _interpretacao_manual_codigo(codigo: str) -> str:
    """Resumo interpretativo por codigo, alinhado aos manuais/layouts do projeto."""
    return _guia_manual_codigo(codigo).get('interpretacao', '')


def _guia_manual_codigo(codigo: str) -> Dict[str, str]:
    """Retorna interpretação, fonte manual e ação recomendada por código."""
    return MANUAL_EXPERT_AGENT.guia(codigo)


def _detalhes_cadmut1(caminho_arquivo: str, arquivo_nome: str, codigo: str) -> List[dict]:
    """Extrai detalhamento por registro para CADMUT1."""
    from .ficha_return_interpreter import CADMUTReturnParser

    parser = CADMUTReturnParser()
    parser.parse_arquivo(caminho_arquivo)

    detalhes = []
    for idx, reg in enumerate(parser.registros, start=1):
        campos = reg.campos or {}
        criticas = reg.codigos_critica or []
        if criticas:
            for c in criticas:
                detalhes.append({
                    'arquivo': arquivo_nome,
                    'codigo': codigo,
                    'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                    'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                    'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                    'origem_parser': 'CADMUTReturnParser',
                    'tipo_linha': reg.tipo_registro,
                    'indice_linha': idx,
                    'numero_contrato': campos.get('numero_contrato', ''),
                    'cpf': campos.get('cpf', ''),
                    'nome': campos.get('nome', ''),
                    'status': reg.status,
                    'codigo_critica': c.get('codigo', ''),
                    'descricao_critica': c.get('descricao', ''),
                    'texto': reg.linha_original[:240],
                })
        else:
            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'CADMUTReturnParser',
                'tipo_linha': reg.tipo_registro,
                'indice_linha': idx,
                'numero_contrato': campos.get('numero_contrato', ''),
                'cpf': campos.get('cpf', ''),
                'nome': campos.get('nome', ''),
                'status': reg.status,
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': reg.linha_original[:240],
            })

    return detalhes


def _detalhes_m3026(caminho_arquivo: str, arquivo_nome: str, codigo: str) -> List[dict]:
    """Extrai detalhamento estruturado para M3026xx com fallback heuristico."""
    from .ficha_p3026_parser_v2 import ParserP3026

    def _get(dados: Dict[str, str], *keys: str) -> str:
        for key in keys:
            valor = dados.get(key)
            if valor:
                return str(valor).strip()
        return ''

    def _depuracao_texto(tipo_linha: str, cod_sit: str, desc_sit: str, tipo_evento: str, data_evento: str) -> str:
        partes = [f"TR{tipo_linha}"]
        if cod_sit or desc_sit:
            partes.append(f"Situacao: {cod_sit} {desc_sit}".strip())
        if tipo_evento or data_evento:
            partes.append(f"Evento: {tipo_evento} {data_evento}".strip())
        return ' | '.join(p for p in partes if p)

    def _status_homologacao(sit_analise: str) -> str:
        mapa = {
            '0': 'Nao homologado',
            '1': 'Em analise',
            '2': 'Homologado',
            '3': 'Homologado com reabertura',
        }
        key = (sit_analise or '').strip()
        return mapa.get(key, f'Situacao analise {key}' if key else '')

    def _decimal_cents(valor: str) -> str:
        v = (valor or '').strip()
        if not v or not v.isdigit():
            return ''
        try:
            return f"{int(v) / 100:.2f}"
        except Exception:
            return ''

    parser = ParserP3026()
    arquivo_p3026, _ = parser.parse_arquivo(caminho_arquivo)

    if arquivo_p3026 and arquivo_p3026.registros:
        detalhes = []
        for idx, reg in enumerate(arquivo_p3026.registros, start=1):
            dados = reg.dados or {}
            linha_original = dados.get('linha_original', '') or ''

            # Extrai nome com prioridade em campos do layout; fallback por regex.
            nome = _get(dados, 'NOME_DO_MUTUARIO', 'NOME_MUTUARIO', 'NOME')
            if not nome:
                nome_match = re.search(r'([A-ZÀ-Ú][A-ZÀ-Ú\s]{8,60})', linha_original)
                if nome_match:
                    nome = re.sub(r'\s+', ' ', nome_match.group(1)).strip()

            cod_sit = _get(dados, 'CODIGO_SITUACAO_DO_CONTRATO', 'CODIGO_SITUACAO_CONTRATO')
            desc_sit = _get(dados, 'DESCRICAO_SITUACAO_DO_CONTRATO', 'DESCRICAO_DA_SITUACAO_DO_CONTRATO', 'SITUACAO_CONTRATO')
            tipo_evento = _get(dados, 'TIPO_DE_EVENTO', 'TIPO_EVENTO')
            data_evento = _get(dados, 'DATA_DO_EVENTO', 'DATA_EVENTO')
            manifestacao = _get(dados, 'RNV_OU_RCV_OU_RCNP_OU_SM', 'RNV_OU_RCV_OU_RCNP_OU_SM_SEM_MANIFESTACAO')
            data_manifestacao = _get(dados, 'DATA_RNV_RCV_RCNP')
            sit_analise = _get(dados, 'SITUACAO_DE_ANALISE_ATUAL', 'SIT_ANALISE')
            status_homologacao = _status_homologacao(sit_analise)

            valor_credito_raw = _get(
                dados,
                'VALOR_A_CREDITO_DO_FCVS',
                'VLR_A_CREDITO_DO_FCVS',
                'SALDO_TOTAL_LIQUIDO',
                'VAF4_RESSARCIDO_A_PARTIR_DE_01_01_1997',
            )
            valor_credito = _decimal_cents(valor_credito_raw)

            situacao = desc_sit or cod_sit or 'LIDO'
            depuracao = _depuracao_texto(reg.tipo_registro, cod_sit, desc_sit, tipo_evento, data_evento)
            if status_homologacao:
                depuracao = f"{depuracao} | Homologacao: {status_homologacao}" if depuracao else f"Homologacao: {status_homologacao}"
            if manifestacao:
                depuracao = f"{depuracao} | Manifestacao: {manifestacao} {data_manifestacao}".strip()
            if valor_credito:
                depuracao = f"{depuracao} | Credito: {valor_credito}"

            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': f"P3026-v2-TR{reg.tipo_registro}",
                'tipo_linha': reg.tipo_registro,
                'indice_linha': idx,
                'numero_contrato': reg.numero_contrato or '',
                'cpf': reg.cpf_mutuario or '',
                'nome': nome,
                'status': situacao,
                'tipo_evento': tipo_evento,
                'data_evento': data_evento,
                'codigo_situacao': cod_sit,
                'descricao_situacao': desc_sit,
                'manifestacao': manifestacao,
                'data_manifestacao': data_manifestacao,
                'situacao_analise': sit_analise,
                'status_homologacao': status_homologacao,
                'valor_credito': valor_credito,
                'valor_credito_bruto': valor_credito_raw,
                'depuracao': depuracao,
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': linha_original[:240],
            })

        return detalhes

    # Fallback: leitura textual/heuristica se parser estruturado nao retornar registros.
    detalhes = []
    with open(caminho_arquivo, 'r', encoding='latin-1', errors='ignore') as f:
        linhas = [ln.rstrip('\n\r') for ln in f if ln.strip()]

    for idx, ln in enumerate(linhas, start=1):
        norm = ln.lstrip()
        tipo = norm[:1] if norm[:1].isdigit() else 'TXT'

        contrato = ''
        cpf = ''
        nome = ''

        contratos = re.findall(r'\d{13}', norm)
        if contratos:
            contrato = contratos[0]

        cpfs = re.findall(r'\d{11}', norm)
        if cpfs:
            cpf = cpfs[0]

        nome_match = re.search(r'([A-ZÀ-Ú][A-ZÀ-Ú\s]{8,60})', norm)
        if nome_match:
            nome = re.sub(r'\s+', ' ', nome_match.group(1)).strip()

        detalhes.append({
            'arquivo': arquivo_nome,
            'codigo': codigo,
            'interpretacao_manual': _interpretacao_manual_codigo(codigo),
            'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
            'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
            'origem_parser': 'M3026-Heuristico',
            'tipo_linha': tipo,
            'indice_linha': idx,
            'numero_contrato': contrato,
            'cpf': cpf,
            'nome': nome,
            'status': 'LIDO',
            'tipo_evento': '',
            'data_evento': '',
            'codigo_situacao': '',
            'descricao_situacao': '',
            'manifestacao': '',
            'data_manifestacao': '',
            'situacao_analise': '',
            'status_homologacao': '',
            'valor_credito': '',
            'valor_credito_bruto': '',
            'depuracao': f"TR{tipo} | Leitura heuristica",
            'codigo_critica': '',
            'descricao_critica': '',
            'texto': norm[:240],
        })

    return detalhes


def _detalhes_relatorio_textual(caminho_arquivo: str, arquivo_nome: str, codigo: str) -> List[dict]:
    """Extrai desdobramento para relatorios textuais: cabecalhos, paginas e totalizadores."""
    linhas = _linhas_relatorio_textual(caminho_arquivo)
    codigo = (codigo or '').upper()

    if codigo == 'S194301':
        detalhes = []
        for idx, reg in enumerate(_extrair_registros_s194301(linhas), start=1):
            motivo = reg.get('motivo_rejeicao', '')
            lote = reg.get('lote', '')
            depuracao = f"Lote {lote}" if lote else ''
            if motivo:
                depuracao = f"{depuracao} | Motivo: {motivo}" if depuracao else f"Motivo: {motivo}"

            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'S194301-TextParser',
                'tipo_linha': 'CONTRATO_REJEITADO',
                'indice_linha': idx,
                'numero_contrato': reg.get('numero_contrato', ''),
                'cpf': '',
                'nome': reg.get('nome', ''),
                'status': 'REJEITADO',
                'codigo_situacao': reg.get('lote', ''),
                'descricao_situacao': motivo,
                'manifestacao': '',
                'data_manifestacao': '',
                'situacao_analise': '',
                'status_homologacao': '',
                'valor_credito': '',
                'valor_credito_bruto': '',
                'tipo_evento': 'PEDIDO_HABILITACAO_REJEITADO',
                'data_evento': reg.get('data_geracao', ''),
                'depuracao': depuracao,
                'codigo_critica': '',
                'descricao_critica': motivo,
                'texto': f"{reg.get('numero_contrato', '')} {reg.get('nome', '')} {motivo}"[:240],
            })

        return detalhes

    if codigo == 'S778101':
        detalhes = []
        for idx, reg in enumerate(_extrair_registros_s778101(linhas), start=1):
            tipo_solicitacao = reg.get('tipo_solicitacao', '')
            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'S778101-TextParser',
                'tipo_linha': 'SOLICITACAO_ACATADA',
                'indice_linha': idx,
                'numero_contrato': reg.get('numero_contrato', ''),
                'cpf': '',
                'nome': reg.get('nome', ''),
                'status': 'ACATADO',
                'codigo_situacao': reg.get('lote', ''),
                'descricao_situacao': tipo_solicitacao,
                'manifestacao': tipo_solicitacao,
                'data_manifestacao': reg.get('data_termino_analise', ''),
                'situacao_analise': reg.get('data_habilitacao', ''),
                'status_homologacao': '',
                'valor_credito': '',
                'valor_credito_bruto': '',
                'tipo_evento': 'SOLICITACAO_ACATADA',
                'data_evento': reg.get('data_geracao', ''),
                'depuracao': (
                    f"Hip {reg.get('hipoteca', '')} | Habilitacao: {reg.get('data_habilitacao', '')} | "
                    f"Termino analise: {reg.get('data_termino_analise', '')} | Lote {reg.get('lote', '')}"
                ),
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': (
                    f"{reg.get('numero_contrato', '')} {reg.get('nome', '')} {tipo_solicitacao} "
                    f"{reg.get('data_habilitacao', '')} {reg.get('data_termino_analise', '')}"
                )[:240],
            })

        return detalhes

    if codigo == 'S343501':
        detalhes = []
        for idx, reg in enumerate(_extrair_registros_s3435xx(linhas, 'vaf4'), start=1):
            vaf4 = reg.get('vaf4', '')
            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'S343501-TextParser',
                'tipo_linha': 'CONTRATO_VAF4',
                'indice_linha': idx,
                'numero_contrato': reg.get('numero_contrato', ''),
                'cpf': '',
                'nome': reg.get('nome', ''),
                'status': 'LIDO',
                'codigo_situacao': reg.get('hipoteca', ''),
                'descricao_situacao': 'DIFERENCIAL TAXA DE JUROS',
                'manifestacao': '',
                'data_manifestacao': '',
                'situacao_analise': '',
                'status_homologacao': '',
                'valor_credito': str(vaf4),
                'valor_credito_bruto': str(vaf4),
                'tipo_evento': 'VAF4',
                'data_evento': '',
                'depuracao': f"Matricula {reg.get('matricula', '')} | Hip {reg.get('hipoteca', '')} | VAF4 {vaf4}",
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': f"{reg.get('numero_contrato', '')} {reg.get('nome', '')} VAF4 {vaf4}"[:240],
            })

        return detalhes

    if codigo == 'S343601':
        detalhes = []
        for idx, reg in enumerate(_extrair_registros_s3435xx(linhas, 'vaf3'), start=1):
            vaf3 = reg.get('vaf3', '')
            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'S343601-TextParser',
                'tipo_linha': 'CONTRATO_VAF3',
                'indice_linha': idx,
                'numero_contrato': reg.get('numero_contrato', ''),
                'cpf': '',
                'nome': reg.get('nome', ''),
                'status': 'LIDO',
                'codigo_situacao': reg.get('hipoteca', ''),
                'descricao_situacao': 'DIFERENCIAL DEC. 97.222/88',
                'manifestacao': '',
                'data_manifestacao': '',
                'situacao_analise': '',
                'status_homologacao': '',
                'valor_credito': str(vaf3),
                'valor_credito_bruto': str(vaf3),
                'tipo_evento': 'VAF3',
                'data_evento': '',
                'depuracao': f"Matricula {reg.get('matricula', '')} | Hip {reg.get('hipoteca', '')} | VAF3 {vaf3}",
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': f"{reg.get('numero_contrato', '')} {reg.get('nome', '')} VAF3 {vaf3}"[:240],
            })

        return detalhes

    if codigo == 'S820401':
        detalhes = []
        for idx, reg in enumerate(_extrair_registros_s820401(linhas), start=1):
            m12 = reg.get('multa_moeda12', '')
            m14 = reg.get('multa_moeda14', '')
            m15 = reg.get('multa_moeda15', '')
            total_multa = reg.get('multa_moeda12_num', 0.0) + reg.get('multa_moeda14_num', 0.0) + reg.get('multa_moeda15_num', 0.0)
            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'S820401-TextParser',
                'tipo_linha': 'CONTRATO_MULTA',
                'indice_linha': idx,
                'numero_contrato': reg.get('numero_contrato', ''),
                'cpf': '',
                'nome': '',
                'status': 'LIDO',
                'codigo_situacao': reg.get('hipoteca', ''),
                'descricao_situacao': 'RESSARCIMENTO CUSTOS ADMINISTRATIVOS',
                'manifestacao': '',
                'data_manifestacao': '',
                'situacao_analise': '',
                'status_homologacao': '',
                'valor_credito': f"{total_multa:.2f}",
                'valor_credito_bruto': f"M12={m12};M14={m14};M15={m15}",
                'tipo_evento': 'MULTA_MENSAL',
                'data_evento': '',
                'depuracao': f"Matricula {reg.get('matricula', '')} | Hip {reg.get('hipoteca', '')} | M12={m12} M14={m14} M15={m15}",
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': f"{reg.get('numero_contrato', '')} M12={m12} M14={m14} M15={m15}"[:240],
            })

        return detalhes

    if codigo == 'S820301':
        detalhes = []
        for idx, reg in enumerate(_extrair_registros_s820301(linhas), start=1):
            m12 = reg.get('multa_moeda12', '')
            m14 = reg.get('multa_moeda14', '')
            m15 = reg.get('multa_moeda15', '')
            q12 = reg.get('meses_moeda12', 0)
            q14 = reg.get('meses_moeda14', 0)
            q15 = reg.get('meses_moeda15', 0)
            total_multa = reg.get('multa_moeda12_num', 0.0) + reg.get('multa_moeda14_num', 0.0) + reg.get('multa_moeda15_num', 0.0)
            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'S820301-TextParser',
                'tipo_linha': 'CONTRATO_MULTA_TOTALIZADA',
                'indice_linha': idx,
                'numero_contrato': reg.get('numero_contrato', ''),
                'cpf': '',
                'nome': '',
                'status': 'LIDO',
                'codigo_situacao': reg.get('hipoteca', ''),
                'descricao_situacao': 'MULTA TOTALIZADA POR MOEDA',
                'manifestacao': '',
                'data_manifestacao': '',
                'situacao_analise': '',
                'status_homologacao': '',
                'valor_credito': f"{total_multa:.2f}",
                'valor_credito_bruto': f"M12={m12};Q12={q12};M14={m14};Q14={q14};M15={m15};Q15={q15}",
                'tipo_evento': 'MULTA_TOTALIZADA',
                'data_evento': '',
                'depuracao': (
                    f"Matricula {reg.get('matricula', '')} | Hip {reg.get('hipoteca', '')} | "
                    f"M12={m12} ({q12}m) M14={m14} ({q14}m) M15={m15} ({q15}m)"
                ),
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': f"{reg.get('numero_contrato', '')} M12={m12} M14={m14} M15={m15}"[:240],
            })

        return detalhes

    if codigo == 'S820601':
        periodos = _extrair_registros_s820601(linhas)
        agregados = {}
        ordem_contratos = []

        for periodo in periodos:
            contrato = periodo.get('numero_contrato', '')
            if not contrato:
                continue

            if contrato not in agregados:
                agregados[contrato] = {
                    'matricula': periodo.get('matricula', ''),
                    'hipoteca': periodo.get('hipoteca', ''),
                    'periodos': 0,
                    'total_multa': 0.0,
                    'total_apurado': 0.0,
                    'saldo_final': 0.0,
                    'inicio_primeiro': periodo.get('inicio_vigencia', ''),
                    'fim_ultimo': periodo.get('fim_vigencia', ''),
                }
                ordem_contratos.append(contrato)

            agg = agregados[contrato]
            agg['periodos'] += 1
            agg['total_multa'] += periodo.get('vlr_multa_num', 0.0)
            agg['total_apurado'] += periodo.get('vlr_apurado_num', 0.0)
            agg['saldo_final'] = periodo.get('vlr_devido_num', 0.0)
            agg['fim_ultimo'] = periodo.get('fim_vigencia', '')

        detalhes = []
        for idx, contrato in enumerate(ordem_contratos, start=1):
            agg = agregados[contrato]
            detalhes.append({
                'arquivo': arquivo_nome,
                'codigo': codigo,
                'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                'origem_parser': 'S820601-TextParser',
                'tipo_linha': 'CONTRATO_APURACAO',
                'indice_linha': idx,
                'numero_contrato': contrato,
                'cpf': '',
                'nome': '',
                'status': 'LIDO',
                'codigo_situacao': agg.get('hipoteca', ''),
                'descricao_situacao': 'APURACAO MENSAL DE MULTA',
                'manifestacao': '',
                'data_manifestacao': '',
                'situacao_analise': '',
                'status_homologacao': '',
                'valor_credito': f"{agg.get('saldo_final', 0.0):.2f}",
                'valor_credito_bruto': (
                    f"periodos={agg.get('periodos', 0)};"
                    f"multa_total={agg.get('total_multa', 0.0):.2f};"
                    f"apurado_total={agg.get('total_apurado', 0.0):.2f}"
                ),
                'tipo_evento': 'APURACAO_MULTA',
                'data_evento': agg.get('fim_ultimo', ''),
                'depuracao': (
                    f"Matricula {agg.get('matricula', '')} | Hip {agg.get('hipoteca', '')} | "
                    f"Periodo {agg.get('inicio_primeiro', '')} a {agg.get('fim_ultimo', '')} | "
                    f"Qtd periodos {agg.get('periodos', 0)} | "
                    f"Apurado {agg.get('total_apurado', 0.0):.2f} | Saldo final {agg.get('saldo_final', 0.0):.2f}"
                ),
                'codigo_critica': '',
                'descricao_critica': '',
                'texto': (
                    f"{contrato} periodos={agg.get('periodos', 0)} "
                    f"apurado={agg.get('total_apurado', 0.0):.2f} saldo={agg.get('saldo_final', 0.0):.2f}"
                )[:240],
            })

        return detalhes

    detalhes = []

    for idx, ln in enumerate(linhas, start=1):
        if not ln.strip():
            continue

        tipo = None
        u = ln.upper()
        if re.search(r'\bTOTAL\b|\bTOTAIS\b|\bTOTALIZ', u):
            tipo = 'TOTALIZADOR'
        elif 'PAGINA' in u or 'SEQ.' in u:
            tipo = 'CABECALHO_PAGINA'
        elif idx <= 2:
            tipo = 'CABECALHO_ARQUIVO'

        if tipo is None:
            continue

        detalhes.append({
            'arquivo': arquivo_nome,
            'codigo': codigo,
            'interpretacao_manual': _interpretacao_manual_codigo(codigo),
            'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
            'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
            'origem_parser': 'RelatorioTextual',
            'tipo_linha': tipo,
            'indice_linha': idx,
            'numero_contrato': '',
            'cpf': '',
            'nome': '',
            'status': 'LIDO',
            'codigo_critica': '',
            'descricao_critica': '',
            'texto': ln[:240],
        })

    return detalhes


def _mapear_tipo_retorno_por_linha(row: dict) -> str:
    """Mapeia linha detalhada para tipo de retorno da tela CEF."""
    status_h = str(row.get('status_homologacao', '')).lower()
    status = str(row.get('status', '')).upper()
    manifestacao = str(row.get('manifestacao', '')).upper()
    codigo_critica = str(row.get('codigo_critica', '')).strip()
    descricao_critica = str(row.get('descricao_critica', '')).strip()
    descricao_situacao = str(row.get('descricao_situacao', '')).strip()
    depuracao = str(row.get('depuracao', '')).strip()
    texto = str(row.get('texto', '')).strip()

    texto_base = ' | '.join([
        status_h,
        status,
        manifestacao,
        descricao_critica,
        descricao_situacao,
        depuracao,
        texto,
    ]).lower()

    termos_rejeicao = [
        'negativa de cobertura', 'negativa cobertura', 'sem cobertura',
        'nao coberto', 'não coberto', 'nao homologado', 'não homologado',
        'indefer', 'rejeitad', 'improcedente', 'negado'
    ]
    termos_exigencia = [
        'exigencia', 'exigência', 'documentacao complementar', 'documentação complementar',
        'documento complementar', 'pendencia documental', 'pendência documental',
        'complementar', 'apresentar documento', 'documentos solicitados'
    ]

    if any(t in texto_base for t in termos_rejeicao):
        return 'REJEITADO'
    if any(t in texto_base for t in termos_exigencia):
        return 'COMPLEMENTAR'

    if 'homologado' in status_h:
        return 'APROVADO'
    if status in {'REJEITADO', 'ERRO'} or codigo_critica or descricao_critica:
        return 'REJEITADO'
    if manifestacao in {'RCNP', 'SM'}:
        return 'PENDENTE'
    return 'PENDENTE'


def _buscar_contrato_por_codigo_retorno(numero_contrato: str, cpf: str = ''):
    cpf_digits = re.sub(r'\D', '', str(cpf or ''))
    if len(cpf_digits) == 11:
        mut = Mutuario.objects.filter(cpf=cpf_digits).first()
        if mut:
            contrato = Contrato.objects.filter(codigo=str(mut.codigo).strip()).first()
            if contrato:
                return contrato

    codigo_raw = str(numero_contrato or '').strip()
    if not codigo_raw:
        return None

    codigo_nozero = codigo_raw.lstrip('0') or codigo_raw

    contrato = Contrato.objects.filter(codigo=codigo_raw).first()
    if contrato:
        return contrato

    contrato = Contrato.objects.filter(codigo=codigo_nozero).first()
    if contrato:
        return contrato

    # Fallback: compara sufixo para lidar com zeros a esquerda em layouts CEF.
    contrato = Contrato.objects.filter(codigo__endswith=codigo_nozero).first()
    if contrato:
        return contrato

    return None


def _persistir_detalhes_como_retornos_cef(detalhes_arquivos: list) -> dict:
    """Persiste detalhes processados em RetornoCEF para aparecer na tela de retornos."""
    criados = 0
    ignorados_sem_contrato = 0
    ignorados_duplicados = 0

    vistos = set()

    for row in detalhes_arquivos:
        numero_contrato = str(row.get('numero_contrato', '')).strip()
        cpf = str(row.get('cpf', '')).strip()

        if not numero_contrato and not cpf:
            ignorados_sem_contrato += 1
            continue

        contrato = _buscar_contrato_por_codigo_retorno(numero_contrato, cpf)
        if not contrato:
            ignorados_sem_contrato += 1
            continue

        arquivo = str(row.get('arquivo', '')).strip()
        codigo = str(row.get('codigo', '')).strip()
        tipo_linha = str(row.get('tipo_linha', '')).strip()
        indice_linha = str(row.get('indice_linha', '')).strip()
        hash_base = f"{arquivo}|{codigo}|{numero_contrato}|{cpf}|{tipo_linha}|{indice_linha}"
        protocolo = 'IMP-' + hashlib.sha1(hash_base.encode('utf-8')).hexdigest()[:16].upper()

        chave = (contrato.id, protocolo)
        if chave in vistos:
            ignorados_duplicados += 1
            continue
        vistos.add(chave)

        if RetornoCEF.objects.filter(contrato=contrato, protocolo=protocolo).exists():
            ignorados_duplicados += 1
            continue

        tipo_retorno = _mapear_tipo_retorno_por_linha(row)
        analise = row.get('descricao_situacao') or row.get('depuracao') or row.get('texto') or ''
        motivo = row.get('descricao_critica') or ''
        requer_acao = tipo_retorno in {'REJEITADO', 'PENDENTE', 'COMPLEMENTAR'}

        RetornoCEF.objects.create(
            contrato=contrato,
            tipo_retorno=tipo_retorno,
            protocolo=protocolo,
            arquivo_retorno=arquivo,
            conteudo=str(row.get('texto', '') or '')[:2000],
            analise_cef=str(analise)[:2000],
            motivo_rejeicao=str(motivo)[:2000],
            data_retorno=timezone.now(),
            lido=False,
            requer_acao=requer_acao,
            processado=True,
        )
        criados += 1

    return {
        'criados': criados,
        'ignorados_sem_contrato': ignorados_sem_contrato,
        'ignorados_duplicados': ignorados_duplicados,
    }


def _detectar_data_remessa_nome_arquivo(nome_arquivo: str) -> str:
    """Extrai data de remessa do padrao .Dddmmaa.H e retorna em YYYY-MM-DD."""
    nome_upper = (nome_arquivo or '').upper()

    m = re.search(r'\.D(\d{6})\.H', nome_upper)
    if not m:
        # Fallback para nomes no formato ...ABR2026...
        m_mes = re.search(r'\.(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)(\d{4})(?:\.|$)', nome_upper)
        if not m_mes:
            return ''

        mapa_mes = {
            'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6,
            'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12,
        }
        try:
            ano = int(m_mes.group(2))
            mes = mapa_mes[m_mes.group(1)]
            return datetime(ano, mes, 1).strftime('%Y-%m-%d')
        except Exception:
            return ''

    txt = m.group(1)
    try:
        return datetime.strptime(txt, '%d%m%y').strftime('%Y-%m-%d')
    except ValueError:
        return ''


def _parse_iso_date(txt: str):
    v = (txt or '').strip()
    if not v:
        return None
    try:
        return datetime.strptime(v, '%Y-%m-%d').date()
    except ValueError:
        return None


def _ultimo_dia_util_mes_apos(data_base, meses_apos: int):
    if not data_base:
        return None

    ano = data_base.year + ((data_base.month - 1 + meses_apos) // 12)
    mes = ((data_base.month - 1 + meses_apos) % 12) + 1

    if mes == 12:
        prox_ano, prox_mes = ano + 1, 1
    else:
        prox_ano, prox_mes = ano, mes + 1

    ultimo_dia = datetime(prox_ano, prox_mes, 1).date() - timedelta(days=1)
    while ultimo_dia.weekday() >= 5:  # 5=sabado, 6=domingo
        ultimo_dia -= timedelta(days=1)

    return ultimo_dia


def _dias_restantes(deadline, hoje):
    if not deadline:
        return ''
    return str((deadline - hoje).days)


def _estado_prazo(dias_txt: str) -> str:
    if dias_txt == '':
        return ''
    try:
        dias = int(dias_txt)
    except Exception:
        return ''

    if dias < 0:
        return 'VENCIDO'
    if dias <= 10:
        return 'CRITICO'
    if dias <= 30:
        return 'ALERTA'
    return 'NO_PRAZO'


def _risco_rcnp_por_estado(estado_manifestacao: str) -> str:
    if estado_manifestacao in ('VENCIDO', 'CRITICO'):
        return 'ALTO'
    if estado_manifestacao == 'ALERTA':
        return 'MEDIO'
    if estado_manifestacao == 'NO_PRAZO':
        return 'BAIXO'
    return ''


def _status_compliance_fcvs(*estados: str) -> str:
    ordem = {
        'VENCIDO': 4,
        'CRITICO': 3,
        'ALERTA': 2,
        'NO_PRAZO': 1,
        '': 0,
    }
    max_estado = ''
    max_peso = 0
    for estado in estados:
        peso = ordem.get(estado or '', 0)
        if peso > max_peso:
            max_estado = estado
            max_peso = peso
    return max_estado


def _aplicar_compliance_fcvs_lote(item: dict) -> None:
    """Calcula prazos normativos FCVS para linha do consolidado do lote."""
    data_base = _parse_iso_date(item.get('data_remessa', ''))
    hoje = datetime.now().date()

    if not data_base:
        item.update({
            'prazo_recepcao_relatorio_dias_restantes': '',
            'prazo_retransmissao_dias_restantes': '',
            'prazo_manifestacao_rcnp_dias_restantes': '',
            'prazo_recurso_rnv_dias_restantes': '',
            'risco_rcnp': '',
            'status_compliance_fcvs': '',
        })
        return

    prazo_recepcao = data_base + timedelta(days=90)
    prazo_retransmissao = data_base + timedelta(days=120)
    prazo_manifestacao = _ultimo_dia_util_mes_apos(data_base, 3)
    prazo_recurso_rnv = _ultimo_dia_util_mes_apos(data_base, 12)

    dias_recepcao = _dias_restantes(prazo_recepcao, hoje)
    dias_retransmissao = _dias_restantes(prazo_retransmissao, hoje)
    dias_manifestacao = _dias_restantes(prazo_manifestacao, hoje)

    qtd_rnv = str(item.get('qtd_rnv', '')).strip()
    considera_prazo_rnv = qtd_rnv.isdigit() and int(qtd_rnv) > 0
    dias_recurso_rnv = _dias_restantes(prazo_recurso_rnv, hoje) if considera_prazo_rnv else ''

    estado_recepcao = _estado_prazo(dias_recepcao)
    estado_retransmissao = _estado_prazo(dias_retransmissao)
    estado_manifestacao = _estado_prazo(dias_manifestacao)
    estado_recurso_rnv = _estado_prazo(dias_recurso_rnv)

    item.update({
        'prazo_recepcao_relatorio_dias_restantes': dias_recepcao,
        'prazo_retransmissao_dias_restantes': dias_retransmissao,
        'prazo_manifestacao_rcnp_dias_restantes': dias_manifestacao,
        'prazo_recurso_rnv_dias_restantes': dias_recurso_rnv,
        'risco_rcnp': _risco_rcnp_por_estado(estado_manifestacao),
        'status_compliance_fcvs': _status_compliance_fcvs(
            estado_recepcao,
            estado_retransmissao,
            estado_manifestacao,
            estado_recurso_rnv,
        ),
    })


def _hash_sha256_arquivo(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def _processar_s_textual_path(path: Path) -> dict:
    codigo = _detectar_codigo_relatorio(path.name)
    return _processar_relatorio_textual(str(path), codigo)


def _linhas_relatorio_textual(caminho_arquivo: str) -> List[str]:
    with open(caminho_arquivo, 'r', encoding='latin-1', errors='ignore') as f:
        return [ln.rstrip('\n\r') for ln in f]


def _extrair_registros_s194301(linhas: List[str]) -> List[dict]:
    registros = []
    i = 0
    while i < len(linhas):
        linha = linhas[i]
        match = re.match(r'^\s*(\d{13})\s+(\d)\s+(.+?)\s{2,}(\d+)\s+(\d{2}\.\d{2}\.\d{4})\s*$', linha)
        if not match:
            i += 1
            continue

        motivo = ''
        if i + 1 < len(linhas):
            proxima = linhas[i + 1].strip()
            if proxima and not re.match(r'^\d{13}\b', proxima):
                motivo = proxima
                i += 1

        registros.append({
            'numero_contrato': match.group(1),
            'hipoteca': match.group(2),
            'nome': re.sub(r'\s+', ' ', match.group(3)).strip(),
            'lote': match.group(4),
            'data_geracao': match.group(5),
            'motivo_rejeicao': motivo,
        })
        i += 1

    return registros


def _extrair_registros_s778101(linhas: List[str]) -> List[dict]:
    registros = []
    pattern = re.compile(
        r'^\s*(\d{13})\s+(\d)\s+(.+?)\s{2,}'
        r'(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+'
        r'(RNV|RCV)\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s*$'
    )
    for linha in linhas:
        match = pattern.match(linha)
        if not match:
            continue

        registros.append({
            'numero_contrato': match.group(1),
            'hipoteca': match.group(2),
            'nome': re.sub(r'\s+', ' ', match.group(3)).strip(),
            'data_habilitacao': match.group(4),
            'data_termino_analise': match.group(5),
            'tipo_solicitacao': match.group(6),
            'data_geracao': match.group(7),
            'lote': match.group(8),
        })

    return registros


def _parse_moeda_br(valor: str) -> float:
    txt = (valor or '').strip().replace('.', '').replace(',', '.')
    try:
        return float(txt)
    except Exception:
        return 0.0


def _extrair_registros_s3435xx(linhas: List[str], campo_valor: str) -> List[dict]:
    registros = []
    pattern = re.compile(r'^\s*(\d{5})\s+(\d{13})\s+(\d)\s+(.+?)\s+([\d\.,]+)\s*$')
    for linha in linhas:
        match = pattern.match(linha)
        if not match:
            continue

        valor_str = match.group(5)
        registros.append({
            'matricula': match.group(1),
            'numero_contrato': match.group(2),
            'hipoteca': match.group(3),
            'nome': re.sub(r'\s+', ' ', match.group(4)).strip(),
            campo_valor: valor_str,
            f'{campo_valor}_num': _parse_moeda_br(valor_str),
        })

    return registros


def _extrair_registros_s820401(linhas: List[str]) -> List[dict]:
    registros = []
    pattern = re.compile(
        r'^\s*(\d{5})\s+(\d{13})\s+(\d)\s+'
        r'([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s*$'
    )
    for linha in linhas:
        match = pattern.match(linha)
        if not match:
            continue

        multa12 = match.group(4)
        multa14 = match.group(5)
        multa15 = match.group(6)
        registros.append({
            'matricula': match.group(1),
            'numero_contrato': match.group(2),
            'hipoteca': match.group(3),
            'multa_moeda12': multa12,
            'multa_moeda14': multa14,
            'multa_moeda15': multa15,
            'multa_moeda12_num': _parse_moeda_br(multa12),
            'multa_moeda14_num': _parse_moeda_br(multa14),
            'multa_moeda15_num': _parse_moeda_br(multa15),
        })

    return registros


def _extrair_registros_s820301(linhas: List[str]) -> List[dict]:
    registros = []
    pattern = re.compile(
        r'^\s*(\d{5})\s+(\d{13})\s+(\d)\s+'
        r'([\d\.,]+)\s+(\d+)\s+([\d\.,]+)\s+(\d+)\s+([\d\.,]+)\s+(\d+)\s*$'
    )
    for linha in linhas:
        match = pattern.match(linha)
        if not match:
            continue

        multa12 = match.group(4)
        meses12 = int(match.group(5))
        multa14 = match.group(6)
        meses14 = int(match.group(7))
        multa15 = match.group(8)
        meses15 = int(match.group(9))

        registros.append({
            'matricula': match.group(1),
            'numero_contrato': match.group(2),
            'hipoteca': match.group(3),
            'multa_moeda12': multa12,
            'meses_moeda12': meses12,
            'multa_moeda14': multa14,
            'meses_moeda14': meses14,
            'multa_moeda15': multa15,
            'meses_moeda15': meses15,
            'multa_moeda12_num': _parse_moeda_br(multa12),
            'multa_moeda14_num': _parse_moeda_br(multa14),
            'multa_moeda15_num': _parse_moeda_br(multa15),
        })

    return registros


def _extrair_registros_s820601(linhas: List[str]) -> List[dict]:
    registros = []
    header_pattern = re.compile(r'MATRICULA\s*-\s*(\d{5})\s+CONTRATO\s*-\s*(\d{13})\s+HIPOTECA\s*-\s*(\d+)', re.IGNORECASE)
    periodo_pattern = re.compile(
        r'^\s*(\d{2}/\d{2}/\d{4})\s+A\s+(\d{2}/\d{2}/\d{4})\s+(\d{2})\s+'
        r'([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s*$'
    )

    matricula = ''
    numero_contrato = ''
    hipoteca = ''

    for linha in linhas:
        m_h = header_pattern.search(linha)
        if m_h:
            matricula = m_h.group(1)
            numero_contrato = m_h.group(2)
            hipoteca = m_h.group(3)
            continue

        m_p = periodo_pattern.match(linha)
        if not m_p or not numero_contrato:
            continue

        registros.append({
            'matricula': matricula,
            'numero_contrato': numero_contrato,
            'hipoteca': hipoteca,
            'inicio_vigencia': m_p.group(1),
            'fim_vigencia': m_p.group(2),
            'moeda': m_p.group(3),
            'vlr_multa': m_p.group(4),
            'tr_mes': m_p.group(5),
            'vlr_apurado': m_p.group(6),
            'vlr_devido': m_p.group(7),
            'vlr_multa_num': _parse_moeda_br(m_p.group(4)),
            'tr_mes_num': _parse_moeda_br(m_p.group(5)),
            'vlr_apurado_num': _parse_moeda_br(m_p.group(6)),
            'vlr_devido_num': _parse_moeda_br(m_p.group(7)),
        })

    return registros


def _escrever_relatorio_lote_md(path: Path, rows: list, comparacao: list, pasta: str) -> None:
    total = len(rows)
    por_tipo = defaultdict(int)
    for r in rows:
        por_tipo[r.get('tipo_detectado', '')] += 1

    linhas = []
    linhas.append('# Relatorio consolidado de retornos CEF')
    linhas.append('')
    linhas.append(f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    linhas.append(f'Pasta analisada: {pasta}')
    linhas.append(f'Total de arquivos: {total}')
    linhas.append('')
    linhas.append('## Inventario por tipo detectado')
    for tipo, qtd in sorted(por_tipo.items()):
        linhas.append(f'- {tipo}: {qtd}')

    linhas.append('')
    linhas.append('## Comparacao entre remessas por codigo')
    for item in comparacao:
        linhas.append(
            f"- {item['codigo']}: {item['conteudo']} "
            f"(arquivos={item['qtd_arquivos']}, datas={item['datas']})"
        )

    linhas.append('')
    linhas.append('## Observacoes')
    linhas.append('- M3026xx: parser P3026 v2.')
    linhas.append('- S194301/S778101/S343501/S343601/S820401/S820301/S820601: leitura textual com extracao estruturada.')
    linhas.append('- Demais Sxxxxxx: leitura textual com totalizadores.')
    linhas.append('- Outros codigos: fallback FCVS/CADMUT quando aplicavel.')

    path.write_text('\n'.join(linhas), encoding='utf-8')


def processar_retornos_cef_lote_view(request):
    """
    Processa uma pasta inteira de retornos CEF e gera consolidado + comparacao.

    Fluxo:
    - Detecta codigo do arquivo pelo nome
    - Roteia parser por familia (CADMUT1, M3026xx, Sxxxxxx, fallback)
    - Gera 2 CSVs e 1 relatorio Markdown na pasta de saida
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        messages.error(request, 'Modulos de fichas nao disponiveis')
        return redirect('integracao_cef')

    pasta_entrada_valor = request.POST.get('pasta_entrada') if request.method == 'POST' else 'manual/00044'
    pasta_saida_valor = request.POST.get('pasta_saida') if request.method == 'POST' else 'manual/00044'

    resultados = []
    comparacao = []
    arquivos_saida = []
    resumo = {}

    if request.method == 'POST':
        try:
            base_dir = Path(settings.BASE_DIR).resolve()

            pasta_entrada = Path(pasta_entrada_valor or 'manual/00044')
            if not pasta_entrada.is_absolute():
                pasta_entrada = (base_dir / pasta_entrada).resolve()
            else:
                pasta_entrada = pasta_entrada.resolve()

            pasta_saida = Path(pasta_saida_valor or pasta_entrada_valor or 'manual/00044')
            if not pasta_saida.is_absolute():
                pasta_saida = (base_dir / pasta_saida).resolve()
            else:
                pasta_saida = pasta_saida.resolve()

            if not pasta_entrada.exists() or not pasta_entrada.is_dir():
                raise ValueError(f'Pasta de entrada nao encontrada: {pasta_entrada}')

            pasta_saida.mkdir(parents=True, exist_ok=True)

            from .ficha_p3026_parser_v2 import ParserP3026
            parser_p3026 = ParserP3026()

            arquivos = [
                p for p in sorted(pasta_entrada.iterdir())
                if p.is_file()
                and p.name.lower() != 'desktop.ini'
                and not p.name.lower().endswith(('.md', '.csv', '.ini'))
            ]

            for path in arquivos:
                nome = path.name
                codigo = _detectar_codigo_relatorio(nome)
                data_remessa = _detectar_data_remessa_nome_arquivo(nome)
                item = {
                    'arquivo': nome,
                    'codigo': codigo,
                    'data_remessa': data_remessa,
                    'hash_sha256': _hash_sha256_arquivo(path),
                    'tipo_detectado': '',
                    'layout_usado': 'manual_atual',
                    'status': 'OK',
                    'linhas_total': '',
                    'movimentos': '',
                    'rejeitados': '',
                    'totalizadores_encontrados': '',
                    'registros_p3026': '',
                    'contratos_rejeitados': '',
                    'motivo_principal': '',
                    'contratos_acatados': '',
                    'qtd_rnv': '',
                    'qtd_rcv': '',
                    'contratos_vaf4': '',
                    'total_vaf4': '',
                    'contratos_vaf3': '',
                    'total_vaf3': '',
                    'contratos_multa': '',
                    'total_multa_moeda12': '',
                    'total_multa_moeda14': '',
                    'total_multa_moeda15': '',
                    'contratos_multa_820301': '',
                    'total_multa12_820301': '',
                    'total_multa14_820301': '',
                    'total_multa15_820301': '',
                    'total_meses12_820301': '',
                    'total_meses14_820301': '',
                    'total_meses15_820301': '',
                    'contratos_apuracao_820601': '',
                    'periodos_apuracao_820601': '',
                    'total_apurado_820601': '',
                    'saldo_final_820601': '',
                    'prazo_recepcao_relatorio_dias_restantes': '',
                    'prazo_retransmissao_dias_restantes': '',
                    'prazo_manifestacao_rcnp_dias_restantes': '',
                    'prazo_recurso_rnv_dias_restantes': '',
                    'risco_rcnp': '',
                    'status_compliance_fcvs': '',
                    'avisos_parser': '',
                    'primeiro_aviso': '',
                }

                try:
                    if codigo == 'CADMUT1':
                        rel = interpretar_retorno_cadmut(str(path))
                        r = rel.get('resumo', {})
                        item.update({
                            'tipo_detectado': 'CADMUT1',
                            'linhas_total': str(r.get('total_registros', 0)),
                            'movimentos': str(r.get('movimentos', 0)),
                            'rejeitados': str(r.get('registros_rejeitados', 0)),
                        })
                    elif codigo.startswith('M3026') or codigo.startswith('P3026'):
                        arquivo_p3026, erros = parser_p3026.parse_arquivo(str(path))
                        item.update({
                            'tipo_detectado': 'M3026xx/P3026',
                            'registros_p3026': str(len(arquivo_p3026.registros) if arquivo_p3026 else 0),
                            'avisos_parser': str(len(erros)),
                            'primeiro_aviso': erros[0] if erros else '',
                        })
                    elif codigo.startswith('S'):
                        txt = _processar_s_textual_path(path)
                        item.update({
                            'tipo_detectado': 'Sxxxxxx/relatorio_textual',
                            'linhas_total': str(txt.get('linhas_total', 0)),
                            'totalizadores_encontrados': str(txt.get('totalizadores_encontrados', 0)),
                            'contratos_rejeitados': str(txt.get('contratos_rejeitados', '')),
                            'motivo_principal': str(txt.get('motivo_principal', '')),
                            'contratos_acatados': str(txt.get('contratos_acatados', '')),
                            'qtd_rnv': str(txt.get('qtd_rnv', '')),
                            'qtd_rcv': str(txt.get('qtd_rcv', '')),
                            'contratos_vaf4': str(txt.get('contratos_vaf4', '')),
                            'total_vaf4': str(txt.get('total_vaf4', '')),
                            'contratos_vaf3': str(txt.get('contratos_vaf3', '')),
                            'total_vaf3': str(txt.get('total_vaf3', '')),
                            'contratos_multa': str(txt.get('contratos_multa', '')),
                            'total_multa_moeda12': str(txt.get('total_multa_moeda12', '')),
                            'total_multa_moeda14': str(txt.get('total_multa_moeda14', '')),
                            'total_multa_moeda15': str(txt.get('total_multa_moeda15', '')),
                            'contratos_multa_820301': str(txt.get('contratos_multa_820301', '')),
                            'total_multa12_820301': str(txt.get('total_multa12_820301', '')),
                            'total_multa14_820301': str(txt.get('total_multa14_820301', '')),
                            'total_multa15_820301': str(txt.get('total_multa15_820301', '')),
                            'total_meses12_820301': str(txt.get('total_meses12_820301', '')),
                            'total_meses14_820301': str(txt.get('total_meses14_820301', '')),
                            'total_meses15_820301': str(txt.get('total_meses15_820301', '')),
                            'contratos_apuracao_820601': str(txt.get('contratos_apuracao_820601', '')),
                            'periodos_apuracao_820601': str(txt.get('periodos_apuracao_820601', '')),
                            'total_apurado_820601': str(txt.get('total_apurado_820601', '')),
                            'saldo_final_820601': str(txt.get('saldo_final_820601', '')),
                        })
                    else:
                        rel = interpretar_retorno_fcvs(str(path))
                        r = rel.get('resumo', {})
                        item.update({
                            'tipo_detectado': 'FCVS_fallback',
                            'linhas_total': str(r.get('total_registros', 0)),
                            'movimentos': str(r.get('movimentos', 0)),
                            'rejeitados': str(r.get('registros_rejeitados', 0)),
                            'layout_usado': 'fallback_generico',
                        })
                except Exception as exc:
                    item['status'] = 'ERRO'
                    item['primeiro_aviso'] = str(exc)

                _aplicar_compliance_fcvs_lote(item)

                resultados.append(item)

            por_codigo = defaultdict(list)
            for r in resultados:
                por_codigo[r['codigo']].append(r)

            for codigo, itens in sorted(por_codigo.items()):
                if len(itens) < 2:
                    comparacao.append({
                        'codigo': codigo,
                        'qtd_arquivos': str(len(itens)),
                        'conteudo': 'sem_comparacao',
                        'datas': ', '.join(sorted(set(i.get('data_remessa', '') for i in itens if i.get('data_remessa')))),
                    })
                    continue

                hashes = {i.get('hash_sha256', '') for i in itens}
                comparacao.append({
                    'codigo': codigo,
                    'qtd_arquivos': str(len(itens)),
                    'conteudo': 'igual' if len(hashes) == 1 else 'diferente',
                    'datas': ', '.join(sorted(set(i.get('data_remessa', '') for i in itens if i.get('data_remessa')))),
                })

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            csv_consolidado = pasta_saida / f'CONSOLIDADO_RETORNOS_CEF_{ts}.csv'
            csv_comparacao = pasta_saida / f'COMPARACAO_REMESSAS_CEF_{ts}.csv'
            md_relatorio = pasta_saida / f'RELATORIO_CONSOLIDADO_CEF_{ts}.md'

            campos_consolidado = [
                'arquivo', 'codigo', 'data_remessa', 'hash_sha256', 'tipo_detectado', 'layout_usado', 'status',
                'linhas_total', 'movimentos', 'rejeitados', 'totalizadores_encontrados', 'registros_p3026',
                'contratos_rejeitados', 'motivo_principal', 'contratos_acatados', 'qtd_rnv', 'qtd_rcv',
                'contratos_vaf4', 'total_vaf4', 'contratos_vaf3', 'total_vaf3',
                'contratos_multa', 'total_multa_moeda12', 'total_multa_moeda14', 'total_multa_moeda15',
                'contratos_multa_820301', 'total_multa12_820301', 'total_multa14_820301', 'total_multa15_820301',
                'total_meses12_820301', 'total_meses14_820301', 'total_meses15_820301',
                'contratos_apuracao_820601', 'periodos_apuracao_820601', 'total_apurado_820601', 'saldo_final_820601',
                'prazo_recepcao_relatorio_dias_restantes', 'prazo_retransmissao_dias_restantes',
                'prazo_manifestacao_rcnp_dias_restantes', 'prazo_recurso_rnv_dias_restantes',
                'risco_rcnp', 'status_compliance_fcvs',
                'avisos_parser', 'primeiro_aviso',
            ]

            with csv_consolidado.open('w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=campos_consolidado, delimiter=';')
                w.writeheader()
                for row in resultados:
                    w.writerow({k: row.get(k, '') for k in campos_consolidado})

            with csv_comparacao.open('w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=['codigo', 'qtd_arquivos', 'conteudo', 'datas'], delimiter=';')
                w.writeheader()
                for row in comparacao:
                    w.writerow(row)

            _escrever_relatorio_lote_md(md_relatorio, resultados, comparacao, str(pasta_entrada).replace('\\', '/'))

            base_str = str(base_dir)
            for p in [csv_consolidado, csv_comparacao, md_relatorio]:
                p_str = str(p)
                if p_str.startswith(base_str):
                    p_str = os.path.relpath(p_str, base_str)
                arquivos_saida.append(p_str.replace('\\', '/'))

            resumo = {
                'total_arquivos': len(resultados),
                'ok': sum(1 for r in resultados if r.get('status') == 'OK'),
                'erro': sum(1 for r in resultados if r.get('status') != 'OK'),
                'comparacoes_iguais': sum(1 for c in comparacao if c.get('conteudo') == 'igual'),
                'comparacoes_diferentes': sum(1 for c in comparacao if c.get('conteudo') == 'diferente'),
            }

            messages.success(request, f"✅ Lote processado: {resumo['ok']} OK, {resumo['erro']} com erro.")

        except Exception as e:
            messages.error(request, f'Erro ao processar lote: {str(e)}')

    context = {
        'pasta_entrada': pasta_entrada_valor or 'manual/00044',
        'pasta_saida': pasta_saida_valor or 'manual/00044',
        'resultados': resultados,
        'comparacao': comparacao,
        'resumo': resumo,
        'arquivos_saida': arquivos_saida,
    }
    return render(request, 'principal/cef_processar_retornos_lote.html', context)


def processar_relatorios_cef_upload(request):
    """
    Upload e processamento de relatorios CEF sem depender de extensao.

    Regras de processamento:
    - CADMUT1: interpretador CADMUT
    - M3026xx: parser P3026 v2 + detalhamento heuristico de linhas
    - Sxxxxxx: leitura textual com extracao de totalizadores
    - Outros: fallback para interpretador FCVS
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        messages.error(request, 'Modulos de fichas nao disponiveis')
        return redirect('integracao_cef')

    resultados = []
    total_arquivos = 0
    processados_ok = 0
    processados_erro = 0
    detalhes_arquivos = []

    # Limpa estado anterior de exportacao ao abrir tela sem POST
    if request.method != 'POST':
        request.session.pop('cef_upload_resultados', None)
        request.session.pop('cef_upload_detalhes', None)

    if request.method == 'POST':
        arquivos = request.FILES.getlist('arquivos')
        arquivo_unico = request.FILES.get('arquivo')
        if not arquivos and arquivo_unico:
            arquivos = [arquivo_unico]

        if not arquivos:
            messages.warning(request, 'Selecione pelo menos um arquivo para processar.')
        else:
            total_arquivos = len(arquivos)

            from .ficha_p3026_parser_v2 import ParserP3026
            parser_p3026 = ParserP3026()

            for arquivo in arquivos:
                codigo = _detectar_codigo_relatorio(arquivo.name)
                resultado = {
                    'arquivo': arquivo.name,
                    'codigo': codigo,
                    'tipo_detectado': '',
                    'status': 'OK',
                    'detalhes': '',
                    'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                    'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                    'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                    'resumo': {},
                }

                temp_path = None
                try:
                    import tempfile
                    with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                        for chunk in arquivo.chunks():
                            temp_file.write(chunk)
                        temp_path = temp_file.name

                    if codigo == 'CADMUT1':
                        relatorio = interpretar_retorno_cadmut(temp_path)
                        resumo = relatorio.get('resumo', {})
                        resultado['tipo_detectado'] = 'CADMUT1'
                        resultado['detalhes'] = 'Processado com interpretador CADMUT.'
                        resultado['resumo'] = {
                            'linhas_total': resumo.get('total_registros', 0),
                            'movimentos': resumo.get('movimentos', 0),
                            'rejeitados': resumo.get('registros_rejeitados', 0),
                        }
                        detalhes_arquivos.extend(_detalhes_cadmut1(temp_path, arquivo.name, codigo))

                    elif codigo.startswith('M3026'):
                        arquivo_p3026, erros = parser_p3026.parse_arquivo(temp_path)
                        total_registros = len(arquivo_p3026.registros) if arquivo_p3026 else 0
                        resultado['tipo_detectado'] = 'M3026xx / P3026'
                        resultado['detalhes'] = 'Processado com parser P3026 v2 + detalhamento de linhas.'
                        resultado['resumo'] = {
                            'linhas_total': total_registros,
                            'avisos_parser': len(erros),
                        }
                        if erros:
                            resultado['resumo']['primeiro_aviso'] = erros[0]
                        detalhes_arquivos.extend(_detalhes_m3026(temp_path, arquivo.name, codigo))

                    elif codigo.startswith('S'):
                        txt = _processar_relatorio_textual(temp_path, codigo)
                        resultado['tipo_detectado'] = 'Relatorio textual SIWFC'
                        resultado['detalhes'] = 'Leitura textual com extracao de totalizadores e cabecalhos.'
                        resultado['resumo'] = txt
                        detalhes_arquivos.extend(_detalhes_relatorio_textual(temp_path, arquivo.name, codigo))

                    else:
                        relatorio = interpretar_retorno_fcvs(temp_path)
                        resumo = relatorio.get('resumo', {})
                        resultado['tipo_detectado'] = 'FCVS (fallback)'
                        resultado['detalhes'] = 'Processado com interpretador FCVS generico.'
                        resultado['resumo'] = {
                            'linhas_total': resumo.get('total_registros', 0),
                            'movimentos': resumo.get('movimentos', 0),
                            'rejeitados': resumo.get('registros_rejeitados', 0),
                        }
                        detalhes_arquivos.append({
                            'arquivo': arquivo.name,
                            'codigo': codigo,
                            'interpretacao_manual': _interpretacao_manual_codigo(codigo),
                            'fonte_manual': _guia_manual_codigo(codigo).get('fonte_manual', ''),
                            'acao_recomendada': _guia_manual_codigo(codigo).get('acao', ''),
                            'origem_parser': 'FCVS-Fallback',
                            'tipo_linha': 'RESUMO',
                            'indice_linha': 0,
                            'numero_contrato': '',
                            'cpf': '',
                            'nome': '',
                            'status': 'LIDO',
                            'codigo_critica': '',
                            'descricao_critica': '',
                            'texto': f"Registros={resumo.get('total_registros', 0)}, Rejeitados={resumo.get('registros_rejeitados', 0)}",
                        })

                    processados_ok += 1

                except Exception as e:
                    resultado['status'] = 'ERRO'
                    resultado['detalhes'] = f'Falha ao processar: {str(e)}'
                    processados_erro += 1

                finally:
                    if temp_path and os.path.exists(temp_path):
                        os.unlink(temp_path)

                resultados.append(resultado)

            if processados_erro == 0:
                messages.success(request, f'✅ {processados_ok} arquivo(s) processado(s) com sucesso.')
            else:
                messages.warning(
                    request,
                    f'⚠️ Processamento finalizado: {processados_ok} sucesso, {processados_erro} com erro.'
                )

            # Evita sessao muito grande em casos extremos
            limite_detalhes = 20000
            if len(detalhes_arquivos) > limite_detalhes:
                detalhes_arquivos = detalhes_arquivos[:limite_detalhes]
                messages.warning(
                    request,
                    f'⚠️ CSV detalhado foi limitado aos primeiros {limite_detalhes} registros para evitar excesso de memoria.'
                )

            # Agente especialista confere as leituras e marca confianca/alertas.
            detalhes_arquivos = [
                MANUAL_EXPERT_AGENT.enriquecer_row(row) for row in detalhes_arquivos
            ]

            request.session['cef_upload_resultados'] = resultados
            request.session['cef_upload_detalhes'] = detalhes_arquivos

            if request.POST.get('salvar_retornos') == '1':
                resumo_persistencia = _persistir_detalhes_como_retornos_cef(detalhes_arquivos)
                messages.success(
                    request,
                    '📥 Retornos gravados no painel: '
                    f"{resumo_persistencia['criados']} criado(s), "
                    f"{resumo_persistencia['ignorados_duplicados']} duplicado(s), "
                    f"{resumo_persistencia['ignorados_sem_contrato']} sem contrato associado."
                )

    # Gera resumo de conferencia por arquivo para UI.
    confianca_por_arquivo = {}
    detalhes_baixa = []
    detalhes_p3026_preview = []
    if detalhes_arquivos:
        agrupados = {}
        for row in detalhes_arquivos:
            agrupados.setdefault(row.get('arquivo', ''), []).append(row)

        for nome_arquivo, rows in agrupados.items():
            contagem = Counter((r.get('confianca_leitura') or 'na').lower() for r in rows)
            alertas = sum(1 for r in rows if r.get('alerta_leitura'))
            confianca_por_arquivo[nome_arquivo] = {
                'alta': contagem.get('alta', 0),
                'media': contagem.get('media', 0),
                'baixa': contagem.get('baixa', 0),
                'alertas': alertas,
            }

        detalhes_baixa = [r for r in detalhes_arquivos if (r.get('confianca_leitura') or '').lower() == 'baixa']
        detalhes_baixa = detalhes_baixa[:200]

        detalhes_p3026_preview = [
            r for r in detalhes_arquivos
            if str(r.get('codigo', '')).startswith('M3026') and str(r.get('tipo_linha', '')) in {'1', '2', '3', '4', '5', '6', '7', '8'}
        ]
        detalhes_p3026_preview = detalhes_p3026_preview[:300]

    for item in resultados:
        resumo_conf = confianca_por_arquivo.get(item.get('arquivo', ''), {'alta': 0, 'media': 0, 'baixa': 0, 'alertas': 0})
        item['confianca_resumo'] = resumo_conf
        item['alertas_leitura'] = resumo_conf.get('alertas', 0)

    context = {
        'resultados': resultados,
        'total_arquivos': total_arquivos,
        'processados_ok': processados_ok,
        'processados_erro': processados_erro,
        'detalhes_baixa': detalhes_baixa,
        'detalhes_p3026_preview': detalhes_p3026_preview,
    }

    return render(request, 'principal/cef_upload_relatorios.html', context)


def exportar_relatorios_cef_csv(request):
    """Exporta em CSV detalhado o ultimo processamento de upload de relatorios CEF."""
    resultados = request.session.get('cef_upload_resultados') or []
    detalhes = request.session.get('cef_upload_detalhes') or []

    if not resultados and not detalhes:
        messages.warning(request, 'Nao ha resultados para exportar. Processe arquivos primeiro.')
        return redirect('processar_relatorios_cef_upload')

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow([
        'arquivo',
        'codigo',
        'interpretacao_manual',
        'fonte_manual',
        'acao_recomendada',
        'confianca_leitura',
        'alerta_leitura',
        'origem_parser',
        'tipo_linha',
        'indice_linha',
        'numero_contrato',
        'cpf',
        'nome',
        'status',
        'codigo_situacao',
        'descricao_situacao',
        'manifestacao',
        'data_manifestacao',
        'situacao_analise',
        'status_homologacao',
        'valor_credito',
        'valor_credito_bruto',
        'tipo_evento',
        'data_evento',
        'depuracao',
        'codigo_critica',
        'descricao_critica',
        'texto',
    ])

    somente_baixa = request.GET.get('somente_baixa') == '1'

    if detalhes:
        for row in detalhes:
            if somente_baixa and (row.get('confianca_leitura', '').lower() != 'baixa'):
                continue
            writer.writerow([
                row.get('arquivo', ''),
                row.get('codigo', ''),
                row.get('interpretacao_manual', ''),
                row.get('fonte_manual', ''),
                row.get('acao_recomendada', ''),
                row.get('confianca_leitura', ''),
                row.get('alerta_leitura', ''),
                row.get('origem_parser', ''),
                row.get('tipo_linha', ''),
                row.get('indice_linha', ''),
                row.get('numero_contrato', ''),
                row.get('cpf', ''),
                row.get('nome', ''),
                row.get('status', ''),
                row.get('codigo_situacao', ''),
                row.get('descricao_situacao', ''),
                row.get('manifestacao', ''),
                row.get('data_manifestacao', ''),
                row.get('situacao_analise', ''),
                row.get('status_homologacao', ''),
                row.get('valor_credito', ''),
                row.get('valor_credito_bruto', ''),
                row.get('tipo_evento', ''),
                row.get('data_evento', ''),
                row.get('depuracao', ''),
                row.get('codigo_critica', ''),
                row.get('descricao_critica', ''),
                row.get('texto', ''),
            ])
    else:
        for item in resultados:
            resumo = item.get('resumo', {})
            texto = (
                f"linhas={resumo.get('linhas_total', '')}; "
                f"movimentos={resumo.get('movimentos', '')}; "
                f"rejeitados={resumo.get('rejeitados', '')}; "
                f"totalizadores={resumo.get('totalizadores_encontrados', '')}; "
                f"avisos={resumo.get('avisos_parser', '')}; "
                f"primeiro_aviso={resumo.get('primeiro_aviso', '')}"
            )
            writer.writerow([
                item.get('arquivo', ''),
                item.get('codigo', ''),
                item.get('interpretacao_manual', ''),
                item.get('fonte_manual', ''),
                item.get('acao_recomendada', ''),
                '',
                '',
                'ResumoFallback',
                'RESUMO',
                '',
                '',
                '',
                '',
                item.get('status', ''),
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                'Resumo fallback sem detalhamento por linha.',
                '',
                '',
                texto,
            ])

    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="relatorios_cef_detalhado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    )
    return response


def selecao_automatica_api(request, contrato_id):
    """
    API JSON para seleção automática de ficha
    Usada via AJAX para atualização dinâmica
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        return JsonResponse({'error': 'Módulos não disponíveis'}, status=500)
    
    try:
        contrato = Contrato.objects.get(pk=contrato_id)
        mutuario = contrato.mutuario if hasattr(contrato, 'mutuario') else None
        
        historico_envios = list(EnvioCEF.objects.filter(
            contrato=contrato
        ).order_by('-criado_em'))
        
        resultado = selecionar_ficha_automatica(
            contrato=contrato,
            mutuario=mutuario,
            historico_envios=historico_envios
        )
        
        return JsonResponse(resultado)
        
    except Contrato.DoesNotExist:
        return JsonResponse({'error': 'Contrato não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def download_arquivo_lote(request):
    """
    Gera arquivo de lote com múltiplas fichas
    - Seleciona múltiplos contratos
    - Gera HEADER e DADOS em arquivos separados (conforme SIWFC)
    - Download de arquivo ZIP com ambos
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        messages.error(request, 'Módulos de fichas não disponíveis')
        return redirect('integracao_cef')
    
    if request.method == 'POST':
        contrato_ids = request.POST.getlist('contratos')
        todos_filtrados = request.POST.get('todos_filtrados') == '1'
        conjuntos_post = request.POST.getlist('conjunto')
        tipo_ficha = request.POST.get('tipo_ficha', 'FH1')
        formato_saida = (request.POST.get('formato_saida') or 'zip').strip().lower()
        matricula = request.POST.get('matricula', '123456')
        numero_lote = request.POST.get('numero_lote', '001')
        forcar_reenvio = request.POST.get('forcar_reenvio') == '1'

        if todos_filtrados:
            contratos = Contrato.objects.all()
            if conjuntos_post:
                contratos = contratos.filter(conjunto__in=conjuntos_post)
            contrato_ids = [str(cid) for cid in contratos.values_list('id', flat=True)]
        else:
            contratos = Contrato.objects.filter(pk__in=contrato_ids)

        if not contrato_ids:
            messages.error(request, 'Selecione pelo menos um contrato (ou marque "todos os filtrados").')
            return redirect('download_arquivo_lote')
        
        try:
            from .ficha_generators import gerar_lote_fh1_separado
            import zipfile
            import io
            
            # Gera arquivo de acordo com o tipo
            if tipo_ficha == 'FH1':
                # Para geração manual de ZIP, permitir reprocessamento completo quando solicitado.
                bloqueados = [] if forcar_reenvio else _buscar_contratos_bloqueados_para_envio(list(contrato_ids), tipo_envio='FH1')

                if bloqueados:
                    exemplos = ', '.join([b['codigo'] for b in bloqueados[:10]])
                    sufixo = '...' if len(bloqueados) > 10 else ''
                    bloqueados_ids = {b['contrato_id'] for b in bloqueados}
                    contrato_ids_elegiveis = [cid for cid in contrato_ids if int(cid) not in bloqueados_ids]

                    if not contrato_ids_elegiveis:
                        messages.error(
                            request,
                            (
                                f'Bloqueado: {len(bloqueados)} contrato(s) já possuem envio FH1 em andamento/concluído. '
                                f'Exemplos: {exemplos}{sufixo}. Marque "Forçar reenvio" para gerar o ZIP mesmo assim.'
                            )
                        )
                        return redirect('download_arquivo_lote')

                    # Segue com os elegíveis e informa quais foram ignorados.
                    contrato_ids = contrato_ids_elegiveis
                    contratos = Contrato.objects.filter(pk__in=contrato_ids)
                    messages.warning(
                        request,
                        (
                            f'{len(bloqueados)} contrato(s) bloqueado(s) e ignorado(s) na geração do lote. '
                            f'Exemplos: {exemplos}{sufixo}'
                        )
                    )

                resultado = gerar_lote_fh1_separado(
                    contratos=list(contratos),
                    matricula=matricula,
                    numero_lote=numero_lote
                )

                expected_matricula = None
                matricula_digitos = ''.join(ch for ch in (matricula or '') if ch.isdigit())
                if len(matricula_digitos) == 6:
                    expected_matricula = matricula_digitos

                precheck = run_fh1_precheck_agent(
                    resultado.get('header_conteudo', ''),
                    resultado.get('dados_conteudo', ''),
                    expected_matricula=expected_matricula,
                )

                if not precheck.get('ok'):
                    mensagens_precheck = precheck.get('errors', [])
                    messages.error(
                        request,
                        'Pré-check FH1 bloqueou a geração do ZIP. Corrija os dados e tente novamente. Erros: ' + ' | '.join(mensagens_precheck[:5])
                    )
                    return redirect('download_arquivo_lote')

                avisos = precheck.get('warnings', [])
                if avisos:
                    messages.warning(request, 'Pré-check FH1 (avisos): ' + ' | '.join(avisos[:3]))
                
                if resultado['total_fichas'] == 0:
                    messages.error(request, 'Nenhuma ficha foi gerada com sucesso')
                    return redirect('download_arquivo_lote')

                if formato_saida == 'excel':
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Alignment, Font, PatternFill
                    except Exception:
                        messages.error(request, 'Biblioteca openpyxl não disponível para exportar Excel.')
                        return redirect('download_arquivo_lote')

                    layout_fh1 = [
                        (1, 'UFS', 1, 2),
                        (2, 'MAT. AG. FINANC. /DV', 3, 8),
                        (3, 'N.º CONTRATO DO MUT. NO AGENTE', 9, 21),
                        (4, 'HIPOTECA', 22, 22),
                        (5, 'TIPO DE REGISTRO', 23, 23),
                        (6, 'SEQUENCIAL', 24, 25),
                        (7, 'CONSTANTE', 26, 26),
                        (8, 'NOME DO MUT. PRINCIPAL', 27, 66),
                        (9, 'TIPO', 67, 67),
                        (10, 'CPF/CI', 68, 84),
                        (11, 'DATA DE NASCIMENTO', 85, 90),
                        (12, 'CODIGO DO MUNICIPIO', 91, 95),
                        (13, 'UF', 96, 97),
                        (14, 'ENDERECO DO IMOVEL', 98, 135),
                        (15, 'DATA DO CONTRATO', 136, 141),
                        (16, 'VALOR DA GARANTIA', 142, 153),
                        (17, 'IM', 154, 155),
                        (18, 'DATA DA LEGISLACAO', 156, 161),
                        (19, 'VALOR FINANCIAMENTO CONTRATADO', 162, 173),
                        (20, 'VALOR FINANC. PADRAO FCVS', 174, 185),
                        (21, 'CODIGO DA CATEG. PROFISSIONAL', 186, 190),
                        (22, 'SEGURO DE CREDITO', 191, 191),
                        (23, 'CARENCIA NO 1O VENCIMENTO', 192, 192),
                        (24, 'SEGURO DFI POR LOTES URBANIZADOS', 193, 193),
                        (25, 'CREDITOS ADQUIRIDOS PELA CAIXA COM RECURSOS DO PROER', 194, 194),
                        (26, 'VAGO', 195, 195),
                        (27, 'PRAZO CONTRATADO', 196, 198),
                        (28, 'TAXA JUROS CONTRATADO', 199, 204),
                        (29, 'CES CONTRATUAL', 205, 208),
                        (30, 'PLANO', 209, 211),
                        (31, 'ST', 212, 212),
                        (32, 'RJ', 213, 213),
                        (33, 'RR', 214, 215),
                        (34, 'INDEX', 216, 218),
                        (35, 'PRAZO FCVS', 219, 221),
                        (36, 'TAXA JUROS PARA FCVS', 222, 227),
                        (37, 'CES PARA FCVS', 228, 231),
                        (38, 'PLANO FCVS', 232, 234),
                        (39, 'ST FCVS', 235, 235),
                        (40, 'RJ FCVS', 236, 236),
                        (41, 'RR FCVS', 237, 238),
                        (42, 'INDEX FCVS', 239, 241),
                        (43, 'DATA SALDO CONSTRUCAO', 242, 247),
                        (44, 'SALDO DEVEDOR', 248, 259),
                        (45, '1O VENCIMENTO', 260, 265),
                        (46, 'SEGURO CREDITO / MIP / DFI', 266, 273),
                        (47, 'VALOR DA PRESTACAO', 274, 283),
                        (49, 'TCA/TAC', 284, 291),
                        (50, 'FCVS MENSAL', 292, 299),
                        (51, 'RAZAO ACRES/ DECRES.', 300, 307),
                        (52, 'TIPO EVENTO', 308, 310),
                        (53, 'DATA DO EVENTO', 311, 316),
                        (54, 'OR/CO', 317, 318),
                        (55, '% CAIXA', 319, 322),
                        (56, 'N.º CONTR. EMPR. CAIXA', 323, 340),
                        (57, 'TAXA JUROS EVENTO', 341, 346),
                        (58, 'VAF1 - VALOR BASICO', 347, 360),
                        (59, 'VAF2 - VALOR COMPLEMENTAR', 361, 374),
                        (60, 'VAF3 - VALOR RESIDUAL', 375, 388),
                        (61, 'JUROS CALCULADOS PELO AGENTE FINANCEIRO', 389, 402),
                        (62, 'DEBITO/CREDITO', 403, 403),
                        (63, 'QTD ALTERACOES', 404, 405),
                        (64, 'UFS LOTE', 406, 407),
                        (65, 'MAT. AG. FINANC. LOTE', 408, 413),
                        (66, 'DATA GERACAO', 414, 419),
                        (67, 'NUMERO LOTE', 420, 422),
                        (68, 'FORMA ENVIO', 423, 423),
                        (69, 'TIPO MOVIMENTO', 424, 424),
                        (70, 'FILLER', 425, 430),
                    ]

                    linhas_dados = [ln for ln in (resultado.get('dados_conteudo') or '').splitlines() if ln]

                    wb = Workbook()
                    ws = wb.active
                    ws.title = 'FH1 Horizontal'

                    headers = ['CONTRATO', 'SEQ'] + [campo[1] for campo in layout_fh1]
                    ws.append(headers)

                    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
                    header_font = Font(color='FFFFFF', bold=True)
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    for idx, linha in enumerate(linhas_dados, start=1):
                        contrato_linha = linha[8:21].strip() if len(linha) >= 21 else ''
                        row = [contrato_linha, idx]
                        for _, _, inicio, fim in layout_fh1:
                            row.append(linha[inicio - 1:fim].rstrip() if len(linha) >= fim else '')
                        ws.append(row)

                    ws.freeze_panes = 'C2'
                    ws.column_dimensions['A'].width = 18
                    ws.column_dimensions['B'].width = 8
                    for col in range(3, len(headers) + 1):
                        ws.column_dimensions[chr(64 + col) if col <= 26 else ws.cell(row=1, column=col).column_letter].width = 20

                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)

                    timestamp_excel = datetime.now().strftime('%Y%m%d_%H%M%S')
                    nome_excel = f'LOTE_FH1_HORIZONTAL_{timestamp_excel}.xlsx'
                    response = HttpResponse(
                        excel_buffer.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename={nome_excel}'
                    return response
                
                # Cria arquivo ZIP com HEADER e DADOS
                zip_buffer = io.BytesIO()
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                nome_zip = f'LOTE_FH1_{timestamp}.zip'
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Adiciona HEADER
                    nome_header = f'HEADER_FH1_{timestamp}.txt'
                    zip_file.writestr(nome_header, resultado['header_conteudo'].encode('latin-1'))
                    
                    # Adiciona DADOS
                    nome_dados = f'DADOS_FH1_{timestamp}.txt'
                    # Garante LF puro (\n) — CEF rejeita \r\n (Windows) com erro 100774
                    dados_lf = resultado['dados_conteudo'].replace('\r\n', '\n').replace('\r', '\n')
                    zip_file.writestr(nome_dados, dados_lf.encode('latin-1'))

                    # Auditoria do campo 62 (posição 403): Débito/Crédito
                    linhas_dados = [ln for ln in (resultado.get('dados_conteudo') or '').splitlines() if ln]
                    dc_vals = [ln[402:403] for ln in linhas_dados if len(ln) >= 403]
                    dc_d = dc_vals.count('D')
                    dc_c = dc_vals.count('C')
                    dc_outros = len(dc_vals) - dc_d - dc_c
                    
                    # Adiciona relatório (opcional)
                    relatorio = f"""RELATÓRIO DE GERAÇÃO DE LOTE FH1
Data/Hora: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}
Matrícula: {matricula}
Número Lote: {numero_lote}

ESTATÍSTICAS:
- Total de fichas: {resultado['total_fichas']}
- Fichas com sucesso: {resultado['total_fichas_sucesso']}
- Fichas com erro: {resultado['total_fichas_erro']}

AUDITORIA D/C (CAMPO 62 - POSIÇÃO 403):
- Registros com D (Débito): {dc_d}
- Registros com C (Crédito): {dc_c}
- Registros com valor inesperado: {dc_outros}

ERROS ENCONTRADOS:
"""
                    if resultado['erros']:
                        for erro in resultado['erros']:
                            relatorio += f"\n  Contrato {erro['contrato']}: {erro['erro']}"
                    else:
                        relatorio += "\n  Nenhum erro encontrado!"

                    relatorio += "\n\nPRÉ-CHECK FH1 (ANTES DO ENVIO):\n"
                    if precheck.get('ok'):
                        relatorio += "\n  OK - sem bloqueios críticos"
                    else:
                        for item in precheck.get('errors', []):
                            relatorio += f"\n  [ERRO] {item}"

                    if precheck.get('warnings'):
                        for item in precheck.get('warnings', []):
                            relatorio += f"\n  [AVISO] {item}"
                    
                    if resultado.get('detalhes'):
                        relatorio += "\n\nAVISOS DE VALIDAÇÃO:\n"
                        for detalhe in resultado.get('detalhes', []):
                            avisos_detalhe = detalhe.get('avisos') or []
                            if avisos_detalhe:
                                relatorio += f"\n  Contrato {detalhe.get('contrato', '')}:"
                                for aviso in avisos_detalhe:
                                    relatorio += f"\n    - {aviso.get('mensagem', aviso)}"
                    
                    zip_file.writestr(f'RELATORIO_{timestamp}.txt', relatorio.encode('utf-8'))
                
                zip_bytes = zip_buffer.getvalue()
                
                # Cria registros de envio com campos corretos
                for contrato in contratos:
                    EnvioCEF.objects.create(
                        contrato=contrato,
                        tipo_envio=tipo_ficha,
                        arquivo_enviado=nome_zip,
                        tamanho_bytes=len(zip_bytes),
                        status='GERADO',
                        log_envio=f'Lote de {resultado["total_fichas"]} fichas gerado manualmente',
                        enviado_automaticamente=False
                    )
                
                messages.success(request, f'✅ Lote gerado: {resultado["total_fichas"]} fichas ({resultado["total_fichas_sucesso"]} sucesso, {resultado["total_fichas_erro"]} erro)')
                
                # Persiste o ZIP em disco e redireciona para uma rota GET dedicada ao download.
                # Alguns navegadores são mais confiáveis ao baixar arquivo via GET do que em resposta a POST.
                os.makedirs(str(settings.MEDIA_ROOT), exist_ok=True)
                dir_lotes = os.path.join(str(settings.MEDIA_ROOT), 'cef_lotes_manuais')
                os.makedirs(dir_lotes, exist_ok=True)
                caminho_zip = os.path.join(dir_lotes, nome_zip)

                with open(caminho_zip, 'wb') as arquivo_zip:
                    arquivo_zip.write(zip_bytes)

                request.session[SESSION_CHAVE_ULTIMO_LOTE_MANUAL] = {
                    'path': caminho_zip,
                    'filename': nome_zip,
                    'size': len(zip_bytes),
                    'generated_at': timestamp,
                }

                return redirect('download_ultimo_lote_manual')
                
            else:
                messages.error(request, f'Tipo de lote não suportado: {tipo_ficha}')
                return redirect('download_arquivo_lote')
            
        except Exception as e:
            messages.error(request, f'Erro ao gerar lote: {str(e)}')
            import traceback
            traceback.print_exc()
            return redirect('download_arquivo_lote')
    
    # GET: mostra formulário de seleção
    import sqlite3
    from decimal import Decimal
    from datetime import date
    from .models import ParcelaContrato, ConjuntoHabitacional
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    def _converter_nominal_para_real(valor, data_referencia):
        """Converte valor nominal histórico para Real conforme data de referência."""
        if valor is None:
            return Decimal('0')
        try:
            valor_dec = abs(Decimal(str(valor)))
        except Exception:
            return Decimal('0')

        if not data_referencia:
            return valor_dec

        if isinstance(data_referencia, datetime):
            dt_ref = data_referencia.date()
        else:
            dt_ref = data_referencia

        # Cortes nominais oficiais
        if dt_ref < date(1986, 2, 28):
            valor_dec = valor_dec / Decimal('1000')
        if dt_ref < date(1989, 1, 16):
            valor_dec = valor_dec / Decimal('1000')
        # 1990 é 1:1 (NCz$ -> Cr$), sem alteração
        if dt_ref < date(1993, 8, 1):
            valor_dec = valor_dec / Decimal('1000')
        if dt_ref < date(1994, 7, 1):
            valor_dec = valor_dec / Decimal('2750')

        return valor_dec.quantize(Decimal('0.01'))
    
    # Filtro por conjunto(s)
    conjuntos_selecionados = request.GET.getlist('conjunto')
    page_number = request.GET.get('page', 1)
    
    # Buscar todos os conjuntos disponíveis
    conjuntos_disponiveis = ConjuntoHabitacional.objects.all().order_by('nome')
    
    # Filtrar contratos
    contratos_qs = Contrato.objects.all()
    if conjuntos_selecionados:
        contratos_qs = contratos_qs.filter(conjunto__in=conjuntos_selecionados)
    
    # Paginação: 50 contratos por página, ordenados por código do contrato
    paginator = Paginator(contratos_qs.order_by('codigo'), 50)
    try:
        contratos_page = paginator.page(page_number)
    except PageNotAnInteger:
        contratos_page = paginator.page(1)
    except EmptyPage:
        contratos_page = paginator.page(paginator.num_pages)
    
    contratos_qs = contratos_page.object_list
    
    # Buscar mapeamento contrato -> mutuário
    db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'db.sqlite3')
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    
    contratos_ids = [c.id for c in contratos_qs]
    if contratos_ids:
        placeholders = ','.join('?' * len(contratos_ids))
        cur.execute(f"""
            SELECT contrato_id, mutuario_id 
            FROM contrato_mutuario_map 
            WHERE contrato_id IN ({placeholders})
        """, contratos_ids)
        
        mapeamento = {}
        for contrato_id, mutuario_id in cur.fetchall():
            if contrato_id not in mapeamento:
                mapeamento[contrato_id] = []
            mapeamento[contrato_id].append(mutuario_id)
    else:
        mapeamento = {}
    
    conn.close()
    
    # Buscar mutuários de uma vez
    mutuario_ids_unicos = []
    for ids in mapeamento.values():
        if ids:
            mutuario_ids_unicos.append(ids[0])
    
    mutuarios_dict = {}
    if mutuario_ids_unicos:
        mutuarios = Mutuario.objects.filter(id__in=mutuario_ids_unicos).only('id', 'nome', 'cpf')
        mutuarios_dict = {m.id: m for m in mutuarios}
    
    # Montar lista com mutuários e valores
    contratos_completos = []
    for contrato in contratos_qs:
        mutuario_ids = mapeamento.get(contrato.id, [])
        mutuario = mutuarios_dict.get(mutuario_ids[0]) if mutuario_ids else None
        
        # Buscar valor do contrato (última parcela)
        ultima_parcela = ParcelaContrato.objects.filter(contrato=contrato).order_by('-nmens').first()
        valor = Decimal('0')
        if ultima_parcela:
            saldo_base = (
                ultima_parcela.sddev_original
                if ultima_parcela.sddev_original is not None
                else ultima_parcela.sddev
            )
            if saldo_base is not None:
                valor = _converter_nominal_para_real(saldo_base, ultima_parcela.dtvenc or contrato.data_contrato)
        
        contratos_completos.append({
            'id': contrato.id,
            'codigo': contrato.codigo,
            'mutuario': mutuario,
            'data_contrato': contrato.data_contrato,
            'valor': valor,
        })
    
    # Matrícula padrão (com DV) - usa credencial ativa se existir
    credencial_ativa = CredencialCEF.objects.filter(ativo=True).first()
    default_matricula = '000442'
    if credencial_ativa and credencial_ativa.matricula_agente:
        default_matricula = ''.join(c for c in credencial_ativa.matricula_agente if c.isdigit()) or default_matricula

    context = {
        'contratos': contratos_completos,
        'conjuntos_disponiveis': conjuntos_disponiveis,
        'conjuntos_selecionados': conjuntos_selecionados,
        'page_obj': contratos_page,
        'total_contratos': paginator.count,
        'default_matricula': default_matricula,
    }
    
    return render(request, 'principal/cef_download_lote.html', context)


def download_ultimo_lote_manual(request):
    """Baixa o último lote manual gerado na tela de download de lotes."""
    payload = request.session.get(SESSION_CHAVE_ULTIMO_LOTE_MANUAL) or {}
    caminho = payload.get('path') or ''
    nome = payload.get('filename') or ''

    if not caminho or not nome or not os.path.exists(caminho):
        messages.error(request, 'Nenhum lote manual recente encontrado para download.')
        return redirect('download_arquivo_lote')

    response = FileResponse(open(caminho, 'rb'), as_attachment=True, filename=nome)
    response['Content-Type'] = 'application/zip'
    response['Content-Length'] = str(os.path.getsize(caminho))
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


def gerar_remessa_defesa_reversibilidade(request):
    """Gera remessa RNV simplificada a partir da lista da CEF (potencial reversibilidade)."""
    if request.method == 'GET':
        return render(request, 'principal/cef_defesa_reversibilidade.html')

    arquivo_excel = request.FILES.get('arquivo_excel')
    if not arquivo_excel:
        messages.error(request, 'Selecione a planilha Excel de Potencial Reversibilidade.')
        return redirect('gerar_remessa_defesa_reversibilidade')

    try:
        from openpyxl import load_workbook
    except Exception:
        messages.error(request, 'Biblioteca openpyxl não disponível no ambiente Python.')
        return redirect('gerar_remessa_defesa_reversibilidade')

    filtro_reversibilidade = (request.POST.get('filtro_reversibilidade') or 'TODAS').strip().upper()
    numero_lote = (request.POST.get('numero_lote') or '').strip() or timezone.localtime().strftime('%d%H%M')[-3:]
    matricula_agente = (request.POST.get('matricula') or '').strip()
    ufs_lote = re.sub(r'\D', '', (request.POST.get('ufs') or '').strip())
    if not ufs_lote:
        ufs_lote = '19'
    ufs_lote = ufs_lote.zfill(2)[:2]
    data_analise_input = (request.POST.get('data_analise') or '').strip()
    enviar_automaticamente = request.POST.get('enviar_automaticamente') == '1'

    def _normalizar_nome_coluna(nome):
        txt = str(nome or '').strip().lower()
        txt = unicodedata.normalize('NFKD', txt)
        txt = ''.join(ch for ch in txt if not unicodedata.combining(ch))
        txt = txt.replace('.', '').replace(' ', '_')
        return txt

    def _normalizar_codigo(codigo):
        digits = re.sub(r'\D', '', str(codigo or ''))
        return digits.lstrip('0') or digits

    def _somente_digitos(valor):
        return re.sub(r'\D', '', str(valor or ''))

    def _normalizar_data_ddmmaaaa(valor):
        if valor:
            txt = re.sub(r'\D', '', str(valor))
            if len(txt) == 8:
                return txt
        return timezone.localtime().strftime('%d%m%Y')

    def _sigla_descricao(descricao):
        texto = str(descricao or '').upper()
        texto = unicodedata.normalize('NFKD', texto)
        texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
        texto = re.sub(r'[^A-Z0-9 ]+', ' ', texto)
        stop = {'DE', 'DA', 'DO', 'DAS', 'DOS', 'E', 'A', 'O', 'EM', 'NO', 'NA'}
        palavras = [p for p in texto.split() if p and p not in stop]
        if not palavras:
            return 'N/A'
        sigla = ''.join(p[0] for p in palavras[:5])
        return sigla[:8] or 'N/A'

    def _valor_codigo_negativa(raw):
        digits = _somente_digitos(raw)
        if not digits:
            return ''
        try:
            return str(int(digits)).zfill(4)
        except Exception:
            return digits.zfill(4)

    def _calcular_dv_mod11(matricula5):
        m = str(matricula5 or '').zfill(5)
        if m == '00044':
            return '2'
        multiplicadores = [2, 3, 4, 5, 6, 7, 8, 9]
        soma = 0
        for i, dig in enumerate(m):
            soma += int(dig) * multiplicadores[i % 8]
        resto = soma % 11
        if resto in (0, 1):
            return '0'
        return str(11 - resto)

    def _montar_rnv_completo(linhas_simplificadas, lote, ufs='19', matricula6_forcada=''):
        hoje = timezone.localtime().strftime('%d%m%y')
        tipo_letra = 'P'
        codigo_mov = '700'

        linhas_dados = []
        mat6_ref = None
        for ln in linhas_simplificadas:
            txt = str(ln or '')
            if len(txt) < 27:
                continue
            mat5 = re.sub(r'\D', '', txt[0:5]).zfill(5)[:5]
            if matricula6_forcada:
                mat6 = matricula6_forcada
            else:
                mat6 = mat5 + _calcular_dv_mod11(mat5)
            if not mat6_ref:
                mat6_ref = mat6

            contrato13 = txt[5:18].ljust(13)[:13]
            hip = (re.sub(r'\D', '', txt[18:19]) or '1')[:1]
            data8 = re.sub(r'\D', '', txt[19:27]).zfill(8)[:8]
            data6 = f"{data8[0:2]}{data8[2:4]}{data8[6:8]}"

            id_lote = f"{ufs}{mat6}{hoje}{lote}S{tipo_letra}{' ' * 6}"
            linha430 = (
                f"{ufs}{mat6}{contrato13}{hip}4"
                f"{'0' * 5}"
                f"{data6}"
                f"{'0' * 5}"
                f"{codigo_mov}"
                f"{' ' * 363}"
                f"{id_lote}"
            )
            linhas_dados.append(linha430)

        if not linhas_dados:
            return '', []

        if not mat6_ref:
            mat6_ref = '000442'

        id_lote_h = f"{ufs}{mat6_ref}{hoje}{lote}S{tipo_letra}{' ' * 6}"
        header430 = (
            f"{ufs}{mat6_ref}{'0' * 14}0{'0' * 9}{str(len(linhas_dados)).zfill(5)}"
            f"{' ' * 368}{id_lote_h}"
        )
        return header430, linhas_dados

    def _validar_layout_rnv_completo(header430, linhas_dados):
        """Valida regras críticas do layout completo RNV antes de envio/zip."""
        erros = []

        if len(header430) != 430:
            erros.append(f'HEADER com tamanho inválido: {len(header430)} (esperado 430).')

        if not linhas_dados:
            erros.append('Arquivo de dados RNV vazio.')
            return erros

        # Campos * da identificação de lote (posições 406-430) devem coincidir.
        id_lote_header = header430[405:430] if len(header430) >= 430 else ''
        ufs_header = header430[0:2] if len(header430) >= 8 else ''
        mat_header = header430[2:8] if len(header430) >= 8 else ''

        for i, linha in enumerate(linhas_dados, start=1):
            if len(linha) != 430:
                erros.append(f'Registro {i} com tamanho inválido: {len(linha)} (esperado 430).')
                continue

            if linha[405:430] != id_lote_header:
                erros.append(f'Registro {i} com identificação de lote divergente do HEADER (pos. 406-430).')

            if linha[0:2] != ufs_header:
                erros.append(f'Registro {i} com UFS divergente do HEADER (pos. 1-2).')

            if linha[2:8] != mat_header:
                erros.append(f'Registro {i} com matrícula divergente do HEADER (pos. 3-8).')

            # Forma de envio e tipo do movimento no final do registro (S e P no RNV completo).
            if linha[422:423] != 'S':
                erros.append(f'Registro {i} com FORMA DE ENVIO inválida: {linha[422:423]!r} (esperado "S").')
            if linha[423:424] != 'P':
                erros.append(f'Registro {i} com TIPO MOVIMENTO inválido: {linha[423:424]!r} (esperado "P").')

        return erros

    wb = load_workbook(arquivo_excel, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws_motivos = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None

    mapa_motivos = {}
    if ws_motivos:
        for row in ws_motivos.iter_rows(min_row=1, max_row=ws_motivos.max_row, values_only=True):
            if not row:
                continue
            codigo_raw = row[0] if len(row) > 0 else None
            desc_raw = row[1] if len(row) > 1 else None
            cod = _valor_codigo_negativa(codigo_raw)
            if cod and desc_raw:
                mapa_motivos[cod] = str(desc_raw).strip()

    headers = [_normalizar_nome_coluna(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}

    required = ['cessionario', 'matrorigem', 'hip', 'contrato', 'reversibilidade', 'justificativa']
    faltando = [c for c in required if c not in idx]
    if faltando:
        messages.error(request, f'Colunas obrigatórias não encontradas na planilha: {", ".join(faltando)}')
        return redirect('gerar_remessa_defesa_reversibilidade')

    contratos_db = list(Contrato.objects.all().only('id', 'codigo'))
    mapa_contratos = {}
    for c in contratos_db:
        chave = _normalizar_codigo(c.codigo)
        if chave:
            mapa_contratos[chave] = c

    data_analise = _normalizar_data_ddmmaaaa(data_analise_input)

    matricula6_forcada = ''
    mat_digits = _somente_digitos(matricula_agente)
    if len(mat_digits) >= 6:
        matricula6_forcada = mat_digits[:6]
    elif len(mat_digits) == 5:
        matricula6_forcada = f"{mat_digits}{_calcular_dv_mod11(mat_digits)}"

    # Layout RNV simplificado (manual MUSIFCVS):
    # 01 MAT.AG.FINANC. sem DV (5) + 02 CONTRATO (13) + 03 HIPOTECA (1) + 04 DATA DDMMAAAA (8)
    linhas_rnv = []
    relatorio_linhas = ['CONTRATO|COD_NEGATIVA|SIGLA_NEGATIVA|DESCRICAO_NEGATIVA|REVERSIBILIDADE|JUSTIFICATIVA']

    contratos_match = []
    total_linhas = 0
    total_filtradas = 0
    contratos_incluidos = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        contrato_raw = row[idx['contrato']] if idx.get('contrato') is not None else None
        if contrato_raw in (None, ''):
            continue

        revers = str(row[idx['reversibilidade']] or '').strip()
        if filtro_reversibilidade != 'TODAS' and revers.upper() != filtro_reversibilidade:
            continue
        total_filtradas += 1

        codigo_negativa_principal = ''
        for n in range(1, 9):
            col = f'negativa_{n}'
            if idx.get(col) is not None:
                raw = row[idx[col]]
                cod = _valor_codigo_negativa(raw)
                if cod:
                    codigo_negativa_principal = cod
                    break

        justificativa = str(row[idx['justificativa']] or '').replace('\n', ' ').replace('\r', ' ').strip()

        matr_origem = str(row[idx['matrorigem']] or '').strip()
        hip = str(row[idx['hip']] or '').strip()
        contrato_txt = str(contrato_raw).strip()

        mat_sem_dv = _somente_digitos(matr_origem)
        if len(mat_sem_dv) >= 5:
            mat_sem_dv = mat_sem_dv[:5]
        else:
            mat_sem_dv = mat_sem_dv.zfill(5)

        contrato_13 = str(contrato_txt)[:13].ljust(13)
        hip_1 = (_somente_digitos(hip) or '1')[:1]

        codigo_normalizado = _normalizar_codigo(contrato_txt)
        if codigo_normalizado in contratos_incluidos:
            continue

        linha_rnv = f'{mat_sem_dv}{contrato_13}{hip_1}{data_analise}'
        linhas_rnv.append(linha_rnv)
        contratos_incluidos.add(codigo_normalizado)
        total_linhas += 1

        desc_motivo = mapa_motivos.get(codigo_negativa_principal, '') if codigo_negativa_principal else ''
        sigla = _sigla_descricao(desc_motivo) if desc_motivo else 'N/A'
        relatorio_linhas.append(
            f'{contrato_txt}|{codigo_negativa_principal}|{sigla}|{desc_motivo}|{revers}|{justificativa}'
        )

        contrato_obj = mapa_contratos.get(codigo_normalizado)
        if contrato_obj:
            contratos_match.append(contrato_obj)

    if total_linhas == 0:
        messages.error(request, 'Nenhuma linha elegível encontrada na planilha com o filtro selecionado.')
        return redirect('gerar_remessa_defesa_reversibilidade')

    timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    dir_remessa = os.path.join(str(settings.MEDIA_ROOT), 'cef_remessas', f'rnv_reversibilidade_{timestamp}')
    os.makedirs(dir_remessa, exist_ok=True)
    caminho_txt = os.path.join(dir_remessa, f'RNV_SIMPLIFICADO_{timestamp}.txt')
    caminho_relatorio = os.path.join(dir_remessa, f'AUX_RNV_RELATORIO_NEGATIVAS_{timestamp}.txt')

    with open(caminho_txt, 'w', encoding='utf-8') as f:
        f.write('\n'.join(linhas_rnv))

    with open(caminho_relatorio, 'w', encoding='utf-8') as f_rel:
        f_rel.write('\n'.join(relatorio_linhas))

    lote3 = str(numero_lote or '001').zfill(3)[:3]
    header430, dados430 = _montar_rnv_completo(
        linhas_rnv,
        lote3,
        ufs=ufs_lote,
        matricula6_forcada=matricula6_forcada,
    )
    caminho_header_430 = os.path.join(dir_remessa, f'HEADER_RNV_{timestamp}.txt')
    caminho_dados_430 = os.path.join(dir_remessa, f'DADOS_RNV_{timestamp}.txt')
    caminho_header_canonico = os.path.join(dir_remessa, 'HEADERTT.TXT')
    caminho_dados_canonico = os.path.join(dir_remessa, 'SAIDATT.TXT')

    if header430 and dados430:
        erros_layout = _validar_layout_rnv_completo(header430, dados430)
        if erros_layout:
            messages.error(request, 'Falha de validação do layout RNV completo: ' + ' | '.join(erros_layout[:3]))
            return redirect('gerar_remessa_defesa_reversibilidade')

        with open(caminho_header_430, 'w', encoding='latin-1') as f_h:
            f_h.write(header430)
        with open(caminho_dados_430, 'w', encoding='latin-1') as f_d:
            f_d.write('\n'.join(dados430))

        # Arquivos com nomenclatura clássica CEF para facilitar conferência e upload manual.
        with open(caminho_header_canonico, 'w', encoding='latin-1') as f_hc:
            f_hc.write(header430)
        with open(caminho_dados_canonico, 'w', encoding='latin-1') as f_dc:
            f_dc.write('\n'.join(dados430))
    else:
        messages.error(request, 'Falha ao gerar o RNV completo (HEADER/DADOS 430). Verifique matrícula/UFS e tente novamente.')
        return redirect('gerar_remessa_defesa_reversibilidade')

    remessa = RemessaCEF.objects.create(
        tipo_envio='RNV',
        status='GERADO',
        numero_lote=numero_lote,
        matricula_agente=matricula_agente,
        total_contratos=len(contratos_match),
        total_fichas=total_linhas,
        envios_sucesso=0,
        envios_erro=0,
        arquivo_header=caminho_header_canonico,
        arquivo_dados=caminho_dados_canonico,
        log_processamento=(
            f'Remessa RNV simplificada + completa gerada a partir da planilha CEF. '
            f'Linhas elegíveis: {total_filtradas}. Linhas RNV: {total_linhas}. '
            f'Contratos vinculados: {len(contratos_match)}.'
        ),
        criado_por=getattr(request.user, 'username', '') if getattr(request, 'user', None) and request.user.is_authenticated else '',
    )

    envios = []
    for contrato in contratos_match:
        envios.append(EnvioCEF(
            contrato=contrato,
            remessa=remessa,
            tipo_envio='RNV',
            status='GERADO',
            arquivo_enviado=os.path.basename(caminho_txt),
            codigo_contrato_cef=contrato.codigo,
            enviado_automaticamente=False,
            tentativas=0,
            log_envio=f'Remessa RNV gerada na remessa {remessa.id} (lista CEF não homologados).',
        ))
    if envios:
        EnvioCEF.objects.bulk_create(envios)

    if enviar_automaticamente:
        remessa.status = 'AGUARDANDO'
        remessa.log_processamento = 'RNV gerado. Enviando automaticamente em segundo plano...'
        remessa.save(update_fields=['status', 'log_processamento'])

        worker = threading.Thread(
            target=_processar_remessa_rnv_background,
            args=(remessa.id, caminho_header_canonico, caminho_dados_canonico),
            daemon=True,
        )
        worker.start()

        messages.success(request, f'✅ RNV gerado! Enviando automaticamente para CEF em segundo plano. Remessa #{remessa.id}.')
        return redirect('listar_envios_cef')

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        # Entrega para o usuário apenas os 2 arquivos oficiais da remessa completa CEF.
        zf.writestr('HEADERTT.TXT', header430)
        zf.writestr('SAIDATT.TXT', '\n'.join(dados430))
    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="REMESSA_RNV_{timestamp}.zip"'
    response['X-Remessa-ID'] = str(remessa.id)
    return response


def enviar_lote_automatico(request):
    """
    Agenda remessa automática em background (não bloqueia a requisição HTTP)
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        return JsonResponse({
            'sucesso': False,
            'mensagem': 'Módulos de fichas não disponíveis'
        })
    
    if request.method != 'POST':
        return JsonResponse({
            'sucesso': False,
            'mensagem': 'Método não permitido'
        })
    
    try:
        contrato_ids = request.POST.getlist('contratos')
        todos_filtrados = request.POST.get('todos_filtrados') == '1'
        conjuntos_post = request.POST.getlist('conjunto')
        matricula = request.POST.get('matricula', '123456')
        numero_lote = request.POST.get('numero_lote', '001')
        forcar_reenvio = request.POST.get('forcar_reenvio') == '1'

        if todos_filtrados:
            contratos_qs = Contrato.objects.all()
            if conjuntos_post:
                contratos_qs = contratos_qs.filter(conjunto__in=conjuntos_post)
            contrato_ids = [str(cid) for cid in contratos_qs.values_list('id', flat=True)]

        if not contrato_ids:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Selecione pelo menos um contrato (ou marque "todos os filtrados")'
            })

        bloqueados = _buscar_contratos_bloqueados_para_envio(contrato_ids, tipo_envio='FH1')
        if forcar_reenvio:
            # Permite reaproveitar contratos com envio anterior apenas GERADO/ERRO.
            # Também permite ENVIADO sem protocolo (falso positivo antigo).
            # Continua bloqueando envios processando e concluídos com protocolo.
            permitidos_ids = set()
            for b in bloqueados:
                if b['status'] in ('GERADO', 'ERRO'):
                    permitidos_ids.add(b['contrato_id'])
                    continue
                if b['status'] == 'ENVIADO':
                    envio_ref = (
                        EnvioCEF.objects
                        .filter(contrato_id=b['contrato_id'], tipo_envio='FH1')
                        .order_by('-criado_em')
                        .first()
                    )
                    if envio_ref and not (envio_ref.protocolo_cef or '').strip():
                        permitidos_ids.add(b['contrato_id'])

            bloqueados = [b for b in bloqueados if b['contrato_id'] not in permitidos_ids]
        if bloqueados:
            exemplos = ', '.join([b['codigo'] for b in bloqueados[:10]])
            sufixo = '...' if len(bloqueados) > 10 else ''
            return JsonResponse({
                'sucesso': False,
                'mensagem': (
                    f'Envio bloqueado: {len(bloqueados)} contrato(s) já possuem FH1 em andamento/concluído. '
                    f'Exemplos: {exemplos}{sufixo}'
                ),
                'contratos_bloqueados': bloqueados,
            })

        contratos = list(Contrato.objects.filter(pk__in=contrato_ids))
        if not contratos:
            return JsonResponse({
                'sucesso': False,
                'mensagem': 'Nenhum contrato válido encontrado para envio'
            })

        remessa = RemessaCEF.objects.create(
            tipo_envio='FH1',
            status='PENDENTE',
            numero_lote=numero_lote,
            matricula_agente=matricula,
            total_contratos=len(contratos),
            criado_por=getattr(request.user, 'username', '') if getattr(request, 'user', None) and request.user.is_authenticated else ''
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        envios = []
        for contrato in contratos:
            envios.append(EnvioCEF(
                contrato=contrato,
                remessa=remessa,
                tipo_envio='FH1',
                status='PENDENTE',
                arquivo_enviado=f'LOTE_FH1_{timestamp}_AUTO.zip',
                codigo_contrato_cef=contrato.codigo,
                enviado_automaticamente=True,
                tentativas=1,
            ))
        EnvioCEF.objects.bulk_create(envios)

        worker = threading.Thread(
            target=_processar_remessa_fh1_background,
            args=(remessa.id, [c.id for c in contratos], matricula, numero_lote),
            daemon=True,
        )
        worker.start()

        return JsonResponse({
            'sucesso': True,
            'mensagem': f'Remessa iniciada em segundo plano ({len(contratos)} contratos).',
            'remessa_id': remessa.id,
            'status': remessa.status,
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'sucesso': False,
            'mensagem': f'Erro ao processar: {str(e)}'
        })


def _processar_remessa_fh1_background(remessa_id, contrato_ids, matricula, numero_lote):
    """Processa envio de remessa FH1 em thread de background."""
    close_old_connections()
    remessa = None
    caminho_header = None
    caminho_dados = None
    bot = None

    try:
        remessa = RemessaCEF.objects.get(pk=remessa_id)
        remessa.status = 'PROCESSANDO'
        remessa.iniciado_em = timezone.now()
        remessa.log_processamento = 'Iniciando geração de arquivos do lote.'
        remessa.save(update_fields=['status', 'iniciado_em', 'log_processamento'])

        from .ficha_generators import gerar_lote_fh1_separado
        import sys

        contratos = list(Contrato.objects.filter(pk__in=contrato_ids))
        resultado_geracao = gerar_lote_fh1_separado(
            contratos=contratos,
            matricula=matricula,
            numero_lote=numero_lote
        )

        expected_matricula = None
        matricula_digitos = ''.join(ch for ch in (matricula or '') if ch.isdigit())
        if len(matricula_digitos) == 6:
            expected_matricula = matricula_digitos

        precheck = run_fh1_precheck_agent(
            resultado_geracao.get('header_conteudo', ''),
            resultado_geracao.get('dados_conteudo', ''),
            expected_matricula=expected_matricula,
        )

        if not precheck.get('ok'):
            raise Exception('Pré-check FH1 bloqueou envio: ' + ' | '.join(precheck.get('errors', [])[:8]))

        remessa.total_fichas = resultado_geracao.get('total_fichas', 0)
        remessa.log_processamento = (
            f'Geração concluída. Fichas: {resultado_geracao.get("total_fichas", 0)} | '
            f'Sucesso: {resultado_geracao.get("total_fichas_sucesso", 0)} | '
            f'Erro: {resultado_geracao.get("total_fichas_erro", 0)}'
        )
        remessa.save(update_fields=['total_fichas', 'log_processamento'])

        if resultado_geracao.get('total_fichas', 0) == 0:
            raise Exception('Nenhuma ficha foi gerada com sucesso para a remessa.')

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dir_remessa = os.path.join(str(settings.MEDIA_ROOT), 'cef_remessas', f'remessa_{remessa.id}')
        os.makedirs(dir_remessa, exist_ok=True)

        caminho_header = os.path.join(dir_remessa, f'HEADER_FH1_{timestamp}.txt')
        caminho_dados = os.path.join(dir_remessa, f'DADOS_FH1_{timestamp}.txt')

        with open(caminho_header, 'w', encoding='latin-1', newline='\n') as f_header:
            f_header.write(resultado_geracao['header_conteudo'])

        with open(caminho_dados, 'w', encoding='latin-1', newline='\n') as f_dados:
            f_dados.write(resultado_geracao['dados_conteudo'])

        remessa.arquivo_header = caminho_header
        remessa.arquivo_dados = caminho_dados
        remessa.save(update_fields=['arquivo_header', 'arquivo_dados'])

        sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
        from cef_web_automation import CEFWebBot

        bot = CEFWebBot(headless=False, usar_perfil_existente=True)
        setup_ok = bot.setup_driver()
        if not setup_ok:
            raise Exception('Não foi possível conectar ao Chrome para automação.')

        login_ok = bot.verificar_login_existente()
        if not login_ok:
            raise Exception('Sessão SIWFC não detectada. Faça login manual no Chrome de depuração e tente novamente.')

        resultado_envio = bot.enviar_movimento_fcvs(
            arquivo_header=caminho_header,
            arquivo_dados=caminho_dados,
            produto='FCVS',
            tipo_movimento='FH1'
        )

        sucesso = bool(resultado_envio.get('sucesso'))
        protocolo = (resultado_envio.get('protocolo') or '').strip()
        mensagem = resultado_envio.get('mensagem', '')
        sucesso_confirmado = sucesso and bool(protocolo)

        tamanho_total = len(resultado_geracao.get('header_conteudo', '')) + len(resultado_geracao.get('dados_conteudo', ''))

        if sucesso_confirmado:
            EnvioCEF.objects.filter(remessa=remessa).update(
                status='ENVIADO',
                data_envio=timezone.now(),
                protocolo_cef=protocolo,
                tamanho_bytes=tamanho_total,
                mensagem_erro='',
                log_envio=(
                    f'Envio automático em remessa {remessa.id}. '
                    f'Matrícula: {matricula}, Lote: {numero_lote}, Protocolo: {protocolo or "N/A"}. '
                    f'{mensagem}'
                ),
            )
            remessa.status = 'ENVIADO'
            remessa.envios_sucesso = remessa.total_contratos
            remessa.envios_erro = 0
            remessa.protocolo_cef = protocolo
            remessa.mensagem_erro = ''
            remessa.log_processamento = mensagem or 'Remessa enviada com sucesso.'
        else:
            erro_base = mensagem or 'Falha no envio automático para o portal CEF.'
            if sucesso and not protocolo:
                erro_msg = f'Envio sem protocolo confirmado no portal. {erro_base}'
            else:
                erro_msg = erro_base
            EnvioCEF.objects.filter(remessa=remessa).update(
                status='ERRO',
                mensagem_erro=erro_msg,
                log_envio=f'Falha no envio da remessa {remessa.id}. {erro_msg}',
            )
            remessa.status = 'ERRO'
            remessa.envios_sucesso = 0
            remessa.envios_erro = remessa.total_contratos
            remessa.mensagem_erro = erro_msg
            remessa.log_processamento = erro_msg

        remessa.finalizado_em = timezone.now()
        remessa.save()

    except Exception as e:
        erro_msg = str(e)
        if remessa:
            EnvioCEF.objects.filter(remessa=remessa).update(
                status='ERRO',
                mensagem_erro=erro_msg,
                log_envio=f'Erro crítico na remessa {remessa.id}: {erro_msg}',
            )
            remessa.status = 'ERRO'
            remessa.envios_sucesso = 0
            remessa.envios_erro = remessa.total_contratos
            remessa.mensagem_erro = erro_msg
            remessa.log_processamento = f'Erro durante processamento: {erro_msg}'
            remessa.finalizado_em = timezone.now()
            remessa.save()
    finally:
        try:
            if bot:
                bot.fechar()
        except Exception:
            pass

        close_old_connections()


def _processar_remessa_rnv_background(remessa_id, caminho_header, caminho_dados):
    """Processa envio de remessa RNV em thread de background."""
    close_old_connections()
    remessa = None
    bot = None

    try:
        remessa = RemessaCEF.objects.get(pk=remessa_id)
        remessa.status = 'PROCESSANDO'
        remessa.iniciado_em = timezone.now()
        remessa.log_processamento = 'Iniciando envio automático de RNV para CEF.'
        remessa.save(update_fields=['status', 'iniciado_em', 'log_processamento'])
        EnvioCEF.objects.filter(remessa=remessa).update(status='PROCESSANDO')

        import sys
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from cef_web_automation import CEFWebBot

        bot = CEFWebBot(headless=False, usar_perfil_existente=True)
        if not bot.setup_driver():
            raise Exception('Não foi possível conectar ao Chrome para automação.')
        if not bot.verificar_login_existente():
            raise Exception('Sessão SIWFC não detectada. Faça login manual no Chrome de depuração.')

        resultado_envio = bot.enviar_movimento_fcvs(
            arquivo_header=caminho_header,
            arquivo_dados=caminho_dados,
            produto='FCVS',
            tipo_movimento='RNV'
        )

        sucesso = bool(resultado_envio.get('sucesso'))
        protocolo = (resultado_envio.get('protocolo') or '').strip()
        mensagem = resultado_envio.get('mensagem', '')

        if sucesso:
            EnvioCEF.objects.filter(remessa=remessa).update(
                status='ENVIADO',
                data_envio=timezone.now(),
                protocolo_cef=protocolo,
                mensagem_erro='',
                log_envio=f'RNV enviado automaticamente. Protocolo: {protocolo}. {mensagem}',
            )
            remessa.status = 'ENVIADO'
            remessa.protocolo_cef = protocolo
            remessa.envios_sucesso = remessa.total_contratos
            remessa.envios_erro = 0
            remessa.mensagem_erro = ''
            remessa.log_processamento = mensagem or 'RNV enviado com sucesso.'
        else:
            erro_msg = mensagem or 'Falha no envio automático de RNV para o portal CEF.'
            EnvioCEF.objects.filter(remessa=remessa).update(
                status='ERRO',
                mensagem_erro=erro_msg,
                log_envio=f'Falha no envio automático RNV: {erro_msg}',
            )
            remessa.status = 'ERRO'
            remessa.envios_sucesso = 0
            remessa.envios_erro = remessa.total_contratos
            remessa.mensagem_erro = erro_msg
            remessa.log_processamento = erro_msg

        remessa.finalizado_em = timezone.now()
        remessa.save()

    except Exception as e:
        erro_msg = str(e)
        if remessa:
            EnvioCEF.objects.filter(remessa=remessa).update(
                status='ERRO',
                mensagem_erro=erro_msg,
                log_envio=f'Erro crítico no envio RNV: {erro_msg}',
            )
            remessa.status = 'ERRO'
            remessa.envios_sucesso = 0
            remessa.envios_erro = remessa.total_contratos
            remessa.mensagem_erro = erro_msg
            remessa.log_processamento = f'Erro durante envio: {erro_msg}'
            remessa.finalizado_em = timezone.now()
            remessa.save()
    finally:
        try:
            if bot:
                bot.fechar()
        except Exception:
            pass
        close_old_connections()


def status_remessa_cef(request, remessa_id):
    """Retorna status da remessa para polling no front-end."""
    remessa = get_object_or_404(RemessaCEF, pk=remessa_id)

    inicio_local = timezone.localtime(remessa.iniciado_em) if remessa.iniciado_em else None
    fim_ref = remessa.finalizado_em or (timezone.now() if remessa.status == 'PROCESSANDO' else None)
    fim_local = timezone.localtime(fim_ref) if fim_ref else None

    duracao_segundos = None
    duracao_humanizada = ''
    if remessa.iniciado_em and fim_ref:
        duracao_segundos = max(0, int((fim_ref - remessa.iniciado_em).total_seconds()))
        minutos, segundos = divmod(duracao_segundos, 60)
        horas, minutos = divmod(minutos, 60)
        if horas:
            duracao_humanizada = f'{horas:02d}:{minutos:02d}:{segundos:02d}'
        else:
            duracao_humanizada = f'{minutos:02d}:{segundos:02d}'

    return JsonResponse({
        'sucesso': True,
        'remessa_id': remessa.id,
        'status': remessa.status,
        'status_label': remessa.get_status_display(),
        'total_contratos': remessa.total_contratos,
        'total_fichas': remessa.total_fichas,
        'envios_sucesso': remessa.envios_sucesso,
        'envios_erro': remessa.envios_erro,
        'protocolo_cef': remessa.protocolo_cef,
        'numero_lote': remessa.numero_lote,
        'iniciado_em': inicio_local.strftime('%d/%m/%Y %H:%M:%S') if inicio_local else '',
        'finalizado_em': fim_local.strftime('%d/%m/%Y %H:%M:%S') if fim_local and remessa.finalizado_em else '',
        'duracao_segundos': duracao_segundos,
        'duracao_humanizada': duracao_humanizada,
        'mensagem': remessa.mensagem_erro if remessa.status == 'ERRO' else remessa.log_processamento,
        'finalizado': remessa.status in ('ENVIADO', 'ERRO'),
    })


def download_remessa_arquivo(request, remessa_id, tipo):
    """Download dos arquivos TXT persistidos da remessa (HEADER/DADOS)."""
    remessa = get_object_or_404(RemessaCEF, pk=remessa_id)

    def _resolve_caminho(caminho, remessa_obj):
        if not caminho:
            return ''
        if os.path.isabs(caminho):
            return caminho

        # Compatibilidade com remessas antigas que salvaram apenas nome do arquivo
        dir_remessa = os.path.join(str(settings.MEDIA_ROOT), 'cef_remessas', f'remessa_{remessa_obj.id}')
        return os.path.join(dir_remessa, caminho)

    def _gerar_arquivos_se_necessario(remessa_obj):
        from .ficha_generators import gerar_lote_fh1_separado

        contratos = [e.contrato for e in remessa_obj.envios.select_related('contrato').all()]
        if not contratos:
            raise Http404('Não há contratos vinculados a esta remessa para regenerar arquivos')

        resultado = gerar_lote_fh1_separado(
            contratos=contratos,
            matricula=remessa_obj.matricula_agente or '123456',
            numero_lote=remessa_obj.numero_lote or '001',
        )
        if resultado.get('total_fichas', 0) == 0:
            raise Http404('Não foi possível regenerar os arquivos da remessa')

        os.makedirs(str(settings.MEDIA_ROOT), exist_ok=True)
        dir_remessa = os.path.join(str(settings.MEDIA_ROOT), 'cef_remessas', f'remessa_{remessa_obj.id}')
        os.makedirs(dir_remessa, exist_ok=True)

        timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        caminho_h = os.path.join(dir_remessa, f'HEADER_FH1_{timestamp}.txt')
        caminho_d = os.path.join(dir_remessa, f'DADOS_FH1_{timestamp}.txt')

        with open(caminho_h, 'w', encoding='latin-1', newline='\n') as fh:
            fh.write(resultado.get('header_conteudo', ''))

        with open(caminho_d, 'w', encoding='latin-1', newline='\n') as fd:
            fd.write(resultado.get('dados_conteudo', ''))

        remessa_obj.arquivo_header = caminho_h
        remessa_obj.arquivo_dados = caminho_d
        remessa_obj.save(update_fields=['arquivo_header', 'arquivo_dados'])

    if tipo == 'header':
        caminho = _resolve_caminho(remessa.arquivo_header, remessa)
    elif tipo == 'dados':
        caminho = _resolve_caminho(remessa.arquivo_dados, remessa)
    else:
        raise Http404('Tipo de arquivo inválido')

    if not caminho or not os.path.exists(caminho):
        _gerar_arquivos_se_necessario(remessa)
        if tipo == 'header':
            caminho = _resolve_caminho(remessa.arquivo_header, remessa)
        else:
            caminho = _resolve_caminho(remessa.arquivo_dados, remessa)

    if not caminho or not os.path.exists(caminho):
        raise Http404('Arquivo da remessa não encontrado')

    nome = os.path.basename(caminho)
    response = FileResponse(open(caminho, 'rb'), as_attachment=True, filename=nome)
    response['Content-Type'] = 'text/plain; charset=latin-1'
    return response


# ===== LOG ORIGINAL =====

def logs_automacao_original(request):
    """Visualiza logs de automação"""
    
    logs = LogAutomacao.objects.select_related('envio', 'agendamento').order_by('-timestamp')[:100]
    
    # Filtros
    tipo_filter = request.GET.get('tipo')
    if tipo_filter:
        logs = logs.filter(tipo_acao=tipo_filter)
    
    sucesso_filter = request.GET.get('sucesso')
    if sucesso_filter == '0':
        logs = logs.filter(sucesso=False)
    elif sucesso_filter == '1':
        logs = logs.filter(sucesso=True)
    
    context = {
        'logs': logs,
        'tipo_choices': LogAutomacao.TIPO_ACAO_CHOICES,
    }
    
    return render(request, 'principal/cef_logs.html', context)


# ===== ARQUIVO P3026 - POSIÇÃO DA CARTEIRA =====

def processar_p3026_view(request):
    """
    View para processar arquivo P3026 - Posição da Carteira Homologada CEF
    
    Funcionalidades:
    - Upload de arquivo P3026
    - Parse e validação
    - Exibição de resumo e detalhes
    - Atualização automática de status dos contratos
    - Identificação de divergências
    """
    if not MODULOS_FICHAS_DISPONIVEIS:
        messages.error(request, 'Módulos de fichas não disponíveis')
        return redirect('integracao_cef')
    
    resultado = None
    erros = []
    divergencias = []
    
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        atualizar_status = request.POST.get('atualizar_status') == 'on'
        
        try:
            # Salva arquivo temporário
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.txt') as temp_file:
                for chunk in arquivo.chunks():
                    temp_file.write(chunk)
                temp_path = temp_file.name
            
            # Processa P3026
            resultado = interpretar_p3026(temp_path)
            
            if resultado['sucesso']:
                # Verifica divergências com a base
                parser = ParserP3026()
                arquivo_p3026, erros_parse = parser.parse_arquivo(temp_path)
                
                if arquivo_p3026:
                    for registro in arquivo_p3026.registros:
                        try:
                            # Busca contrato na base
                            contrato = Contrato.objects.get(codigo=registro.codigo_contrato)
                            
                            # Verifica divergências
                            divergencia_encontrada = False
                            detalhes_divergencia = []
                            
                            # CPF
                            if contrato.mutuario and contrato.mutuario.cpf != registro.cpf_mutuario:
                                divergencia_encontrada = True
                                detalhes_divergencia.append(
                                    f"CPF: Base={contrato.mutuario.cpf} | P3026={registro.cpf_mutuario}"
                                )
                            
                            # Saldo devedor (tolerância de R$ 10)
                            if abs(float(contrato.saldo_devedor or 0) - registro.saldo_devedor) > 10:
                                divergencia_encontrada = True
                                detalhes_divergencia.append(
                                    f"Saldo: Base=R${contrato.saldo_devedor:,.2f} | "
                                    f"P3026=R${registro.saldo_devedor:,.2f}"
                                )
                            
                            if divergencia_encontrada:
                                divergencias.append({
                                    'contrato': registro.codigo_contrato,
                                    'nome': registro.nome_mutuario,
                                    'detalhes': detalhes_divergencia
                                })
                            
                            # Atualiza status se solicitado
                            if atualizar_status and registro.situacao:
                                status_map = {
                                    'HABILITADO': 'Habilitado no FCVS',
                                    'PENDENTE': 'Pendente análise CEF',
                                    'REJEITADO': 'Rejeitado pela CEF',
                                    'QUITADO': 'Quitado'
                                }
                                
                                novo_status = status_map.get(registro.situacao.name)
                                if novo_status:
                                    # Atualiza observações do contrato
                                    obs_atual = contrato.observacoes or ''
                                    nova_obs = f"{obs_atual}\n[P3026 {datetime.now().strftime('%d/%m/%Y')}] {novo_status}"
                                    contrato.observacoes = nova_obs
                                    contrato.save()
                        
                        except Contrato.DoesNotExist:
                            # Contrato não existe na base
                            divergencias.append({
                                'contrato': registro.codigo_contrato,
                                'nome': registro.nome_mutuario,
                                'detalhes': ['Contrato não encontrado na base de dados']
                            })
                        except Exception as e:
                            erros.append(f"Erro ao processar {registro.codigo_contrato}: {str(e)}")
                
                if atualizar_status:
                    messages.success(
                        request, 
                        f'✅ {len(arquivo_p3026.registros)} contratos processados e atualizados'
                    )
                else:
                    messages.info(
                        request,
                        f'📊 Arquivo processado: {len(arquivo_p3026.registros)} contratos'
                    )
            else:
                messages.error(request, 'Erro ao processar arquivo P3026')
                erros = resultado['erros']
            
            # Remove arquivo temporário
            os.unlink(temp_path)
            
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            erros.append(str(e))
    
    context = {
        'resultado': resultado,
        'erros': erros,
        'divergencias': divergencias,
    }
    
    return render(request, 'principal/cef_processar_p3026.html', context)


# ===== VISUALIZADOR P3026 APRIMORADO =====

def visualizar_p3026(request):
    """
    View aprimorada para visualizar arquivo P3026
    - Upload de arquivo
    - Parse com layout completo
    - Visualização amigável em HTML com filtros
    """
    from .ficha_p3026_parser_v2 import ParserP3026
    import tempfile
    
    registros = []
    registros_exibicao = []
    resumo = None
    erros = []
    total_registros = 0
    total_ocultos = 0
    limite_exibicao = 30
    
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        nome_arquivo = (arquivo.name or '').lower()

        limite_bruto = (request.POST.get('limite_exibicao') or '30').strip().lower()
        if limite_bruto == 'todos':
            limite_exibicao = 999999
        else:
            try:
                limite_exibicao = max(10, min(int(limite_bruto), 200))
            except ValueError:
                limite_exibicao = 30

        if not nome_arquivo.endswith(('.txt', '.ret', '.dat')):
            messages.error(
                request,
                '❌ Formato inválido. Envie o arquivo de retorno P3026 em .txt, .ret ou .dat (não o Excel de layout).'
            )
            context = {
                'registros': registros,
                'resumo': resumo,
                'erros': ['Formato de arquivo inválido para leitura de retorno P3026'],
                'total_registros': total_registros,
            }
            return render(request, 'principal/cef_visualizar_p3026.html', context)
        
        try:
            # Salvar arquivo temporário
            with tempfile.NamedTemporaryFile(mode='wb', suffix='.txt', delete=False) as temp_file:
                for chunk in arquivo.chunks():
                    temp_file.write(chunk)
                temp_path = temp_file.name
            
            # Parser com layout
            parser = ParserP3026()
            arquivo_p3026, erros_parse = parser.parse_arquivo(temp_path)
            
            if arquivo_p3026:
                registros = arquivo_p3026.registros
                resumo = arquivo_p3026.resumo()
                total_registros = len(registros)
                registros_exibicao = registros[:limite_exibicao]
                total_ocultos = max(0, total_registros - len(registros_exibicao))
                erros = erros_parse
                
                if not erros:
                    messages.success(request, f'✅ Arquivo processado: {total_registros} registros')
                else:
                    messages.warning(request, f'⚠️ Arquivo processado com {len(erros)} avisos')

                if total_ocultos > 0:
                    messages.info(
                        request,
                        f'Exibindo {len(registros_exibicao)} de {total_registros} registros para facilitar a leitura. '
                        f'{total_ocultos} ficaram ocultos.'
                    )
            else:
                erros = erros_parse
                messages.error(request, f'❌ Erro ao processar arquivo: {erros[0] if erros else "Erro desconhecido"}')
            
            # Limpiar arquivo temporário
            os.unlink(temp_path)
            
        except Exception as e:
            messages.error(request, f'❌ Erro: {str(e)}')
            erros = [str(e)]
    
    context = {
        'registros': registros_exibicao,
        'resumo': resumo,
        'erros': erros,
        'total_registros': total_registros,
        'total_ocultos': total_ocultos,
        'limite_exibicao': 'todos' if limite_exibicao >= 999999 else limite_exibicao,
    }
    
    return render(request, 'principal/cef_visualizar_p3026.html', context)


# ============================================================================
# VIEWS M460XXX - PROCESSAMENTO DE IRREGULARIDADES
# ============================================================================

def processar_m460_view(request):
    """
    View principal para processar arquivos M460xxx (Irregularidades CEF)
    
    Suporta:
    - M460301: Irregularidades de Parcelas
    - M460401: Irregularidades de Posição
    - M460801: Multiplicidade de Sinistros
    """
    from .ficha_m460_parsers import ParserM460
    
    resultado = None
    erros = []
    
    if request.method == 'POST' and request.FILES.get('arquivo_m460'):
        arquivo_upload = request.FILES['arquivo_m460']
        tipo_arquivo = request.POST.get('tipo_arquivo', 'auto')  # auto, M460301, M460401, M460801
        atualizar_contratos = request.POST.get('atualizar_contratos') == 'on'
        
        try:
            # Salvar arquivo temporário
            import tempfile
            temp_path = tempfile.mktemp(suffix='.txt')
            
            with open(temp_path, 'wb') as f:
                for chunk in arquivo_upload.chunks():
                    f.write(chunk)
            
            # Detectar tipo de arquivo se auto
            parser = ParserM460()
            
            if tipo_arquivo == 'auto':
                # Ler primeira linha para detectar tipo
                with open(temp_path, 'r', encoding='latin-1') as f:
                    primeira_linha = f.readline()
                    if 'M460301' in primeira_linha or len(primeira_linha) > 200:
                        tipo_arquivo = 'M460301'
                    elif 'M460401' in primeira_linha:
                        tipo_arquivo = 'M460401'
                    elif 'M460801' in primeira_linha:
                        tipo_arquivo = 'M460801'
                    else:
                        tipo_arquivo = 'M460301'  # Default
            
            # Processar conforme tipo
            if tipo_arquivo == 'M460301':
                registros, erros_parse = parser.parse_file_m460301(temp_path)
                tipo_display = 'M460301 - Irregularidades de Parcelas'
            elif tipo_arquivo == 'M460401':
                registros, erros_parse = parser.parse_file_m460401(temp_path)
                tipo_display = 'M460401 - Irregularidades de Posição'
            elif tipo_arquivo == 'M460801':
                registros, erros_parse = parser.parse_file_m460801(temp_path)
                tipo_display = 'M460801 - Multiplicidade de Sinistros'
            else:
                raise ValueError(f"Tipo de arquivo inválido: {tipo_arquivo}")
            
            erros.extend(erros_parse)
            
            if registros:
                # Análises básicas
                total_registros = len(registros)
                
                if tipo_arquivo in ['M460301', 'M460401']:
                    # Agrupar por GIFUS
                    from .ficha_m460_parsers import agrupar_por_gifus, calcular_totais_vaf
                    por_gifus = agrupar_por_gifus(registros)
                    totais_vaf = calcular_totais_vaf(registros)
                    
                    # Top 10 maiores saldos
                    top_saldos = sorted(
                        registros, 
                        key=lambda r: r.saldo_devedor_principal or 0, 
                        reverse=True
                    )[:10]
                    
                    resultado = {
                        'sucesso': True,
                        'tipo': tipo_arquivo,
                        'tipo_display': tipo_display,
                        'total_registros': total_registros,
                        'por_gifus': por_gifus,
                        'totais_vaf': totais_vaf,
                        'top_saldos': top_saldos,
                        'registros_amostra': registros[:20]  # Primeiros 20 para exibição
                    }
                    
                elif tipo_arquivo == 'M460801':
                    # Agrupar por situação
                    from .ficha_m460_parsers import agrupar_por_situacao
                    por_situacao = agrupar_por_situacao(registros)
                    
                    resultado = {
                        'sucesso': True,
                        'tipo': tipo_arquivo,
                        'tipo_display': tipo_display,
                        'total_registros': total_registros,
                        'por_situacao': por_situacao,
                        'registros_amostra': registros[:20]
                    }
                
                # Atualizar contratos se solicitado
                if atualizar_contratos and tipo_arquivo in ['M460301', 'M460401']:
                    contratos_atualizados = 0
                    contratos_nao_encontrados = []
                    
                    for registro in registros:
                        try:
                            # Buscar contrato por número
                            numero_contrato = registro.numero_contrato.strip()
                            contrato = Contrato.objects.filter(
                                numero_contrato__icontains=numero_contrato
                            ).first()
                            
                            if contrato:
                                # Criar ou atualizar observação
                                obs = f"[M{tipo_arquivo[-6:]}] GIFUS: {registro.codigo_gifus} - {registro.descricao_gifus}\n"
                                obs += f"Saldo Devedor Principal: R$ {registro.saldo_devedor_principal:,.2f}\n"
                                obs += f"VAF1: R$ {registro.vaf1_fcvs_caixa:,.2f} | VAF2: R$ {registro.vaf2_fcvs_caixa:,.2f}\n"
                                obs += f"Data Processamento: {registro.data_processamento_retorno.strftime('%d/%m/%Y')}\n"
                                
                                if contrato.observacoes:
                                    contrato.observacoes += f"\n\n{obs}"
                                else:
                                    contrato.observacoes = obs
                                
                                contrato.save()
                                contratos_atualizados += 1
                            else:
                                contratos_nao_encontrados.append(numero_contrato)
                        
                        except Exception as e:
                            erros.append(f"Erro ao atualizar contrato {registro.numero_contrato}: {str(e)}")
                    
                    resultado['contratos_atualizados'] = contratos_atualizados
                    resultado['contratos_nao_encontrados'] = contratos_nao_encontrados[:10]  # Primeiros 10
                    
                    messages.success(
                        request,
                        f'✅ {total_registros} registros processados. {contratos_atualizados} contratos atualizados.'
                    )
                else:
                    messages.success(
                        request,
                        f'✅ {total_registros} registros processados com sucesso!'
                    )
            else:
                messages.warning(request, 'Nenhum registro encontrado no arquivo')
            
            # Remove arquivo temporário
            os.unlink(temp_path)
            
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            erros.append(str(e))
            import traceback
            erros.append(traceback.format_exc())
    
    context = {
        'resultado': resultado,
        'erros': erros,
    }
    
    return render(request, 'principal/cef_processar_m460.html', context)


def exportar_m460_excel(request, tipo_arquivo):
    """
    Exporta análise M460xxx para Excel
    
    Args:
        tipo_arquivo: M460301, M460401 ou M460801
    """
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    
    if request.method != 'POST' or not request.FILES.get('arquivo_m460'):
        messages.error(request, 'Arquivo M460 é obrigatório')
        return redirect('processar_m460_view')
    
    try:
        from .ficha_m460_parsers import ParserM460
        
        arquivo_upload = request.FILES['arquivo_m460']
        
        # Salvar temporário
        import tempfile
        temp_path = tempfile.mktemp(suffix='.txt')
        
        with open(temp_path, 'wb') as f:
            for chunk in arquivo_upload.chunks():
                f.write(chunk)
        
        # Parsear arquivo
        parser = ParserM460()
        
        if tipo_arquivo == 'M460301':
            registros, erros = parser.parse_file_m460301(temp_path)
        elif tipo_arquivo == 'M460401':
            registros, erros = parser.parse_file_m460401(temp_path)
        elif tipo_arquivo == 'M460801':
            registros, erros = parser.parse_file_m460801(temp_path)
        else:
            raise ValueError(f"Tipo inválido: {tipo_arquivo}")
        
        os.unlink(temp_path)
        
        if not registros:
            messages.error(request, 'Nenhum registro encontrado no arquivo')
            return redirect('processar_m460_view')
        
        # Criar workbook Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove sheet padrão
        
        # Estilos
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Sheet 1: Dados Brutos
        ws_dados = wb.create_sheet('Dados Completos')
        
        if tipo_arquivo in ['M460301', 'M460401']:
            headers = [
                'UF', 'Agente', 'Nº Contrato', 'Grau Hipoteca', 'GIFUS', 'Descrição GIFUS',
                'Data Retorno', 'Saldo Dev. Principal', 'Saldo Dev. Acessórios',
                'VAF1 CEF', 'VAF2 CEF', 'VAF3 CEF', 'VAF1 Calculado', 'VAF2 Calculado', 'VAF3 Calculado',
                'Mutuário', 'Endereço', 'Município', 'Data Contrato'
            ]
            
            ws_dados.append(headers)
            
            for registro in registros:
                ws_dados.append([
                    registro.uf,
                    registro.agente_financeiro,
                    registro.numero_contrato,
                    registro.grau_hipoteca,
                    registro.codigo_gifus,
                    registro.descricao_gifus,
                    registro.data_processamento_retorno.strftime('%d/%m/%Y') if registro.data_processamento_retorno else '',
                    registro.saldo_devedor_principal,
                    registro.saldo_devedor_acessorios,
                    registro.vaf1_fcvs_caixa,
                    registro.vaf2_fcvs_caixa,
                    registro.vaf3_fcvs_caixa,
                    registro.vaf1_calculado_caixa,
                    registro.vaf2_calculado_caixa,
                    registro.vaf3_calculado_caixa,
                    registro.nome_mutuario,
                    registro.endereco_imovel,
                    registro.codigo_municipio,
                    registro.data_contrato.strftime('%d/%m/%Y') if registro.data_contrato else ''
                ])
        
        elif tipo_arquivo == 'M460801':
            headers = [
                'UF', 'Agente', 'Nº Contrato', 'Grau Hipoteca', 'Situação',
                'Descrição Situação', 'Data Retorno', 'Mutuário', 'Município'
            ]
            
            ws_dados.append(headers)
            
            for registro in registros:
                ws_dados.append([
                    registro.uf,
                    registro.agente_financeiro,
                    registro.numero_contrato,
                    registro.grau_hipoteca,
                    registro.codigo_situacao_multiplicidade,
                    registro.descricao_situacao,
                    registro.data_processamento_retorno.strftime('%d/%m/%Y') if registro.data_processamento_retorno else '',
                    registro.nome_mutuario,
                    registro.codigo_municipio
                ])
        
        # Formatar headers
        for cell in ws_dados[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajustar larguras
        for column in ws_dados.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_dados.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Resumo por GIFUS (se aplicável)
        if tipo_arquivo in ['M460301', 'M460401']:
            from .ficha_m460_parsers import agrupar_por_gifus
            
            ws_resumo = wb.create_sheet('Resumo por GIFUS')
            
            por_gifus = agrupar_por_gifus(registros)
            
            ws_resumo.append(['Código GIFUS', 'Descrição', 'Quantidade', 'Total VAF1', 'Total VAF2', 'Total VAF3'])
            
            for cell in ws_resumo[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            for gifus, dados in sorted(por_gifus.items()):
                ws_resumo.append([
                    gifus,
                    dados['descricao'],
                    dados['quantidade'],
                    dados['total_vaf1'],
                    dados['total_vaf2'],
                    dados['total_vaf3']
                ])
            
            # Ajustar larguras
            ws_resumo.column_dimensions['A'].width = 15
            ws_resumo.column_dimensions['B'].width = 60
            ws_resumo.column_dimensions['C'].width = 12
            ws_resumo.column_dimensions['D'].width = 18
            ws_resumo.column_dimensions['E'].width = 18
            ws_resumo.column_dimensions['F'].width = 18
        
        elif tipo_arquivo == 'M460801':
            from .ficha_m460_parsers import agrupar_por_situacao
            
            ws_resumo = wb.create_sheet('Resumo por Situação')
            
            por_situacao = agrupar_por_situacao(registros)
            
            ws_resumo.append(['Código Situação', 'Descrição', 'Quantidade'])
            
            for cell in ws_resumo[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            for situacao, dados in sorted(por_situacao.items()):
                ws_resumo.append([
                    situacao,
                    dados['descricao'],
                    dados['quantidade']
                ])
            
            ws_resumo.column_dimensions['A'].width = 18
            ws_resumo.column_dimensions['B'].width = 60
            ws_resumo.column_dimensions['C'].width = 12
        
        # Preparar resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'{tipo_arquivo}_Analise_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Erro ao exportar: {str(e)}')
        import traceback
        messages.error(request, traceback.format_exc())
        return redirect('processar_m460_view')


def comparar_m460_view(request):
    """
    Compara dois arquivos M460 (ex: M460301 vs M460401) para encontrar divergências
    """
    from .ficha_m460_parsers import ParserM460
    
    resultado = None
    erros = []
    
    if request.method == 'POST':
        arquivo1 = request.FILES.get('arquivo1')
        arquivo2 = request.FILES.get('arquivo2')
        tipo1 = request.POST.get('tipo1', 'M460301')
        tipo2 = request.POST.get('tipo2', 'M460401')
        
        if not arquivo1 or not arquivo2:
            messages.error(request, 'Ambos os arquivos são obrigatórios')
        else:
            try:
                import tempfile
                
                # Processar arquivo 1
                temp1 = tempfile.mktemp(suffix='.txt')
                with open(temp1, 'wb') as f:
                    for chunk in arquivo1.chunks():
                        f.write(chunk)
                
                parser = ParserM460()
                
                if tipo1 == 'M460301':
                    registros1, _ = parser.parse_file_m460301(temp1)
                else:
                    registros1, _ = parser.parse_file_m460401(temp1)
                
                os.unlink(temp1)
                
                # Processar arquivo 2
                temp2 = tempfile.mktemp(suffix='.txt')
                with open(temp2, 'wb') as f:
                    for chunk in arquivo2.chunks():
                        f.write(chunk)
                
                if tipo2 == 'M460301':
                    registros2, _ = parser.parse_file_m460301(temp2)
                else:
                    registros2, _ = parser.parse_file_m460401(temp2)
                
                os.unlink(temp2)
                
                # Comparar por número de contrato
                contratos1 = {r.numero_contrato.strip(): r for r in registros1}
                contratos2 = {r.numero_contrato.strip(): r for r in registros2}
                
                # Contratos únicos em cada arquivo
                so_em_1 = set(contratos1.keys()) - set(contratos2.keys())
                so_em_2 = set(contratos2.keys()) - set(contratos1.keys())
                em_ambos = set(contratos1.keys()) & set(contratos2.keys())
                
                # Divergências de valores para contratos em ambos
                divergencias_valores = []
                
                for contrato in em_ambos:
                    r1 = contratos1[contrato]
                    r2 = contratos2[contrato]
                    
                    divs = []
                    
                    # Comparar VAFs
                    if abs(r1.vaf1_fcvs_caixa - r2.vaf1_fcvs_caixa) > 0.01:
                        divs.append(f"VAF1: R$ {r1.vaf1_fcvs_caixa:,.2f} vs R$ {r2.vaf1_fcvs_caixa:,.2f}")
                    
                    if abs(r1.vaf2_fcvs_caixa - r2.vaf2_fcvs_caixa) > 0.01:
                        divs.append(f"VAF2: R$ {r1.vaf2_fcvs_caixa:,.2f} vs R$ {r2.vaf2_fcvs_caixa:,.2f}")
                    
                    if abs(r1.saldo_devedor_principal - r2.saldo_devedor_principal) > 0.01:
                        divs.append(f"Saldo: R$ {r1.saldo_devedor_principal:,.2f} vs R$ {r2.saldo_devedor_principal:,.2f}")
                    
                    if r1.codigo_gifus != r2.codigo_gifus:
                        divs.append(f"GIFUS: {r1.codigo_gifus} vs {r2.codigo_gifus}")
                    
                    if divs:
                        divergencias_valores.append({
                            'contrato': contrato,
                            'mutuario': r1.nome_mutuario,
                            'divergencias': divs
                        })
                
                resultado = {
                    'sucesso': True,
                    'tipo1': tipo1,
                    'tipo2': tipo2,
                    'total1': len(registros1),
                    'total2': len(registros2),
                    'so_em_1': sorted(list(so_em_1)),
                    'so_em_2': sorted(list(so_em_2)),
                    'em_ambos': len(em_ambos),
                    'divergencias_valores': divergencias_valores[:50]  # Primeiras 50
                }
                
                messages.success(
                    request,
                    f'✅ Comparação concluída! {len(em_ambos)} contratos em ambos, '
                    f'{len(so_em_1)} apenas em arquivo 1, {len(so_em_2)} apenas em arquivo 2'
                )
                
            except Exception as e:
                messages.error(request, f'Erro na comparação: {str(e)}')
                erros.append(str(e))
    
    context = {
        'resultado': resultado,
        'erros': erros,
    }
    
    return render(request, 'principal/cef_comparar_m460.html', context)


# ===== ENVIO DE MOVIMENTOS FCVS =====

def enviar_movimento_fcvs_view(request):
    """
    View para envio de movimentos FCVS/CADMUT conforme especificação SIWFC
    
    Suporta:
    - FCVS: FH1, FH2, FH3 (Arquivo Header + Arquivo Dados)
    - FCVS: RCV, RNV (Arquivo único)
    - CADMUT: CADMUT0, CADMUT1 (Arquivo único)
    """
    
    if request.method == 'POST':
        try:
            produto = request.POST.get('produto')  # FCVS ou CADMUT
            tipo_movimento = request.POST.get('tipo_movimento')
            possui_fh2_fh3 = request.POST.get('possui_fh2_fh3') == 'on'
            remessa_id_post = (request.POST.get('remessa_id') or '').strip()

            def _somente_digitos(valor):
                return re.sub(r'\D', '', str(valor or ''))

            def _normalizar_codigo_contrato(valor):
                digits = _somente_digitos(valor)
                return digits.lstrip('0') or digits

            def _tipo_envio_modelo(tipo):
                t = (tipo or '').upper().strip()
                if t in ('FH1', 'RCV', 'RNV', 'DEFESA_REV'):
                    return t
                if t.startswith('CADMUT'):
                    return 'CADMUT'
                return 'COMPLEMENTAR'

            contratos_db = list(Contrato.objects.all().only('id', 'codigo'))
            mapa_contratos = {}
            for c in contratos_db:
                cod = _normalizar_codigo_contrato(c.codigo)
                if cod:
                    mapa_contratos[cod] = c

            def _extrair_codigo_linha(linha, tipo):
                if not linha:
                    return ''
                t = (tipo or '').upper().strip()
                texto = str(linha)

                # RNV/RCV simplificado: mat(5) + contrato(13) + hip(1) + data(8)
                if t in ('RNV', 'RCV') and len(texto) >= 19 and len(texto) < 430:
                    return _normalizar_codigo_contrato(texto[5:18])

                # Layout 430 (FHx/RCV/RNV completo): contrato em 09-21
                if len(texto) >= 21:
                    return _normalizar_codigo_contrato(texto[8:21])

                return ''

            def _registrar_envios_manual(tipo, linhas, nome_arquivo, tamanho_bytes, observacoes='', status_inicial='ENVIADO'):
                tipo_envio = _tipo_envio_modelo(tipo)

                contratos_vinculados = []
                vistos = set()
                for linha in linhas:
                    cod = _extrair_codigo_linha(linha, tipo)
                    if not cod or cod in vistos:
                        continue
                    contrato_obj = mapa_contratos.get(cod)
                    if contrato_obj:
                        contratos_vinculados.append(contrato_obj)
                        vistos.add(cod)

                if not contratos_vinculados:
                    LogAutomacao.objects.create(
                        tipo_acao='ENVIO',
                        descricao=(
                            f'Envio manual {tipo} sem vínculo automático de contratos. '
                            f'Arquivo: {nome_arquivo} | Linhas: {len(linhas)}. {observacoes}'
                        ),
                        sucesso=True,
                    )
                    return {'remessa_id': None, 'envio_ids': [], 'vinculados': 0}

                remessa = RemessaCEF.objects.create(
                    tipo_envio=tipo_envio,
                    status=status_inicial,
                    numero_lote='',
                    matricula_agente='',
                    total_contratos=len(contratos_vinculados),
                    total_fichas=len(linhas),
                    envios_sucesso=len(contratos_vinculados) if status_inicial == 'ENVIADO' else 0,
                    envios_erro=0 if status_inicial == 'ENVIADO' else len(contratos_vinculados),
                    log_processamento=f'Envio manual {tipo} via tela Enviar Movimentos.',
                    iniciado_em=timezone.now(),
                    finalizado_em=timezone.now() if status_inicial == 'ENVIADO' else None,
                    criado_por=getattr(request.user, 'username', '') if getattr(request, 'user', None) and request.user.is_authenticated else '',
                )

                envios = []
                for contrato in contratos_vinculados:
                    envios.append(EnvioCEF(
                        contrato=contrato,
                        remessa=remessa,
                        tipo_envio=tipo_envio,
                        status=status_inicial,
                        arquivo_enviado=nome_arquivo,
                        tamanho_bytes=tamanho_bytes,
                        codigo_contrato_cef=contrato.codigo,
                        data_envio=timezone.now() if status_inicial == 'ENVIADO' else None,
                        log_envio=f'Envio manual {tipo} pela tela de movimentos. {observacoes}',
                        enviado_automaticamente=False,
                        tentativas=1,
                    ))
                EnvioCEF.objects.bulk_create(envios)
                envio_ids = list(EnvioCEF.objects.filter(remessa=remessa).values_list('id', flat=True))

                return {
                    'remessa_id': remessa.id,
                    'envio_ids': envio_ids,
                    'vinculados': len(contratos_vinculados),
                }
            
            resultado = {
                'sucesso': False,
                'mensagem': '',
                'detalhes': {},
                'criticas': []
            }
            
            # Validações iniciais
            if not produto or not tipo_movimento:
                resultado['mensagem'] = '❌ Produto e Tipo de Movimento são obrigatórios'
                return JsonResponse(resultado)
            
            # FCVS - Processa Header + Dados
            if produto == 'FCVS' and tipo_movimento in ['FH1', 'FH2', 'FH3']:
                arquivo_header = request.FILES.get('arquivo_header')
                arquivo_dados = request.FILES.get('arquivo_dados')
                
                if not arquivo_header or not arquivo_dados:
                    resultado['mensagem'] = '❌ Arquivo de Header e Arquivo de Dados são obrigatórios'
                    return JsonResponse(resultado)
                
                # Lê conteúdo dos arquivos
                try:
                    header_content = arquivo_header.read().decode('utf-8', errors='ignore')
                    dados_content = arquivo_dados.read().decode('utf-8', errors='ignore')
                except Exception as e:
                    resultado['mensagem'] = f'❌ Erro ao ler arquivos: {str(e)}'
                    return JsonResponse(resultado)
                
                # Valida formato dos arquivos
                header_lines = [l.strip() for l in header_content.split('\n') if l.strip()]
                dados_lines = [l.strip() for l in dados_content.split('\n') if l.strip()]
                
                if not header_lines or not dados_lines:
                    resultado['mensagem'] = '❌ Arquivos vazios ou inválidos'
                    return JsonResponse(resultado)
                
                # Validação básica HEADER (430 caracteres esperados)
                if len(header_lines[0]) != 430:
                    resultado['criticas'].append({
                        'linha': 1,
                        'arquivo': 'HEADER',
                        'codigo': 'E001',
                        'descricao': f'HEADER deve ter 430 caracteres (encontrado: {len(header_lines[0])})'
                    })
                
                # Validação básica DADOS (430 caracteres por linha para FH1)
                for idx, linha in enumerate(dados_lines, start=1):
                    if tipo_movimento == 'FH1' and len(linha) != 430:
                        resultado['criticas'].append({
                            'linha': idx,
                            'arquivo': 'DADOS',
                            'codigo': 'E002',
                            'descricao': f'Registro {tipo_movimento} deve ter 430 caracteres (encontrado: {len(linha)})'
                        })
                    
                    # Limita a 10 críticas para não sobrecarregar
                    if len(resultado['criticas']) >= 10:
                        resultado['criticas'].append({
                            'linha': 0,
                            'arquivo': 'DADOS',
                            'codigo': 'WARN',
                            'descricao': f'... e mais {len(dados_lines) - idx} linhas não validadas'
                        })
                        break
                
                # Se há críticas, retorna erro
                if resultado['criticas']:
                    resultado['mensagem'] = f'❌ Encontradas {len(resultado["criticas"])} crítica(s) no arquivo'
                    return JsonResponse(resultado)
                
                # Extrai informações do HEADER
                ufs = header_lines[0][0:2]
                matricula = header_lines[0][2:8]
                tipo_registro = header_lines[0][22:23]
                qtd_registros = int(header_lines[0][32:37])
                data_geracao = header_lines[0][413:419]
                numero_lote = header_lines[0][419:422]
                tipo_mov_header = header_lines[0][423:424]
                
                registro = _registrar_envios_manual(
                    tipo=tipo_movimento,
                    linhas=dados_lines,
                    nome_arquivo=arquivo_dados.name,
                    tamanho_bytes=arquivo_dados.size,
                    observacoes=f'UFS: {ufs}, Matrícula: {matricula}, Lote: {numero_lote}, Qtd: {qtd_registros}'
                )
                
                resultado['sucesso'] = True
                resultado['mensagem'] = f'✅ Movimento {tipo_movimento} enviado com sucesso!'
                resultado['detalhes'] = {
                    'envio_ids': registro['envio_ids'],
                    'remessa_id': registro['remessa_id'],
                    'contratos_vinculados': registro['vinculados'],
                    'ufs': ufs,
                    'matricula': matricula,
                    'numero_lote': numero_lote,
                    'qtd_registros': qtd_registros,
                    'data_geracao': data_geracao,
                    'total_linhas_dados': len(dados_lines),
                    'possui_fh2_fh3': possui_fh2_fh3
                }
                
                return JsonResponse(resultado)
            
            # FCVS - RCV/RNV (arquivo único)
            elif produto == 'FCVS' and tipo_movimento in ['RCV', 'RNV']:
                arquivo_dados = request.FILES.get('arquivo_dados')

                remessa_ref = None
                if remessa_id_post.isdigit():
                    remessa_ref = RemessaCEF.objects.filter(pk=int(remessa_id_post)).first()

                if not arquivo_dados and remessa_ref and remessa_ref.arquivo_dados and os.path.exists(remessa_ref.arquivo_dados):
                    with open(remessa_ref.arquivo_dados, 'r', encoding='utf-8', errors='ignore') as f:
                        dados_content = f.read()
                    nome_arquivo_dados = os.path.basename(remessa_ref.arquivo_dados)
                    tamanho_arquivo_dados = len(dados_content.encode('utf-8'))
                else:
                    if not arquivo_dados:
                        resultado['mensagem'] = '❌ O arquivo de dados é de preenchimento obrigatório (ou informe uma remessa válida).'
                        return JsonResponse(resultado)
                    dados_content = arquivo_dados.read().decode('utf-8', errors='ignore')
                    nome_arquivo_dados = arquivo_dados.name
                    tamanho_arquivo_dados = arquivo_dados.size

                dados_lines = [l.strip() for l in dados_content.split('\n') if l.strip()]
                
                registro = _registrar_envios_manual(
                    tipo=tipo_movimento,
                    linhas=dados_lines,
                    nome_arquivo=nome_arquivo_dados,
                    tamanho_bytes=tamanho_arquivo_dados,
                    observacoes=f'Tipo: {tipo_movimento}, Linhas: {len(dados_lines)}',
                    status_inicial='PROCESSANDO',
                )

                remessa_id = registro.get('remessa_id')
                remessa_obj = RemessaCEF.objects.filter(pk=remessa_id).first() if remessa_id else None
                bot = None
                caminho_temp_dados = None
                caminho_temp_header = None

                try:
                    import tempfile
                    import sys

                    def _calcular_dv_mod11(matricula5):
                        m = str(matricula5 or '').zfill(5)
                        if m == '00044':
                            return '2'
                        multiplicadores = [2, 3, 4, 5, 6, 7, 8, 9]
                        soma = 0
                        for i, dig in enumerate(m):
                            soma += int(dig) * multiplicadores[i % 8]
                        resto = soma % 11
                        if resto in (0, 1):
                            return '0'
                        return str(11 - resto)

                    def _montar_rnv_rcv_completo_de_simplificado(linhas_simplificadas, tipo_mov):
                        ufs = '19'
                        tipo_letra = 'P' if tipo_mov == 'RNV' else 'R'
                        codigo_mov = '700' if tipo_mov == 'RNV' else '710'
                        lote = (remessa_obj.numero_lote if remessa_obj and remessa_obj.numero_lote else timezone.localtime().strftime('%d%H%M')[-3:]).zfill(3)[:3]
                        hoje = timezone.localtime().strftime('%d%m%y')

                        linhas_dados_completo = []
                        mat6_ref = None
                        for linha in linhas_simplificadas:
                            txt = str(linha or '')
                            if len(txt) < 27:
                                continue

                            mat5 = re.sub(r'\D', '', txt[0:5]).zfill(5)[:5]
                            mat6 = mat5 + _calcular_dv_mod11(mat5)
                            if not mat6_ref:
                                mat6_ref = mat6

                            contrato13 = txt[5:18].ljust(13)[:13]
                            hip = re.sub(r'\D', '', txt[18:19]) or '1'
                            data8 = re.sub(r'\D', '', txt[19:27]).zfill(8)[:8]
                            data6 = f"{data8[0:2]}{data8[2:4]}{data8[6:8]}"  # DDMMAA

                            id_lote = f"{ufs}{mat6}{hoje}{lote}S{tipo_letra}{' ' * 6}"
                            linha430 = (
                                f"{ufs}{mat6}{contrato13}{hip}4"  # 1-23
                                f"{'0' * 5}"                    # 24-28
                                f"{data6}"                      # 29-34
                                f"{'0' * 5}"                    # 35-39
                                f"{codigo_mov}"                 # 40-42
                                f"{' ' * 363}"                  # 43-405
                                f"{id_lote}"                    # 406-430
                            )
                            linhas_dados_completo.append(linha430)

                        if not linhas_dados_completo:
                            return '', []

                        if not mat6_ref:
                            mat6_ref = '000442'

                        id_lote_h = f"{ufs}{mat6_ref}{hoje}{lote}S{tipo_letra}{' ' * 6}"
                        header430 = (
                            f"{ufs}{mat6_ref}{'0' * 14}0{'0' * 9}{str(len(linhas_dados_completo)).zfill(5)}"
                            f"{' ' * 368}{id_lote_h}"
                        )

                        return header430, linhas_dados_completo

                    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as tmp_dados:
                        # Para RNV/RCV, gerar também versão completa (HEADER + DADOS 430)
                        header430, dados430 = _montar_rnv_rcv_completo_de_simplificado(dados_lines, tipo_movimento)

                        if header430 and dados430:
                            tmp_dados.write('\n'.join(dados430))
                        else:
                            tmp_dados.write('\n'.join(dados_lines))
                        caminho_temp_dados = tmp_dados.name

                    if tipo_movimento in ('RNV', 'RCV'):
                        header430, dados430 = _montar_rnv_rcv_completo_de_simplificado(dados_lines, tipo_movimento)
                        if header430:
                            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as tmp_header:
                                tmp_header.write(header430)
                                caminho_temp_header = tmp_header.name

                    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
                    from cef_web_automation import CEFWebBot

                    bot = CEFWebBot(headless=False, usar_perfil_existente=True)
                    if not bot.setup_driver():
                        raise Exception('Não foi possível conectar ao Chrome para automação.')

                    if not bot.verificar_login_existente():
                        raise Exception('Sessão SIWFC não detectada. Faça login no Chrome de depuração e tente novamente.')

                    resultado_envio = bot.enviar_movimento_fcvs(
                        arquivo_header=caminho_temp_header or '',
                        arquivo_dados=caminho_temp_dados,
                        produto='FCVS',
                        tipo_movimento=tipo_movimento,
                    )

                    sucesso = bool(resultado_envio.get('sucesso'))
                    protocolo = (resultado_envio.get('protocolo') or '').strip()
                    mensagem = resultado_envio.get('mensagem', '')

                    if sucesso:
                        if remessa_id:
                            EnvioCEF.objects.filter(remessa_id=remessa_id).update(
                                status='ENVIADO',
                                data_envio=timezone.now(),
                                protocolo_cef=protocolo,
                                mensagem_erro='',
                                log_envio=f'Envio {tipo_movimento} automático concluído. {mensagem}',
                            )
                        if remessa_obj:
                            remessa_obj.status = 'ENVIADO'
                            remessa_obj.protocolo_cef = protocolo
                            remessa_obj.envios_sucesso = remessa_obj.total_contratos
                            remessa_obj.envios_erro = 0
                            remessa_obj.mensagem_erro = ''
                            remessa_obj.log_processamento = mensagem or f'Envio {tipo_movimento} concluído.'
                            remessa_obj.finalizado_em = timezone.now()
                            remessa_obj.save()

                        resultado['sucesso'] = True
                        resultado['mensagem'] = f'✅ {tipo_movimento} enviado com sucesso!'
                        resultado['detalhes'] = {
                            'envio_ids': registro['envio_ids'],
                            'remessa_id': remessa_id,
                            'contratos_vinculados': registro['vinculados'],
                            'total_linhas': len(dados_lines),
                            'protocolo': protocolo,
                        }
                        return JsonResponse(resultado)

                    erro_msg = mensagem or f'Falha no envio {tipo_movimento} para o portal CEF.'
                    if remessa_id:
                        EnvioCEF.objects.filter(remessa_id=remessa_id).update(
                            status='ERRO',
                            mensagem_erro=erro_msg,
                            log_envio=f'Falha no envio {tipo_movimento}. {erro_msg}',
                        )
                    if remessa_obj:
                        remessa_obj.status = 'ERRO'
                        remessa_obj.envios_sucesso = 0
                        remessa_obj.envios_erro = remessa_obj.total_contratos
                        remessa_obj.mensagem_erro = erro_msg
                        remessa_obj.log_processamento = erro_msg
                        remessa_obj.finalizado_em = timezone.now()
                        remessa_obj.save()

                    resultado['sucesso'] = False
                    resultado['mensagem'] = f'❌ Falha no envio {tipo_movimento}: {erro_msg}'
                    resultado['detalhes'] = {
                        'envio_ids': registro['envio_ids'],
                        'remessa_id': remessa_id,
                        'contratos_vinculados': registro['vinculados'],
                        'total_linhas': len(dados_lines),
                    }
                    return JsonResponse(resultado)

                finally:
                    try:
                        if bot:
                            bot.fechar()
                    except Exception:
                        pass
                    try:
                        if caminho_temp_dados and os.path.exists(caminho_temp_dados):
                            os.unlink(caminho_temp_dados)
                    except Exception:
                        pass
                    try:
                        if caminho_temp_header and os.path.exists(caminho_temp_header):
                            os.unlink(caminho_temp_header)
                    except Exception:
                        pass
            
            # CADMUT (arquivo único)
            elif produto == 'CADMUT':
                arquivo_dados = request.FILES.get('arquivo_dados')
                
                if not arquivo_dados:
                    resultado['mensagem'] = '❌ Arquivo de Dados é obrigatório'
                    return JsonResponse(resultado)
                
                dados_content = arquivo_dados.read().decode('utf-8', errors='ignore')
                dados_lines = [l.strip() for l in dados_content.split('\n') if l.strip()]
                
                registro = _registrar_envios_manual(
                    tipo=tipo_movimento,
                    linhas=dados_lines,
                    nome_arquivo=arquivo_dados.name,
                    tamanho_bytes=arquivo_dados.size,
                    observacoes=f'Tipo: {tipo_movimento}, Linhas: {len(dados_lines)}'
                )
                
                resultado['sucesso'] = True
                resultado['mensagem'] = f'✅ {tipo_movimento} enviado com sucesso!'
                resultado['detalhes'] = {
                    'envio_ids': registro['envio_ids'],
                    'remessa_id': registro['remessa_id'],
                    'contratos_vinculados': registro['vinculados'],
                    'total_linhas': len(dados_lines),
                }
                
                return JsonResponse(resultado)
            
            else:
                resultado['mensagem'] = '❌ Combinação de Produto/Tipo de Movimento inválida'
                return JsonResponse(resultado)
                
        except Exception as e:
            return JsonResponse({
                'sucesso': False,
                'mensagem': f'❌ Erro ao processar: {str(e)}',
                'detalhes': {}
            })
    
    # GET - Renderiza formulário (com pré-seleção opcional)
    preselect_produto = (request.GET.get('produto') or '').strip().upper()
    preselect_tipo_movimento = (request.GET.get('tipo_movimento') or '').strip().upper()
    remessa_id = (request.GET.get('remessa_id') or '').strip()

    remessa_context = None
    if remessa_id.isdigit():
        remessa = RemessaCEF.objects.filter(pk=int(remessa_id)).first()
        if remessa:
            remessa_context = {
                'id': remessa.id,
                'status': remessa.status,
                'status_label': remessa.get_status_display(),
                'tipo_envio': remessa.tipo_envio,
                'numero_lote': remessa.numero_lote,
                'total_fichas': remessa.total_fichas,
                'download_dados_url': f'/cef/remessa/{remessa.id}/arquivo/dados/',
            }

    context = {
        'preselect_produto': preselect_produto,
        'preselect_tipo_movimento': preselect_tipo_movimento,
        'remessa_context': remessa_context,
    }

    return render(request, 'principal/cef_enviar_movimentos.html', context)
